"""
STN-A设备巡检系统 v2.8
使用前需手动安装模块：pip install openpyxl pytz paramiko tqdm colorama pyinstaller
更新说明：
- 修复若干BUG
        
作者：杨茂森

最后更新：2025-6-29
"""
# 导入必要的库
from openpyxl.styles import PatternFill, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, wait
import queue
import select
import paramiko
import openpyxl
from colorama import Fore, Style
from datetime import timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment
import pytz
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Font, Alignment
from threading import Lock
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import csv
import time
import re
from datetime import datetime
import socket
import random
from contextlib import ContextDecorator
from itertools import count
from operator import truediv
from tkinter import SW
from warnings import catch_warnings
import paramiko        # SSH连接库
from tqdm import tqdm
from colorama import Fore, Style, init
import threading
from collections import defaultdict
import json
import os
import logging
from datetime import datetime
import sys
from itertools import cycle
import pytz  # 需要导入 pytz 来处理时区
from openpyxl.cell.cell import MergedCell
import shutil
import openpyxl


workbook = openpyxl.Workbook()  # Creates a new workbook

# 初始化 colorama
init(autoreset=True)
# 初始化打印锁
print_lock = Lock()


def input_with_timeout(prompt, default, timeout=10):
    print(f"{Fore.CYAN}{prompt}{Style.RESET_ALL}", end='')
    print(f"{Fore.GREEN}默认值：{default} | 超时：{timeout}s{Style.RESET_ALL}", flush=True)
    result = [default]

    def get_input():
        try:
            inp = input().strip()
            result[0] = inp if inp else default
        except:
            pass
    t = threading.Thread(target=get_input)
    t.daemon = True
    t.start()
    t.join(timeout)
    if t.is_alive():
        print(f"\n{Fore.YELLOW}⏱️ 输入超时，已使用默认值：{default}{Style.RESET_ALL}")
    return result[0]


def getinput(defval, inputval, timeout=None):
    if timeout is not None:
        return input_with_timeout(inputval, defval, timeout)
    else:
        userinput = input(inputval)
        return defval if userinput == '' else userinput


def string_to_number(s):
    s = s.strip()
    if not s or s == '---':  # 处理空字符串和无效占位符
        return 0.0
    try:
        return float(s)
    except ValueError:
        if s.startswith('-'):  # 处理负号开头的无效字符串（如"--"）
            return -abs(string_to_number(s.lstrip('-')))
        return 0.0


def keep_digits(s):

    return re.sub(r'\D', '', s)


def find_char(string, char):

    return string.find(char)


def delete_after(string, char):

    return string.split(char, 1)[0] if char in string else string


def splitdot(content, char, index):

    return content.split(char, 1)[index]


def splitstr(str):

    cleaned = str.replace('\r', '').split('\n')
    # 移除空行和特定关键词行
    return [line for line in cleaned if line and 'show' not in line and 'screen' not in line]


ERROR_MAPPING = {
    '10054': '设备安全策略阻断连接（建议检查ACL/连接频率限制）远程主机强迫关闭了一个现有的连接',
    'Error reading SSH protocol banner': '远程主机强迫关闭了一个现有的连接（设备脱管）',
    'timed out': '协议协商超时',
    'No existing session': 'SSH会话已过期（需重新认证）',
    'Authentication failed': '认证失败（检查用户名/密码）'
}


def pretty_error(e):
    for code, msg in ERROR_MAPPING.items():
        if code in str(e):
            return f"{Fore.RED}{msg}{Style.RESET_ALL}"
    return f"{Fore.RED}错误：{str(e)}{Style.RESET_ALL}"


# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='network_automation.log'
)


def client_close(client):
    """安全关闭SSH客户端"""
    if client:
        try:
            client.close()
        except:
            pass


def config_host(channel, filename, revfile, ipaddr='', device_name='', cmd_delay=1):
    # 禁用分页
    execute_some_command(channel, 'screen-length 0',
                         timeout=2, command_delay=1, device_name=device_name, ip=ipaddr)
    try:
        with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
            for i, cmd in enumerate(csv.reader(csvFile)):
                # 为每条命令之间添加延迟
                if i > 0:
                    print(f"[INFO] 等待 {cmd_delay} 秒后执行下一命令...")
                    time.sleep(cmd_delay)

                result = execute_some_command(
                    channel, cmd[0]+'\n', timeout=5, command_delay=1.5,
                    device_name=device_name, ip=ipaddr)

                for line in splitstr(result):
                    try:
                        revfile.write(f"{ipaddr} , {cmd[0]} , {line}\n")
                    except UnicodeEncodeError:
                        encoded = line.encode(
                            'gbk', errors='replace').decode('gbk')
                        revfile.write(f"{ipaddr} , {cmd[0]} , {encoded}\n")
                    except Exception as e:
                        print(f"执行命令 {cmd} 时出错: {e}")
                        continue
    finally:
        # 执行完命令后等待一段时间再恢复分页设置
        time.sleep(2)
        # 恢复分页设置
        execute_some_command(channel, 'screen-length 25',
                             timeout=2, command_delay=1, device_name=device_name, ip=ipaddr)


# 执行一跳CLI指令并保存结果


def send_cmd(channel, strCmd, revfile, ipaddr=''):
    sResult = execute_some_command(channel, strCmd + '\n')
    ret = splitstr(sResult)  # 结果除去杂项
    for item in ret:
        try:
            # 尝试使用utf-8编码写入文件
            revfile.writelines(ipaddr + " , " + strCmd + " , " +
                               item.replace('\ufffd', '?').replace('\r', '') + '\n')
        except UnicodeEncodeError:
            # 如果仍然失败，可以选择使用其他编码方式尝试编码，比如gbk
            revfile.writelines((ipaddr + " , " + strCmd + " , " + item.replace(
                '\ufffd', '?').replace('\r', '')).encode('gbk', errors='replace').decode('gbk') + '\n')


def wash_cpu(content):

    if find_char(content, 'CPU utilization') != -1:
        strtemp = content.replace('for five seconds:', ',')
        strtemp1 = strtemp.replace(': fifteen  minutes :', ',')
        return strtemp1
    return ''


def wash_device(content):
    if find_char(content, 'system info') != -1:
        strtemp = content.replace(':', ',')
        strtemp1 = strtemp.replace('(', ',')
        return strtemp1
    return ''


def wash_mem(content):
    if find_char(content, 'Memory Using Percentage :') != -1:
        strtemp = content.replace(':', ',')
        return strtemp
    return ''


def wash_laser(content):
    items = content.split(',')
    if len(items) >= 3 and 'laser' in items[1]:
        # 提取关键字段并确保格式正确
        cleaned = re.sub(r'\s{2,}', ',', items[2].split('[')[0].strip())
        return f"{items[0]},{items[1]},{cleaned}\n"
    return ''


def wash_ldp(content):
    items = content.split(',')
    strtemp = ""
    if len(items) >= 2:
        if find_char(items[1], 'ldp') != -1:
            if find_char(items[2], 'remote') != -1 or find_char(items[2], 'local') != -1:
                strtemp = re.sub(r"\s\s+", ",", content, count=20) + "\n"
                strtemp = re.sub(r"\n+", "\n", strtemp)
                return strtemp

    return ''


def wash_ospf(content):
    items = content.split(',')
    strtemp = ""
    # 确保至少有3个元素且第三个字段足够长
    if len(items) >= 3 and len(items[2]) >= 2:
        if find_char(items[1], 'ospf') != -1:
            # 检查第二个字符是否为数字
            if items[2][1].isdigit():
                strtemp = re.sub(r"50GE|25GE|xgigabitethernet", " ", content)
                strtemp = re.sub(r"\s\s+", ",", strtemp, count=20) + "\n"
                return strtemp
    return ''


def wash_vc(content):
    items = content.split(',')
    if len(items) >= 2 and 'l2vc' in items[1]:
        stmps = items[2].split()
        if stmps and stmps[0].isdigit():
            # 使用更严格的分割方式
            cleaned = re.sub(r'\s{2,}', ',', content.strip())  # 多个空格替换为逗号
            cleaned = cleaned.replace('  ', ',').replace(',,', ',')
            return cleaned + '\n'
    return ''


def wash_BFD(content):
    items = content.split(',')
    strtemp = ""
    if len(items) >= 2:
        if find_char(items[1], 'bfd session') != -1:
            stmps = items[2].split()
            if find_char(items[2], 'BFD For') != -1:
                strtemp = re.sub(r"\s\s+", ",", content, count=20) + "\n"
                strtemp = re.sub(r",,", ",", strtemp, count=20)
                return strtemp
    return ''


def wash_temperature(content):
    items = content.split(',')
    strtemp = ""
    if len(items) >= 2:
        if find_char(items[1], 'temperature') != -1:
            stemps = items[2].split()
            if len(stemps) == 4:
                if stemps[3][0].isdigit():
                    return items[0]+','+items[1]+','+stemps[3] + '\n'
    return ''

# 清洗 show interface main, 可以关联上下文


def wash_int_main(srcfile, retfile):
    sFile = open(srcfile, "r", encoding='gbk', errors='ignore')
    if sFile == None:
        return -1
    dFile = open(retfile, mode="a", newline='')
    reader = csv.reader(sFile)
    content = sFile.readline()
    content1 = ""
    strtmp = ""
    while content:
        items = content.split(',')
        if len(items) >= 2:
            if find_char(items[1], 'main') != -1:
                if find_char(items[2], 'current state') != -1:
                    strtmp = content.replace("\n", "")
                    content1 = strtmp.replace("current state :", ",")
                if find_char(items[2], 'CRC') != -1:
                    strtmp = items[2].replace(" ", "")
                    content1 = content1 + "," + \
                        strtmp.replace("packets", "")+"\n"
                    dFile.write(content1)
                    content1 = ""
        content = sFile.readline()
    sFile.close()
    dFile.close()

# 通用清洗方法结果文件清洗


def wash_result(srcfile, retfile, cmd):
    try:
        sFile = open(srcfile, "r", encoding='gbk', errors='ignore')
    except Exception as e:
        print(f"{Fore.RED}{srcfile} 文件不存在: {e}{Style.RESET_ALL}")
        exit()
    except PermissionError:
        print(f"{Fore.RED}⛔ 无法写入文件 {retfile}，请检查文件是否被其他程序占用{Style.RESET_ALL}")
        exit()

    if sFile is None:
        return -1
    # n = 0
    dFile = open(retfile, mode="a", newline='')
    reader = csv.reader(sFile)
    content = sFile.readline()
    content1 = ""
    strtemp = ""
    lines = 0
    while content:
        if cmd == 1:
            strtemp = wash_cpu(content)
        if cmd == 2:
            strtemp = wash_mem(content)
        if cmd == 3:
            strtemp = wash_device(content)
        if cmd == 4:
            strtemp = wash_laser(content)
        if cmd == 5:
            strtemp = wash_ospf(content)
        if cmd == 6:
            strtemp = wash_ldp(content)
        if cmd == 7:
            strtemp = wash_temperature(content)
        if cmd == 8:
            strtemp = wash_vc(content)
        if cmd == 9:
            strtemp = wash_BFD(content)
        if cmd == 10:
            strtemp = wash_l2vc_brief(content)

        if strtemp != '':
            dFile.write(strtemp)
            print('.', end="")
            lines = lines + 1
        content = sFile.readline()
    sFile.close()
    dFile.close()
    print(f"\n{Fore.GREEN}✅ 数据清洗完成！结果已保存至 {retfile}{Style.RESET_ALL}")


def fish_slot_cmd(filename, ret_name, max_workers=40):
    """槽位检查采集函数，支持多线程并行采集多台设备的槽位信息"""
    import os
    import csv
    import time
    import re
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    # 初始化输出文件
    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "w") as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_slot_device, ip, user, pwd, revFile, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 空闲槽位检查进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 采集异常: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ 槽位数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_slot_device(ip, user, pwd, revFile, fail_log):
    """处理单个设备的槽位信息采集"""
    import os
    import csv
    import time
    import re
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            return

        execute_some_command(channel, "screen-length 0", 1)
        output = execute_some_command(channel, "show install package", 3)
        with file_lock:  # 线程安全写入
            for line in splitstr(output):
                revFile.write(f"{ip} , show install package , {line}\n")

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def wash_slot_info(content_line, device_names):
    """解析槽位信息，支持设备名称提取"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names

    device_ip, cmd, output = parts

    # 捕获设备名称
    if cmd.strip().lower() == "show install package":
        name_match = re.search(
            r'<([^>]+)>\s*(?:show install package)?',  # 适配命令可能换行的情况
            output,
            re.IGNORECASE
        )
        if name_match:
            device_name = name_match.group(1).strip()
            device_names[device_ip] = device_name  # 更新字典中的设备名称

    # 解析槽位信息
    if cmd.strip().lower() == "show install package":
        slot_match = re.search(
            r'slot\s*(\d+)\s*:\s*([^,]+?)\s*,\s*real\s*:\s*(\S+)',
            output,
            re.IGNORECASE
        )
        if slot_match:
            slot_num = slot_match.group(1)
            slot_type = slot_match.group(2).strip()
            real_type = slot_match.group(3).strip()
            is_idle = (real_type.upper() == 'NULL') or (slot_type != real_type)
            current_name = device_names.get(device_ip, "Unknown")
            return (device_ip, current_name, slot_num, slot_type, real_type, is_idle), device_names

    return None, device_names


def generate_slot_report(src_file, dst_file, host_list_file):
    """生成带槽位状态的准确报告"""
    device_names = defaultdict(str)  # 存储IP到设备名称的映射
    slot_data = defaultdict(list)

    with open(src_file, 'r', encoding='utf-8') as f_in:
        for line in f_in:
            line = line.strip()
            data, device_names = wash_slot_info(line, device_names)
            if data:
                ip, dev_name, slot_num, st, rt, is_idle = data
                slot_data[ip].append((slot_num, st, rt, is_idle))

    # 读取所有设备IP
    with open(host_list_file, 'r', encoding='gbk') as f:
        all_devices = [row[0].strip() for row in csv.reader(f) if row]

    # 处理连接失败的IP
    failure_ips = []
    if os.path.exists("failure_ips.tmp"):
        with open("failure_ips.tmp", 'r') as f:
            failure_ips = [line.strip() for line in f]

    # 生成报告
    with open(dst_file, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['设备IP', '设备名称', '空闲槽位', '连接状态'])

        for ip in all_devices:
            if ip in failure_ips:
                writer.writerow([ip, "N/A", "-", "连接失败"])
                continue

            slots = slot_data.get(ip, [])
            idle_slots = [
                f"slot {s[0]} : {s[1]} | real: {s[2]}" for s in slots if s[3]]
            dev_name = device_names.get(ip, "Unknown")

            if idle_slots:
                for slot in idle_slots:
                    writer.writerow([ip, dev_name, slot, "成功"])
            elif slots:  # 有槽位数据但无空闲
                writer.writerow([ip, dev_name, "-", "槽位满"])
            else:  # 无槽位数据但连接成功
                writer.writerow([ip, dev_name, "-", "未检测"])

    # 清理临时文件
    if os.path.exists("failure_ips.tmp"):
        os.remove("failure_ips.tmp")

    print(f"{Fore.GREEN}✅ 报告已生成，共处理{len(all_devices)}台设备{Style.RESET_ALL}")


def report_result(srcfile, retfile, h_name):
    try:
        sFile = open(srcfile, "r", encoding='gbk', errors='ignore')
    except FileNotFoundError:
        print(f"文件 {srcfile} 不存在")
        return -1
    resultlist = []

    reader = csv.reader(sFile)
    for host in reader:
        if host:  # 跳过空行
            resultlist.append(host)
    sFile.close()

    try:
        dFile = open(retfile, mode="a", newline='', encoding='gbk')
    except Exception as e:
        print(f"无法打开输出文件 {retfile}: {e}")
        return -1

    try:
        hFile = open(h_name, "r", encoding='gbk', errors='ignore')
    except FileNotFoundError:
        print(f"文件 {h_name} 不存在")
        dFile.close()
        return -1

    if sFile is None or hFile is None:
        dFile.close()
        return -1

    tstr = 'host,CPU,内存,版本,温度,CRC,端口,收光,ospf,ldp,vc,bfd\n'
    dFile.write(tstr)

    reader = csv.reader(hFile)
    device_count = 0
    for host in reader:
        if not host or len(host) < 1 or not host[0].strip():
            continue
        device_count += 1
        strhost = host[0].strip()
        newList = [x for x in resultlist if x and x[0].strip() == strhost]

        strReport = ['-'] * 12
        strReport[0] = strhost
        crcCount = 0
        upCount = 0
        downCount = 0
        laserCount = 0
        proc31 = [0, 0]
        proc65534 = [0, 0]
        proc_ldp = ["", ""]
        vc_total = 0
        vc_down = 0
        bfd_total = 0
        bfd_down = 0
        ospf_all_full = 0

        for y in newList:
            if len(y) < 2:
                continue
            if find_char(y[1], 'temperature') != -1 and len(y) > 2:
                strReport[4] = y[2]
            elif find_char(y[1], 'device') != -1 and len(y) > 4:
                if find_char(y[2], 'CPU') != -1:
                    strReport[1] = f"{y[3]}:{y[4]}"
                elif find_char(y[2], 'Memory') != -1:
                    strReport[2] = y[3]
                elif find_char(y[2], 'info') != -1:
                    strReport[3] = f"{y[3]} {y[4]}"
            elif find_char(y[1], 'interface main') != -1:
                if len(y) > 4 and y[4].strip():
                    num_str = keep_digits(y[4])
                    if num_str:
                        try:
                            if int(num_str) > 0:
                                crcCount += 1
                        except ValueError:
                            print(f"⚠️ 无效的 CRC 数据: {y[4]} 在行 {y}")
                if len(y) > 3:
                    if y[3].strip() == 'UP':
                        upCount += 1
                    elif y[3].strip() == 'DOWN':
                        downCount += 1
            elif find_char(y[1], 'laser') != -1 and len(y) > 3:
                val = string_to_number(y[3].strip())
                if val is not None and val < -10 and val != -40:
                    laserCount += 1
            elif find_char(y[1], 'ospf') != -1 and len(y) > 8:
                if find_char(y[4], 'Full') != -1:
                    ospf_all_full += 1
                    ospffull = 1 if find_char(y[4], 'Full') != -1 else 0
                    proc = splitdot(y[7], '.', 1)
                    if proc == '31':
                        proc31[0] += 1
                        proc31[1] += ospffull
                    elif proc == '4094':
                        proc65534[0] += 1
                        proc65534[1] += ospffull
            elif find_char(y[1], 'ldp session') != -1 and len(y) > 6:
                if find_char(y[4], '.31') != -1 and find_char(y[2], 'remote') != -1 and find_char(y[6], 'OPER') != -1:
                    idx = 0 if proc_ldp[0] == '' else 1
                    proc_ldp[idx] = y[4]
            elif find_char(y[1], 'l2vc') != -1:
                vc_total += 1
                if len(y) > 5:
                    vc_down += 1 if y[5].strip().lower() == 'down' else 0
                else:
                    print(f"⚠️ VC数据异常：字段不足 {y}")
            elif find_char(y[1], 'bfd session') != -1:
                bfd_total += 1
                if len(y) > 5:
                    bfd_down += 1 if y[5].strip() == 'Down' else 0
                else:
                    print(f"⚠️ BFD数据异常：{y}")

        strReport[5] = f"{crcCount} port crc err"
        strReport[6] = f"{upCount}:up {downCount}:down"
        strReport[7] = f"{laserCount} Rx low"
        strReport[8] = f"31进程 总:{proc31[0]}-full:{proc31[1]} 65534进程 总:{proc65534[0]}-full:{proc65534[1]}"
        remote_peer = "远端operational:0" if not proc_ldp[0] and not proc_ldp[
            1] else "远端operational:1" if not proc_ldp[1] else "远端operational:2"
        ldploop = "LDP 成环" if proc_ldp[0] and proc_ldp[0] != proc_ldp[1] else "LDP 未成环"
        strReport[9] = f"{remote_peer} {ldploop}"
        strReport[10] = f"up:{vc_total-vc_down} down:{vc_down}"
        strReport[11] = f"up:{bfd_total-bfd_down} down:{bfd_down}"

        tstr = ','.join([field.strip() for field in strReport]) + '\n'
        dFile.write(tstr)

    hFile.close()
    dFile.close()
    print(f"\n📊 本次报告共统计 {device_count} 台设备")
    return 0


def dynamic_colored_divider(color_code=36, symbol='―', enable_timestamp=True):
    timestamp = time.strftime("%H:%M:%S") if enable_timestamp else ""
    line_length = 60 - len(timestamp) - 3
    line = f"[{timestamp}] " if enable_timestamp else ""
    line += symbol * line_length
    print(f"\033[{color_code}m{line}\033[0m")


def parse_interface_description(content_line, interface_map):
    """Parse show inter description output to map interfaces to full business names"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return interface_map

    device_ip, cmd, output = parts
    if cmd.strip().lower() != "show inter description":
        return interface_map

    # Skip header or irrelevant lines
    if "Interface" in output or "Physical" in output or "PHY:" in output:
        return interface_map

    # Parse interface and description
    fields = re.split(r'\s{2,}', output.strip())
    if len(fields) >= 4:
        interface = fields[0].strip()
        description = ' '.join(fields[3:]).strip().replace(
            '\n', ' ').replace('\r', '')
        # Clean up multiple spaces
        description = re.sub(r'\s+', ' ', description)
        interface_map[(device_ip, interface)] = description

    return interface_map


def wash_l2vc_brief(content_line, device_names, interface_map):
    """Enhanced business data parsing with full name mapping"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names

    device_ip, cmd, output = parts

    # Extract device name
    if cmd.strip().lower() == "show mpls l2vc brief":
        for line in output.split('\n'):
            name_match = re.search(r'<([^>]+)>', line.strip())
            if name_match:
                device_name = name_match.group(1).strip()
                if 'show' not in device_name.lower():
                    device_names[device_ip] = device_name
                    break

    # Process L2VC data
    if 'show mpls l2vc brief' not in cmd:
        return None, device_names

    if any(x in output for x in ['VC-ID', '----', 'Total LDP VC']):
        return None, device_names

    cleaned = re.sub(r'[\t\xa0]+', '  ', output)
    items = re.split(r'\s{2,}', cleaned.strip())
    if len(items) < 6:
        return None, device_names

    try:
        vcid = items[0].strip()
        destination = items[1].strip()
        service_name = items[2].strip()
        status = items[3].strip().lower()
        interface = items[4].strip()
        vc_type = items[5].strip()

        # Get full business name from interface description
        full_service_name = interface_map.get(
            (device_ip, interface), service_name)
        # Limit to 40 characters if necessary, remove extra spaces
        full_service_name = re.sub(r'\s+', ' ', full_service_name.strip())[:40]

        role = 'N/A'
        for item in items:
            if item.strip().lower() in ['master', 'backup']:
                role = item.strip()
                break

        current_name = device_names.get(device_ip, "Unknown")
        return f"{device_ip},{current_name},{vcid},{destination},{full_service_name},{status},{interface},{vc_type},{role}\n", device_names
    except IndexError as e:
        print(f"字段解析异常：{items}")
        return None, device_names


def fish_cmd(filename, ret_name, strCmd, para0=0, para1=0, max_workers=40):  # max_workers=线程数

    from concurrent.futures import ThreadPoolExecutor, as_completed
    """增强型业务统计函数，支持多线程并行采集多台设备的命令输出"""
    # 清除之前的失败记录
    if os.path.exists("failure_ips.tmp"):
        os.remove("failure_ips.tmp")

    with open(ret_name, mode="w", newline='', encoding='utf-8') as revFile, \
            open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
        try:
            reader = csv.reader(csvFile)
            hostip = list(reader)
            total_devices = len(hostip)

            # 使用线程池并行处理设备
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = []
                for row in hostip:
                    ip = row[0].strip()
                    user = row[1].strip()
                    pwd = row[2].strip()
                    futures.append(executor.submit(
                        process_device, ip, user, pwd, strCmd, para0, revFile))

                # 使用tqdm显示进度
                with tqdm(total=total_devices, desc="📡 采集业务数据", unit="台") as pbar:
                    for future in as_completed(futures):
                        try:
                            future.result()  # 获取结果，触发异常处理
                        except Exception as e:
                            print(f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                        pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 采集过程中发生意外错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ 业务数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_device(ip, user, pwd, strCmd, para0, revFile):
    import time
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    # 文件写入锁，确保线程安全
    file_lock = Lock()
    """处理单个设备的连接和命令执行"""
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        channel = None
        try:
            channel = create_channel(ip, user, pwd)
            if not channel:
                raise Exception(f"连接 {ip} 失败")

            # 设置屏幕长度以避免分页
            execute_some_command(channel, "screen-length 512", 1)

            # 采集 show mpls l2vc brief
            strCmd1 = strCmd + (para0 if para0 else "")
            sResult1 = execute_some_command(channel, strCmd1, 3)
            with file_lock:  # 线程安全写入
                for item in splitstr(sResult1):
                    revFile.write(f"{ip} , {strCmd1} , {item}\n")

            # 采集 show inter description
            sResult2 = execute_some_command(
                channel, "show inter description", 3)
            with file_lock:  # 线程安全写入
                for item in splitstr(sResult2):
                    revFile.write(f"{ip} , show inter description , {item}\n")

            # 重置屏幕长度
            execute_some_command(channel, "screen-length 25", 1)
            break  # 成功，退出重试循环

        except Exception as e:
            print(f"🔄 {ip} 尝试 {attempt}/{max_retries} 失败: {str(e)}")
            if attempt == max_retries:
                print(f"⛔ {ip} 已达最大重试次数，跳过该设备")
                with file_lock:  # 线程安全写入失败记录
                    with open("failure_ips.tmp", "a") as f:
                        f.write(f"{ip}\n")
            else:
                time.sleep(2)
        finally:
            if channel:
                channel.close()


def wash_l2vc_brief(content_line, device_names, interface_map):
    """增强型业务数据解析，支持完整名称映射"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names

    device_ip, cmd, output = parts

    # 提取设备名称
    if cmd.strip().lower() == "show mpls l2vc brief":
        for line in output.split('\n'):
            name_match = re.search(r'<([^>]+)>', line.strip())
            if name_match:
                device_name = name_match.group(1).strip()
                if 'show' not in device_name.lower():
                    device_names[device_ip] = device_name
                    break

    # 处理 L2VC 数据
    if 'show mpls l2vc brief' not in cmd:
        return None, device_names

    if any(x in output for x in ['VC-ID', '----', 'Total LDP VC']):
        return None, device_names

    cleaned = re.sub(r'[\t\xa0]+', '  ', output)
    items = re.split(r'\s{2,}', cleaned.strip())
    if len(items) < 6:
        return None, device_names

    try:
        vcid = items[0].strip()
        destination = items[1].strip()
        service_name = items[2].strip()
        status = items[3].strip().lower()
        interface = items[4].strip()
        vc_type = items[5].strip()

        # 从接口描述中获取完整业务名称
        full_service_name = interface_map.get(
            (device_ip, interface), service_name)
        full_service_name = re.sub(r'\s+', ' ', full_service_name.strip())[:40]

        role = 'N/A'
        for item in items:
            if item.strip().lower() in ['master', 'backup']:
                role = item.strip()
                break

        current_name = device_names.get(device_ip, "Unknown")
        return f"{device_ip},{current_name},{vcid},{destination},{full_service_name},{status},{interface},{vc_type},{role}\n", device_names
    except IndexError as e:
        print(f"字段解析异常：{items}")
        return None, device_names


def generate_vc_report(src_file, dst_file, host_list_file):
    """生成增强型VC报告，包含完整业务名称"""
    device_names = defaultdict(str)
    interface_map = {}
    vc_entries = []
    failure_ips = []

    # 加载失败的IP
    if os.path.exists("failure_ips.tmp"):
        with open("failure_ips.tmp", 'r') as f:
            failure_ips = [line.strip() for line in f]

    # 解析原始数据
    with open(src_file, 'r', encoding='utf-8') as f:
        total_lines = sum(1 for _ in f)
        f.seek(0)
        with tqdm(total=total_lines, desc="📥 解析原始数据", unit="行") as pbar:
            for line in f:
                line = line.strip()
                interface_map = parse_interface_description(
                    line, interface_map)
                cleaned_line, device_names = wash_l2vc_brief(
                    line, device_names, interface_map)
                if cleaned_line:
                    vc_entries.append(cleaned_line)
                pbar.update(1)

    # 加载所有设备
    with open(host_list_file, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        all_devices = [row[0].strip() for row in reader if row]

    # 生成报告
    with open(dst_file, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow([
            '设备IP', '设备名称', 'VCID', '目的地址', '业务名称', '状态',
            '接口', '业务类型', '角色', '连接状态'
        ])

        with tqdm(total=len(all_devices), desc="📤 生成报告", unit="台",
                  bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]") as pbar:
            for ip in all_devices:
                if ip in failure_ips:
                    writer.writerow([ip, "N/A", "-", "-", "-",
                                    "-", "-", "-", "-", "连接失败"])
                    pbar.update(1)
                    continue

                current_name = device_names.get(ip, "Unknown")
                device_vcs = [e for e in vc_entries if e.startswith(f"{ip},")]

                if not device_vcs:
                    writer.writerow(
                        [ip, current_name, "-", "-", "-", "-", "-", "-", "-", "无业务数据"])
                    pbar.update(1)
                    continue

                vc_count = 0
                for entry in device_vcs:
                    fields = entry.strip().split(',')
                    if len(fields) < 9:
                        continue
                    if fields[1] == "Unknown" and current_name != "Unknown":
                        fields[1] = current_name
                    fields.append("成功连接")
                    writer.writerow(fields)
                    vc_count += 1

                pbar.set_postfix_str(f"{ip[:15]} ({vc_count}业务)")
                pbar.update(1)

    # 更新设备名称映射
    with open("device_name_mapping.csv", 'w', encoding='utf-8') as f_map:
        f_map.write("IP地址,设备名称\n")
        for ip, name in device_names.items():
            f_map.write(f"{ip},{name}\n")

    print(f"\n{Fore.GREEN}✅ 报告生成完成！设备总数：{len(all_devices)} | 业务条目：{len(vc_entries)}")
    print(f"📁 主报告文件：{os.path.abspath(dst_file)}")
    print(
        f"📜 设备名称映射文件：{os.path.abspath('device_name_mapping.csv')}{Style.RESET_ALL}")


def fish(filename, ret_name, max_workers=40):
    """全量采集函数，支持多线程并行连接设备并执行命令"""
    import csv
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    success_count = 0
    failure_count = 0
    total_attempts = 0

    with open(ret_name, "w", encoding='utf-8') as revFile:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for index, row in enumerate(hostip, start=1):
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        cmd = row[3].strip()
                        futures.append(executor.submit(
                            process_device1,
                            ip, user, pwd, cmd, index, total_devices, revFile,
                            total_attempts
                        ))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="处理设备", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                result = future.result()
                                if result["success"]:
                                    success_count += 1
                                else:
                                    failure_count += 1
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                                failure_count += 1
                            pbar.update(1)

                # 最终统计输出
                print(
                    f"\n📊 成功 {success_count} 台设备 | 失败 {failure_count} 台设备 | 总共 {total_devices} 台设备"
                )

        except Exception as global_error:
            print(f"{Fore.RED}⛔ 全局错误导致进程终止: {str(global_error)}{Style.RESET_ALL}")


def process_device1(ip, user, pwd, cmd, index, total_devices, revFile, total_attempts):
    """处理单个设备的连接和命令执行"""
    from threading import Lock
    # 文件写入锁，确保线程安全
    file_lock = Lock()
    result = {"success": False}
    try:
        print("\n")
        with file_lock:  # 确保分隔线打印线程安全
            dynamic_colored_divider(
                color_code=36, symbol='#', enable_timestamp=False)

        # 连接设备
        channel = create_channel(
            ip=ip,
            username=user,     # Changed from user to username
            password=pwd,      # Changed from pwd to password
            retry_count=3
        )

        # 处理连接结果
        if channel is not None:
            try:
                # 执行配置命令
                config_host(channel, cmd, revFile, ip)
                result["success"] = True
            except Exception as cmd_error:
                print(
                    f"{Fore.RED}🔴 设备 {ip} 命令执行失败: {str(cmd_error)}{Style.RESET_ALL}")
            finally:
                try:
                    channel.close()
                except Exception:
                    pass  # 确保关闭操作不会引发异常
        else:
            print(f"{Fore.RED}🔴 设备 {ip} 连接失败{Style.RESET_ALL}")

        # 打印剩余设备提示（避免频繁打印，简化输出）
        remaining = total_devices - index
        if remaining > 0:
            print(f"{Fore.YELLOW}⚠️  发现 {remaining} 台设备未完成处理{Style.RESET_ALL}")

    except Exception as device_error:
        print(f"{Fore.RED}🔴 设备 {ip} 处理过程中发生未捕获异常: {str(device_error)}{Style.RESET_ALL}")

    return result


def fish_port_cmd(filename, ret_name, max_workers=40):
    """空闲端口检查数据收集函数，支持多线程并行采集多台设备的槽位和端口信息"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_port_device, ip, user, pwd, revFile, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 空闲端口检查进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ 端口数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_port_device(ip, user, pwd, revFile, fail_log):
    """处理单个设备的槽位和端口信息采集"""
    from colorama import Fore, Style
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            return

        execute_some_command(channel, "screen-length 0", 1)

        # 采集槽位信息
        try:
            slot_output = execute_some_command(
                channel, "show install package", 3)
            with file_lock:  # 线程安全写入
                for line in splitstr(slot_output):
                    revFile.write(f"{ip} , show install package , {line}\n")
        except Exception as slot_error:
            print(f"{Fore.YELLOW}⚠️ 设备 {ip} 采集槽位信息失败: {slot_error}{Style.RESET_ALL}")

        # 采集端口信息
        try:
            port_output = execute_some_command(channel, "show install port", 3)
            with file_lock:  # 线程安全写入
                for line in splitstr(port_output):
                    revFile.write(f"{ip} , show install port , {line}\n")
        except Exception as port_error:
            print(f"{Fore.YELLOW}⚠️ 设备 {ip} 采集端口信息失败: {port_error}{Style.RESET_ALL}")

    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def wash_port_info(content_line, device_names, slot_info):
    """清洗采集的槽位和端口数据并提取设备名称"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names, slot_info

    device_ip, cmd, output = parts

    # 提取设备名称
    name_match = re.search(r'<([^>]+)>', output, re.IGNORECASE)
    if name_match:
        device_name = name_match.group(1).strip()
        device_names[device_ip] = device_name  # 更新设备名称映射

    # 处理槽位信息
    if cmd.strip().lower() == "show install package":
        slot_match = re.search(
            r'slot\s*(\d+)\s*:\s*([^,]+?)\s*,\s*real\s*:\s*(\S+)', output, re.IGNORECASE)
        if slot_match:
            slot_num = slot_match.group(1)
            slot_type = slot_match.group(2).strip()
            real_type = slot_match.group(3).strip()
            slot_info[device_ip][slot_num] = (slot_type, real_type)
        return None, device_names, slot_info

    # 处理端口信息
    if cmd.strip().lower() == "show install port":
        # 检查是否为槽位分隔行
        slot_header = re.search(
            r'\*{5,}\s*slot\s*(\d+)\s*\*{5,}', output, re.IGNORECASE)
        if slot_header:
            slot_info[device_ip]['current_slot'] = slot_header.group(1)
            return None, device_names, slot_info

        # 解析端口信息
        port_match = re.search(
            r'port\s*(\d+)\s*:\s*([^,]+?)\s*,\s*real\s*:\s*(\S+)', output, re.IGNORECASE)
        if port_match:
            port_num = port_match.group(1)
            port_type = port_match.group(2).strip()
            real_type = port_match.group(3).strip()
            is_idle = (real_type.upper() == 'NULL') or (port_type != real_type)
            current_slot = slot_info[device_ip].get('current_slot', '未知槽位')
            slot_type, slot_real = slot_info[device_ip].get(
                current_slot, ('未知类型', '未知实际类型'))
            current_name = device_names.get(device_ip, "未知设备")
            if is_idle:
                return (device_ip, current_name, current_slot, slot_type, port_num, port_type, real_type), device_names, slot_info
        return None, device_names, slot_info

    return None, device_names, slot_info


def generate_port_report(src_file, dst_file, host_list_file):
    """生成精确的端口状态报告，包括槽位信息"""

    device_names = defaultdict(str)  # IP与设备名称映射表
    # IP -> {slot_num: (slot_type, real_type), 'current_slot': str}
    slot_info = defaultdict(lambda: {})
    port_data = defaultdict(list)

    # 处理原始数据
    with open(src_file, 'r', encoding='utf-8') as f_in:
        for line in f_in:
            line = line.strip()
            data, device_names, slot_info = wash_port_info(
                line, device_names, slot_info)
            if data:
                ip, dev_name, slot_num, slot_type, port_num, pt, rt = data
                port_data[ip].append((slot_num, slot_type, port_num, pt, rt))

    # 加载所有设备IP
    with open(host_list_file, 'r', encoding='gbk') as f:
        all_devices = [row[0].strip() for row in csv.reader(f) if row]

    # 加载连接失败IP
    failure_ips = []
    if os.path.exists("failure_ips.tmp"):
        with open("failure_ips.tmp", 'r', encoding='utf-8') as f:
            failure_ips = [line.strip() for line in f]

    # 生成CSV报告
    with open(dst_file, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['设备IP', '设备名称', '空闲端口', '连接状态'])

        for ip in all_devices:
            if ip in failure_ips:
                writer.writerow([ip, "数据异常", "-", "连接失败"])
                continue

            ports = port_data.get(ip, [])
            idle_ports = [f"槽位{slot_num}: {slot_type} | 端口{port_num}: {pt} | real: {rt}"
                          for slot_num, slot_type, port_num, pt, rt in ports]
            dev_name = device_names.get(ip, "未知设备")

            if idle_ports:
                for port in idle_ports:
                    writer.writerow([ip, dev_name, port, "检测成功"])
            elif ports:  # 有端口数据但无空闲（此处逻辑上不适用，因只收集空闲端口）
                writer.writerow([ip, dev_name, "-", "端口满载"])
            else:  # 无端口数据但连接成功
                writer.writerow([ip, dev_name, "-", "未检测到"])

    print(f"{Fore.GREEN}✅ 报告生成完成，共处理 {len(all_devices)} 台设备{Style.RESET_ALL}")


def wash_board_info(content_line, device_names, board_counts):
    """清洗采集的槽位数据并统计业务板卡数量"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names, board_counts

    device_ip, cmd, output = parts

    # 提取设备名称
    name_match = re.search(r'<([^>]+)>', output)
    if name_match:
        device_names[device_ip] = name_match.group(1).strip()

    # 提取并统计板卡信息
    if cmd.strip().lower() == "show install package":
        slot_match = re.search(r'slot\s*\d+\s*:\s*(\w+)', output)
        if slot_match:
            board_type = slot_match.group(1).strip()
            if board_type != "NULL":
                if device_ip not in board_counts:
                    board_counts[device_ip] = defaultdict(int)
                board_counts[device_ip][board_type] += 1
        return None, device_names, board_counts

    return None, device_names, board_counts


def fish_board_cmd(filename, ret_name, max_workers=40):
    """业务板卡统计数据收集函数，支持多线程并行采集多台设备的板卡信息"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    failure_ips = []  # 记录连接失败的IP

    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_board_device, ip, user, pwd, revFile, fail_log, failure_ips))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 业务板卡统计进度", unit="台", ncols=100) as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ 板卡数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_board_device(ip, user, pwd, revFile, fail_log, failure_ips):
    """处理单个设备的板卡信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    # 文件写入锁，确保线程安全
    file_lock = Lock()

    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
                failure_ips.append(ip)
            return

        execute_some_command(channel, "screen-length 0", 1)
        output = execute_some_command(channel, "show install package", 3)
        with file_lock:  # 线程安全写入
            for line in splitstr(output):
                revFile.write(f"{ip} , show install package , {line}\n")

    except Exception as cmd_error:
        print(f"{Fore.RED}⚠️ 设备 {ip} 执行命令失败: {str(cmd_error)}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
            failure_ips.append(ip)
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {str(close_error)}{Style.RESET_ALL}")


def generate_board_report(src_file, dst_file, host_list_file):
    """生成业务板卡统计报告"""
    device_names = defaultdict(str)  # IP与设备名称映射
    board_counts = {}  # IP -> {board_type: count}

    # 处理采集数据
    with open(src_file, 'r', encoding='utf-8') as f_in:
        for line in f_in:
            line = line.strip()
            _, device_names, board_counts = wash_board_info(
                line, device_names, board_counts)

    # 加载所有设备IP
    with open(host_list_file, 'r', encoding='gbk') as f:
        all_devices = [row[0].strip() for row in csv.reader(f) if row]

    # 加载连接失败IP
    failure_ips = []
    if os.path.exists("failure_ips.tmp"):
        with open("failure_ips.tmp", 'r', encoding='utf-8') as f:
            failure_ips = [line.strip() for line in f]

    # 生成报告
    with open(dst_file, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['设备IP', '设备名称', '业务板卡统计', '数量', '连接状态'])

        for ip in all_devices:
            if ip in failure_ips:
                writer.writerow([ip, "数据异常", "-", "-", "连接失败"])
                continue

            dev_name = device_names.get(ip, "未知设备")
            if ip in board_counts:
                for board_type, count in board_counts[ip].items():
                    writer.writerow([ip, dev_name, board_type, count, "检测成功"])
            else:
                writer.writerow([ip, dev_name, "-", "-", "未检测到"])

    print(f"{Fore.GREEN}✅ 报告生成完成，共处理 {len(all_devices)} 台设备{Style.RESET_ALL}")


def fish_port_usage_cmd(filename, ret_name, max_workers=40):
    """采集端口使用率数据，支持多线程并行采集多台设备的端口信息"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_port_usage_device, ip, user, pwd, revFile, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 端口使用率统计进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ 端口使用率数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_port_usage_device(ip, user, pwd, revFile, fail_log):
    """处理单个设备的端口使用率信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            return

        execute_some_command(channel, "screen-length 0", 1)
        output = execute_some_command(channel, "show interface brief main", 3)
        with file_lock:  # 线程安全写入
            for line in splitstr(output):
                revFile.write(f"{ip} , show interface brief main , {line}\n")

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def wash_port_usage_info(content_line, device_names, port_data):
    """清洗端口数据并统计使用率，支持字段缺失情况"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names, port_data
    device_ip, cmd, output = parts

    # 提取设备名称
    name_match = re.search(r'<([^>]+)>', output)
    if name_match:
        device_names[device_ip] = name_match.group(1).strip()

    if cmd.strip().lower() == "show interface brief main":
        # 跳过头信息或空行
        if any(x in output for x in ["Interface", "PHY:", "InUti/OutUti", "down:", "*down", "(l):", "(b):", "(d):"]):
            return None, device_names, port_data

        # 分割字段
        fields = re.split(r'\s{2,}', output.strip())
        if len(fields) < 7:  # 至少需要7个字段（接口、物理状态、Auto-Neg、速率、带宽、输入、输出）
            print(f"警告：设备 {device_ip} 输出字段不足: {output}")
            return None, device_names, port_data

        # 初始化默认值
        interface = phy_status = speed = in_uti = out_uti = lag = crc_count = "字段不足"
        status = "未知"

        try:
            interface = fields[0].strip() if len(fields) > 0 else "字段不足"
            phy_status = fields[1].strip() if len(fields) > 1 else "字段不足"
            speed = fields[3].split()[0].strip() if len(fields) > 3 else "字段不足"
            in_uti_str = fields[5].strip().rstrip(
                '%') if len(fields) > 5 else "字段不足"
            out_uti_str = fields[6].strip().rstrip(
                '%') if len(fields) > 6 else "字段不足"
            lag = fields[7].strip() if len(fields) > 7 else "-"
            crc_count = fields[9].strip() if len(fields) > 9 else "0"

            # 处理使用率
            if in_uti_str != "字段不足" and out_uti_str != "字段不足":
                if (in_uti_str.replace('.', '').replace('-', '').isdigit() and
                        out_uti_str.replace('.', '').replace('-', '').isdigit()):
                    in_uti = float(in_uti_str) if in_uti_str != '-' else 0.0
                    out_uti = float(out_uti_str) if out_uti_str != '-' else 0.0
                    status = "警告" if in_uti > 80 or out_uti > 80 else "良好"
                else:
                    print(
                        f"警告：设备 {device_ip} 接口 {interface} 使用率数据异常: {in_uti_str}/{out_uti_str}")
                    in_uti = out_uti = "数据异常"
                    status = "异常"
            else:
                in_uti = out_uti = "字段不足"
                status = "未知"

            # 处理CRC
            try:
                crc_count = int(crc_count)
            except ValueError:
                crc_count = 0

            # 存储数据
            if device_ip not in port_data:
                port_data[device_ip] = []
            port_data[device_ip].append(
                (interface, speed, lag, in_uti, out_uti, status, phy_status, crc_count))

        except Exception as e:
            print(f"警告：设备 {device_ip} 接口解析异常: {output} | 错误: {e}")

    return None, device_names, port_data


def generate_port_usage_report(src_file, dst_file, host_list_file):
    """生成端口使用率统计报告，支持字段缺失"""
    device_names = {}
    port_data = {}

    # 处理原始数据
    with open(src_file, 'r', encoding='utf-8') as f_in:
        for line in f_in:
            line = line.strip()
            _, device_names, port_data = wash_port_usage_info(
                line, device_names, port_data)

    # 加载所有设备IP
    with open(host_list_file, 'r', encoding='gbk') as f:
        all_devices = [row[0].strip() for row in csv.reader(f) if row]

    # 加载连接失败IP
    failure_ips = []
    if os.path.exists("failure_ips.tmp"):
        with open("failure_ips.tmp", 'r', encoding='utf-8') as f:
            failure_ips = [line.strip() for line in f]

    # 生成报告
    with open(dst_file, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['设备IP', '设备名称', '端口', '端口速率', 'Lag',
                        '输入', '输出', '状态', '物理状态', 'CRC', '连接状态'])

        for ip in all_devices:
            if ip in failure_ips:
                writer.writerow([ip, "数据异常", "-", "-", "-",
                                "-", "-", "-", "-", "-", "连接失败"])
                continue

            dev_name = device_names.get(ip, "未知设备")
            if ip in port_data:
                for port in port_data[ip]:
                    interface, speed, lag, in_uti, out_uti, status, phy_status, crc_count = port
                    in_uti_display = f"{in_uti:.2f}%" if isinstance(
                        in_uti, (int, float)) else in_uti
                    out_uti_display = f"{out_uti:.2f}%" if isinstance(
                        out_uti, (int, float)) else out_uti
                    writer.writerow([
                        ip, dev_name, interface, speed, lag,
                        in_uti_display, out_uti_display,
                        status, phy_status, crc_count, "检测成功"
                    ])
            else:
                writer.writerow([ip, dev_name, "-", "-", "-",
                                "-", "-", "-", "-", "-", "未检测到"])

    print(f"{Fore.GREEN}✅ 报告生成完成，共处理 {len(all_devices)} 台设备{Style.RESET_ALL}")


def fish_crc_cmd(filename, ret_name, max_workers=40):
    """采集CRC数据，支持多线程并行采集多台设备的CRC信息"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict

# 文件写入锁，确保线程安全
    file_lock = Lock()
    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_crc_device, ip, user, pwd, revFile, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 CRC检查进度", unit="台", ncols=100) as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 采集异常: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ CRC数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_crc_device(ip, user, pwd, revFile, fail_log):
    """处理单个设备的CRC信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict

# 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            return

        execute_some_command(channel, "screen-length 512", 1)
        sResult = execute_some_command(channel, "show interface brief main", 3)
        execute_some_command(channel, "screen-length 25", 1)

        with file_lock:  # 线程安全写入
            for item in splitstr(sResult):
                revFile.write(f"{ip} , show interface brief main , {item}\n")

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def wash_crc_info(content_line, device_names):
    """清洗CRC数据并提取设备名称"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names

    device_ip, cmd, output = parts

    # 提取设备名称
    name_match = re.search(r'<([^>]+)>', output, re.IGNORECASE)
    if name_match:
        device_name = name_match.group(1).strip()
        device_names[device_ip] = device_name

    # 处理端口信息
    if cmd.strip().lower() == "show interface brief main":
        # 跳过头信息或空行
        if "Interface" in output or not output.strip():
            return None, device_names

        # 解析端口数据
        fields = re.split(r'\s{2,}', output.strip())
        if len(fields) >= 10:  # 确保有足够的字段（某些组件的索引为 9 处有 Input-CRC）
            interface = fields[0].strip()
            phy_status = fields[1].strip()
            speed = fields[3].split()[0].strip()  # 提取速率
            # CRC 可能在某些输出中缺失
            crc_errors = fields[9].strip() if len(fields) > 9 else "0"
            try:
                crc_count = int(crc_errors)
            except ValueError:
                crc_count = 0
            current_name = device_names.get(device_ip, "未知设备")
            return (device_ip, current_name, interface, speed, crc_count, phy_status), device_names
    return None, device_names


def generate_crc_report(src_file, dst_file, host_list_file):
    """生成CRC检查报告"""
    device_names = defaultdict(str)
    crc_data = defaultdict(list)

    # 处理原始数据
    with open(src_file, 'r', encoding='utf-8') as f_in:
        for line in f_in:
            line = line.strip()
            data, device_names = wash_crc_info(line, device_names)
            if data:
                ip, dev_name, interface, speed, crc_count, phy_status = data
                crc_data[ip].append((interface, speed, crc_count, phy_status))

    # 加载所有设备IP
    with open(host_list_file, 'r', encoding='gbk') as f:
        all_devices = [row[0].strip() for row in csv.reader(f) if row]

    # 加载连接失败IP
    failure_ips = []
    if os.path.exists("failure_ips.tmp"):
        with open("failure_ips.tmp", 'r') as f:
            failure_ips = [line.strip() for line in f]

    # 生成报告
    with open(dst_file, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['设备IP', '设备名称', '端口', '端口速率', 'CRC', '物理状态', '连接状态'])

        for ip in all_devices:
            if ip in failure_ips:
                writer.writerow([ip, "N/A", "-", "-", "-", "-", "连接失败"])
                continue

            ports = crc_data.get(ip, [])
            dev_name = device_names.get(ip, "未知设备")

            if ports:
                for port in ports:
                    interface, speed, crc_count, phy_status = port
                    writer.writerow(
                        [ip, dev_name, interface, speed, crc_count, phy_status, "检测成功"])
            else:
                writer.writerow([ip, dev_name, "-", "-", "-", "-", "未检测"])

    # 清理临时文件
    if os.path.exists("failure_ips.tmp"):
        os.remove("failure_ips.tmp")

    print(f"{Fore.GREEN}✅ 报告已生成，共处理{len(all_devices)}台设备{Style.RESET_ALL}")


def fish_lldp_neighbor_cmd(filename, ret_name, max_workers=40):
    """采集LLDP邻居数据，支持多线程并行采集多台设备的LLDP信息"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock

# 文件写入锁，确保线程安全
    file_lock = Lock()
    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_lldp_device, ip, user, pwd, revFile, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 LLDP邻居检查进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ LLDP邻居数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_lldp_device(ip, user, pwd, revFile, fail_log):
    """处理单个设备的LLDP邻居信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            return

        # 设置屏幕长度以避免分页
        execute_some_command(channel, "screen-length 512", 1)
        output = execute_some_command(channel, "show lldp neighbor", 3)
        with file_lock:  # 线程安全写入
            for line in splitstr(output):
                revFile.write(f"{ip} , show lldp neighbor , {line}\n")
        # 重置屏幕长度为默认值
        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def wash_lldp_neighbor_info(content_line, device_names, lldp_data):
    """清洗LLDP邻居数据并提取设备名称"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names, lldp_data

    device_ip, cmd, output = parts

    # 提取设备名称
    name_match = re.search(r'<([^>]+)>', output)
    if name_match:
        device_names[device_ip] = name_match.group(1).strip()

    if cmd.strip().lower() != "show lldp neighbor":
        return None, device_names, lldp_data

    # 初始化设备 IP 的 LLDP 数据
    if device_ip not in lldp_data:
        lldp_data[device_ip] = []

    # 提取接口信息
    interface_match = re.search(
        r"Interface '([^']+)' has\s+(\d+)\s+LLDP Neighbors:", output)
    if interface_match:
        interface = interface_match.group(1)
        lldp_data[device_ip].append({"interface": interface, "neighbors": []})
        return None, device_names, lldp_data

    # 检测新的邻居条目
    neighbor_match = re.search(r"Neighbor \d+:", output)
    if neighbor_match and lldp_data[device_ip]:
        lldp_data[device_ip][-1]["neighbors"].append({})
        return None, device_names, lldp_data

    # 提取系统名称
    system_name_match = re.search(r"System Name:\s*(.+)", output)
    if system_name_match and lldp_data[device_ip] and lldp_data[device_ip][-1]["neighbors"]:
        system_name = system_name_match.group(1).strip()
        lldp_data[device_ip][-1]["neighbors"][-1]["system_name"] = system_name
        return None, device_names, lldp_data

    # 提取端口 ID
    port_id_match = re.search(r"Port ID:.*?-\s*(.+)", output)
    if port_id_match and lldp_data[device_ip] and lldp_data[device_ip][-1]["neighbors"]:
        port_id = port_id_match.group(1).strip()
        lldp_data[device_ip][-1]["neighbors"][-1]["port_id"] = port_id
        return None, device_names, lldp_data

    # 提取管理 IP 地址
    mgmt_ip_match = re.search(
        r"Management Address: IPv4 - (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", output)
    if mgmt_ip_match and lldp_data[device_ip] and lldp_data[device_ip][-1]["neighbors"]:
        mgmt_ip = mgmt_ip_match.group(1)
        lldp_data[device_ip][-1]["neighbors"][-1]["mgmt_ip"] = mgmt_ip
        return None, device_names, lldp_data

    # 提取邻居系统详情（System Description）
    system_desc_match = re.search(r"System Description:\s*(.+)", output)
    if system_desc_match and lldp_data[device_ip] and lldp_data[device_ip][-1]["neighbors"]:
        system_desc = system_desc_match.group(1).strip()
        lldp_data[device_ip][-1]["neighbors"][-1]["system_desc"] = system_desc
        return None, device_names, lldp_data

    return None, device_names, lldp_data


def generate_lldp_neighbor_report(src_file, dst_file, host_list_file):
    """生成LLDP邻居检查报告"""
    device_names = {}
    lldp_data = {}
    connection_failures = set()

    # 读取连接失败的设备 IP
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as conn_fail:
            connection_failures = set(line.strip() for line in conn_fail)
    except FileNotFoundError:
        pass  # 如果文件不存在，则继续执行，列表为空

    # 解析采集数据
    with open(src_file, "r", encoding='utf-8') as f:
        for line in f:
            wash_lldp_neighbor_info(line, device_names, lldp_data)

    # 读取设备清单
    with open(host_list_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]

    # 生成报告
    with open(dst_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow(["本端设备IP", "本端设备名称", "本端接口", "对端邻居系统名称",
                         "对端邻居端口", "邻居系统IP", "邻居系统详情", "连接状态"])

        for ip in host_ips:
            if ip in connection_failures:
                # 连接失败的设备
                writer.writerow([ip, "连接失败", "-", "-", "-", "-", "-", "连接失败"])
            elif ip in device_names:
                # 成功采集到数据的设备
                if ip in lldp_data and lldp_data[ip]:
                    for entry in lldp_data[ip]:
                        for neighbor in entry["neighbors"]:
                            row = [
                                ip,
                                device_names[ip],
                                entry["interface"],
                                neighbor.get("system_name", "-"),
                                neighbor.get("port_id", "-"),
                                neighbor.get("mgmt_ip", "-"),
                                neighbor.get("system_desc", "-"),
                                "成功"
                            ]
                            writer.writerow(row)
                else:
                    # 无邻居数据
                    row = [ip, device_names[ip], "-",
                           "-", "-", "-", "-", "无邻居数据"]
                    writer.writerow(row)
            else:
                # 未采集到数据但不在失败列表中的设备
                row = [ip, "未知设备", "-", "-", "-", "-", "-", "数据异常"]
                writer.writerow(row)

    print(f"{Fore.GREEN}✅ 报告生成完成，共处理 {len(host_ips)} 台设备{Style.RESET_ALL}")


def fish_arp_cmd(filename, ret_name, max_workers=40):
    """采集基站和业务上报IP数据，支持多线程并行采集多台设备的ARP信息"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    with open(ret_name, "w", encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_arp_device, ip, user, pwd, revFile, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 ARP数据采集进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}✅ ARP数据已保存至 {ret_name}{Style.RESET_ALL}")


def process_arp_device(ip, user, pwd, revFile, fail_log):
    """处理单个设备的ARP信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            return

        # 设置屏幕长度以避免分页
        execute_some_command(channel, "screen-length 512", 1)
        # 采集 "show user-arp detail" 输出
        output1 = execute_some_command(channel, "show user-arp detail", 3)
        with file_lock:  # 线程安全写入
            for line in splitstr(output1):
                revFile.write(f"{ip} , show user-arp detail , {line}\n")
        # 采集 "show arp all" 输出
        output2 = execute_some_command(channel, "show arp all", 3)
        with file_lock:  # 线程安全写入
            for line in splitstr(output2):
                revFile.write(f"{ip} , show arp all , {line}\n")
        # 重置屏幕长度为默认值
        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def parse_arp_data(content_line, device_names, arp_data):
    """解析ARP数据并提取设备名称"""
    parts = content_line.split(' , ', 2)
    if len(parts) != 3:
        return None, device_names, arp_data

    device_ip, cmd, output = parts

    # 提取设备名称
    name_match = re.search(r'<([^>]+)>', output)
    if name_match:
        device_names[device_ip] = name_match.group(1).strip()

    if device_ip not in arp_data:
        arp_data[device_ip] = []

    if cmd.strip().lower() == "show user-arp detail":
        # 解析MAC、VLAN、端口和IP
        arp_match = re.search(
            r'(\w{4}\.\w{4}\.\w{4})\s+(\d+)\s+([\w\s/]+\d+/\d+/\d+(?:\.\d+)?)\s+(\d+\.\d+\.\d+\.\d+)',
            output
        )
        if arp_match:
            mac = arp_match.group(1)
            vlan = arp_match.group(2)
            port = arp_match.group(3).strip()
            ip_addr = arp_match.group(4)
            # 根据VLAN设置接口和模型
            if vlan.isdigit():
                model = "VLAN专线"
                interface = f"{port}.{vlan}"  # 显示VLAN子接口
            else:
                model = "纯通道"
                interface = port  # 显示物理端口
            arp_data[device_ip].append({
                "mac": mac,
                "vlan": vlan if vlan.isdigit() else "-",
                "port": interface,
                "ip": ip_addr,
                "model": model
            })

    elif cmd.strip().lower() == "show arp all":
        # 解析IP、MAC和接口
        arp_all_match = re.search(
            r'(\d+\.\d+\.\d+\.\d+)\s+(\w{4}\.\w{4}\.\w{4})\s+\w+\s+\d+\s+([\w\s/]+\d+/\d+/\d+(?:\.\d+)?)',
            output
        )
        if arp_all_match:
            ip_addr = arp_all_match.group(1)
            mac = arp_all_match.group(2)
            port = arp_all_match.group(3).strip()
            # 默认无VLAN信息
            vlan = "-"
            model = "纯通道"
            interface = port  # 显示物理端口
            arp_data[device_ip].append({
                "mac": mac,
                "vlan": vlan,
                "port": interface,
                "ip": ip_addr,
                "model": model
            })
    return None, device_names, arp_data


def generate_arp_report(src_file, dst_file, host_list_file):
    """生成ARP统计报告"""
    device_names = {}
    arp_data = defaultdict(list)
    connection_failures = set()

    # 读取连接失败的设备IP
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as conn_fail:
            connection_failures = set(line.strip() for line in conn_fail)
    except FileNotFoundError:
        pass

    # 解析采集数据
    with open(src_file, "r", encoding='utf-8') as f:
        for line in f:
            parse_arp_data(line, device_names, arp_data)

    # 读取设备清单
    with open(host_list_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]

    # 生成报告
    with open(dst_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow(["设备IP", "设备名称", "业务MAC", "业务接口",
                        "业务VLAN", "业务上报IP", "业务模型", "连接状态"])

        for ip in host_ips:
            if ip in connection_failures:
                writer.writerow([ip, "连接失败", "-", "-", "-", "-", "-", "连接失败"])
            elif ip in device_names:
                if ip in arp_data and arp_data[ip]:
                    for entry in arp_data[ip]:
                        row = [
                            ip,
                            device_names[ip],
                            entry["mac"],
                            entry["port"],  # 根据VLAN调整后的接口
                            entry["vlan"],
                            entry["ip"],
                            entry["model"],  # 根据VLAN调整后的模型
                            "成功"
                        ]
                        writer.writerow(row)
                else:
                    row = [ip, device_names[ip], "-",
                           "-", "-", "-", "-", "无ARP数据"]
                    writer.writerow(row)
            else:
                row = [ip, "未知设备", "-", "-", "-", "-", "-", "数据异常"]
                writer.writerow(row)

    print(f"{Fore.GREEN}✅ 报告生成完成，共处理 {len(host_ips)} 台设备{Style.RESET_ALL}")


def generate_topology_html(lldp_report_file, output_html_file):
    """根据LLDP邻居报告生成HTML拓扑图"""
    # 读取LLDP邻居报告
    with open(lldp_report_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        lldp_data = list(reader)

    # 构建设备节点和连接边
    nodes = {}
    edges = []
    # 定义设备类型到Font Awesome图标代码的映射
    group_icons = {
        'OPTEL': '\uf108',      # desktop
        'Huawei': '\uf109',     # laptop
        'Fiberhome': '\uf0a0',  # hdd
        'Other': '\uf1cb'       # network-wired
    }
    for row in lldp_data:
        local_ip = row['本端设备IP']
        local_name = row['本端设备名称']
        local_port = row['本端接口']
        neighbor_name = row['对端邻居系统名称']
        neighbor_port = row['对端邻居端口']
        neighbor_ip = row['邻居系统IP']
        neighbor_details = row['邻居系统详情']

        # 添加本地设备节点
        if local_ip not in nodes:
            group = get_device_group(neighbor_details)
            nodes[local_ip] = {
                'id': local_ip,
                'label': local_name,
                'title': neighbor_details,
                'shape': 'icon',
                'icon': {
                    'face': 'FontAwesome',
                    # 默认使用network-wired图标
                    'code': group_icons.get(group, '\uf1cb'),
                    'size': 50,
                    'color': '#2B7CE9'
                }
            }

        # 添加邻居设备节点
        if neighbor_ip not in nodes:
            group = get_device_group(neighbor_details)
            nodes[neighbor_ip] = {
                'id': neighbor_ip,
                'label': neighbor_name,
                'title': neighbor_details,
                'shape': 'icon',
                'icon': {
                    'face': 'FontAwesome',
                    'code': group_icons.get(group, '\uf1cb'),
                    'size': 50,
                    'color': '#2B7CE9'
                }
            }

        # 添加连接边
        edges.append({
            'from': local_ip,
            'to': neighbor_ip,
            'label': f"{local_port} -- {neighbor_port}",
            'title': f"{local_port} -- {neighbor_port}"
        })

    # 转换为Vis.js格式
    vis_nodes = list(nodes.values())
    vis_edges = edges

    # 生成HTML和JavaScript代码
    html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>网络拓扑图</title>
    <!-- 引入Font Awesome CSS -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css">
    <script type="text/javascript" src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>
    <style>
        #mynetwork {{
            width: 100%;
            height: 100vh;
            border: 1px solid lightgray;
            background-color: #f5f5f5;
        }}
    </style>
</head>
<body>
    <div id="mynetwork"></div>
    <script type="text/javascript">
        // 初始化节点和边数据
        var nodes = new vis.DataSet({json.dumps(vis_nodes, ensure_ascii=False)});
        var edges = new vis.DataSet({json.dumps(vis_edges, ensure_ascii=False)});
        var container = document.getElementById('mynetwork');
        var data = {{
            nodes: nodes,
            edges: edges
        }};
        var options = {{
            layout: {{
                hierarchical: {{
                    enabled: true,
                    levelSeparation: 150,  // 垂直间距
                    nodeSpacing: 100,      // 水平间距，满足最小间隔100px
                    treeSpacing: 200,
                    direction: 'UD',       // 从上到下
                    sortMethod: 'directed'
                }}
            }},
            physics: false,  // 禁用物理效果
            nodes: {{
                shape: 'box',  // 使用矩形形状
                size: 20,
                font: {{
                    size: 14,
                    color: '#333333'
                }},
                borderWidth: 2,
                shadow: true
            }},
            edges: {{
                smooth: {{
                    enabled: true,
                    type: 'continuous',
                    roundness: 0.5
                }},
                color: {{
                    color: '#848484',
                    highlight: '#ff4500'
                }},
                arrows: 'to'
            }},
            interaction: {{
                hover: true,
                dragNodes: true,
                dragView: true,
                zoomView: true,
                tooltipDelay: 200
            }}
        }};
        var network = new vis.Network(container, data, options);
    </script>
</body>
</html>
    """

    # 写入HTML文件
    with open(output_html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"✅ 拓扑图已生成：{output_html_file}")


def get_device_group(details):
    """根据邻居系统详情判断设备类型"""
    if 'OPTEL' in details:
        return 'OPTEL'
    elif 'Huawei' in details:
        return 'Huawei'
    elif 'Fiberhome' in details:
        return 'Fiberhome'
    else:
        return 'Other'


def fish_ospf_neighbor_cmd(filename, ret_name, max_workers=40):
    """采集OSPF邻居接口信息，支持多线程并行采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    print(f"[START] 开始采集OSPF邻居接口信息，输入文件: {filename}, 输出文件: {ret_name}")
    with open(ret_name, "w", encoding='utf-8', newline='') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)  # 创建csv.writer对象用于写入
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_ospf_device1, ip, user, pwd, writer, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 OSPF邻居接口速率采集进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
    print("[END] 数据采集完成")


def process_ospf_device1(ip, user, pwd, writer, fail_log):
    """处理单个设备的OSPF邻居接口信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        print(f"✅ 成功连接设备 {ip}")
        execute_some_command(channel, "screen-length 512", 1)
        output = execute_some_command(channel, "show ospf neighbor brief", 3)
        # 线程安全写入，使用csv.writer
        with file_lock:
            for line in splitstr(output):
                writer.writerow([ip, "show ospf neighbor brief", line])
        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def parse_ospf_neighbor_data(device_ip, cmd, output, device_names, ospf_data):
    """解析OSPF邻居接口数据"""
    print(f"[parse_ospf_neighbor_data] 设备IP: {device_ip}, 命令: {cmd}")
    if device_ip not in device_names:
        name_match = re.search(r'<([^>]+)>', output)
        device_names[device_ip] = name_match.group(1) if name_match else "未知设备"

    if cmd.strip().lower() == 'show ospf neighbor brief':
        lines = output.split('\n')
        current_process = None
        for line in lines:
            process_match = re.search(r'OSPF process (\d+):', line)
            if process_match:
                current_process = process_match.group(1)
                print(f"[DEBUG] 设备 {device_ip} 找到进程: {current_process}")
                continue
            if current_process:  # 只处理有进程号的行
                neighbor_match = re.search(
                    r'(\d+\.\d+\.\d+\.\d+)\s+\d+\s+([\w/ -]+)\s+([\d:]+)\s+\d+\.\d+\.\d+\.\d+\s+(\S+)\s+\d+',
                    line.strip()
                )
                if neighbor_match:
                    neighbor_id = neighbor_match.group(1)
                    state = neighbor_match.group(2).strip()
                    uptime = neighbor_match.group(3)
                    interface = neighbor_match.group(4)
                    # 提取物理接口名称，例如 "50GE 0/6/1.31" -> "50GE 0/6/1"
                    physical_intf = interface.split(
                        '.')[0] if '.' in interface else interface
                    if device_ip not in ospf_data:
                        ospf_data[device_ip] = {}
                    if interface not in ospf_data[device_ip]:
                        ospf_data[device_ip][interface] = []
                    ospf_data[device_ip][interface].append({
                        "process": current_process,
                        "neighbor_id": neighbor_id,
                        "state": state,
                        "uptime": uptime,
                        "physical_intf": physical_intf
                    })
                    print(
                        f"[DEBUG] 设备 {device_ip} 解析邻居: {neighbor_id}, 接口: {interface}")
    return device_names, ospf_data


def generate_ospf_neighbor_report(src_file, dst_file, host_list_file):
    """生成OSPF邻居接口报告"""
    print(
        f"\n[generate_ospf_neighbor_report] 开始生成报告，源文件: {src_file}, 目标文件: {dst_file}")
    device_names = {}  # 存储设备名称
    ospf_data = defaultdict(list)  # 存储OSPF邻居数据
    connection_failures = set()  # 存储连接失败的设备IP

    # 读取连接失败的设备IP
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as conn_fail:
            connection_failures = set(line.strip() for line in conn_fail)
            print(f"[DEBUG] 读取连接失败设备: {len(connection_failures)} 台")
    except FileNotFoundError:
        print("[INFO] 未找到failure_ips.tmp文件")

    # 第一步：按设备IP分组输出行
    device_outputs = defaultdict(list)
    with open(src_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)  # 使用csv.reader读取
        for row in reader:
            if len(row) == 3:
                device_ip, cmd, output_line = row
                if cmd.strip().lower() == "show ospf neighbor brief":
                    device_outputs[device_ip].append(output_line.strip())
        print(f"[DEBUG] 数据分组完成，共 {len(device_outputs)} 台设备")

    # 第二步：解析每组输出
    for device_ip, lines in device_outputs.items():
        # 提取设备名称（假设提示符在最后一行）
        if lines:
            last_line = lines[-1]
            name_match = re.search(r'<([^>]+)>', last_line)
            if name_match:
                device_names[device_ip] = name_match.group(1).strip()
                print(f"[DEBUG] 设备 {device_ip} 名称: {device_names[device_ip]}")

        current_process = None  # 当前OSPF进程号
        for line in lines:
            # 匹配OSPF进程行
            process_match = re.search(r'OSPF process (\d+):', line)
            if process_match:
                current_process = process_match.group(1)
                print(f"[DEBUG] 设备 {device_ip} 找到进程: {current_process}")
                continue
            # 匹配邻居信息行
            neighbor_match = re.search(
                r'(\d+\.\d+\.\d+\.\d+)\s+(\d+)\s+(\w+/\s*-)\s+([\d:]+)\s+(\d+\.\d+\.\d+\.\d+)\s+(\w+\s+\d+/\d+/\d+\.\d+)\s+\d+',
                line.strip()
            )
            if neighbor_match and current_process:
                neighbor_id = neighbor_match.group(1)  # 邻居ID
                priority = neighbor_match.group(2)    # 优先级
                state = neighbor_match.group(3)       # OSPF状态
                uptime = neighbor_match.group(4)      # 在线时间
                address = neighbor_match.group(5)     # 邻居接口IP
                interface = neighbor_match.group(6)   # 接口
                port_speed = interface.split()[0]     # 端口速率（接口类型）
                ospf_data[device_ip].append({
                    "process": current_process,
                    "neighbor_id": neighbor_id,
                    "state": state,
                    "uptime": uptime,
                    "address": address,
                    "interface": interface,
                    "port_speed": port_speed
                })
                print(
                    f"[DEBUG] 设备 {device_ip} 解析邻居: {neighbor_id}, 接口: {interface}")

    # 第三步：读取设备清单
    with open(host_list_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(f"[DEBUG] 读取设备清单: {len(host_ips)} 台设备")

    # 第四步：生成报告
    with open(dst_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow(["设备IP", "设备名称", "互联端口速率", "端口",
                        "OSPF进程", "OSPF状态", "uptime", "邻居IP", "邻居接口IP"])
        print(f"[DEBUG] 写入报告表头")

        for ip in host_ips:
            if ip in connection_failures:
                writer.writerow(
                    [ip, "连接失败", "-", "-", "-", "-", "-", "-", "-"])
                print(f"[DEBUG] 设备 {ip}: 连接失败")
            elif ip in device_names:
                if ip in ospf_data and ospf_data[ip]:
                    for entry in ospf_data[ip]:
                        writer.writerow([
                            ip,
                            device_names[ip],
                            entry["port_speed"],
                            entry["interface"],
                            entry["process"],
                            entry["state"],
                            entry["uptime"],
                            entry["neighbor_id"],
                            entry["address"]
                        ])
                        print(f"[DEBUG] 设备 {ip} 写入邻居: {entry['neighbor_id']}")
                else:
                    writer.writerow(
                        [ip, device_names[ip], "-", "-", "-", "-", "-", "-", "-"])
                    print(f"[DEBUG] 设备 {ip}: 无OSPF邻居数据")
            else:
                writer.writerow(
                    [ip, "未知设备", "-", "-", "-", "-", "-", "-", "-"])
                print(f"[DEBUG] 设备 {ip}: 未采集到数据")
    print(f"✅ 报告生成完成，共处理 {len(host_ips)} 台设备")


def extract_device_name(output):
    """从命令输出中提取设备名称"""
    match = re.search(r'<([^>]+)>', output)
    return match.group(1).strip() if match else None


def export_running_config(host_file, output_dir_base="设备运行配置导出", max_workers=40):
    """导出设备运行配置到.sh文件，支持多线程并行处理多台设备"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from threading import Lock
    # 获取当前日期并创建输出文件夹
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_dir = f"{output_dir_base}-{current_date}"
    os.makedirs(output_dir, exist_ok=True)

    # 读取设备清单
    with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
        reader = csv.reader(csvFile)
        hostip = list(reader)
        total_devices = len(hostip)

        # 使用线程池并行处理设备
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for row in hostip:
                ip = row[0].strip()
                user = row[1].strip()
                pwd = row[2].strip()
                futures.append(executor.submit(
                    process_device_config, ip, user, pwd, output_dir))

            # 使用tqdm显示进度
            with tqdm(total=total_devices, desc="🔍 导出设备运行配置", unit="台") as pbar:
                for future in as_completed(futures):
                    try:
                        future.result()  # 获取结果，触发异常处理
                    except Exception as e:
                        print(f"线程执行出错: {str(e)}")
                    pbar.update(1)

    print(f"\n✅ 设备运行配置导出完成，文件保存至 {output_dir}")


def process_device_config(ip, user, pwd, output_dir):
    """处理单个设备的运行配置导出"""
    import os
    import csv
    import datetime
    import time
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from threading import Lock
    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                with open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
                    fail_log.write(ip + '\n')
            print(f"⚠️ 无法连接到设备 {ip}")
            return

        # 设置屏幕长度以避免分页
        execute_some_command(channel, "screen-length 512", 1)
        # 执行show running-configuration命令
        output = execute_some_command(channel, "show running-configuration", 5)
        # 提取设备名称
        device_name = extract_device_name(output)
        if device_name:
            # 保存输出到以设备名称命名的.sh文件
            file_path = os.path.join(output_dir, f"{device_name}.sh")
            with file_lock:  # 线程安全写入文件
                with open(file_path, "w", encoding='utf-8') as f:
                    f.write(output)
        else:
            print(f"⚠️ 无法从 {ip} 的输出中提取设备名称")
            with file_lock:  # 线程安全写入失败记录
                with open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
                    fail_log.write(ip + '\n')
        # 重置屏幕长度
        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"⚠️ 设备 {ip} 执行命令失败: {cmd_error}")
        with file_lock:  # 线程安全写入失败记录
            with open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
                fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(f"⚠️ 关闭 {ip} 连接时出错: {close_error}")


def fish_interface_optical_cmd(filename, ret_name, max_workers=40):
    """采集接口光功率与CRC信息，支持多线程并行采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    print(f"[START] 开始采集数据，输入文件: {filename}, 输出文件: {ret_name}")
    with open(ret_name, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)  # 使用csv.writer写入
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_optical_device1, ip, user, pwd, writer, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 接口光功率与CRC采集进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
    print("[END] 数据采集完成")


def process_optical_device1(ip, user, pwd, writer, fail_log):
    """处理单个设备的接口光功率与CRC信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        execute_some_command(channel, "screen-length 512", 1)
        laser_output = execute_some_command(channel, "show laser", 5)
        # 清理输出，移除空行和命令回显
        clean_laser = "\n".join([
            line.strip()
            for line in laser_output.split('\n')
            if line.strip() and line.strip() != "show laser"
        ])
        # 使用csv.writer写入，线程安全
        with file_lock:
            writer.writerow([ip, "show laser", clean_laser])

        # 提取接口并获取详细信息
        interfaces = []
        for line in clean_laser.split('\n'):
            line = line.strip()
            if not line or line.startswith(('Interface', '(')):
                continue
            match = re.match(r'^(\S+\s?\d+/\d+/\d+)\s+', line)
            if match:
                intf = match.group(1).replace(' ', '')
                interfaces.append(intf)
                print(f"[DEBUG] 采集到接口: {intf}")

        for intf in interfaces:
            cmd = f"show interface {intf}"
            int_output = execute_some_command(channel, cmd, 3)
            # 清理接口输出
            clean_intf = "\n".join([
                line.strip()
                for line in int_output.split('\n')
                if line.strip() and line.strip() != cmd
            ])
            with file_lock:
                writer.writerow([ip, cmd, clean_intf])

        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def parse_optical_data(device_ip, cmd, output, device_names, optical_data, crc_data):
    """解析光功率与CRC数据"""
    print(f"[parse_optical_data] 设备IP: {device_ip}, 命令: {cmd}")

    # 提取设备名称
    if device_ip not in device_names:
        name_match = re.search(r'<([^>]+)>', output)
        device_names[device_ip] = name_match.group(1) if name_match else "未知设备"

    # 解析光功率数据 (show laser)
    if cmd.strip().lower() == 'show laser':
        lines = output.split('\n')
        header = lines[0].strip() if lines else ""

        # 根据格式定义正则表达式
        if 'temperature' in header.lower():
            optical_pattern = (
                r'^(\S+[\s/-]+\d+/\d+/\d+)\s+'  # 接口名称
                r'\d+\s+'                       # 温度
                r'(-?\d+\.\d+)\s+'              # 接收功率
                r'\[\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]\s+'  # 接收告警范围
                r'\[\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]\s+'  # 接收警告范围
                r'(-?\d+\.\d+)\s+'              # 发送功率
                r'\[\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]\s+'  # 发送告警范围
                r'\[\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]'   # 发送警告范围
            )
        else:
            optical_pattern = (
                r'^(\S+[\s/-]+\d+/\d+/\d+)\s+'  # 接口名称
                r'(-?\d+\.\d+)\s+'              # 接收功率
                r'\[\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]\s+'  # 接收告警范围
                r'\[\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]\s+'  # 接收警告范围
                r'(-?\d+\.\d+)'                 # 发送功率
            )

        interfaces_found = False
        for line in lines[2:]:  # 跳过头部行
            line = line.strip()
            if not line or line.startswith(('<', '(', 'dBm', '--')):
                continue
            match = re.search(optical_pattern, line)
            if match:
                interfaces_found = True
                interface = match.group(1).replace(' ', '')  # 标准化接口名称
                rx_pwr = float(match.group(2))
                rx_alarm_low = float(match.group(3))
                rx_alarm_high = float(match.group(4))
                if 'temperature' in header.lower():
                    tx_pwr = float(match.group(7))
                    tx_alarm_low = float(match.group(8))
                    tx_alarm_high = float(match.group(9))
                else:
                    tx_pwr = float(match.group(7))
                    tx_alarm_low = float(match.group(5))
                    tx_alarm_high = float(match.group(6))

                # 接收功率状态
                if abs(rx_pwr - (-40.0)) < 0.001:
                    rx_status = "收无光"
                else:
                    rx_status = "良好" if rx_alarm_low <= rx_pwr <= rx_alarm_high else (
                        "过弱" if rx_pwr < rx_alarm_low else "过强")

                # 发送功率状态
                tx_status = "良好" if tx_alarm_low <= tx_pwr <= tx_alarm_high else (
                    "过弱" if tx_pwr < tx_alarm_low else "过强")

                # 存储数据
                if device_ip not in optical_data:
                    optical_data[device_ip] = {}
                optical_data[device_ip][interface] = {
                    "rx_pwr": rx_pwr,
                    "rx_alarm_range": f"{rx_alarm_low}~{rx_alarm_high}",
                    "tx_pwr": tx_pwr,
                    "tx_alarm_range": f"{tx_alarm_low}~{tx_alarm_high}",
                    "rx_status": rx_status,
                    "tx_status": tx_status
                }
                print(
                    f"[parse_optical_data] 接口 {interface}: Rx={rx_pwr}, Tx={tx_pwr}")

        if not interfaces_found:
            print(f"[parse_optical_data] 设备 {device_ip} 没有光功率数据")

    # 解析CRC数据 (show interface)
    elif cmd.strip().lower().startswith('show interface'):
        interface = cmd[14:].strip().replace(' ', '')
        crc_match = re.search(r'CRC(?:\s*errors)?\s*:\s*(\d+)', output)
        if device_ip not in crc_data:
            crc_data[device_ip] = {}
        crc_data[device_ip][interface] = crc_match.group(
            1) if crc_match else "N/A"
        print(
            f"[parse_optical_data] 接口 {interface} CRC: {crc_data[device_ip][interface]}")

    return device_names, optical_data, crc_data


def generate_optical_report(src_file, dst_file, host_list_file):
    """生成接口光功率与CRC检查报告"""
    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的failure_ips.tmp文件{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除failure_ips.tmp失败: {str(e)}{Style.RESET_ALL}")
    print(
        f"\n[generate_optical_report] 开始生成报告，源文件: {src_file}, 目标文件: {dst_file}")
    device_names = {}
    optical_data = defaultdict(dict)
    crc_data = defaultdict(dict)
    connection_failures = set()

    # 读取连接失败的设备
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
    except FileNotFoundError:
        print("[generate_optical_report] 未找到failure_ips.tmp文件")

    # 解析原始数据
    with open(src_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        for line_num, row in enumerate(reader, 1):
            if len(row) != 3:
                print(f"[generate_optical_report] 第 {line_num} 行数据格式错误，跳过")
                continue
            device_ip, cmd, output = row
            print(
                f"\n[generate_optical_report] 解析第 {line_num} 行: 设备IP={device_ip}, 命令={cmd}")
            device_names, optical_data, crc_data = parse_optical_data(
                device_ip, cmd, output, device_names, optical_data, crc_data
            )

    # 读取设备清单
    with open(host_list_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]

    # 生成报告
    with open(dst_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow(["设备IP", "设备名称", "接口", "输入光功率", "输入光功率告警范围",
                         "输出光功率", "输出光功率告警阈值", "输入光功率状态", "输出光功率状态", "CRC"])

        for ip in host_ips:
            if ip in connection_failures:
                writer.writerow(
                    [ip, "连接失败", "-", "-", "-", "-", "-", "-", "-", "-"])
            else:
                device_name = device_names.get(ip, "未知设备")
                interfaces = optical_data.get(ip, {})
                if not interfaces:
                    writer.writerow(
                        [ip, device_name, "无光功率数据", "-", "-", "-", "-", "-", "-", "-"])
                else:
                    for intf, data in interfaces.items():
                        crc = crc_data.get(ip, {}).get(intf, "N/A")
                        writer.writerow([
                            ip,
                            device_name,
                            intf,
                            data.get("rx_pwr", "-"),
                            data.get("rx_alarm_range", "-"),
                            data.get("tx_pwr", "-"),
                            data.get("tx_alarm_range", "-"),
                            data.get("rx_status", "-"),
                            data.get("tx_status", "-"),
                            crc
                        ])
    print(f"✅ 报告生成完成，共处理 {len(host_ips)} 台设备")


def fish_optical_cmd(filename, ret_name, max_workers=40):
    """采集光模块性能数据，支持多线程并行采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    print(f"[START] 开始采集数据，输入文件: {filename}, 输出文件: {ret_name}")
    with open(ret_name, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_optical_device, ip, user, pwd, writer, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 光模块性能数据采集进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
    print("[END] 数据采集完成")


def process_optical_device(ip, user, pwd, writer, fail_log):
    """处理单个设备的光模块性能数据采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        execute_some_command(channel, "screen-length 512", 1)
        brief_output = execute_some_command(
            channel, "show interface brief main", 5)
        # 清理输出，移除空行和命令回显
        clean_brief = "\n".join([
            line.strip()
            for line in brief_output.split('\n')
            if line.strip() and line.strip() != "show interface brief main"
        ])
        # 使用csv.writer写入，线程安全
        with file_lock:
            writer.writerow([ip, "show interface brief main", clean_brief])

        # 提取接口列表
        interfaces = []
        for line in clean_brief.split('\n'):
            line = line.strip()
            if not line or line.startswith(('Interface', 'PHY')):
                continue
            match = re.match(r'^(\S+\s?\d+/\d+/\d+)\s+', line)
            if match:
                intf = match.group(1).replace(' ', '')
                interfaces.append(intf)
                print(f"[DEBUG] 采集到接口: {intf}")

        # 采集每个接口的详细信息
        for intf in interfaces:
            cmd = f"show interface {intf}"
            int_output = execute_some_command(channel, cmd, 3)
            clean_intf = "\n".join([
                line.strip()
                for line in int_output.split('\n')
                if line.strip() and line.strip() != cmd
            ])
            with file_lock:
                writer.writerow([ip, cmd, clean_intf])

        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def parse_optical_module_data(device_ip, cmd, output, device_names, optical_module_data):
    """解析光模块性能数据"""
    print(f"[parse_optical_module_data] 设备IP: {device_ip}, 命令: {cmd}")

    # 提取设备名称
    if device_ip not in device_names:
        name_match = re.search(r'<([^>]+)>', output)
        device_names[device_ip] = name_match.group(1) if name_match else "未知设备"

    # 跳过brief命令的解析，因为接口已提取
    if cmd.strip().lower() == 'show interface brief main':
        return device_names, optical_module_data

    # 解析详细接口数据
    elif cmd.strip().lower().startswith('show interface'):
        interface = cmd[14:].strip().replace(' ', '')
        patterns = {
            "vendor_name": r'The Vendor Name : (\S+)',
            "vendor_pn": r'The Vendor PN : (\S+)',
            "transceiver": r'Transceiver Identifier: (\S+)',
            "mode": r'Transceiver Mode: (\S+)',
            "wavelength": r'WaveLength: (\S+)',
            "distance": r'Transmission Distance: (\S+)',
            "rx_power": r'Rx Power: (-?\d+\.\d+)dBm',
            "tx_power": r'Tx Power: (-?\d+\.\d+)dBm',
            "bias": r'Bias: (\S+)',
            "voltage": r'Voltage: (\S+)',
            "temperature": r'temperature: (\S+).*?°C',
            "port_bw": r'Port BW: (\S+)'
        }

        data = {}
        for key, pattern in patterns.items():
            match = re.search(pattern, output)
            data[key] = match.group(1) if match else "N/A"

        if device_ip not in optical_module_data:
            optical_module_data[device_ip] = {}
        optical_module_data[device_ip][interface] = data
        print(f"[parse_optical_module_data] 接口 {interface} 数据已解析: {data}")

    return device_names, optical_module_data


def generate_optical_module_report(src_file, dst_file, host_list_file):
    """生成光模块性能统计报告"""
    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的failure_ips.tmp文件{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除failure_ips.tmp失败: {str(e)}{Style.RESET_ALL}")
    print(
        f"\n[generate_optical_module_report] 开始生成报告，源文件: {src_file}, 目标文件: {dst_file}")
    device_names = {}
    optical_module_data = defaultdict(dict)
    connection_failures = set()

    # 读取连接失败的设备
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
    except FileNotFoundError:
        print("[generate_optical_module_report] 未找到failure_ips.tmp文件")

    # 解析原始数据
    with open(src_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        for line_num, row in enumerate(reader, 1):
            if len(row) != 3:
                print(
                    f"[generate_optical_module_report] 第 {line_num} 行数据格式错误，跳过")
                continue
            device_ip, cmd, output = row
            print(
                f"\n[generate_optical_module_report] 解析第 {line_num} 行: 设备IP={device_ip}, 命令={cmd}")
            device_names, optical_module_data = parse_optical_module_data(
                device_ip, cmd, output, device_names, optical_module_data
            )

    # 读取设备清单
    with open(host_list_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]

    # 生成CSV报告
    with open(dst_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow([
            "设备IP", "设备名称", "接口", "光模块速率", "厂商名称", "型号", "模块类型",
            "工作模式", "波长nm", "传输距离M", "接收光功率dBm", "发送光功率dBm", "偏置电流mV",
            "电压mV", "温度°C"
        ])

        for ip in host_ips:
            if ip in connection_failures:
                writer.writerow([ip, "连接失败", "-", "-", "-", "-",
                                "-", "-", "-", "-", "-", "-", "-", "-"])
            else:
                device_name = device_names.get(ip, "未知设备")
                interfaces = optical_module_data.get(ip, {})
                if not interfaces:
                    writer.writerow(
                        [ip, device_name, "无光模块数据", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"])
                else:
                    for intf, data in interfaces.items():
                        writer.writerow([
                            ip,
                            device_name,
                            intf,
                            data.get("port_bw", "-"),
                            data.get("vendor_name", "-"),
                            data.get("vendor_pn", "-"),
                            data.get("transceiver", "-"),
                            data.get("mode", "-"),
                            data.get("wavelength", "-"),
                            data.get("distance", "-"),
                            data.get("rx_power", "-"),
                            data.get("tx_power", "-"),
                            data.get("bias", "-"),
                            data.get("voltage", "-"),
                            data.get("temperature", "-")
                        ])
    print(f"✅ 报告生成完成，共处理 {len(host_ips)} 台设备")


def fish_custom_cmd(host_file, raw_file, commands, max_workers=40):
    """采集自定义指令数据 - 多线程版本 (Collect Custom Command Data - Multithreaded)"""
    import logging
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style

    logging.basicConfig(filename='custom_cmd.log', level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)s - %(message)s')

    print(f"[START] 开始采集自定义指令数据，输入文件: {host_file}, 输出文件: {raw_file}")
    print(f"[INFO] 自定义指令列表: {commands}")
    print(f"[INFO] 最大线程数: {max_workers}")

    with open(raw_file, "w", encoding='utf-8', newline='') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    ip_to_future = {}

                    for row_idx, row in enumerate(hostip):
                        # 检查行数据完整性
                        if len(row) < 3:
                            print(f"[WARNING] 第 {row_idx+1} 行数据不完整，跳过: {row}")
                            continue

                        ip = row[0].strip()
                        username = row[1].strip()
                        password = row[2].strip()

                        future = executor.submit(
                            process_custom_commands, ip, username, password, commands, writer, fail_log)
                        futures.append(future)
                        ip_to_future[future] = ip

                    with tqdm(total=len(futures), desc="🔍 自定义指令采集进度", unit="台", dynamic_ncols=True) as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result(timeout=180)  # 每个任务最多 180 秒
                            except TimeoutError:
                                logging.error(
                                    f"设备 {ip_to_future[future]} 任务超时")
                                print(
                                    f"{Fore.YELLOW}⚠️ 设备 {ip_to_future[future]} 任务超时{Style.RESET_ALL}")
                            except Exception as e:
                                logging.error(
                                    f"设备 {ip_to_future[future]} 线程执行出错: {str(e)}")
                                print(
                                    f"{Fore.RED}⚠️ 设备 {ip_to_future[future]} 线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except FileNotFoundError as file_error:
            logging.error(f"主机文件读取错误: {file_error}")
            print(f"{Fore.RED}⛔ 主机文件读取错误: {file_error}{Style.RESET_ALL}")
        except Exception as e:
            logging.error(f"数据采集错误: {str(e)}")
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")

    print("[END] 自定义指令数据采集完成")

#


def detect_device_type(channel, ip):
    """检测设备类型，判断是否为A3设备"""
    try:
        # 清空缓冲区
        while channel.recv_ready():
            channel.recv(4096).decode('utf-8', 'ignore')

        # 发送回车获取提示符
        channel.send('\n')
        time.sleep(1)

        output = ""
        if channel.recv_ready():
            output = channel.recv(4096).decode('utf-8', 'ignore')

        with print_lock:
            print(f"[DEBUG] 设备 {ip} 提示符检测: {repr(output[:200])}")

        # 检查是否包含A3标识
        is_a3_device = "MssEdge25-S10-3" in output or "A3-" in output

        with print_lock:
            print(
                f"[DEBUG] 设备 {ip} 检测结果: {'A3设备' if is_a3_device else '非A3设备'}")

        return is_a3_device

    except Exception as e:
        with print_lock:
            print(f"[WARNING] 设备 {ip} 类型检测失败: {e}")
        return False


def process_custom_commands(ip, username, password, commands, writer, fail_log):
    """处理单个设备的自定义指令采集 - 修复版本（保持PC时间功能）"""
    from threading import Lock
    import time
    from datetime import datetime

    # 文件写入锁和打印锁，确保线程安全
    file_lock = Lock()
    channel = None

    try:
        with print_lock:
            print(f"\n[DEBUG] {'='*40}")
            print(f"[DEBUG] 开始处理设备: {ip}")
            print(f"[DEBUG] 尝试连接设备 {ip}...")

        channel = create_channel(ip, username, password)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(f"{ip}\n")
                fail_log.flush()
            with print_lock:
                print(f"[ERROR] 设备 {ip} 连接失败")
            return

        with print_lock:
            print(f"[SUCCESS] 设备 {ip} 连接成功")

        try:
            # 增加初始等待时间，确保设备完全准备好
            time.sleep(2)

            # 清空初始登录信息
            while channel.recv_ready():
                channel.recv(65535).decode('utf-8', 'ignore')
                time.sleep(0.1)

            # 发送一个回车，获取干净的提示符
            channel.send('\n')
            time.sleep(1)

            # 再次清空缓冲区
            while channel.recv_ready():
                channel.recv(65535).decode('utf-8', 'ignore')
                time.sleep(0.1)

            # 检测设备类型
            is_a3_device = detect_device_type(channel, ip)

            # 根据设备类型设置屏幕长度
            if is_a3_device:
                screen_length_cmd = "screen-length 512"
                restore_cmd = "screen-length 25"
            else:
                screen_length_cmd = "screen-length 0"
                restore_cmd = "screen-length 25"

            with print_lock:
                print(f"[DEBUG] 设备 {ip} 设置命令: {screen_length_cmd}")

            execute_some_command(channel, screen_length_cmd,
                                 3, ip=ip, is_a3=is_a3_device)

            # 设置屏幕长度后等待一下
            time.sleep(1)

            # 执行所有自定义命令
            for cmd_idx, cmd in enumerate(commands):
                try:
                    with print_lock:
                        print(
                            f"[COMMAND] 设备 {ip} 执行命令 ({cmd_idx+1}/{len(commands)}): {cmd}")

                    # 在每个命令执行前清空缓冲区
                    while channel.recv_ready():
                        channel.recv(65535).decode('utf-8', 'ignore')
                        time.sleep(0.1)

                    # 在真正开始执行命令时记录PC时间（确保每个设备时间不同）
                    pc_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # 执行命令，如果出现语法错误则重试
                    output = execute_command_with_retry(
                        channel, cmd, ip, is_a3_device)

                    # 验证输出是否包含命令回显
                    if cmd not in output:
                        with print_lock:
                            print(
                                f"[WARNING] 设备 {ip} 命令 '{cmd}' 的输出可能不完整，未找到命令回显")

                    # 清理输出内容
                    clean_output = "\n".join([
                        line.strip()
                        for line in output.split('\n')
                        if line.strip() and line.strip() != cmd
                    ])

                    # 将PC时间附加到输出中（保持PC时间功能）
                    clean_output_with_time = f"PC_TIME: {pc_time}\n{clean_output}"

                    with print_lock:
                        print(
                            f"[OUTPUT] 设备 {ip} 命令 {cmd} 输出长度: {len(clean_output)} 字符")
                        # 显示设备输出内容预览（前1000字符）
                        if clean_output:
                            output_preview = clean_output[:1000]
                            print(f"[OUTPUT-PREVIEW] 设备 {ip} 命令 '{cmd}' 输出预览:")
                            print("-" * 60)
                            print(output_preview)
                            if len(clean_output) > 1000:
                                print(
                                    f"... (还有 {len(clean_output) - 1000} 字符)")
                            print("-" * 60)

                    with file_lock:  # 线程安全写入
                        writer.writerow([ip, cmd, clean_output_with_time])

                    with print_lock:
                        print(f"[DEBUG] 设备 {ip} 命令 {cmd} 处理完成")

                    # 命令间增加延迟，避免设备处理不及
                    time.sleep(1)

                except Exception as single_cmd_error:
                    error_msg = f"命令执行失败: {str(single_cmd_error)}"
                    with print_lock:
                        print(
                            f"[ERROR] 设备 {ip} 命令 '{cmd}' 执行异常: {str(single_cmd_error)}")
                    with file_lock:
                        writer.writerow([ip, cmd, error_msg])

            # 恢复屏幕长度设置
            try:
                with print_lock:
                    print(f"[DEBUG] 设备 {ip} 恢复命令: {restore_cmd}")
                execute_some_command(channel, restore_cmd,
                                     3, ip=ip, is_a3=is_a3_device)
            except Exception as restore_error:
                with print_lock:
                    print(f"[WARNING] 设备 {ip} 恢复屏幕长度设置失败: {restore_error}")

        except Exception as cmd_error:
            with print_lock:
                print(f"[ERROR] 设备 {ip} 命令执行全局异常: {str(cmd_error)}")
            with file_lock:
                fail_log.write(f"{ip} - {str(cmd_error)}\n")
                fail_log.flush()

    except Exception as connection_error:
        with print_lock:
            print(f"[ERROR] 设备 {ip} 连接异常: {str(connection_error)}")
        with file_lock:
            fail_log.write(f"{ip} - {str(connection_error)}\n")
            fail_log.flush()

    finally:
        # 确保连接被正确关闭
        if channel:
            try:
                with print_lock:
                    print(f"[DEBUG] 设备 {ip} 尝试关闭连接")
                channel.close()
                with print_lock:
                    print(f"[DEBUG] 设备 {ip} 连接已关闭")
            except Exception as close_error:
                with print_lock:
                    print(f"[WARNING] 关闭设备 {ip} 连接时出错: {close_error}")


def execute_some_command(channel, command, timeout=5, max_retries=3, command_delay=1.5, device_name="", ip="", is_a3=False):
    """
    执行命令并正确处理分页输出 - 修复版本，确保命令和输出正确对应
    """
    import select
    import socket
    import sys
    import logging
    import time
    from colorama import Fore, Style

    if not channel:
        return ""

    # 在发送命令前添加延迟
    time.sleep(command_delay)

    for attempt in range(1, max_retries + 1):
        try:
            with print_lock:
                print(f"[DEBUG] 设备 {ip} 尝试第 {attempt} 次执行命令: {command}")

            # 步骤1: 彻底清空缓冲区，包括发送回车并等待
            cleared_data = ""
            clear_count = 0

            # 先发送一个回车，获取一个干净的提示符
            channel.send('\n')
            time.sleep(0.5)

            # 清空所有缓冲数据
            while channel.recv_ready():
                data = channel.recv(65535).decode('utf-8', 'ignore')
                cleared_data += data
                clear_count += 1
                if clear_count > 50:  # 防止无限循环
                    break
                time.sleep(0.1)

            if cleared_data:
                with print_lock:
                    print(f"[DEBUG] 设备 {ip} 清空缓冲区，数据长度: {len(cleared_data)}")

            # 步骤2: 发送命令并立即加上换行符
            channel.send((command + '\n').encode('utf-8'))

            # 步骤3: 等待命令被设备接收和处理
            # A3设备需要更长的等待时间
            initial_wait = 2.0 if is_a3 else 1.0
            time.sleep(initial_wait)

            output = ""
            command_echo_found = False
            start_time = time.time()
            last_data_time = time.time()
            pagination_count = 0
            consecutive_no_data = 0

            # 步骤4: 确保捕获到命令回显，这样才能确保后续是命令的输出
            while time.time() - start_time < timeout:
                readable, _, _ = select.select([channel], [], [], 0.3)

                if readable:
                    try:
                        # 使用更大的缓冲区
                        data = channel.recv(65535).decode('utf-8', 'ignore')

                        if data:
                            output += data
                            last_data_time = time.time()
                            consecutive_no_data = 0

                            # 检查是否包含命令回显
                            if not command_echo_found and command in output:
                                command_echo_found = True
                                with print_lock:
                                    print(f"[DEBUG] 设备 {ip} 找到命令回显: {command}")

                            # 简化调试输出：只在有分页符时显示
                            if '----MORE----' in data:
                                with print_lock:
                                    print(
                                        f"[DEBUG] 设备 {ip} 检测到分页符，数据长度: {len(data)}")

                            # 处理A3设备的分页符
                            if is_a3:
                                while '----MORE----' in data:
                                    pagination_count += 1
                                    if pagination_count % 10 == 1:  # 每10次显示一次进度
                                        with print_lock:
                                            print(
                                                f"[DEBUG] 设备 {ip} 检测到A3分页符 (第{pagination_count}次)")

                                    # 发送空格继续
                                    channel.send(' ')
                                    time.sleep(0.5)  # A3设备需要更长的等待时间

                                    # 等待并读取更多数据
                                    wait_start = time.time()
                                    new_data = ""
                                    while time.time() - wait_start < 3:  # 等待最多3秒
                                        if channel.recv_ready():
                                            new_data = channel.recv(
                                                65535).decode('utf-8', 'ignore')
                                            if new_data:
                                                output += new_data
                                                data = new_data  # 更新data以继续检查分页
                                                break
                                        time.sleep(0.1)

                                    if not new_data:
                                        with print_lock:
                                            print(
                                                f"[DEBUG] 设备 {ip} 分页后无新数据，退出分页处理")
                                        break

                                    # 防止无限分页
                                    if pagination_count > 200:
                                        with print_lock:
                                            print(
                                                f"[WARNING] 设备 {ip} 分页次数过多({pagination_count})，强制结束")
                                        channel.send('\x03')  # Ctrl+C
                                        time.sleep(1)
                                        break
                            else:
                                # 非A3设备的其他分页符处理保持原样
                                while '----MORE----' in data or '--More--' in data or '-- More --' in data:
                                    pagination_count += 1
                                    if pagination_count % 10 == 1:  # 每10次显示一次
                                        with print_lock:
                                            print(
                                                f"[DEBUG] 设备 {ip} 检测到分页符 (第{pagination_count}次)")

                                    channel.send(' ')
                                    time.sleep(0.2)

                                    wait_start = time.time()
                                    new_data = ""
                                    while time.time() - wait_start < 2:
                                        if channel.recv_ready():
                                            new_data = channel.recv(
                                                65535).decode('utf-8', 'ignore')
                                            if new_data:
                                                output += new_data
                                                data = new_data
                                                break
                                        time.sleep(0.1)

                                    if not new_data:
                                        break

                                    if pagination_count > 100:
                                        with print_lock:
                                            print(
                                                f"[WARNING] 设备 {ip} 分页次数过多，强制结束")
                                        channel.send('\x03')
                                        time.sleep(1)
                                        break

                            # 检查主备切换
                            if "Start Switching" in output:
                                error_msg = f"⚠️⚠️⚠️ 检测到设备 {device_name}({ip}) 发生主备切换! 终止操作! ⚠️⚠️⚠️"
                                for i in range(3):
                                    print(
                                        f"{Fore.RED}{error_msg}{Style.RESET_ALL}")
                                    logging.critical(error_msg)
                                sys.exit(1)

                    except socket.error as e:
                        with print_lock:
                            print(f"[WARNING] 设备 {ip} 读取数据错误: {e}")
                        break
                else:
                    consecutive_no_data += 1

                    # 步骤5: 改进命令完成检测逻辑
                    if command_echo_found and output.strip():  # 确保已经找到命令回显
                        lines = output.split('\n')
                        if len(lines) > 2:
                            # 检查最后几行是否有提示符
                            for line in lines[-5:]:  # 增加检查行数
                                line_stripped = line.strip()
                                if line_stripped and (line_stripped.endswith('>') or
                                                      line_stripped.endswith('#') or
                                                      line_stripped.endswith('$')):
                                    # 确保最近没有分页提示
                                    recent_output = output[-1000:] if len(
                                        output) > 1000 else output
                                    if ('----MORE----' not in recent_output and
                                        '--More--' not in recent_output and
                                            '-- More --' not in recent_output):
                                        with print_lock:
                                            print(
                                                f"[DEBUG] 设备 {ip} 检测到命令完成，提示符: {repr(line_stripped)}")
                                        return output

                    # 检查是否超时无新数据
                    if consecutive_no_data > 15:  # 连续15次没有数据 (约4.5秒)
                        with print_lock:
                            print(f"[DEBUG] 设备 {ip} 连续无新数据，结束命令执行")
                        break

            # 步骤6: 验证输出的有效性
            if not command_echo_found:
                with print_lock:
                    print(f"[WARNING] 设备 {ip} 未找到命令回显，可能输出不完整")
                # 如果没有找到命令回显，尝试重新执行
                if attempt < max_retries:
                    time.sleep(3)
                    continue

            # 最后检查输出
            with print_lock:
                print(f"[DEBUG] 设备 {ip} 命令执行完成，输出长度: {len(output)}")

            return output

        except Exception as ex:
            with print_lock:
                print(
                    f"[ERROR] 设备 {ip} 执行命令异常 (尝试{attempt}/{max_retries}): {ex}")
            if attempt < max_retries:
                time.sleep(3)
                continue
            return f"命令执行错误: {ex}"

    return output if 'output' in locals() else ""

#


def generate_custom_cmd_report(raw_file, report_file, host_file):
    """生成自定义指令报告 (Generate Custom Command Report) - 排除连接失败的设备"""
    print(
        f"\n🐛 [DEBUG] 进入 generate_custom_cmd_report 函数，参数: raw_file={raw_file}, report_file={report_file}, host_file={host_file}")

    connection_failures = set()
    try:
        print(f"🐛 [DEBUG] 正在读取连接失败记录 failure_ips.tmp")
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            for line in f:
                ip = line.strip()  # 修改：直接使用strip()，因为现在只记录IP
                if ip:
                    connection_failures.add(ip)
            print(f"🐛 [DEBUG] 读取到 {len(connection_failures)} 个连接失败的IP")
    except FileNotFoundError:
        print(f"🐛 [DEBUG] 未找到 failure_ips.tmp 文件，跳过连接失败记录")

    # 读取主机列表
    host_ips = []
    try:
        print(f"🐛 [DEBUG] 正在读取主机列表文件: {host_file}")
        with open(host_file, "r", encoding='gbk', errors='ignore') as f:
            reader = csv.reader(f)
            host_ips = [row[0].strip() for row in reader if len(row) > 0]
            print(f"🐛 [DEBUG] 主机列表共 {len(host_ips)} 台设备")
    except Exception as e:
        print(f"🐛 [DEBUG] ⚠️ 读取主机文件异常: {e}")
        return

    # 过滤掉连接失败的设备，只分析成功连接的设备
    successful_host_ips = [
        ip for ip in host_ips if ip not in connection_failures]
    print(
        f"🐛 [DEBUG] 过滤后，参与分析的设备数量: {len(successful_host_ips)} 台 (排除了 {len(connection_failures)} 台连接失败的设备)")

    # 读取原始数据
    data = []
    try:
        print(f"🐛 [DEBUG] 正在读取原始数据文件: {raw_file}")
        with open(raw_file, "r", encoding='utf-8') as f:
            reader = csv.reader(f)
            data = list(reader)
            print(f"🐛 [DEBUG] 读取到 {len(data)} 条原始数据记录")
    except Exception as e:
        print(f"🐛 [DEBUG] ⚠️ 读取原始数据文件异常: {e}")
        return

    # 生成报告 - 只包含成功连接的设备
    try:
        print(f"🐛 [DEBUG] 正在生成报告文件: {report_file}")
        with open(report_file, "w", encoding='utf-8', newline='') as report:
            writer = csv.writer(report)
            writer.writerow(["设备IP", "设备名称", "运行指令", "执行状态", "设备输出"])
            processed_ips = set()
            processed_count = 0

            print(f"🐛 [DEBUG] 开始处理原始数据...")
            for idx, row in enumerate(data):
                if len(row) != 3:
                    print(f"🐛 [DEBUG] 第 {idx+1} 行数据格式异常，跳过处理: {row}")
                    continue

                device_ip, cmd, output = row

                # 跳过连接失败的设备记录
                if device_ip in connection_failures:
                    print(f"🐛 [DEBUG] 跳过连接失败设备 {device_ip} 的记录")
                    continue

                print(
                    f"🐛 [DEBUG] 正在处理 {device_ip} 的第 {idx+1} 条记录，命令: {cmd[:20]}...")

                # 安全地显示输出内容
                output_preview = output[:800] if len(output) > 800 else output
                print(f"🐛 [DEBUG] {device_ip} 输出内容: {output_preview}...")

                processed_ips.add(device_ip)

                # 从输出中提取设备名称
                device_name = "未知设备"
                try:
                    name_match = re.search(
                        r'^\[([^\]]+)\]', output, re.MULTILINE)
                    if name_match:
                        device_name = name_match.group(1).strip()
                        print(f"🐛 [DEBUG] 从输出中提取设备名称成功: {device_name}")
                    else:
                        print(f"🐛 [DEBUG] 未匹配到设备名称，使用默认值")
                except Exception as name_error:
                    print(f"🐛 [DEBUG] ⚠️ 提取设备名称时出错: {name_error}")

                # 判断执行状态
                status = "执行成功"
                if any(keyword in output.lower() for keyword in ["error", "failed", "timeout", "连接失败", "执行失败", "连接异常"]):
                    status = "执行失败"
                    print(f"🐛 [DEBUG] {device_ip} 的命令执行状态为失败")
                elif output.startswith(("执行失败", "连接失败", "连接异常")):
                    status = "执行失败"
                    print(f"🐛 [DEBUG] {device_ip} 的命令执行状态为失败（异常记录）")
                else:
                    print(f"🐛 [DEBUG] {device_ip} 的命令执行状态为成功")

                writer.writerow([device_ip, device_name, cmd, status, output])
                processed_count += 1
                print(f"🐛 [DEBUG] 已写入报告第 {processed_count} 行数据")

            # 检查是否有成功连接但没有在原始数据中的设备（理论上不应该出现）
            missing_successful_devices = [
                ip for ip in successful_host_ips if ip not in processed_ips]
            if missing_successful_devices:
                print(
                    f"🐛 [DEBUG] ⚠️ 发现 {len(missing_successful_devices)} 台成功连接但没有数据的设备: {missing_successful_devices}")
                for ip in missing_successful_devices:
                    writer.writerow(
                        [ip, "数据异常", "-", "数据缺失", "设备连接成功但无命令执行数据"])
                    processed_count += 1

        print(f"✅ 自定义指令报告生成完成")
        print(f"📊 统计信息:")
        print(f"   - 总设备数量: {len(host_ips)} 台")
        print(f"   - 连接失败设备: {len(connection_failures)} 台")
        print(f"   - 参与分析设备: {len(successful_host_ips)} 台")
        print(f"   - 实际处理记录: {processed_count} 条")

    except Exception as e:
        print(f"🐛 [DEBUG] ⚠️ 生成报告时异常: {e}")
        print(f"⛔ 报告生成错误: {e}")


def create_channel(ip, username, password, port=22, timeout=10, retry_count=3, retry_delay=2):
    """
    创建SSH通道连接，增加了重试机制和更好的错误处理

    Args:
        ip: 设备IP地址
        username: 用户名
        password: 密码
        port: SSH端口，默认22
        timeout: 连接超时时间(秒)
        retry_count: 重试次数
        retry_delay: 重试间隔(秒)

    Returns:
        成功返回SSH通道，失败返回None
    """
    client = None
    for attempt in range(1, retry_count + 1):
        try:
            with print_lock:
                print(
                    f"\n{Fore.CYAN}🔄 正在连接设备 {ip} (尝试 {attempt}/{retry_count})...{Style.RESET_ALL}")
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

            client.connect(
                hostname=ip,
                port=port,
                username=username,
                password=password,
                timeout=timeout,
                allow_agent=False,
                look_for_keys=False,
                banner_timeout=timeout
            )

            channel = client.invoke_shell()
            channel.settimeout(timeout)
            if channel.recv_ready():
                _ = channel.recv(4096).decode('utf-8', 'ignore')

            with print_lock:
                print(f"\n{Fore.GREEN}✅ 设备 {ip} 连接成功{Style.RESET_ALL}")
            return channel

        # except paramiko.AuthenticationException:
        #     client_close(client)
        #     with print_lock:
        #         print(f"\n{Fore.RED}❌ 设备 {ip} 认证失败 - 用户名或密码错误{Style.RESET_ALL}")
        #     logging.error(f"设备 {ip} 认证失败 - 用户名或密码错误")
        #     raise ValueError("认证失败")

        except paramiko.SSHException as ssh_ex:
            client_close(client)
            with print_lock:
                print(
                    f"\n{Fore.YELLOW}⚠️ 设备 {ip} SSH异常: {ssh_ex}{Style.RESET_ALL}")
            logging.warning(f"设备 {ip} SSH异常: {ssh_ex}")

        except socket.timeout:
            client_close(client)
            with print_lock:
                print(
                    f"\n{Fore.YELLOW}⌛ [连接响应超时] {ip} 请检查网络或设备负载{Style.RESET_ALL}")
            logging.warning(f"设备 {ip} 连接超时")

        except socket.error as sock_ex:
            client_close(client)
            with print_lock:
                print(f"\n{Fore.RED}🌐 设备 {ip} 网络错误: {sock_ex}{Style.RESET_ALL}")
            logging.error(f"设备 {ip} 网络错误: {sock_ex}")

        except Exception as ex:
            client_close(client)
            with print_lock:
                print(f"\n{Fore.RED}❗ 设备 {ip} 连接异常: {ex}{Style.RESET_ALL}")
            logging.error(f"设备 {ip} 连接异常: {ex}")

        if attempt < retry_count:
            retry_time = retry_delay * attempt
            with print_lock:
                print(f"\n{Fore.CYAN}⏳ 等待{retry_time}秒后重试...{Style.RESET_ALL}")
            time.sleep(retry_time)
        else:
            logging.error(f"设备 {ip} 连接失败，已达到最大重试次数")
            with print_lock:
                print(f"\n{Fore.RED}🚫 设备 {ip} 连接失败，已达到最大重试次数{Style.RESET_ALL}")

    return None


def execute_command_with_retry(channel, command, ip, is_a3, max_command_retries=3):
    """
    执行命令并检测语法错误，如果出现 'Invalid input detected' 错误则重试

    Args:
        channel: SSH通道
        command: 要执行的命令
        ip: 设备IP地址
        is_a3: 是否为A3设备
        max_command_retries: 最大重试次数，默认3次

    Returns:
        命令执行结果字符串
    """
    import time

    for retry_attempt in range(1, max_command_retries + 1):
        try:
            with print_lock:
                if retry_attempt > 1:
                    print(
                        f"[RETRY] 设备 {ip} 重试执行命令 (第{retry_attempt}次): {command}")

            # 执行命令
            output = execute_some_command(
                channel, command, 30, ip=ip, is_a3=is_a3)

            # 检查是否出现语法错误
            if "ERROR:  Invalid input detected at '^' marker" in output:
                with print_lock:
                    print(f"[ERROR] 设备 {ip} 命令语法错误，尝试重新执行: {command}")
                    print(f"[ERROR-OUTPUT] 错误输出: {output[:500]}")

                if retry_attempt < max_command_retries:
                    # 清理缓冲区后重试
                    with print_lock:
                        print(f"[DEBUG] 设备 {ip} 清理缓冲区，准备重试...")

                    time.sleep(1)

                    # 清空缓冲区
                    clear_count = 0
                    while channel.recv_ready():
                        try:
                            channel.recv(65535).decode('utf-8', 'ignore')
                            clear_count += 1
                            if clear_count > 20:  # 防止无限循环
                                break
                        except:
                            break
                        time.sleep(0.1)

                    with print_lock:
                        print(f"[DEBUG] 设备 {ip} 缓冲区清理完成，清理了 {clear_count} 次数据")

                    continue
                else:
                    with print_lock:
                        print(f"[ERROR] 设备 {ip} 命令重试次数已达上限，返回错误结果")
                    return output
            else:
                # 命令执行成功，返回结果
                if retry_attempt > 1:
                    with print_lock:
                        print(f"[SUCCESS] 设备 {ip} 命令重试成功: {command}")
                return output

        except Exception as e:
            with print_lock:
                print(
                    f"[ERROR] 设备 {ip} 命令执行异常 (重试{retry_attempt}/{max_command_retries}): {e}")

            if retry_attempt < max_command_retries:
                time.sleep(2)
                continue
            else:
                return f"命令执行错误: {e}"

    # 如果所有重试都失败，返回最后的输出或错误信息
    return output if 'output' in locals() else "命令执行失败：所有重试均失败"


def fish_device_info_cmd(host_file, raw_file, max_workers=40):
    import logging
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    logging.basicConfig(filename='device_info.log', level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)s - %(message)s')

    print(f"[START] 开始采集设备信息数据，输入文件: {host_file}, 输出文件: {raw_file}")
    commands = [
        "show device", "show temperature", "show mpls l2vc brief",
        "show ldp session", "show running-configuration include .31",
        "show ospf neighbor brief", "show voltage", "show bfd session brief"
    ]

    with open(raw_file, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    ip_to_future = {}
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        future = executor.submit(
                            process_device_info, ip, user, pwd, commands, writer, fail_log)
                        futures.append(future)
                        ip_to_future[future] = ip

                    with tqdm(total=total_devices, desc="🔍 设备信息采集进度", unit="台", dynamic_ncols=True) as pbar:
                        for future in as_completed(futures):  # 移除总超时
                            try:
                                future.result(timeout=120)  # 每个任务最多 120 秒
                            except TimeoutError:
                                logging.error(
                                    f"设备 {ip_to_future[future]} 任务超时")
                            except Exception as e:
                                logging.error(
                                    f"设备 {ip_to_future[future]} 线程执行出错: {str(e)}")
                            pbar.update(1)

        except Exception as e:
            logging.error(f"数据采集错误: {str(e)}")
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
    print("[END] 数据采集完成")


def process_device_info(ip, user, pwd, commands, writer, fail_log):
    """处理单个设备的设备信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from collections import defaultdict
    from threading import Lock

    # 文件写入锁和打印锁，确保线程安全
    file_lock = Lock()
    channel = None

    try:
        with print_lock:
            print(f"\n[DEBUG] {'='*40}")
            print(f"[DEBUG] 开始处理设备: {ip}")
            print(f"[DEBUG] 尝试连接设备 {ip}...")

        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:
                fail_log.write(ip + '\n')
            with print_lock:
                print(f"[ERROR] 设备 {ip} 连接失败")
            return

        with print_lock:
            print(f"[SUCCESS] 设备 {ip} 连接成功")
            print(f"[DEBUG] 设置 screen-length 512...")
        execute_some_command(channel, "screen-length 512", 1)

        for cmd in commands:
            with print_lock:
                print(f"[COMMAND] 执行命令: {cmd}")
            output = execute_some_command(channel, cmd, 5)
            with print_lock:
                print(f"[OUTPUT] 命令 {cmd} 输出长度: {len(output)} 字符")

            clean_output = "\n".join([
                line.strip()
                for line in output.split('\n')
                if line.strip() and line.strip() != cmd
            ])
            with file_lock:
                writer.writerow([ip, cmd, clean_output])
            with print_lock:
                print(f"[DEBUG] 命令 {cmd} 处理完成")

        with print_lock:
            print(f"[DEBUG] 恢复 screen-length 25...")
        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        with print_lock:
            print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:
            fail_log.write(ip + '\n')
            for cmd in commands:
                writer.writerow([ip, cmd, f"执行失败: {cmd_error}"])
    finally:
        if channel:
            try:
                channel.close()
                with print_lock:
                    print(f"[DEBUG] 设备 {ip} 连接已关闭")
            except Exception as close_error:
                with print_lock:
                    print(
                        f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def generate_device_info_report(raw_file, report_file, host_file):
    """生成设备状态统计报告"""
    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的failure_ips.tmp文件{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除failure_ips.tmp失败: {str(e)}{Style.RESET_ALL}")
    print(f"\n[START] 开始生成报告，源文件: {raw_file}, 目标文件: {report_file}")

    # 加载连接失败的设备
    connection_failures = set()
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
            print(f"[DEBUG] 加载失败设备列表成功，共 {len(connection_failures)} 台")
    except FileNotFoundError:
        print("[INFO] 未找到failure_ips.tmp文件")

    # 加载主机列表
    with open(host_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(f"[DEBUG] 加载主机列表成功，共 {len(host_ips)} 台")

    # 加载原始数据
    with open(raw_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        data = list(reader)
        print(f"[DEBUG] 加载原始数据成功，共 {len(data)} 条记录")

    # 按设备IP分组数据
    device_data = defaultdict(dict)
    for row in data:
        if len(row) != 3:
            continue
        ip, cmd, output = row
        device_data[ip][cmd] = output
    print(f"[DEBUG] 数据分组完成，共 {len(device_data)} 台有效设备")

    # 生成CSV报告
    with open(report_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        headers = [
            "设备IP", "设备名称", "设备MAC", "设备版本", "温度℃", "风扇速度百分比", "CPU使用率",
            "十五分钟内CPU使用率", "内存使用率", "电压V", "vc总数", "vc-up", "vc-down",
            "ldp会话总数", "ldp会话up", "ldp会话down", "ospf总数", "ospf-up", "ospf-down",
            "bfd总数", "bfd-up", "bfd-down"
        ]
        writer.writerow(headers)

        for ip in host_ips:
            print(f"\n[PROCESS] {'-'*40}")
            print(f"[PROCESS] 正在处理设备: {ip}")

            if ip in connection_failures:
                print(f"[SKIP] 设备 {ip} 连接失败，写入连接失败行")
                writer.writerow([ip] + ["连接失败"] * (len(headers) - 1))
                continue

            device_info = device_data.get(ip, {})
            print(f"[DEBUG] 找到 {len(device_info)} 条命令记录")

            # 解析每个命令输出
            print("[STEP] 解析 show device...")
            show_device = parse_show_device(
                device_info.get("show device", "执行失败"))

            print("[STEP] 解析 show temperature...")
            temperature = parse_show_temperature(
                device_info.get("show temperature", "执行失败"))

            print("[STEP] 解析 show voltage...")
            voltage = parse_show_voltage(
                device_info.get("show voltage", "执行失败"))

            print("[STEP] 解析 show mpls l2vc brief...")
            vc_info = parse_show_mpls_l2vc_brief(
                device_info.get("show mpls l2vc brief", "执行失败"))

            print("[STEP] 解析 show ldp session...")
            ldp_info = parse_show_ldp_session(
                device_info.get("show ldp session", "执行失败"))

            print("[STEP] 解析 show running-configuration include .31...")
            ospf_total = parse_show_run_include_31(device_info.get(
                "show running-configuration include .31", "执行失败"))

            print("[STEP] 解析 show ospf neighbor brief...")
            ospf_info = parse_show_ospf_neighbor_brief(
                device_info.get("show ospf neighbor brief", "执行失败"))

            print("[STEP] 解析 show bfd session brief...")
            bfd_info = parse_show_bfd_session_brief(
                device_info.get("show bfd session brief", "执行失败"))

            # 格式化数值
            def format_percent(value):
                return f"{value}%" if value != "连接失败" and value != "N/A" else value

            cpu_usage = format_percent(show_device.get("cpu_usage", "N/A"))
            fifteen_min_cpu = format_percent(
                show_device.get("fifteen_min_cpu_usage", "N/A"))
            memory_usage = format_percent(
                show_device.get("memory_usage", "N/A"))

            # 编译行数据
            row = [
                ip,
                show_device.get("device_name", "N/A"),
                show_device.get("system_mac", "N/A"),
                show_device.get("device_version", "N/A"),
                temperature,
                show_device.get("fan_speed", "N/A"),
                cpu_usage,
                fifteen_min_cpu,
                memory_usage,
                voltage,
                vc_info.get("vc_total", "N/A"),
                vc_info.get("vc_up", "N/A"),
                vc_info.get("vc_down", "N/A"),
                ldp_info.get("ldp_total", "N/A"),
                ldp_info.get("ldp_up", "N/A"),
                ldp_info.get("ldp_down", "N/A"),
                ospf_total,
                ospf_info.get("ospf_up", "N/A"),
                ospf_info.get("ospf_down", "N/A"),
                bfd_info.get("bfd_total", "N/A"),
                bfd_info.get("bfd_up", "N/A"),
                bfd_info.get("bfd_down", "N/A")
            ]

            print("[DEBUG] 生成行数据:", row)
            writer.writerow(row)

    print(f"✅ [SUCCESS] 设备信息报告生成完成，共处理 {len(host_ips)} 台设备")


def parse_show_device(output):
    """Parse 'show device' output for device info"""
    print("\n[DEBUG] 开始解析 show device 输出")
    if output.startswith("执行失败"):
        print("[WARN] 命令执行失败，返回默认值")
        return {key: "N/A" for key in ['device_name', 'system_mac', 'device_version',
                                       'fan_speed', 'cpu_usage',
                                       'fifteen_min_cpu_usage', 'memory_usage']}

    data = {}
    # Device Name
    name_match = re.search(r'<([^>]+)>', output)
    data['device_name'] = name_match.group(1) if name_match else "未知设备"
    print(
        f"[DEBUG] 解析设备名称成功: {data['device_name']}" if name_match else "[WARN] 未找到设备名称")

    # System MAC
    mac_match = re.search(r'System-MAC:\s*([\w.:]+)', output)
    data['system_mac'] = mac_match.group(1) if mac_match else "N/A"
    print(
        f"[DEBUG] 解析MAC地址成功: {data['system_mac']}" if mac_match else "[WARN] 未找到MAC地址")

    # Device Version
    version_match = re.search(r'system info\s*:\s*\S+\s*\(([\w]+)\)', output)
    data['device_version'] = version_match.group(1) if version_match else "N/A"
    print(
        f"[DEBUG] 解析设备版本成功: {data['device_version']}" if version_match else "[WARN] 未找到设备版本")

    # Fan Speed (fan #01)
    fan_match = re.search(r'\[fan #01\]\s+(\d+)%', output)
    data['fan_speed'] = fan_match.group(1) if fan_match else "N/A"
    print(
        f"[DEBUG] 解析风扇速度成功: {data['fan_speed']}%" if fan_match else "[WARN] 未找到风扇信息")

    # CPU Usage
    cpu_match = re.search(r'CPU Usage\s*:\s*(\d+)%', output)
    data['cpu_usage'] = cpu_match.group(1) if cpu_match else "N/A"
    print(
        f"[DEBUG] 解析CPU使用率成功: {data['cpu_usage']}%" if cpu_match else "[WARN] 未找到CPU使用率")

    # 15-minute CPU Usage
    fifteen_min_match = re.search(r'fifteen  minutes :\s*(\d+)%', output)
    data['fifteen_min_cpu_usage'] = fifteen_min_match.group(
        1) if fifteen_min_match else "N/A"
    print(
        f"[DEBUG] 解析15分钟CPU使用率成功: {data['fifteen_min_cpu_usage']}%" if fifteen_min_match else "[WARN] 未找到15分钟CPU使用率")

    # Memory Usage
    memory_match = re.search(r'Memory Using Percentage :\s*(\d+)%', output)
    data['memory_usage'] = memory_match.group(1) if memory_match else "N/A"
    print(
        f"[DEBUG] 解析内存使用率成功: {data['memory_usage']}%" if memory_match else "[WARN] 未找到内存使用率")

    return data


def parse_show_bfd_session_brief(output):
    """Parse 'show bfd session brief' output"""
    print("\n[DEBUG] 开始解析 show bfd session brief 输出")
    if output.startswith("执行失败"):
        print("[WARN] 命令执行失败")
        return {'bfd_total': "N/A", 'bfd_up': "N/A", 'bfd_down': "N/A"}

    match = re.search(
        r'Number of sessions:\s*Sum:\s*(\d+)\s*Up:\s*(\d+)\s*Down&Init:\s*(\d+)', output)
    if match:
        bfd_info = {
            'bfd_total': match.group(1),
            'bfd_up': match.group(2),
            'bfd_down': match.group(3)
        }
        print(
            f"[DEBUG] 解析BFD信息成功: 总数={bfd_info['bfd_total']}, Up={bfd_info['bfd_up']}, Down={bfd_info['bfd_down']}")
        return bfd_info
    print("[WARN] 未找到BFD会话信息")
    return {'bfd_total': "N/A", 'bfd_up': "N/A", 'bfd_down': "N/A"}


def parse_show_temperature(output):
    """解析 'show temperature' 输出"""
    print("\n[DEBUG] 开始解析 show temperature 输出")
    if output.startswith("执行失败"):
        print("[WARN] 命令执行失败")
        return "N/A"

    # 定义正则表达式模式
    four_field_pattern = r'^\s*\d+\s+\d+\s+\d+\s+(\d+)\s*$'  # 四字段，温度在第4列
    # 七字段，温度在第2列
    seven_field_pattern = r'^\s*\d+\s+(\d+)\s+\d+\s+\d+\s+(?:\d+|\-\-)\s+(?:\d+|\-\-)\s+(?:\d+|\-\-)\s*$'
    combined_pattern = f'{four_field_pattern}|{seven_field_pattern}'

    # 按行分割输出
    lines = output.splitlines()
    temperatures = []

    # 调试每一行
    for line in lines:
        match = re.match(combined_pattern, line)
        if match:
            if match.group(1):  # 四字段格式的温度
                temp = int(match.group(1))
                print(f"[DEBUG] 匹配四字段格式: {line}, 温度={temp}℃")
                temperatures.append(temp)
            elif match.group(2):  # 七字段格式的温度
                temp = int(match.group(2))
                print(f"[DEBUG] 匹配七字段格式: {line}, 温度={temp}℃")
                temperatures.append(temp)
        else:
            print(f"[DEBUG] 未匹配: {line}")

    # 如果找到温度，返回最大值
    if temperatures:
        max_temperature = max(temperatures)
        print(f"[DEBUG] 解析温度成功: {max_temperature}℃ (最高温度)")
        return str(max_temperature)
    else:
        print("[WARN] 未找到温度信息")
        return "N/A"


def parse_show_voltage(output):
    """Parse 'show voltage' output"""
    print("\n[DEBUG] 开始解析 show voltage 输出")
    if output.startswith("执行失败"):
        print("[WARN] 命令执行失败")
        return "N/A"

    # 匹配12或13槽的正则表达式，精确匹配行首
    voltage_pattern = re.compile(
        r'^(12|13)\s+\d+\s+\d+\s+\d+\s+(\d+)\s+\d+\.\d+', re.MULTILINE)
    slots_voltage = {}

    for match in voltage_pattern.finditer(output):
        slot = match.group(1)
        voltage_raw = match.group(2)
        slots_voltage[slot] = voltage_raw
        print(f"[DEBUG] 找到槽位 {slot} 的电压值: {voltage_raw}mV")

    # 优先选择12槽，其次13槽
    voltage_raw = None
    if '12' in slots_voltage:
        voltage_raw = slots_voltage['12']
        print("[DEBUG] 使用槽位12的电压值")
    elif '13' in slots_voltage:
        voltage_raw = slots_voltage['13']
        print("[DEBUG] 使用槽位13的电压值")
    else:
        print("[WARN] 未找到12或13槽的电压信息")
        return "N/A"

    try:
        voltage = int(voltage_raw) / 1000  # 转换为V并保留1位小数
        voltage_str = f"{voltage:.1f}"
        print(f"[DEBUG] 转换后电压值: {voltage_str}V")
        return voltage_str+'V'
    except ValueError:
        print("[ERROR] 电压值转换失败")
        return "N/A"


def parse_show_mpls_l2vc_brief(output):
    """Parse 'show mpls l2vc brief' output"""
    if output.startswith("执行失败"):
        return {'vc_total': "N/A", 'vc_up': "N/A", 'vc_down': "N/A"}
    vc_match = re.search(r'Total LDP VC : (\d+), (\d+) up, (\d+) down', output)
    if vc_match:
        return {
            'vc_total': vc_match.group(1),
            'vc_up': vc_match.group(2),
            'vc_down': vc_match.group(3)
        }
    return {'vc_total': "N/A", 'vc_up': "N/A", 'vc_down': "N/A"}


def parse_show_ldp_session(output):
    """Parse 'show ldp session' output"""
    if output.startswith("执行失败"):
        return {'ldp_total': "N/A", 'ldp_up': "N/A", 'ldp_down': "N/A"}
    total_match = re.search(r'Total number\s*:\s*(\d+)', output)
    up_match = re.search(r'OPERATIONAL\s*:\s*(\d+)', output)
    down_match = re.search(r'NON OPERATIONAL:\s*(\d+)', output)
    if total_match and up_match and down_match:
        return {
            'ldp_total': total_match.group(1),
            'ldp_up': up_match.group(1),
            'ldp_down': down_match.group(1)
        }
    return {'ldp_total': "N/A", 'ldp_up': "N/A", 'ldp_down': "N/A"}


def parse_show_run_include_31(output):
    """Parse 'show running-configuration include .31' output"""
    if output.startswith("执行失败"):
        return "N/A"

    # 修复正则表达式：匹配任意接口名称结构，但必须包含.31子接口
    interfaces = re.findall(
        r'^interface \S+ .*?\.31$',  # 关键修复点
        output,
        re.MULTILINE
    )
    return str(len(interfaces))


def parse_show_ospf_neighbor_brief(output):
    """Parse 'show ospf neighbor brief' output"""
    if output.startswith("执行失败"):
        return {'ospf_up': "N/A", 'ospf_down': "N/A"}
    process_31_start = output.find("OSPF process 31:")
    if process_31_start == -1:
        return {'ospf_up': "0", 'ospf_down': "0"}
    next_process = output.find("OSPF process", process_31_start + 1)
    section = output[process_31_start:next_process] if next_process != - \
        1 else output[process_31_start:]
    neighbors = re.findall(
        r'^(\d+\.\d+\.\d+\.\d+)\s+\d+\s+(\w+/\s*\S*)\s+', section, re.MULTILINE)
    up_count = sum(1 for _, state in neighbors if state.startswith("Full"))
    down_count = len(neighbors) - up_count
    return {'ospf_up': str(up_count), 'ospf_down': str(down_count)}


def parse_ospf_interfaces(ospf_output, process_id=31):
    """解析OSPF邻居输出，提取指定进程的接口"""
    interfaces = []
    lines = ospf_output.split('\n')
    in_process = False
    for line in lines:
        if f"OSPF process {process_id}:" in line:
            in_process = True
            continue
        if in_process and line.strip() and not line.startswith('Neighbor ID'):
            # 使用两个或以上空格分割字段
            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) >= 7:
                interface = parts[5].strip()
                interfaces.append(interface)
            else:
                print(f"[WARNING] 行格式异常，跳过: {line}")
        if in_process and line.startswith('OSPF process'):
            break
    return interfaces


def fish_ospf_interface_info_cmd(filename, ret_name, max_workers=40):
    """采集OSPF互联接口信息，支持多线程并行采集多台设备的OSPF接口信息"""
    import os
    import csv
    import time
    import re
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    print(f"[START] 开始采集OSPF互联接口信息，输入文件: {filename}, 输出文件: {ret_name}")
    with open(ret_name, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(filename, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_ospf_device, ip, user, pwd, writer, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 OSPF互联接口信息采集进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
        finally:
            print(f"{Fore.GREEN}[END] 数据采集完成{Style.RESET_ALL}")


def process_ospf_device(ip, user, pwd, writer, fail_log):
    """处理单个设备的OSPF互联接口信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    from collections import defaultdict
    # 文件写入锁，确保线程安全
    file_lock = Lock()

    channel = None
    try:
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        execute_some_command(channel, "screen-length 512", 1)
        ospf_output = execute_some_command(
            channel, "show ospf neighbor brief", 5)
        clean_ospf = "\n".join([
            line.strip()
            for line in ospf_output.split('\n')
            if line.strip() and line.strip() != "show ospf neighbor brief"
        ])
        with file_lock:  # 线程安全写入
            writer.writerow([ip, "show ospf neighbor brief", clean_ospf])

        interfaces = parse_ospf_interfaces(clean_ospf, process_id=31)
        print(f"[DEBUG] 设备 {ip} 解析到的接口: {interfaces}")
        for intf in interfaces:
            physical_intf = intf.split('.')[0]
            print(f"[DEBUG] 设备 {ip} 使用物理接口: {physical_intf}")
            cmd_intf = f"show interface {physical_intf}"
            intf_output = execute_some_command(channel, cmd_intf, 3)
            clean_intf = "\n".join([
                line.strip()
                for line in intf_output.split('\n')
                if line.strip() and line.strip() != cmd_intf
            ])
            with file_lock:  # 线程安全写入
                writer.writerow([ip, cmd_intf, clean_intf])

            cmd_lldp = f"show lldp neighbor interface {physical_intf}"
            lldp_output = execute_some_command(channel, cmd_lldp, 3)
            clean_lldp = "\n".join([
                line.strip()
                for line in lldp_output.split('\n')
                if line.strip() and line.strip() != cmd_lldp
            ])
            with file_lock:  # 线程安全写入
                writer.writerow([ip, cmd_lldp, clean_lldp])

        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            channel.close()


def generate_ospf_interface_report(src_file, dst_file, host_list_file):
    """生成OSPF互联接口信息报告"""
    print(
        f"\n[generate_ospf_interface_report] 开始生成报告，源文件: {src_file}, 目标文件: {dst_file}")
    device_names = {}
    ospf_data = defaultdict(dict)
    optical_module_data = defaultdict(dict)
    lldp_data = defaultdict(dict)
    connection_failures = set()

    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
    except FileNotFoundError:
        print("[generate_ospf_interface_report] 未找到failure_ips.tmp文件")

    # 数据解析部分
    with open(src_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        for line_num, row in enumerate(reader, 1):
            if len(row) != 3:
                print(
                    f"[generate_ospf_interface_report] 第 {line_num} 行数据格式错误，跳过")
                continue
            device_ip, cmd, output = row
            device_names, ospf_data = parse_ospf_neighbor_data1(
                device_ip, cmd, output, device_names, ospf_data)
            device_names, optical_module_data = parse_optical_module_data1(
                device_ip, cmd, output, device_names, optical_module_data)
            lldp_data = parse_lldp_neighbor_data(
                device_ip, cmd, output, lldp_data)

    with open(host_list_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]

    with open(dst_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow([
            "设备IP", "设备名称", "OSPF进程", "OSPF状态", "uptime", "接口",
            "光模块厂商名称", "型号", "模块类型", "工作模式", "波长nm", "传输距离M",
            "接收光功率dBm", "发送光功率dBm", "偏置电流mA", "电压mV", "温度°C", "CRC",
            "对端邻居系统名称", "邻居系统IP", "邻居系统详情"
        ])

        for ip in host_ips:
            if ip in connection_failures:
                writer.writerow([ip, "连接失败"] + ["-"] * 19)
                continue

            device_name = device_names.get(ip, "未知设备")
            interfaces = ospf_data.get(ip, {})

            if not interfaces:
                writer.writerow([ip, device_name] +
                                ["无OSPF数据"] + ["-"] * 18)
                continue

            for intf, neighbors in interfaces.items():
                physical_intf = intf.split('.')[0].strip()
                print(f"[生成报告] 设备 {ip} 接口 {intf} → 物理接口: {physical_intf}")

                optical_data = optical_module_data.get(
                    ip, {}).get(physical_intf, {})
                lldp_info = lldp_data.get(ip, {}).get(physical_intf, {})

                print(f"[光模块数据] {optical_data}")
                print(f"[LLDP数据] {lldp_info}")

                for neighbor in neighbors:
                    writer.writerow([
                        ip,
                        device_name,
                        neighbor.get("process", "-"),
                        neighbor.get("state", "-"),
                        neighbor.get("uptime", "-"),
                        intf,
                        optical_data.get("vendor_name", "-"),
                        optical_data.get("vendor_pn", "-"),
                        optical_data.get("transceiver", "-"),
                        optical_data.get("mode", "-"),
                        optical_data.get("wavelength", "-"),
                        optical_data.get("distance", "-"),
                        optical_data.get("rx_power", "-"),
                        optical_data.get("tx_power", "-"),
                        optical_data.get("bias", "-"),
                        optical_data.get("voltage", "-"),
                        optical_data.get("temperature", "-"),
                        optical_data.get("crc", "-"),
                        lldp_info.get("system_name", "-"),
                        lldp_info.get("management_address", "-"),
                        lldp_info.get("system_description", "-")
                    ])

    print(f"{Fore.GREEN}✅ 报告生成完成，共处理 {len(host_ips)} 台设备{Style.RESET_ALL}")


def parse_ospf_neighbor_data1(device_ip, cmd, output, device_names, ospf_data):
    """解析OSPF邻居信息"""
    print(f"[parse_ospf_neighbor_data] 设备IP: {device_ip}, 命令: {cmd}")
    if device_ip not in device_names:
        name_match = re.search(r'<([^>]+)>', output)
        device_names[device_ip] = name_match.group(1) if name_match else "未知设备"

    if cmd.strip().lower() == 'show ospf neighbor brief':
        lines = output.split('\n')
        current_process = None
        for line in lines:
            process_match = re.search(r'OSPF process (\d+):', line)
            if process_match:
                current_process = process_match.group(1)
                continue
            if current_process == '31':
                parts = re.split(r'\s{2,}', line.strip())
                # 仅处理第一列为IP地址的行，跳过表头
                if len(parts) >= 7 and re.match(r'\d+\.\d+\.\d+\.\d+', parts[0]):
                    neighbor_id = parts[0]
                    state = parts[2]
                    uptime = parts[3]
                    interface = parts[5]
                    if device_ip not in ospf_data:
                        ospf_data[device_ip] = {}
                    if interface not in ospf_data[device_ip]:
                        ospf_data[device_ip][interface] = []
                    ospf_data[device_ip][interface].append({
                        "process": current_process,
                        "neighbor_id": neighbor_id,
                        "state": state,
                        "uptime": uptime
                    })
    return device_names, ospf_data


def parse_optical_module_data1(device_ip, cmd, output, device_names, optical_module_data):
    """解析光模块信息，改进正则表达式和接口处理"""
    if cmd.strip().lower().startswith('show interface'):
        # 精确提取接口名（例如："show interface 50GE 0/6/1" -> "50GE 0/6/1"）
        interface = ' '.join(cmd.split()[2:]).strip()

        # 调试：打印接口名和原始输出
        print(f"[parse_optical_module_data1] 解析接口: {interface}")
        print(f"[RAW OUTPUT] {output[:500]}...")  # 打印部分输出以便调试

        # 改进的正则表达式，处理可能的空格和格式变化
        patterns = {
            "vendor_name": r'The Vendor Name\s*:\s*(.+)',
            "vendor_pn": r'The Vendor PN\s*:\s*(.+)',
            "transceiver": r'Transceiver Identifier\s*:\s*(.+)',
            "mode": r'Transceiver Mode\s*:\s*(.+)',
            "wavelength": r'WaveLength\s*:\s*([\d.]+)\s*nm',  # 允许空格和单位格式
            "distance": r'Transmission Distance\s*:\s*(\d+)\s*m',
            "rx_power": r'Rx Power\s*:\s*(-?[\d.]+)\s*dBm',  # 允许空格
            "tx_power": r'Tx Power\s*:\s*(-?[\d.]+)\s*dBm',
            "bias": r'Bias\s*:\s*(\d+)\s*mA',
            "voltage": r'Voltage\s*:\s*(\d+)\s*mV',
            "temperature": r'temperature\s*:\s*(\d+)\s*°?\s*C',  # 处理可能的°符号
            "crc": r'CRC\s*:\s*(\d+)\s*packets'  # 新增CRC错误计数
        }

        data = {}
        for key, pattern in patterns.items():
            match = re.search(pattern, output, re.IGNORECASE)
            if match:
                data[key] = match.group(1).strip()
                print(f"[成功解析] {key}: {data[key]}")  # 调试输出
            else:
                data[key] = "-"
                print(f"[警告] 未找到 {key} 的数据")     # 调试输出

        # 存储数据，键为设备IP和精确接口名
        if device_ip not in optical_module_data:
            optical_module_data[device_ip] = {}
        optical_module_data[device_ip][interface] = data
        print(f"[存储数据] {device_ip} - {interface}: {data}")  # 调试输出

    return device_names, optical_module_data


def parse_lldp_neighbor_data(device_ip, cmd, output, lldp_data):
    """解析LLDP信息，精确匹配接口和邻居数据"""
    if cmd.strip().lower().startswith('show lldp neighbor interface'):
        # 精确提取接口名（例如："show lldp neighbor interface 50GE 0/6/1" -> "50GE 0/6/1"）
        interface = ' '.join(cmd.split()[4:]).strip()

        # 调试信息
        print(f"[parse_lldp_neighbor_data] 解析接口: {interface}")
        print(f"[RAW OUTPUT] {output[:500]}...")

        data = {
            "system_name": "-",
            "management_address": "-",
            "system_description": "-"
        }

        # 使用多行模式匹配，处理可能换行的字段
        system_name_match = re.search(
            r'System Name:\s*(.+?)\n', output, re.IGNORECASE)
        if system_name_match:
            data["system_name"] = system_name_match.group(1).strip()

        # 精确匹配IPv4地址，忽略后续内容
        mgmt_ip_match = re.search(
            r'Management Address: IPv4 - (\d+\.\d+\.\d+\.\d+)\b', output)
        if mgmt_ip_match:
            data["management_address"] = mgmt_ip_match.group(1).strip()

        # 捕获系统描述的全部内容（可能含换行）
        sys_desc_match = re.search(
            r'System Description:\s*(.+?)(?=\n\S+:|$)', output, re.DOTALL)
        if sys_desc_match:
            data["system_description"] = sys_desc_match.group(
                1).strip().replace('\n', ' ')

        print(f"[LLDP数据] {device_ip} - {interface}: {data}")  # 调试输出

        if device_ip not in lldp_data:
            lldp_data[device_ip] = {}
        lldp_data[device_ip][interface] = data

    return lldp_data


def parse_lldp_neighbor_data(device_ip, cmd, output, lldp_data):
    """解析LLDP信息，精确匹配接口和邻居数据"""
    if cmd.strip().lower().startswith('show lldp neighbor interface'):
        # 精确提取接口名（例如："show lldp neighbor interface 50GE 0/6/1" -> "50GE 0/6/1"）
        interface = ' '.join(cmd.split()[4:]).strip()

        # 调试信息
        print(f"[parse_lldp_neighbor_data] 解析接口: {interface}")
        print(f"[RAW OUTPUT] {output[:500]}...")

        data = {
            "system_name": "-",
            "management_address": "-",
            "system_description": "-"
        }

        # 使用多行模式匹配，处理可能换行的字段
        system_name_match = re.search(
            r'System Name:\s*(.+?)\n', output, re.IGNORECASE)
        if system_name_match:
            data["system_name"] = system_name_match.group(1).strip()

        # 精确匹配IPv4地址，忽略后续内容
        mgmt_ip_match = re.search(
            r'Management Address: IPv4 - (\d+\.\d+\.\d+\.\d+)\b', output)
        if mgmt_ip_match:
            data["management_address"] = mgmt_ip_match.group(1).strip()

        # 捕获系统描述的全部内容（可能含换行）
        sys_desc_match = re.search(
            r'System Description:\s*(.+?)(?=\n\S+:|$)', output, re.DOTALL)
        if sys_desc_match:
            data["system_description"] = sys_desc_match.group(
                1).strip().replace('\n', ' ')

        print(f"[LLDP数据] {device_ip} - {interface}: {data}")  # 调试输出

        if device_ip not in lldp_data:
            lldp_data[device_ip] = {}
        lldp_data[device_ip][interface] = data

    return lldp_data


def fish_lsp_cmd(host_file, raw_file, max_workers=40):
    """采集业务LSP信息，支持多线程并行采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock
    print(f"[START] 开始采集业务LSP信息，输入文件: {host_file}, 输出文件: {raw_file}")
    with open(raw_file, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                # 使用线程池并行处理设备
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_lsp_device, ip, user, pwd, writer, fail_log))

                    # 使用tqdm显示进度
                    with tqdm(total=total_devices, desc="🔍 业务LSP信息采集进度", unit="台") as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result()  # 获取结果，触发异常处理
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
    print("[END] 数据采集完成")


def process_lsp_device(ip, user, pwd, writer, fail_log):
    """处理单个设备的业务LSP信息采集"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    from threading import Lock

    # 文件写入锁，确保线程安全
    file_lock = Lock()
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:  # 线程安全写入失败记录
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        execute_some_command(channel, "screen-length 512", 1)
        lsp_output = execute_some_command(channel, "show mpls lsp brief", 5)
        # 调试：显示部分输出
        print(f"[DEBUG] 设备 {ip} 原始LSP输出: {lsp_output[:800]}...")
        # 清理输出，移除空行和命令回显
        clean_lsp = "\n".join([
            line.strip()
            for line in lsp_output.split('\n')
            if line.strip() and line.strip() != "show mpls lsp brief"
        ])
        # 调试：显示清洗后输出
        print(f"[DEBUG] 设备 {ip} 清洗后LSP输出: {clean_lsp[:800]}...")
        # 使用csv.writer写入，线程安全
        with file_lock:
            writer.writerow([ip, "show mpls lsp brief", clean_lsp])
        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:  # 线程安全写入失败记录
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def parse_lsp_output(output):
    """解析show mpls lsp brief的输出"""
    print(f"[parse_lsp_output] 开始解析LSP输出")
    lines = output.split('\n')
    device_name = None
    lsp_list = []
    in_data_section = False
    for line in lines:
        if not line:
            continue
        if device_name is None:
            match = re.search(r'<([^>]+)>', line)
            if match:
                device_name = match.group(1)
                print(f"[DEBUG] 提取设备名称: {device_name}")
                continue
        if line.startswith('Dest LsrId'):
            in_data_section = True
            print(f"[DEBUG] 找到数据表头: {line}")
            continue
        if in_data_section:
            parts = re.split(r'\s{2,}', line)
            if len(parts) == 8:
                lsp = {
                    'Dest LsrId': parts[0],
                    'Type': parts[1],
                    'Description': parts[2],
                    'Stat': parts[3],
                    'InLabel': parts[4],
                    'OutLabel': parts[5],
                    'OutIntf': parts[6],
                    'Nexthop': parts[7]
                }
                # 判断是否丢标签
                if lsp['Type'] == 'Transit' and (lsp['InLabel'] == '-' or lsp['OutLabel'] == '-'):
                    lsp['丢标签'] = '是'
                else:
                    lsp['丢标签'] = '否'
                print(
                    f"[DEBUG] 解析LSP: Dest LsrId={lsp['Dest LsrId']}, Type={lsp['Type']}, 是否丢标签={lsp['丢标签']}")
                lsp_list.append(lsp)
            else:
                print(f"[WARNING] 行格式异常，跳过: {line}")
    print(f"[DEBUG] 共解析 {len(lsp_list)} 条LSP记录")
    return device_name, lsp_list


def generate_lsp_report(raw_file, report_file, host_file):
    """生成业务LSP报告"""
    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的failure_ips.tmp文件{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除failure_ips.tmp失败: {str(e)}{Style.RESET_ALL}")
    print(
        f"\n[generate_lsp_report] 开始生成报告，源文件: {raw_file}, 目标文件: {report_file}")
    connection_failures = set()
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
            print(f"[DEBUG] 读取连接失败设备: {len(connection_failures)} 台")
    except FileNotFoundError:
        print("[generate_lsp_report] 未找到failure_ips.tmp文件")

    lsp_data = {}
    with open(raw_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) != 3:
                print(f"[WARNING] 数据行格式错误，跳过: {row}")
                continue
            device_ip, cmd, output = row
            if cmd.strip().lower() == 'show mpls lsp brief':
                print(f"[DEBUG] 解析设备 {device_ip} 的LSP数据")
                device_name, lsps = parse_lsp_output(output)
                lsp_data[device_ip] = {
                    'name': device_name if device_name else "未知设备",
                    'lsps': lsps
                }
                print(
                    f"[DEBUG] 设备 {device_ip} 解析结果: 名称={lsp_data[device_ip]['name']}, LSP数量={len(lsps)}")

    with open(host_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(f"[DEBUG] 读取设备清单: {len(host_ips)} 台设备")

    with open(report_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow([
            "设备IP", "设备名称", "目的路由(Dest LsrId)", "类型(Type)", "描述(Description)", "状态(Stat)",
            "入标签(InLabel)", "出标签(OutLabel)", "出接口(OutIntf)", "下一跳 IP / 目标 MAC 地址(Nexthop ip/DstMac)", "是否丢标签"
        ])
        print(f"[DEBUG] 写入报告表头")
        for ip in host_ips:
            if ip in connection_failures:
                writer.writerow([ip, "连接失败"] + ["-"] * 9)
                print(f"[DEBUG] 设备 {ip}: 连接失败")
                continue
            if ip not in lsp_data:
                writer.writerow([ip, "未知设备", "无LSP数据"] + ["-"] * 8)
                print(f"[DEBUG] 设备 {ip}: 无LSP数据或未采集")
                continue
            device_name = lsp_data[ip]['name']
            lsps = lsp_data[ip]['lsps']
            if not lsps:
                writer.writerow([ip, device_name, "无LSP数据"] + ["-"] * 8)
                print(f"[DEBUG] 设备 {ip}: 无LSP记录")
                continue
            for lsp in lsps:
                writer.writerow([
                    ip,
                    device_name,
                    lsp['Dest LsrId'],
                    lsp['Type'],
                    lsp['Description'],
                    lsp['Stat'],
                    lsp['InLabel'],
                    lsp['OutLabel'],
                    lsp['OutIntf'],
                    lsp['Nexthop'],
                    lsp['丢标签']
                ])
                print(
                    f"[DEBUG] 设备 {ip} 写入LSP: Dest LsrId={lsp['Dest LsrId']}, 是否丢标签={lsp['丢标签']}")
    print(f"✅ 报告生成完成，共处理 {len(host_ips)} 台设备")


def fish_alarm_cmd(host_file, raw_file, max_workers=40):
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style
    import csv
    print(f"[START] 开始采集告警信息，输入文件: {host_file}, 输出文件: {raw_file}")
    with open(raw_file, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        futures.append(executor.submit(
                            process_alarm_device, ip, user, pwd, writer, fail_log))

                    # Custom progress bar with enhanced style and information
                    bar_format = "{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}{postfix}]"
                    with tqdm(total=total_devices, desc="🔍 告警信息采集进度", unit="台",
                              bar_format=bar_format, colour='green') as pbar:
                        for future in as_completed(futures):
                            try:
                                # Each thread waits up to 60 seconds
                                future.result(timeout=60)
                            except TimeoutError:
                                print(f"{Fore.RED}线程超时: {ip}{Style.RESET_ALL}")
                            except Exception as e:
                                print(
                                    f"{Fore.RED}线程执行出错: {str(e)}{Style.RESET_ALL}")
                            finally:
                                pbar.update(1)  # Ensure progress bar updates

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")
    print("[END] 数据采集完成")


def process_alarm_device(ip, user, pwd, writer, fail_log):
    from threading import Lock
    file_lock = Lock()
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        execute_some_command(channel, "screen-length 512", 1)
        current_alarm_output = execute_command_with_paging(
            channel, "show alarm current")
        history_alarm_output = execute_command_with_paging(
            channel, "show alarm history")

        # 数据清洗和写入逻辑保持不变
        clean_current_alarm = "\n".join([
            line.strip()
            for line in current_alarm_output.split('\n')
            if line.strip() and line.strip() != "show alarm current"
        ])
        clean_history_alarm = "\n".join([
            line.strip()
            for line in history_alarm_output.split('\n')
            if line.strip() and line.strip() != "show alarm history"
        ])

        with file_lock:
            writer.writerow([ip, "show alarm current", clean_current_alarm])
            writer.writerow([ip, "show alarm history", clean_history_alarm])

        execute_some_command(channel, "screen-length 25", 1)

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                channel.close()
                channel.get_transport().close()  # 确保底层的 transport 也关闭
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def execute_command_with_paging(channel, command, timeout=30):
    import time
    import select

    print(f"[DEBUG] 执行命令: {command}")
    channel.send(command + '\n')
    time.sleep(1)
    output = ''
    start_time = time.time()

    while True:
        rlist, _, _ = select.select([channel], [], [], 5.0)
        if not rlist:
            print(f"[WARNING] 命令 {command} 数据接收超时")
            break

        data = channel.recv(65535).decode('utf-8', errors='ignore')
        output += data
        print(f"[DEBUG] 接收数据长度: {len(data)}")

        if data.strip().endswith('----MORE----'):
            print(f"[DEBUG] 检测到分页提示，发送空格")
            channel.send(' ')
            time.sleep(0.5)
        elif '>' in data or '#' in data or len(data) < 65535:
            print(f"[DEBUG] 命令 {command} 执行完成")
            break

        if time.time() - start_time > timeout:
            print(f"[ERROR] 命令 {command} 分页处理超时")
            break

    print(f"[DEBUG] 命令 {command} 总输出长度: {len(output)}")
    return output


def parse_alarm_output(output, alarm_type):
    """解析 show alarm current 或 show alarm history 的输出"""
    import re
    print(f"[parse_alarm_output] 开始解析 {alarm_type} 告警输出")
    lines = output.split('\n')
    device_name = None
    alarm_summary = {'Total': 0, 'Critical': 0,
                     'Major': 0, 'Minor': 0, 'Warning': 0}
    alarm_list = []
    in_data_section = False

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # 提取设备名称
        if device_name is None:
            match = re.search(r'<([^>]+)>', line)
            if match:
                device_name = match.group(1)
                print(f"[DEBUG] 提取设备名称: {device_name}")
                continue
        # 解析告警统计
        if "Total number:" in line:
            alarm_summary['Total'] = int(line.split(":")[1].strip())
        elif "Critical    :" in line:
            alarm_summary['Critical'] = int(line.split(":")[1].strip())
        elif "Major       :" in line:
            alarm_summary['Major'] = int(line.split(":")[1].strip())
        elif "Minor       :" in line:
            alarm_summary['Minor'] = int(line.split(":")[1].strip())
        elif "Warning     :" in line:
            alarm_summary['Warning'] = int(line.split(":")[1].strip())
        elif line.startswith('Index'):
            in_data_section = True
            continue
        # 解析告警详情
        if in_data_section and line and line[0].isdigit():
            # 使用更鲁棒的方式分割行
            parts = re.split(r'\s{2,}', line.strip())
            # 根据告警类型确定预期字段数
            expected_fields = 5 if alarm_type == 'current' else 6
            if len(parts) >= expected_fields:
                alarm = {
                    'Index': parts[0],
                    'Alarm source': parts[1],
                    'Alarm info': parts[2],
                    'Level': parts[3],
                    'start time': parts[4],
                    'end time': '-' if alarm_type == 'current' else parts[5],
                    'Status': 'Active' if alarm_type == 'current' else 'Cleared'
                }
                alarm_list.append(alarm)
            else:
                print(f"[WARNING] 行格式异常，跳过: {line}")

    print(f"[DEBUG] 共解析 {len(alarm_list)} 条 {alarm_type} 告警记录")
    return device_name, alarm_summary, alarm_list


def generate_alarm_report(raw_file, report_file, host_file):
    import sys
    import csv
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    import os
    import re
    from colorama import Fore, Style
    """生成告警报告"""
    # Increase CSV field size limit
    csv.field_size_limit(sys.maxsize)

    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的 failure_ips.tmp 文件 {Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除 failure_ips.tmp 失败: {str(e)} {Style.RESET_ALL}")
    print(
        f"\n[generate_alarm_report] 开始生成报告，源文件: {raw_file}, 目标文件: {report_file}")

    connection_failures = set()
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
            print(f"[DEBUG] 读取连接失败设备: {len(connection_failures)} 台")
    except FileNotFoundError:
        print("[generate_alarm_report] 未找到 failure_ips.tmp 文件")

    alarm_data = {}
    try:
        with open(raw_file, "r", encoding='utf-8') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader, start=1):
                try:
                    if len(row) != 3:
                        print(f"[WARNING] 数据行格式错误，跳过: {row}")
                        continue
                    device_ip, cmd, output = row
                    if cmd.strip().lower() == 'show alarm current':
                        device_name, summary, alarms = parse_alarm_output(
                            output, 'current')
                        if device_ip not in alarm_data:
                            alarm_data[device_ip] = {'name': device_name, 'current': {
                                'summary': summary, 'alarms': alarms}}
                        else:
                            alarm_data[device_ip]['current'] = {
                                'summary': summary, 'alarms': alarms}
                    elif cmd.strip().lower() == 'show alarm history':
                        device_name, summary, alarms = parse_alarm_output(
                            output, 'history')
                        if device_ip not in alarm_data:
                            alarm_data[device_ip] = {'name': device_name, 'history': {
                                'summary': summary, 'alarms': alarms}}
                        else:
                            alarm_data[device_ip]['history'] = {
                                'summary': summary, 'alarms': alarms}
                except Exception as e:
                    print(f"[ERROR] 处理行 {i} 失败: {str(e)}")
                    continue
    except Exception as e:
        print(f"[ERROR] 读取 raw_file 失败: {str(e)}")
        return

    with open(host_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(f"[DEBUG] 读取设备清单: {len(host_ips)} 台设备")

    # 生成 Excel 报告
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "告警汇总统计"
    ws_main.append(["设备IP", "设备名称", "统计时间范围", "告警总数", "Critical",
                    "Major", "Minor", "Warning", "最后更新时间", "状态", "跳转"])

    for ip in host_ips:
        if ip in connection_failures:
            ws_main.append([ip, "设备连接失败", "-", 0, 0, 0, 0, 0, "-", "连接失败", ""])
            continue

        if ip not in alarm_data:
            print(
                f"[WARNING] 设备 {ip} 不在 alarm_data 中，但也不在 connection_failures 中")
            ws_main.append([ip, "未知设备", "数据异常", 0, 0,
                           0, 0, 0, "数据异常", "异常", ""])
            continue

        device_name = alarm_data[ip]['name'] or "未知设备"
        current_summary = alarm_data[ip].get('current', {}).get('summary', {})
        history_summary = alarm_data[ip].get('history', {}).get('summary', {})

        # 计算告警总数
        total_alarms = current_summary.get(
            'Total', 0) + history_summary.get('Total', 0)
        critical = current_summary.get(
            'Critical', 0) + history_summary.get('Critical', 0)
        major = current_summary.get('Major', 0) + \
            history_summary.get('Major', 0)
        minor = current_summary.get('Minor', 0) + \
            history_summary.get('Minor', 0)
        warning = current_summary.get(
            'Warning', 0) + history_summary.get('Warning', 0)

        # 收集所有告警时间
        all_alarms = alarm_data[ip].get('current', {}).get(
            'alarms', []) + alarm_data[ip].get('history', {}).get('alarms', [])
        all_times = []
        for alarm in all_alarms:
            if 'start time' in alarm and alarm['start time'] and alarm['start time'] != '-':
                all_times.append(alarm['start time'])
            if 'end time' in alarm and alarm['end time'] and alarm['end time'] != '-':
                all_times.append(alarm['end time'])

        # 计算时间范围和最后更新时间
        time_range = "无告警数据"
        last_update = "无告警数据"
        if all_times:
            try:
                time_format = "%Y-%m-%d %H:%M:%S"
                datetime_list = []
                for t in all_times:
                    try:
                        dt = datetime.strptime(t.strip(), time_format)
                        datetime_list.append(dt)
                    except ValueError:
                        print(f"[WARNING] 无效时间格式: {t}")
                        continue

                if datetime_list:
                    datetime_list.sort()
                    earliest = datetime_list[0].strftime("%Y-%m-%d")
                    latest = datetime_list[-1].strftime("%Y-%m-%d")
                    time_range = f"{earliest} ~ {latest}"
                    last_update = datetime_list[-1].strftime(
                        "%Y-%m-%d %H:%M:%S")
                else:
                    print(f"[WARNING] 设备 {ip} 无有效时间数据")
            except Exception as e:
                print(f"[ERROR] 时间处理错误 for {ip}: {str(e)}")
                time_range = "时间解析错误"
                last_update = "时间解析错误"

        # 清理工作表名称，确保合法且唯一
        safe_device_name = re.sub(
            r'[\\\/:*?"<>|\.\s]', '_', device_name.strip())
        safe_device_name = safe_device_name[:25]
        ws_detail_title = f"{safe_device_name}_详情"
        base_title = ws_detail_title
        suffix = 1
        while ws_detail_title in wb.sheetnames:
            ws_detail_title = f"{base_title}_{suffix}"
            suffix += 1

        # 创建子表工作表
        ws_detail = wb.create_sheet(title=ws_detail_title)
        ws_detail.append(["告警ID", "告警源", "告警类型", "告警级别", "发生时间",
                         "结束时间", "状态", "持续时间", "确认状态", "建议措施"])

        # 添加主表数据
        ws_main.append([ip, device_name, time_range, total_alarms,
                       critical, major, minor, warning, last_update, "正常", "跳转"])

        # 设置告警级别颜色
        level_colors = {
            'Critical': PatternFill(start_color='E54545', end_color='E54545', fill_type='solid'),
            'Major': PatternFill(start_color='FF8000', end_color='FF8000', fill_type='solid'),
            'Warning': PatternFill(start_color='FFBB33', end_color='FFBB33', fill_type='solid'),
            'Minor': PatternFill(start_color='4EAFF5', end_color='4EAFF5', fill_type='solid')
        }

        for alarm in all_alarms:
            duration = "-"
            if alarm['Status'] == 'Cleared' and 'start time' in alarm and 'end time' in alarm and alarm['end time'] != '-':
                try:
                    start = datetime.strptime(
                        alarm['start time'], "%Y-%m-%d %H:%M:%S")
                    end = datetime.strptime(
                        alarm['end time'], "%Y-%m-%d %H:%M:%S")
                    delta = end - start
                    duration = str(delta)
                except ValueError:
                    duration = "时间格式错误"

            confirmation_status = "自动清除" if alarm['Status'] == 'Cleared' else "未确认"
            row = [alarm['Index'], alarm['Alarm source'], alarm['Alarm info'], alarm['Level'],
                   alarm['start time'], alarm['end time'], alarm['Status'], duration, confirmation_status, ""]
            ws_detail.append(row)
            level_cell = ws_detail.cell(row=ws_detail.max_row, column=4)
            level_cell.fill = level_colors.get(alarm['Level'], PatternFill())

        link_cell = ws_main.cell(row=ws_main.max_row, column=11)
        link_cell.value = "跳转"
        link_cell.hyperlink = f"#'{ws_detail.title}'!A1"
        link_cell.font = Font(color="0000FF", underline="single")
        print(f"[DEBUG] 设置超链接: #'{'{ws_detail.title}'}'!A1 for IP {ip}")

    wb.save(report_file)
    print(f"✅ 告警报告生成完成，保存在 {report_file}")


def set_system_time_cmd(host_file, raw_file, report_file, max_workers=40):
    """设置系统时间并生成时间同步报告 (Set System Time and Generate Report)"""
    print(
        f"[START] 开始设置系统时间，输入文件: {host_file}, 输出文件: {raw_file}, 报告文件: {report_file}")

    # 清空旧的failure_ips.tmp文件
    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的failure_ips.tmp文件{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除failure_ips.tmp失败: {str(e)}{Style.RESET_ALL}")

    # 文件写入锁，确保线程安全
    file_lock = Lock()

    with open(raw_file, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    ip_to_future = {}
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        future = executor.submit(
                            process_device_time, ip, user, pwd, writer, fail_log, file_lock)
                        futures.append(future)
                        ip_to_future[future] = ip

                    with tqdm(total=total_devices, desc="🔍 系统时间同步进度", unit="台", dynamic_ncols=True) as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result(timeout=60)  # 每个任务最多120秒
                            except TimeoutError:
                                print(
                                    f"{Fore.RED}设备 {ip_to_future[future]} 任务超时{Style.RESET_ALL}")
                            except Exception as e:
                                print(
                                    f"{Fore.RED}设备 {ip_to_future[future]} 线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")

    print(f"[INFO] 开始生成时间同步报告...")
    generate_time_sync_report(raw_file, report_file, host_file)
    print(f"[END] 系统时间同步及报告生成完成")


def process_device_time(ip, user, pwd, writer, fail_log, file_lock):
    """处理单个设备的系统时间同步"""
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        # 设置屏幕长度避免分页
        execute_some_command(channel, "screen-length 512", 1)

        # 执行 con 命令
        con_cmd = "con"
        con_output = execute_some_command(channel, con_cmd, 3)
        print(f"[DEBUG] 设备 {ip} con 输出: {con_output[:800]}...")
        commands_executed = [
            (con_cmd, con_output, "执行成功" if "error" not in con_output.lower() else "执行失败")]

        # 获取设备当前时间
        time_output = execute_some_command(channel, "show cloc", 3)
        print(f"[DEBUG] 设备 {ip} show cloc 输出: {time_output[:800]}...")
        commands_executed.append(("show cloc", time_output, "检查时间"))

        # 解析设备时间
        device_time = parse_device_time(time_output)
        current_time = datetime.now()
        time_diff = abs((current_time - device_time).total_seconds()
                        ) if device_time else float('inf')
        print(f"[DEBUG] 设备 {ip} 时间差: {time_diff}秒")

        if device_time and time_diff <= 60:
            print(f"[INFO] 设备 {ip} 时间差 {time_diff}秒，小于60秒，无需同步")
        else:
            # 设置时区
            timezone_cmd = "clock time-zone add 8:0:0"
            timezone_output = execute_some_command(channel, timezone_cmd, 3)
            commands_executed.append(
                (timezone_cmd, timezone_output, "执行成功" if "error" not in timezone_output.lower() else "执行失败"))
            print(f"[DEBUG] 设备 {ip} 设置时区输出: {timezone_output[:800]}...")

            # 设置时间
            current_time_str = current_time.strftime("%H:%M:%S %Y/%m/%d")
            datetime_cmd = f"cloc datetime {current_time_str}"
            datetime_output = execute_some_command(channel, datetime_cmd, 3)
            commands_executed.append(
                (datetime_cmd, datetime_output, "执行成功" if "error" not in datetime_output.lower() else "执行失败"))
            print(f"[DEBUG] 设备 {ip} 设置时间输出: {datetime_output[:800]}...")

            # 验证时间
            verify_output = execute_some_command(channel, "show cloc", 3)
            commands_executed.append(
                ("show cloc", verify_output, "执行成功" if "error" not in verify_output.lower() else "执行失败"))
            print(f"[DEBUG] 设备 {ip} 验证时间输出: {verify_output[:800]}...")

            # 保存配置
            save_cmd = "sa"
            save_output = execute_some_command(channel, save_cmd, 3)
            commands_executed.append(
                (save_cmd, save_output, "执行成功" if "error" not in save_output.lower() else "执行失败"))
            print(f"[DEBUG] 设备 {ip} 保存配置输出: {save_output[:800]}...")

        # 写入原始数据
        with file_lock:
            for cmd, output, _ in commands_executed:
                writer.writerow([ip, cmd, output])
                print(f"[DEBUG] 设备 {ip} 写入原始数据: 命令={cmd}")

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                execute_some_command(channel, "screen-length 25", 1)
                channel.close()
                print(f"[DEBUG] 设备 {ip} 连接已关闭")
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def parse_device_time(output):
    """解析设备时间输出"""
    try:
        # 匹配时间格式：LOCAL TIME : 2025-04-22 10:53:52
        time_match = re.search(
            r'LOCAL TIME\s*:\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})', output)
        if time_match:
            time_str = time_match.group(1)
            return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        print(f"[WARNING] 无法解析设备时间: {output[:800]}...")
        return None
    except Exception as e:
        print(f"[ERROR] 解析设备时间出错: {str(e)}")
        return None


def generate_time_sync_report(raw_file, report_file, host_file):
    """生成时间同步报告"""
    print(
        f"[generate_time_sync_report] 开始生成报告，源文件: {raw_file}, 目标文件: {report_file}")
    connection_failures = set()
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
            print(f"[DEBUG] 读取连接失败设备: {len(connection_failures)} 台")
    except FileNotFoundError:
        print("[DEBUG] 未找到failure_ips.tmp文件")

    # 读取主机列表
    with open(host_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(f"[DEBUG] 读取设备清单: {len(host_ips)} 台设备")

    # 读取原始数据
    with open(raw_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        data = list(reader)
        print(f"[DEBUG] 读取原始数据: {len(data)} 条记录")

    # 生成报告
    with open(report_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow(["设备IP", "设备名称", "运行指令", "执行状态", "设备输出"])
        print(f"[DEBUG] 写入报告表头")
        processed_ips = set()

        for idx, row in enumerate(data):
            if len(row) != 3:
                print(f"[WARNING] 第 {idx+1} 行数据格式错误，跳过: {row}")
                continue

            device_ip, cmd, output = row
            print(f"[DEBUG] 处理设备 {device_ip} 的第 {idx+1} 条记录，命令: {cmd[:20]}...")
            processed_ips.add(device_ip)

            # 提取设备名称
            name_match = re.search(r'\[([^\]]+)\]', output, re.MULTILINE)
            device_name = name_match.group(1).strip() if name_match else "未知设备"
            print(f"[DEBUG] 设备 {device_ip} 提取设备名称: {device_name}")

            # 判断执行状态
            if "error" in output.lower():
                status = "执行失败"
                print(f"[DEBUG] 设备 {device_ip} 输出包含 'error'，状态为失败")
            # 第一个show cloc或紧跟con后的show cloc
            elif cmd == "show cloc" and idx == 0 or data[idx-1][1] == "con":
                status = "检查时间"
                print(f"[DEBUG] 设备 {device_ip} 命令为初始时间检查")
            elif cmd == "con":
                status = "执行成功"
                print(f"[DEBUG] 设备 {device_ip} 命令为con，状态为成功")
            else:
                status = "执行成功"
                print(f"[DEBUG] 设备 {device_ip} 命令执行状态为成功")

            writer.writerow([device_ip, device_name, cmd, status, output])
            print(f"[DEBUG] 设备 {device_ip} 写入报告第 {idx+1} 行数据")

        # 处理连接失败的设备
        for ip in host_ips:
            if ip not in processed_ips and ip in connection_failures:
                writer.writerow([ip, "连接失败", "-", "连接失败", "-"])
                print(f"[DEBUG] 设备 {ip}: 连接失败")

    print(f"✅ 报告生成完成，共处理 {len(host_ips)} 台设备")


######


def fish_multiple_cmds(host_file, raw_file, commands, max_workers=40):
    """Collect data for multiple commands from devices with debug output."""
    import csv
    import shutil
    import threading
    import time
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from tqdm import tqdm
    from colorama import Fore, Style

    print(
        f"{Fore.CYAN}[START] 开始采集QA巡检数据，输入文件: {host_file}, 输出文件: {raw_file}, 命令: {commands}{Style.RESET_ALL}")

    with open(raw_file, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "w", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(
                    f"{Fore.GREEN}[INFO] 共发现 {total_devices} 台设备{Style.RESET_ALL}")

                # 获取终端宽度
                terminal_width = shutil.get_terminal_size().columns

                # 设置 tqdm 进度条，留出 20 个字符空间给其他信息
                bar_format = "{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}{postfix}]"
                with print_lock:
                    pbar = tqdm(total=total_devices, desc="🔍 QA巡检数据采集进度", unit="台",
                                bar_format=bar_format, colour='green', ncols=terminal_width - 20)

                # 初始化完成计数和锁
                completed_count = 0
                progress_lock = threading.Lock()
                stop_periodic_update = threading.Event()

                # 定义开始时间
                start_time = time.time()

                # 定期更新进度条的函数
                def periodic_update():
                    nonlocal completed_count
                    while not stop_periodic_update.is_set():
                        try:
                            with progress_lock:
                                current_count = completed_count

                            current_time = time.time()
                            elapsed_time = current_time - start_time

                            # 设置手动进度条长度，留出 79 个字符空间给其他文本
                            progress_bar_length = terminal_width - 79
                            progress_bar = "█" * \
                                int(current_count / total_devices *
                                    progress_bar_length)
                            progress_bar += "░" * \
                                (progress_bar_length - len(progress_bar))

                            # 每 5 秒打印一次进度信息
                            eta_info = ""
                            if current_count > 0:
                                avg_time_per_device = elapsed_time / current_count
                                remaining_devices = total_devices - current_count
                                estimated_remaining_time = avg_time_per_device * remaining_devices
                                total_minutes = estimated_remaining_time / 60
                                if total_minutes < 60:
                                    eta_info = f" | 预计耗时: {total_minutes:.1f}分钟"
                                else:
                                    hours = int(total_minutes // 60)
                                    minutes = int(total_minutes % 60)
                                    eta_info = f" | 预计耗时: {hours}小时{minutes}分钟"
                            with print_lock:
                                print(f"{Fore.BLUE}[PROGRESS] [{progress_bar}] {current_count}/{total_devices} "
                                      f"({current_count/total_devices*100:.1f}%) | "
                                      f"已用时: {elapsed_time:.0f}秒 | 剩余:{total_devices - current_count}台{eta_info}{Style.RESET_ALL}")

                            # 如果任务完成，退出循环
                            if current_count >= total_devices:
                                break

                        except Exception as e:
                            print(
                                f"{Fore.YELLOW}[WARNING] 进度更新出错: {e}{Style.RESET_ALL}")

                        stop_periodic_update.wait(5)

                # 在单独线程中启动定期更新
                update_thread = threading.Thread(
                    target=periodic_update, daemon=True)
                update_thread.start()

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        # 调用修改后的 process_custom_commands 函数，保持PC时间功能
                        futures.append(executor.submit(
                            process_custom_commands, ip, user, pwd, commands, writer, fail_log))

                    # 更新完成计数的函数
                    def update_completed_count():
                        nonlocal completed_count
                        completed_count += 1
                        return completed_count

                    try:
                        for future in as_completed(futures):
                            try:
                                result = future.result()
                                with progress_lock:
                                    current_completed = update_completed_count()
                                pbar.update(1)

                            except Exception as e:
                                with progress_lock:
                                    current_completed = update_completed_count()
                                pbar.update(1)
                                print(
                                    f"{Fore.RED}[ERROR] 线程执行出错: {str(e)}{Style.RESET_ALL}")

                    finally:
                        stop_periodic_update.set()
                        pbar.close()

                        if update_thread.is_alive():
                            update_thread.join(timeout=2)

        except Exception as e:
            print(f"{Fore.RED}[ERROR] 数据采集错误: {str(e)}{Style.RESET_ALL}")

    print(f"{Fore.CYAN}[END] QA巡检数据采集完成{Style.RESET_ALL}")


def parse_ldp_lsp(ldp_output, ne_type, ne_name, ne_ip):
    """解析LDP LSP输出，检测链路邻接状态"""
    # 初始化结果列表
    result_data = []

    # 将输出按行分割
    lines = ldp_output.splitlines()

    # 使用正则表达式提取FEC和下一跳地址
    next_hops = set()  # 使用集合存储唯一的下一跳地址
    for line in lines:
        fec_match = re.match(r'FEC IPV4:(\S+) -> (\S+)', line.strip())
        if fec_match:
            next_jump = fec_match.group(2)
            # 忽略 none 和 0.0.0.0
            if next_jump not in ["none", "0.0.0.0"]:
                next_hops.add(next_jump)

    # 如果没有有效下一跳，返回默认行
    if not next_hops:
        return [{
            "网元类型": ne_type,
            "网元名称": ne_name,
            "网元IP": ne_ip,
            "邻接网元数": 0,
            "邻接下一跳地址": "-",
            "Result": "error"
        }]

    # 计算邻接网元数和邻接下一跳地址
    protection_count = len(next_hops)
    next_hop_str = ", ".join(sorted(next_hops))  # 将下一跳地址按顺序拼接
    result = "normal" if protection_count >= 2 else "normal"

    # 生成结果行
    row = {
        "网元类型": ne_type,
        "网元名称": ne_name,
        "网元IP": ne_ip,
        "邻接网元数": protection_count,
        "邻接下一跳地址": next_hop_str,
        "Result": result
    }
    result_data.append(row)

    return result_data


def parse_ptp_clock_status(ptp_output, synce_output, ne_type, device_name, ip):
    """
    Enhanced parser for PTP clock status data, supporting multiple receiving ports
    and merging device-level and port-level information.

    Args:
        ptp_output (str): Output from 'show ptp a' or similar command
        synce_output (str): Output from 'show synce' command
        ne_type (str): Network element type
        device_name (str): Device name
        ip (str): Device IP address

    Returns:
        dict: Structured data with device information and port details
    """
    import re

    result = {
        "设备信息": {
            "网元类型": ne_type,
            "网元名称": device_name,
            "网元IP": ip,
            "时钟标识": "-",
            "PTP状态": "-",
            "时钟模式": "-",
            "域值": "-",
            "从模式": "-",
            "步进模式": "-",
            "BMC优先级1": "-",
            "BMC优先级2": "-",
            "BMC时钟等级": "-",
            "BMC时钟精度": "-",
            "当前时钟源(PTP)": "-",
            "GM时钟标识": "-",
            "父时间标识": "-",
            "父时钟跳数": "-",
            "GM偏移统计(us)": "-",
            "GM时钟源类型": "-",
            "SyncE当前时钟源": "-",
            "SSM控制": "-",
            "SSM输入门限": "-",
            "外部SyncE类型": "-",
            "输出控制": "-",
            "SaBit": "-",
            "输出门限": "-"
        },
        "接收端口": []
    }

    if not ptp_output:
        return result

    lines = ptp_output.splitlines()

    # Step 1: Extract device-level PTP information
    # Using regex for more robust parsing
    for line in lines:
        # Clock identity
        if re.search(r"ClockIdentity\s*:\s*(\S+)", line):
            result["设备信息"]["时钟标识"] = re.search(
                r"ClockIdentity\s*:\s*(\S+)", line).group(1).strip()

        # PTP state and clock mode
        if re.search(r"PTPState\s*:\s*(\S+)", line) and re.search(r"ClockMode\s*:\s*(\S+)", line):
            result["设备信息"]["PTP状态"] = re.search(
                r"PTPState\s*:\s*(\S+)", line).group(1).strip()
            result["设备信息"]["时钟模式"] = re.search(
                r"ClockMode\s*:\s*(\S+)", line).group(1).strip()

        # Domain value and slave mode
        if re.search(r"DomainValue\s*:\s*(\S+)", line) and re.search(r"SlaveMode\s*:\s*(\S+)", line):
            result["设备信息"]["域值"] = re.search(
                r"DomainValue\s*:\s*(\S+)", line).group(1).strip()
            result["设备信息"]["从模式"] = re.search(
                r"SlaveMode\s*:\s*(\S+)", line).group(1).strip()

        # Step mode and port WTR
        if re.search(r"StepMode\s*:\s*(\S+)", line) and re.search(r"PortWTR\s*:\s*(\S+)", line):
            result["设备信息"]["步进模式"] = re.search(
                r"StepMode\s*:\s*(\S+)", line).group(1).strip()
            port_wtr = re.search(r"PortWTR\s*:\s*(\S+)", line).group(1).strip()
            # Store port_wtr for later use with ports

        # BMC priorities
        if re.search(r"BMCPriority1\s*:\s*(\S+)", line) and re.search(r"BMCPriority2\s*:\s*(\S+)", line):
            result["设备信息"]["BMC优先级1"] = re.search(
                r"BMCPriority1\s*:\s*(\S+)", line).group(1).strip()
            result["设备信息"]["BMC优先级2"] = re.search(
                r"BMCPriority2\s*:\s*(\S+)", line).group(1).strip()

        # BMC clock class and accuracy
        if re.search(r"BMCClockClass\s*:\s*(\S+)", line) and re.search(r"BMCClockAccuracy\s*:\s*(\S+)", line):
            result["设备信息"]["BMC时钟等级"] = re.search(
                r"BMCClockClass\s*:\s*(\S+)", line).group(1).strip()
            result["设备信息"]["BMC时钟精度"] = re.search(
                r"BMCClockAccuracy\s*:\s*(\S+)", line).group(1).strip()

        # Select source
        if re.search(r"SelectSource\s*:\s*(\S+)", line):
            result["设备信息"]["当前时钟源(PTP)"] = re.search(
                r"SelectSource\s*:\s*(\S+)", line).group(1).strip()

        # GM clock identity
        if re.search(r"GMClockIdentity\s*:\s*(\S+)", line):
            result["设备信息"]["GM时钟标识"] = re.search(
                r"GMClockIdentity\s*:\s*(\S+)", line).group(1).strip()

        # Parent clock identity
        if re.search(r"ParentClockIdentity\s*:\s*(\S+)", line):
            result["设备信息"]["父时间标识"] = re.search(
                r"ParentClockIdentity\s*:\s*(\S+)", line).group(1).strip()

        # Parent steps removed
        if re.search(r"ParentStepsRemoved\s*:\s*(\S+)", line):
            result["设备信息"]["父时钟跳数"] = re.search(
                r"ParentStepsRemoved\s*:\s*(\S+)", line).group(1).strip()

        # GM offset and time source
        if re.search(r"GMOffsetStats\s*:\s*(\S+)", line) and "GMTimeSource" in line:
            result["设备信息"]["GM偏移统计(us)"] = re.search(
                r"GMOffsetStats\s*:\s*(\S+)", line).group(1).strip()
            result["设备信息"]["GM时钟源类型"] = re.search(
                r"GMTimeSource\s*:\s*(\S+)", line).group(1).strip()
            # Convert GMTimeSource to hex format if needed
            if result["设备信息"]["GM时钟源类型"] == "GPS":
                result["设备信息"]["GM时钟源类型"] = "0x20"  # GPS in hex format
            elif result["设备信息"]["GM时钟源类型"] == "PTP":
                result["设备信息"]["GM时钟源类型"] = "0x40"  # Common mapping for PTP

        # Check for ClockRelation in device output format
        if re.search(r"ClockRelation\s*:\s*(.+?)$", line):
            clock_relation = re.search(
                r"ClockRelation\s*:\s*(.+?)$", line).group(1).strip()
            # This will be used if we find clock relation at the device level

    # Step 2: Extract port-specific information
    port_sections = []
    current_port = None
    for line in lines:
        # Start of new port section
        if re.search(r"Receive number\s*:\s*(.+?)\s+PTPNo\s*:\s*(\d+)", line):
            if current_port:
                port_sections.append(current_port)
            current_port = {"lines": [line]}
        elif current_port:
            current_port["lines"].append(line)

    # Add the last port section if exists
    if current_port:
        port_sections.append(current_port)

    # Process each port section
    for port_section in port_sections:
        port_data = {}

        # First line contains port number and PTP number
        first_line = port_section["lines"][0]
        port_match = re.search(
            r"Receive number\s*:\s*(.+?)\s+PTPNo\s*:\s*(\d+)", first_line)
        if port_match:
            port_data["时钟端口"] = port_match.group(1).strip()
            port_data["PTPNo"] = port_match.group(2).strip()

        # Process remaining port attributes
        for line in port_section["lines"]:
            # Check for clock relation - now using proper search pattern
            if "ClockRelation" in line:
                clock_relation_match = re.search(
                    r"ClockRelation\s*:\s*(.+?)$", line)
                if clock_relation_match:
                    port_data["时钟关系"] = clock_relation_match.group(1).strip()
                else:
                    port_data["时钟关系"] = "-"

            # RealState contains PTP actual status
            if "RealState" in line:
                port_data["PTP实际状态"] = re.search(
                    r"RealState\s*:\s*(.+?)$", line).group(1).strip()

            # Delay mechanism
            if "DelayMechanism" in line:
                port_data["延迟机制"] = re.search(
                    r"DelayMechanism\s*:\s*(.+?)(?:\s+|$)", line).group(1).strip()

            # Announce timeout
            if "AnnounceTimeout" in line:
                port_data["通告超时"] = re.search(
                    r"AnnounceTimeout\s*:\s*(.+?)(?:\s+|$)", line).group(1).strip()

            # Announce interval
            if "AnnounceInterval" in line:
                port_data["通告间隔(s)"] = re.search(
                    r"AnnounceInterval\(s\)\s*:\s*(.+?)(?:\s+|$)", line).group(1).strip()

            # Sync interval
            if "SyncInterval" in line:
                port_data["同步间隔(s)"] = re.search(
                    r"SyncInterval\(s\)\s*:\s*(.+?)(?:\s+|$)", line).group(1).strip()

            # Delay request interval
            if "DelayReqInterval" in line:
                port_data["延迟请求间隔(s)"] = re.search(
                    r"DelayReqInterval\(s\)\s*:\s*(.+?)(?:\s+|$)", line).group(1).strip()

            # Asymmetry direction
            if "AsymmetryDirection" in line:
                port_data["不对称方向"] = re.search(
                    r"AsymmetryDirection\s*:\s*(.+?)(?:\s+|$)", line).group(1).strip()

        # Use the global port WTR if available
        port_data["端口WTR(s)"] = port_wtr if 'port_wtr' in locals(
        ) else "300"  # Default to 300 if not found

        # Step 3: Add default values for missing fields
        expected_fields = [
            "时钟端口", "端口WTR(s)", "时钟关系", "通告超时", "不对称方向",
            "延迟机制", "通告间隔(s)", "同步间隔(s)", "延迟请求间隔(s)", "PTP实际状态"
        ]

        for field in expected_fields:
            if field not in port_data:
                port_data[field] = "-"

        # Step 4: Determine result and remarks
        port_data["Result"] = "normal"
        remarks = []

        # Check for PTP state issues
        if port_data["PTP实际状态"] not in ["MASTER", "SLAVE"]:
            port_data["Result"] = "error"
            remarks.append("PTP实际状态异常")

        # Check for GM offset issues
        if (result["设备信息"]["GM偏移统计(us)"] != "-" and
            result["设备信息"]["GM偏移统计(us)"].isdigit() and
                int(result["设备信息"]["GM偏移统计(us)"]) > 65536):
            port_data["Result"] = "normal"
            remarks.append("GM偏移统计过高")

        # Check BMC priority configuration
        if (result["设备信息"]["BMC优先级1"] != "255" or
                result["设备信息"]["BMC优先级2"] != "255"):
            port_data["Result"] = "normal"
            remarks.append("BMC优先级配置异常")
        if (result["设备信息"]["父时钟跳数"] >= "15"):
            port_data["Result"] = "normal"
            remarks.append("父时钟跳数过高")

        # Check BMC clock class
        # if result["设备信息"]["BMC时钟等级"] != "6":
        #     port_data["Result"] = "error"
        #     remarks.append("BMC时钟等级异常")

        port_data["备注"] = "; ".join(remarks) if remarks else ""

        # Add port data to result
        result["接收端口"].append(port_data)

    # Step 5: Check for additional device-level ClockRelation in the provided sample format
    # This is to handle the case shown in your example output
    for line in lines:
        if "RecvPortNumber" in line and "ClockRelation" in line:
            clock_relation_match = re.search(
                r"ClockRelation\s*:\s*(.+?)$", line)
            if clock_relation_match:
                # If there's at least one port and no clock relation has been set yet
                if result["接收端口"] and result["接收端口"][0]["时钟关系"] == "-":
                    # Set clock relation for all ports that don't have it yet
                    for port in result["接收端口"]:
                        if port["时钟关系"] == "-":
                            port["时钟关系"] = clock_relation_match.group(
                                1).strip()

    # Step 6: Parse SyncE information
    if synce_output:
        synce_lines = synce_output.splitlines()
        for line in synce_lines:
            if "Current Clock Source" in line:
                result["设备信息"]["SyncE当前时钟源"] = re.search(
                    r"Current Clock Source\s*:\s*(.+?)$", line).group(1).strip()
            elif "SSM-control" in line:
                result["设备信息"]["SSM控制"] = re.search(
                    r"SSM-control\s*:\s*(.+?)$", line).group(1).strip()
            elif "input-threshold" in line:
                result["设备信息"]["SSM输入门限"] = re.search(
                    r"input-threshold\s*:\s*(.+?)$", line).group(1).strip()
            elif "extern synce type" in line:
                result["设备信息"]["外部SyncE类型"] = re.search(
                    r"extern synce type\s*:\s*(.+?)$", line).group(1).strip()
            elif "output-control" in line:
                result["设备信息"]["输出控制"] = re.search(
                    r"output-control\s*:\s*(.+?)$", line).group(1).strip()
            elif "sa-bit" in line:
                result["设备信息"]["SaBit"] = re.search(
                    r"sa-bit\s*:\s*(.+?)$", line).group(1).strip()
            elif "output-threshold" in line:
                result["设备信息"]["输出门限"] = re.search(
                    r"output-threshold\s*:\s*(.+?)$", line).group(1).strip()

    # If no ports were found, create a default entry
    if not result["接收端口"]:
        result["接收端口"].append({
            "时钟端口": "-",
            "端口WTR(s)": "-",
            "时钟关系": "-",
            "通告超时": "-",
            "不对称方向": "-",
            "延迟机制": "-",
            "通告间隔(s)": "-",
            "同步间隔(s)": "-",
            "延迟请求间隔(s)": "-",
            "PTP实际状态": "-",
            "Result": "normal",
            "备注": "无接收端口数据"
        })

    return result


def parse_private_network_service(vpls_output, vsi_output, ne_type, ne_name, ne_ip):
    """解析专网业务分析数据"""
    print("Debug: Starting private network service parsing")
    service_data = []

    lines = vsi_output.splitlines()

    # 找到所有VSI块的起始位置
    vsi_block_starts = []
    for i, line in enumerate(lines):
        if "VSI:" in line and "Name:" in line and "MTU:" in line:
            vsi_block_starts.append(i)

    # 如果没有找到VSI块，返回默认行
    if not vsi_block_starts:
        print("Debug: No VSI blocks found")
        return [{
            "网元类型": ne_type,
            "网元名称": ne_name,
            "网元IP": ne_ip,
            "类型": "-",
            "VSI_ID": "-",
            "VSI名称": "-",
            "MTU": "-",
            "目的节点": "-",
            "状态": "-",
            "VC_ID": "-",
            "入标签": "-",
            "出标签": "-",
            "隧道ID": "-",
            "接口": "-",
            "PE VLAN[服务提供商]": "-",
            "CE VLAN[用户侧]": "-",
            "剥离外层 VLAN": "-",
            "HSID": "-",
            "Result": "normal"
        }]

    # 处理每个VSI块
    for block_idx, start_idx in enumerate(vsi_block_starts):
        # 确定块的结束位置
        end_idx = vsi_block_starts[block_idx + 1] if block_idx + \
            1 < len(vsi_block_starts) else len(lines)
        block_lines = lines[start_idx:end_idx]

        # 解析VSI基本信息
        vsi_id = "-"
        vsi_name = "-"
        mtu = "-"
        pw_signal = "-"
        vsi_type = "-"
        mac_learn = "-"
        limit_act = "-"
        limit_num = "-"
        learned_num = "-"

        for line in block_lines:
            # 解析VSI基本信息行
            vsi_match = re.search(
                r'VSI:(\d+)\s+Name:(\S+)\s+MTU:(\d+)\s+PwSignal:(\S+)\s+type:(\S+)', line)
            if vsi_match:
                vsi_id = vsi_match.group(1)
                vsi_name = vsi_match.group(2)
                mtu = vsi_match.group(3)
                pw_signal = vsi_match.group(4)
                vsi_type = vsi_match.group(5)
                print(
                    f"Debug: Parsed VSI - ID: {vsi_id}, Name: {vsi_name}, MTU: {mtu}, PwSignal: {pw_signal}, Type: {vsi_type}")
                continue

            # 解析MAC学习相关信息
            mac_learn_match = re.search(
                r'mac-learn:\s+(\S+)\s+limit-Act:(\S+)\s+limit-num:(\d+)\s+learned-num:(\d+)', line)
            if mac_learn_match:
                mac_learn = "启用" if mac_learn_match.group(1) == "En" else "禁用"
                limit_act = mac_learn_match.group(2)
                limit_num = mac_learn_match.group(3)
                learned_num = mac_learn_match.group(4)
                print(
                    f"Debug: Parsed MAC Learning - Status: {mac_learn}, Limit-Act: {limit_act}, Limit-Num: {limit_num}, Learned-Num: {learned_num}")
                continue

        # 找到当前VSI块中的VC和AC部分
        line_vc = next((i for i, line in enumerate(
            block_lines) if "--VC--" in line), None)
        line_ac = next((i for i, line in enumerate(
            block_lines) if "--AC--" in line and i > line_vc), None) if line_vc is not None else None

        # 获取VC和AC数据行（跳过表头）
        vc_lines = block_lines[line_vc +
                               2:line_ac] if line_vc is not None and line_ac is not None else []
        ac_lines = block_lines[line_ac + 2:] if line_ac is not None else []

        # 解析VC信息
        vc_details = []
        for vc_line in vc_lines:
            vc_match = re.match(
                r'\s*(\d+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\d+)\s+(\d+)\s+(\S+)\s+(\S)\s+(\d+)', vc_line)
            if vc_match:
                status = vc_match.group(3).lower()
                dest_node = vc_match.group(2)
                result = "normal" if status == "up" or dest_node == "1.1.1.1" else "error"

                # 不转换状态为中文，使用原始状态
                status_orig = vc_match.group(3)

                vc_details.append({
                    "VC_ID": vc_match.group(1),
                    "DestNode": dest_node,
                    "Status": status_orig,
                    "PW_Type": vc_match.group(4),
                    "PW_Tpid": vc_match.group(5),
                    "InLabel": vc_match.group(6),
                    "OutLabel": vc_match.group(7),
                    "TunnelID": vc_match.group(8),
                    "CW": vc_match.group(9),
                    "HSID": vc_match.group(10),
                    "Result": result
                })
                print(
                    f"Debug: VC - VSI: {vsi_id}, VC_ID: {vc_match.group(1)}, DestNode: {dest_node}, Status: {status_orig}")

        # 解析AC信息
        ac_details = []
        for ac_line in ac_lines:
            ac_match = re.match(
                r'\s*(\d+)\s+(\S+\s+\S+\s*\S*)\s+(\d+)\s+(\d+)\s+(\S+)\s+(\d+)', ac_line)
            if ac_match:
                ac_details.append({
                    "ID": ac_match.group(1),
                    "Interface": ac_match.group(2).strip(),
                    "PE_VLAN": ac_match.group(3),
                    "CE_VLAN": ac_match.group(4),
                    "StripSvlan": ac_match.group(5).lower(),
                    "HSID": ac_match.group(6)
                })
                print(
                    f"Debug: AC - VSI: {vsi_id}, Interface: {ac_match.group(2)}, PE_VLAN: {ac_match.group(3)}")

        # 添加AC数据
        for ac in ac_details:
            row = {
                "网元类型": ne_type,
                "网元名称": ne_name,
                "网元IP": ne_ip,
                "类型": "AC",
                "VSI_ID": vsi_id,
                "VSI名称": vsi_name,
                "MTU": mtu,
                "目的节点": "PW信令:" + pw_signal,       # 只填写PW信令部分
                "状态": "类型:" + vsi_type,             # 只填写类型部分
                "VC_ID": ac["ID"],                     # 使用AC的ID
                "入标签": "MAC学习: " + mac_learn,       # 只填写MAC学习部分
                "出标签": "限制动作:" + limit_act,       # 只填写限制动作部分
                "隧道ID": "限制数量:" + limit_num + " 已学习数量:" + learned_num,  # 填写限制数量和已学习数量
                "接口": ac["Interface"],
                "PE VLAN[服务提供商]": ac["PE_VLAN"],
                "CE VLAN[用户侧]": ac["CE_VLAN"],
                "剥离外层 VLAN": ac["StripSvlan"],
                "HSID": ac["HSID"],
                "Result": "normal"
            }
            service_data.append(row)

        # 添加VC数据
        for vc in vc_details:
            row = {
                "网元类型": ne_type,
                "网元名称": ne_name,
                "网元IP": ne_ip,
                "类型": "VC",
                "VSI_ID": vsi_id,
                "VSI名称": vsi_name,
                "MTU": mtu,
                "目的节点": vc["DestNode"],
                "状态": vc["Status"],
                "VC_ID": vc["VC_ID"],
                "入标签": vc["InLabel"],
                "出标签": vc["OutLabel"],
                "隧道ID": vc["TunnelID"],
                "接口": "-",
                "PE VLAN[服务提供商]": "-",
                "CE VLAN[用户侧]": "-",
                "剥离外层 VLAN": "-",
                "HSID": vc["HSID"],
                "Result": vc["Result"]
            }

            # 如果有AC数据，填充接口和VLAN信息
            if ac_details:
                # 默认使用第一个AC的信息
                row["接口"] = ac_details[0]["Interface"]
                row["PE VLAN[服务提供商]"] = ac_details[0]["PE_VLAN"]
                row["CE VLAN[用户侧]"] = ac_details[0]["CE_VLAN"]
                row["剥离外层 VLAN"] = ac_details[0]["StripSvlan"]

            service_data.append(row)

    # 如果没有数据，返回默认行
    if not service_data:
        print("Debug: No service data parsed")
        return [{
            "网元类型": ne_type,
            "网元名称": ne_name,
            "网元IP": ne_ip,
            "类型": "-",
            "VSI_ID": "-",
            "VSI名称": "-",
            "MTU": "-",
            "目的节点": "-",
            "状态": "-",
            "VC_ID": "-",
            "入标签": "-",
            "出标签": "-",
            "隧道ID": "-",
            "接口": "-",
            "PE VLAN[服务提供商]": "-",
            "CE VLAN[用户侧]": "-",
            "剥离外层 VLAN": "-",
            "HSID": "-",
            "Result": "normal"
        }]

    print(f"Debug: Parsed {len(service_data)} service entries")
    return service_data


yellow_fill = PatternFill(start_color="FFFF00",
                          end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500",
                          end_color="FFA500", fill_type="solid")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center_alignment = Alignment(horizontal="center", vertical="center")


def parse_snmp_config(trap_output, community_output, mib_view_output, sys_info_output):
    snmp_data = {
        "陷阱计数": 0,
        "陷阱主机IP地址": [],
        "陷阱UDP端口": [],
        "陷阱社区": [],
        "陷阱版本": [],
        "陷阱VPN实例": [],
        "社区计数": 0,
        "社区名称（缩写）": [],
        "社区权限": [],
        "社区访问级别": [],
        "社区绑定IP": [],
        "MIB视图名称": [],
        "MIB子树": [],
        "MIB视图类型": [],
        "MIB视图状态": [],
        "SNMP VPN": "",
        "联系人": "",
        "物理位置": "",
        "SNMP版本": "",
        "Result": "normal",
        "备注": ""
    }

    # Parse trap information
    trap_lines = trap_output.split('\n')
    for line in trap_lines:
        line = line.strip()
        if "Trap number" in line:
            snmp_data["陷阱计数"] += 1
        elif "Trap Host IP" in line:
            snmp_data["陷阱主机IP地址"].append(line.split(':')[1].strip())
        elif "Trap Udp Port" in line:
            snmp_data["陷阱UDP端口"].append(line.split(':')[1].strip())
        elif "Trap Community" in line:
            snmp_data["陷阱社区"].append(line.split(':')[1].strip())
        elif "Trap Version" in line:
            snmp_data["陷阱版本"].append(line.split(':')[1].strip())
        elif "Trap vpn-instance" in line:
            snmp_data["陷阱VPN实例"].append(line.split(':')[1].strip())

    # Parse community information
    community_lines = community_output.split('\n')
    for line in community_lines:
        line = line.strip()
        if "Community number" in line:
            snmp_data["社区计数"] += 1
        elif "Community Name" in line:
            name = line.split(':')[1].strip()
            snmp_data["社区名称（缩写）"].append(
                name[:10] + "..." if len(name) > 10 else name)
        elif "Community Authority" in line:
            auth = "只读" if "read-only" in line.lower() else "读写"
            snmp_data["社区权限"].append(auth)
        elif "Community Visit Level" in line:
            snmp_data["社区访问级别"].append(line.split(':')[1].strip())
        elif "Community Bind IP" in line:
            snmp_data["社区绑定IP"].append(line.split(':')[1].strip())

    # Parse MIB view information
    mib_view_lines = mib_view_output.split('\n')
    current_view = None
    for line in mib_view_lines:
        line = line.strip()
        if "View name:" in line:
            current_view = line.split(':')[1].strip()
            snmp_data["MIB视图名称"].append(current_view)
        elif "MIB Subtree:" in line:
            snmp_data["MIB子树"].append(line.split(':')[1].strip())
        elif "View Type" in line:
            view_type = "包含" if "included" in line.lower() else "排除"
            snmp_data["MIB视图类型"].append(view_type)
        elif "View status:" in line:
            status = "活跃" if "active" in line.lower() else "非活跃"
            snmp_data["MIB视图状态"].append(status)

    # Parse sys-info information
    sys_info_lines = sys_info_output.split('\n')
    for line in sys_info_lines:
        line = line.strip()
        if "SNMP VPN" in line:
            snmp_data["SNMP VPN"] = line.split(' ', 1)[1].strip()
        elif "The contact person" in line:
            snmp_data["联系人"] = line.split(':')[1].strip()
        elif "The physical location" in line:
            snmp_data["物理位置"] = line.split(':')[1].strip()
        elif "SNMP version" in line:
            snmp_data["SNMP版本"] = line.split(':')[1].strip()

    # Apply rules and generate remarks
    remarks = []
    # if snmp_data["SNMP版本"] not in ["v3"]:
    #     remarks.append("SNMP版本未启用v3加密协议，存在安全风险，建议升级至SNMPv3并启用加密认证。")
    # if any(ip == "0.0.0.0" for ip in snmp_data["社区绑定IP"]):
    #     remarks.append("绑定IP为0.0.0.0导致社区字符串暴露于全网，建议限制为特定管理网段（如4.148.32.0/24）。")
    # Additional rules can be added here (e.g., Trap Host IP validity, Community string security)

    if remarks:
        snmp_data["备注"] = "; ".join(remarks)

    return [snmp_data]


def parse_device_accounts(users_output, login_rule_output, logging_user_output):
    account_data = {
        "当前账户数量": 0,
        "用户名": [],
        "权限级别": [],
        "锁定状态": [],
        "锁定分钟数": "",
        "最大尝试次数": "",
        "密码提示天数": "",
        "密码最小长度": "",
        "密码需包含数字": "",
        "密码需包含大写字母": "",
        "密码需包含小写字母": "",
        "密码需包含特殊字符": "",
        "密码重用检查次数": "",
        "当前登录用户": [],
        "登录 Tty": [],
        "登录 Tid": [],
        "Result": "normal",
        "备注": ""
    }

    # 解析用户信息
    users_lines = users_output.split('\n')
    for line in users_lines:
        line = line.strip()
        if "Total Entries" in line:
            account_data["当前账户数量"] = int(line.split(':')[1].strip())
        elif "username" in line or "privilege" in line or "lockstatus" in line or "---" in line:
            continue
        else:
            parts = line.split()
            if len(parts) >= 3 and ("lock" in parts[2].lower() or "unlock" in parts[2].lower()):
                account_data["用户名"].append(parts[0])
                privilege = "super" if "super" in parts[1].lower(
                ) else parts[1]
                account_data["权限级别"].append(privilege)
                lock_status = "未锁定" if "unlock" in parts[2].lower() else "已锁定"
                account_data["锁定状态"].append(lock_status)

    # 解析登录规则信息
    login_rule_lines = login_rule_output.split('\n')
    for line in login_rule_lines:
        line = line.strip()
        if "lock-minutes" in line:
            account_data["锁定分钟数"] = line.split(':')[1].strip()
        elif "lock-try-times" in line:
            account_data["最大尝试次数"] = line.split(':')[1].strip()
        elif "prompt-days" in line:
            account_data["密码提示天数"] = line.split(':')[1].strip()
        elif "pwd-lenth" in line:
            account_data["密码最小长度"] = line.split(':')[1].strip()
        elif "pwd-include-digit" in line:
            status = "启用" if "enable" in line.lower() else "禁用"
            account_data["密码需包含数字"] = status
        elif "pwd-include-ABC" in line:
            status = "启用" if "enable" in line.lower() else "禁用"
            account_data["密码需包含大写字母"] = status
        elif "pwd-include-abc" in line:
            status = "启用" if "enable" in line.lower() else "禁用"
            account_data["密码需包含小写字母"] = status
        elif "pwd-include-special-char" in line:
            status = "启用" if "enable" in line.lower() else "禁用"
            account_data["密码需包含特殊字符"] = status
        elif "pwd-reuse-check-times" in line:
            account_data["密码重用检查次数"] = line.split(':')[1].strip()

    # 解析当前登录用户信息
    logging_user_lines = logging_user_output.split('\n')
    for line in logging_user_lines:
        line = line.strip()
        if "Tty" in line or "Tid" in line or "User Name" in line or "---" in line:
            continue
        parts = line.split()
        if len(parts) >= 3:
            account_data["登录 Tty"].append(parts[0])
            account_data["登录 Tid"].append(parts[1])
            account_data["当前登录用户"].append(parts[2])

    # 应用规则并生成备注
    remarks = []
    # try:
    #     lock_minutes = int(account_data["锁定分钟数"])
    #     if lock_minutes < 30:
    #         remarks.append("锁定时间过短（{}分钟），建议设置为30分钟以上。".format(lock_minutes))
    # except ValueError:
    #     pass
    # try:
    #     max_attempts = int(account_data["最大尝试次数"])
    #     if max_attempts > 5:
    #         remarks.append("最大尝试次数过多（{}次），建议设置为5次以下。".format(max_attempts))
    # except ValueError:
    #     pass
    # try:
    #     reuse_checks = int(account_data["密码重用检查次数"])
    #     if reuse_checks < 3:
    #         remarks.append("密码重用检查次数不足（{}次），建议设置为3次以上。".format(reuse_checks))
    # except ValueError:
    #     pass

    if remarks:
        account_data["备注"] = "; ".join(remarks)

    return [account_data]


def parse_loopback_address(output):
    """Extract IPv4 address from show interface loopback X output"""
    if not output or "CLI PTN_SVC_APP_Qx_Get_RecycleControl ERROR!" in output and "Internet IPV4 Address" not in output:
        return "无条目"
    match = re.search(
        r'Internet IPV4 Address is (\d+\.\d+\.\d+\.\d+/\d+)', output)
    if match:
        # Return IP address without subnet mask
        return match.group(1).split('/')[0]
    return "无条目"


def parse_ospf_routing_table(output):
    """解析 OSPF 路由表，检查 Cost 和 Uptime 异常"""
    routes = []
    lines = output.split('\n')
    routing_section = False

    # 改进的正则表达式，更好地处理空格和接口名
    route_pattern = re.compile(
        r'(\d+\.\d+\.\d+\.\d+/\d+)\s+(\S+)\s+(\d+)\s+(\d+)\s+(\d+\.\d+\.\d+\.\d+)\s+(.*?)\s+(\S+)$'
    )

    def parse_uptime(uptime):
        """将uptime字符串转换为秒数"""
        if ':' in uptime:
            # 处理"HH:MM:SS"格式
            parts = uptime.split(':')
            if len(parts) == 3:
                try:
                    hours, minutes, seconds = map(int, parts)
                    return hours * 3600 + minutes * 60 + seconds
                except ValueError:
                    return 0
            else:
                return 0
        else:
            # 处理"1w2d3h4m5s"格式
            units = {'w': 604800, 'd': 86400, 'h': 3600, 'm': 60, 's': 1}
            uptime_secs = 0
            pattern = re.compile(r'(\d+)([wdhms])')
            matches = pattern.findall(uptime)
            for num, unit in matches:
                uptime_secs += int(num) * units.get(unit, 0)
            return uptime_secs

    print(f"[DEBUG] 开始解析OSPF路由表，总行数: {len(lines)}")

    for i, line in enumerate(lines):
        line = line.strip()

        # 跳过空行
        if not line:
            continue

        # 寻找分隔线，表示路由表开始
        if '---' in line or 'Destination/Mask' in line:
            routing_section = True
            # print(f"[DEBUG] 第{i}行找到路由表开始标志: {line[:50]}...")
            continue

        if not routing_section:
            continue

        # 跳过表头行
        if 'Destination/Mask' in line or 'Proto' in line:
            # print(f"[DEBUG] 跳过表头行: {line}")
            continue

        # 尝试匹配路由行
        match = route_pattern.search(line)
        if match:
            dest_mask, proto, pre, cost, nexthop, interface, uptime = match.groups()

            # 清理接口名（去除多余空格）
            interface = ' '.join(interface.split())

            # print(f"[DEBUG] 匹配到路由: {dest_mask} - {proto} - Cost: {cost}")

            # 只处理 OSPF 相关协议
            if 'OSPF' not in proto and 'ospf' not in proto.lower():
                # print(f"[DEBUG] 跳过非OSPF路由: {proto}")
                continue

            # 检查规则
            remarks = []
            try:
                cost_val = int(cost)
                uptime_secs = parse_uptime(uptime)

                # 规则检查
                if cost_val > 18000:
                    remarks.append(
                        f"Cost值过高（{cost_val} > 18000），可能导致次优路径选择或环路")
                if uptime_secs < 3600:  # 小于1小时
                    uptime_str = str(timedelta(seconds=uptime_secs))
                    remarks.append(f"Uptime < 1小时（{uptime_str}），区域内OSPF有刷新")

                # 只有异常的条目才加入结果
                if remarks:
                    route = {
                        "目的网络/掩码": dest_mask,
                        "协议": proto,
                        "优先级": pre,
                        "开销": cost,
                        "下一跳": nexthop,
                        "接口": interface.strip(),
                        "存活时间": uptime,
                        "Result": "normal",
                        "备注": "; ".join(remarks)
                    }
                    routes.append(route)
                    # print(
                    #     f"[DEBUG] 添加异常路由: {dest_mask} - {'; '.join(remarks)}")
                # else:
                    # print(f"[DEBUG] 路由正常，不添加: {dest_mask}")

            except ValueError as e:
                print(f"[DEBUG] 解析数值失败: {line}, 错误: {e}")
                continue
        else:
            # 如果不匹配，打印调试信息
            if len(line) > 10 and not line.startswith(('Codes:', 'Routing count:', 'RIP', 'BGP', 'OSPF', 'ISIS', 'Total')):
                print(f"[DEBUG] 未匹配的行: {line}")

    print(f"[DEBUG] 解析完成，找到 {len(routes)} 个异常路由")

    # 如果没有异常路由但有路由数据，返回空列表（表示有数据但都正常）
    # 如果完全没有找到路由数据，返回无条目标识
    if not routes and not routing_section:
        print("[DEBUG] 没有找到路由表数据")
        return [{"目的网络/掩码": "无条目"}]
    elif not routes and routing_section:
        print("[DEBUG] 有路由表但没有异常条目")
        return []

    return routes


def parse_ldp_session_status(output):
    """解析 show ldp session 命令输出"""
    sessions = []
    lines = output.split('\n')
    session_section = False

    print(f"[DEBUG] 开始解析LDP会话输出...")
    print(f"[DEBUG] 输出内容: {output[:200]}...")

    for i, line in enumerate(lines):
        line = line.strip()
        print(f"[DEBUG] 处理行 {i}: '{line}'")

        # 找到表头行
        if line.startswith('Peer_type'):
            session_section = True
            print(f"[DEBUG] 找到表头行，开始解析会话数据")
            continue

        if not session_section or not line:
            continue

        # 跳过汇总信息行
        if any(keyword in line for keyword in ['Total number:', 'OPERATIONAL    :', 'NON OPERATIONAL:']):
            print(f"[DEBUG] 跳过汇总行: {line}")
            continue

        # 使用IP地址作为锚点解析数据行
        import re
        ip_match = re.search(r'\b(\d+\.\d+\.\d+\.\d+)\b', line)
        if not ip_match:
            print(f"[DEBUG] 行中未找到IP地址，跳过: {line}")
            continue

        ip_start = ip_match.start()
        ip_end = ip_match.end()
        peer_ip = ip_match.group(1)

        # peer_type 是IP地址之前的所有内容（去除首尾空白）
        peer_type = line[:ip_start].strip()

        # 剩余部分是IP地址之后的内容
        remaining = line[ip_end:].strip()
        print(f"[DEBUG] 解析字段 - Peer_type: '{peer_type}', Peer_IP: '{peer_ip}', 剩余: '{remaining}'")

        # 从右往左解析，因为接口名称可能包含空格
        parts = remaining.split()
        print(f"[DEBUG] 剩余部分分割后: {parts}")

        # 判断是否有UpTime字段（通过检查最后一个字段是否为时间格式或"-"）
        has_uptime = False
        if parts:
            last_field = parts[-1]
            # 如果最后一个字段包含冒号或是"-"，说明有UpTime字段
            if ':' in last_field or last_field == '-':
                has_uptime = True

        if has_uptime:
            # 有UpTime的格式：接口名 角色 状态 KeepAlive UpTime
            if len(parts) < 4:
                print(f"[DEBUG] 字段数量不足，跳过该行: {len(parts)} < 4")
                continue
            
            uptime = parts[-1]
            keepalive = parts[-2]
            state = parts[-3]
            role = parts[-4]
            
            # 找到role在remaining中的位置来提取接口名
            role_pos = remaining.find(role)
            if role_pos > 0:
                interface = remaining[:role_pos].strip()
            else:
                interface = "-"
        else:
            # 没有UpTime的格式：接口名 角色 状态 KeepAlive
            if len(parts) < 3:
                print(f"[DEBUG] 字段数量不足，跳过该行: {len(parts)} < 3")
                continue
            
            uptime = "-"
            keepalive = parts[-1]
            state = parts[-2]
            role = parts[-3]
            
            # 找到role在remaining中的位置来提取接口名
            role_pos = remaining.find(role)
            if role_pos > 0:
                interface = remaining[:role_pos].strip()
            else:
                interface = "-"

        print(f"[DEBUG] 解析结果 - 接口: '{interface}', 角色: '{role}', 状态: '{state}', KeepAlive: '{keepalive}', 运行时间: '{uptime}'")

        # 检查会话状态，设置 Result
        # OPERATIONAL 是正常状态，其他状态（如 NON_EXISTENT）是异常
        result = "normal" if state == "OPERATIONAL" else "error"

        # 设置备注
        remark = "会话正常" if result == "normal" else f"会话状态异常: {state}"

        session = {
            "对端类型": peer_type,
            "对端IP": peer_ip,
            "接口名称": interface,
            "角色": role,
            "会话状态": state,
            "KeepAlive时间": f"{keepalive}(s)",
            "运行时间": uptime,
            "Result": result,
            "备注": remark
        }
        sessions.append(session)
        print(f"[DEBUG] 添加会话: {session}")

    print(f"[DEBUG] 总共解析到 {len(sessions)} 个LDP会话")

    # 返回空列表表示无会话，而不是返回特殊的标记
    return sessions


def parse_ospf_buffers(output):
    buffers = {}
    lines = output.split('\n')
    current_process = None
    for line in lines:
        line = line.strip()
        if line.startswith('OSPF process'):
            current_process = line.split(' ')[2].rstrip(':')
            buffers[current_process] = {}
        elif current_process and ':' in line:
            key, value = line.split(':', 1)
            key = key.strip()
            value = value.strip()
            if 'buffer' in key:
                value = value.split()[0]
            elif 'list' in key:
                value = value.split()[0]
            buffers[current_process][key] = value
    return buffers


def parse_ospf_neighbor_brief(output):
    neighbors_brief = {}
    lines = output.split('\n')
    current_process = None
    # Matches HH:MM:SS or MM:SS
    up_time_pattern = re.compile(r'\d+:\d+(?::\d+)?')

    for line in lines:
        line = line.strip()
        if line.startswith('OSPF process'):
            current_process = line.split(':')[0].split()[-1]
        elif current_process and line and not line.startswith('Neighbor ID'):
            # Split on whitespace and filter out empty parts
            parts = [p for p in line.split() if p]
            if len(parts) >= 7:
                neighbor_id = parts[0]
                # Search for up_time using regex in the entire line
                up_time_match = up_time_pattern.search(line)
                up_time = up_time_match.group() if up_time_match else '-'
                instance_id = parts[-1]
                neighbors_brief[(current_process, neighbor_id)] = {
                    'instance_id': instance_id,
                    'up_time': up_time
                }
    return neighbors_brief


def parse_ospf_neighbor(output, brief_output):
    neighbors = []
    brief_data = parse_ospf_neighbor_brief(
        brief_output) if brief_output else {}
    lines = output.split('\n')
    current_process = None
    current_neighbor = None
    for line in lines:
        line = line.strip()
        if line.startswith('OSPF Process'):
            if current_neighbor:
                neighbors.append(current_neighbor)
            current_process = line.split('with Router ID')[0].split()[-1]
            current_neighbor = None
        elif line.startswith('Neighbor') and 'interface address' in line:
            if current_neighbor:
                neighbors.append(current_neighbor)
            neighbor_id = line.split(',')[0].split()[-1]
            interface_address = line.split('interface address')[-1].strip()
            brief_key = (current_process, neighbor_id)
            current_neighbor = {
                'process': current_process,
                'neighbor_id': neighbor_id,
                'interface_address': interface_address,
                'interface': '-',
                'area': '-',
                'priority': '-',
                'state': '-',
                'dr': '-',
                'bdr': '-',
                'up_time': brief_data.get(brief_key, {}).get('up_time', '-'),
                'ls_request_list': '0',
                'crypt_seq': '-',
                'instance_id': brief_data.get(brief_key, {}).get('instance_id', '0')
            }
        elif current_neighbor and 'In the area' in line:
            parts = line.split('via interface')
            if len(parts) == 2:
                area = parts[0].split('area')[1].strip()
                interface = parts[1].strip()
                current_neighbor['area'] = area
                current_neighbor['interface'] = interface
        elif current_neighbor and 'Neighbor priority is' in line:
            parts = line.split(',')
            priority = parts[0].split('is')[1].strip()
            state = parts[1].split('State is')[1].strip()
            current_neighbor['priority'] = priority
            current_neighbor['state'] = state
        elif current_neighbor and 'DR is' in line:
            dr = line.split('DR is')[1].split(',')[0].strip()
            bdr = line.split('BDR is')[1].strip()
            current_neighbor['dr'] = dr
            current_neighbor['bdr'] = bdr
        elif current_neighbor and 'Link State Request List' in line:
            ls_request_list = line.split('List')[1].strip()
            current_neighbor['ls_request_list'] = ls_request_list
        elif current_neighbor and 'Crypt Sequence Number is' in line:
            crypt_seq = line.split('is')[1].strip()
            current_neighbor['crypt_seq'] = crypt_seq
    if current_neighbor:
        neighbors.append(current_neighbor)
    return neighbors


def check_ospf_neighbor(neighbor, buffers):
    remarks = []
    result = "normal"
    process = neighbor['process']
    if process in buffers:
        buf = buffers[process]
        recv_buf = int(buf.get('Packet RECV buffer', '0'))
        send_buf = int(buf.get('Packet SEND buffer', '0'))
        lsa_buf = int(buf.get('LSA buffer', '0'))
        packet_unused = int(buf.get('Packet unused list', '0/0').split('/')[0])
        lsa_unused = int(buf.get('LSA unused list', '0/0').split('/')[0])
        if recv_buf < 2048:
            remarks.append("接收缓冲区过小")
            result = "error"
        if send_buf < 2048:
            remarks.append("发送缓冲区过小")
            result = "error"
        if lsa_buf < 2048:
            remarks.append("LSA缓冲区过小")
            result = "error"
        # if packet_unused < 2:
        #     remarks.append("未使用包列表过少")
        #     result = "normal"
        # if lsa_unused < 2:
        #     remarks.append("未使用LSA列表过少")
        #     result = "normal"

    # 获取接口名称（默认值设为小写，避免大小写问题）
    interface_name = neighbor.get('interface', '-').lower()

    # 定义需要匹配的接口前缀列表（不区分大小写）
    allowed_interface_prefixes = [
        'gigabitethernet',
        'xgigabitethernet',
        '50ge'  # 覆盖 50GE/50ge/50Ge 等变体
    ]
    # 检查接口是否以允许的前缀开头
    if any(interface_name.startswith(prefix) for prefix in allowed_interface_prefixes):
        # 当接口匹配时，检查 DR/BDR 是否为非零地址
        if neighbor.get('dr', '0.0.0.0') != '0.0.0.0' or neighbor.get('bdr', '0.0.0.0') != '0.0.0.0':
            remarks.append("DR/BDR路由非点到点模式")
            result = "normal"
    state = neighbor.get('state', '-')
    if state != 'Full' and state != '-':
        remarks.append("邻居状态非Full")
        result = "error"
    if int(neighbor.get('ls_request_list', '0')) > 0:
        remarks.append("链路状态请求列表非0")
        result = "error"
    # 增加网络类型判断（需要从设备获取实际网络类型参数）
    network_type = neighbor.get('network_type', 'broadcast')  # 默认广播网络

    return result, "; ".join(remarks) if remarks else "-"


def parse_ospf_neighbor_status(buffers_output, neighbor_output, brief_output=""):
    buffers = parse_ospf_buffers(buffers_output)
    neighbors = parse_ospf_neighbor(neighbor_output, brief_output)
    rows = []
    if not neighbors:
        return [{"OSPF进程": "无条目"}]
    for neighbor in neighbors:
        process = neighbor['process']
        buf = buffers.get(process, {})
        row = {
            'OSPF进程': process,
            '接收缓冲区(字节)': buf.get('Packet RECV buffer', '-'),
            '发送缓冲区(字节)': buf.get('Packet SEND buffer', '-'),
            'LSA缓冲区(字节)': buf.get('LSA buffer', '-'),
            '未使用包列表': buf.get('Packet unused list', '-'),
            '未使用LSA列表': buf.get('LSA unused list', '-'),
            '邻居ID': neighbor['neighbor_id'],
            '优先级': neighbor.get('priority', '-'),
            '状态': neighbor.get('state', '-'),
            '存活时间': neighbor.get('up_time', '-'),
            '接口地址': neighbor.get('interface_address', '-'),
            '接口': neighbor.get('interface', '-'),
            '区域': neighbor.get('area', '-'),
            'DR/BDR': f"{neighbor.get('dr', '-')}/{neighbor.get('bdr', '-')}",
            '链路状态请求列表': neighbor.get('ls_request_list', '-'),
            '加密序列号': neighbor.get('crypt_seq', '-'),
            '实例ID': neighbor.get('instance_id', '0'),
        }
        result, remarks = check_ospf_neighbor(neighbor, buffers)
        row['Result'] = result
        row['备注'] = remarks
        rows.append(row)
    return rows


def parse_up_time(up_time_str):
    """Parse up time string and convert to total seconds."""
    try:
        parts = up_time_str.split(':')
        if len(parts) == 3:
            h, m, s = map(int, parts)
            return h * 3600 + m * 60 + s
        elif len(parts) == 2:
            m, s = map(int, parts)
            return m * 60 + s
        else:
            return 0
    except:
        return 0


def parse_lag(output):
    """Parse 'show lag' output to extract LAG and member port details."""
    lags = []
    lines = output.split('\n')
    in_table = False
    current_lag = None
    for line in lines:
        line = line.strip()
        if line.startswith('id'):
            in_table = True
            continue
        if in_table and line and not line.startswith('--'):
            parts = line.split()
            if len(parts) >= 10:  # Full LAG entry
                if current_lag:  # Save previous LAG if exists
                    lags.append(current_lag)
                lag_id = parts[0]
                name = parts[1]
                mode = parts[2]
                hash_mode = parts[3]
                amc = parts[4]
                rvt = parts[5]
                wtr = parts[6]
                method = parts[7]
                syspri = parts[8]
                members_str = ' '.join(parts[9:])
                members = []
                member_parts = members_str.split('gigabitethernet')
                for mp in member_parts[1:]:
                    mp = 'gigabitethernet' + mp
                    interface, rest = mp.split('(', 1)
                    role, pri, status = rest.rstrip(')').split(',')
                    members.append({
                        'interface': interface.strip(),
                        'role': role.strip(),
                        'priority': pri.strip(),
                        'status': status.strip()
                    })
                current_lag = {
                    'id': lag_id,
                    'name': name,
                    'mode': mode,
                    'hash_mode': hash_mode,
                    'amc': amc,
                    'rvt': rvt,
                    'wtr': wtr,
                    'method': method,
                    'syspri': syspri,
                    'members': members
                }
            # Additional member port
            elif current_lag and line.startswith('gigabitethernet'):
                members_str = line
                member_parts = members_str.split('gigabitethernet')
                for mp in member_parts[1:]:
                    mp = 'gigabitethernet' + mp
                    interface, rest = mp.split('(', 1)
                    role, pri, status = rest.rstrip(')').split(',')
                    current_lag['members'].append({
                        'interface': interface.strip(),
                        'role': role.strip(),
                        'priority': pri.strip(),
                        'status': status.strip()
                    })
    if current_lag:  # Append the last LAG
        lags.append(current_lag)
    return lags


def parse_lacp(output):
    """Parse 'show lacp' output to extract detailed port information per LAG."""
    lacp_data = {}
    lines = output.split('\n')
    current_lag = None
    local_ports = {}
    remote_ports = {}
    local_sys_id = None
    parsing_local = False
    parsing_remote = False

    for line in lines:
        line = line.strip()
        if line.startswith('LAG:'):
            if current_lag:
                lacp_data[current_lag] = {
                    'local_sys_id': local_sys_id,
                    'local': local_ports,
                    'remote': remote_ports
                }
            current_lag = line.split(':')[1].split()[0]
            local_sys_id = line.split('Local SysId:')[1].strip()
            local_ports = {}
            remote_ports = {}
            parsing_local = False
            parsing_remote = False
        elif line.startswith('--Local'):
            parsing_local = True
            parsing_remote = False
            continue
        elif line.startswith('--Remote'):
            parsing_local = False
            parsing_remote = True
            continue
        elif current_lag and line and not line.startswith('=') and not line.startswith('Port'):
            parts = line.split()
            if len(parts) >= 6:  # Ensure enough fields for port data
                port = parts[0].strip()  # e.g., '02/03'
                if parsing_local:
                    local_ports[port] = {
                        'status': parts[1],
                        'port_pri': parts[2],
                        'port_no': parts[3],
                        'port_key': parts[4],
                        'port_state': parts[5]
                    }
                elif parsing_remote and len(parts) >= 7:
                    remote_ports[port] = {
                        'sys_pri': parts[1],
                        'sys_id': parts[2],
                        'port_pri': parts[3],
                        'port_no': parts[4],
                        'port_key': parts[5],
                        'port_state': parts[6]
                    }

    if current_lag:
        lacp_data[current_lag] = {
            'local_sys_id': local_sys_id,
            'local': local_ports,
            'remote': remote_ports
        }
    return lacp_data


def parse_lacp_status(lag_output, lacp_output):
    """Parse LACP member status from 'show lag' and 'show lacp' outputs."""
    lags = parse_lag(lag_output)
    lacp_data = parse_lacp(lacp_output)
    rows = []
    if not lags:
        return [{"聚合组ID": "无条目"}]
    for lag in lags:
        lag_id = lag['id']
        lag_details = lacp_data.get(
            lag_id, {'local': {}, 'remote': {}, 'local_sys_id': '-'})
        for member in lag['members']:
            interface = member['interface']
            # Extract port key from interface name
            try:
                _, slot_port = interface.split(' ')
                parts = slot_port.split('/')
                if len(parts) == 3:  # Format: 0/2/3
                    _, slot, port = parts
                elif len(parts) == 2:  # Format: 2/3
                    slot, port = parts
                else:
                    raise ValueError("Invalid interface format")
                # Normalize to '02/03'
                port_key = f"{int(slot):02d}/{int(port):02d}"
            except (ValueError, IndexError, AttributeError):
                port_key = '-'  # Fallback if parsing fails

            # Retrieve local and remote port details
            local_detail = lag_details['local'].get(port_key, {})
            remote_detail = lag_details['remote'].get(port_key, {})

            # Construct detail strings
            local_str = f"{port_key} ({local_detail.get('status', '-')}, {local_detail.get('port_pri', '-')}, {local_detail.get('port_no', '-')}, {local_detail.get('port_key', '-')}, {local_detail.get('port_state', '-')})"
            remote_str = f"{port_key} ({remote_detail.get('sys_pri', '-')}, {remote_detail.get('sys_id', '-')}, {remote_detail.get('port_pri', '-')}, {remote_detail.get('port_no', '-')}, {remote_detail.get('port_key', '-')}, {remote_detail.get('port_state', '-')})"

            # Build row dictionary
            row = {
                '聚合组ID': lag['id'],
                '聚合组名称': lag['name'],
                '模式': lag['mode'],
                '哈希模式': lag['hash_mode'],
                'AMC': lag['amc'],
                'RVT': lag['rvt'],
                'WTR': lag['wtr'],
                '协议': lag['method'],
                '系统优先级': lag['syspri'],
                '本地系统ID': lag_details['local_sys_id'],
                '成员端口（角色，优先级，状态）': f"{interface} ({member['role']}, {member['priority']}, {member['status']})",
                '本地端口详情': local_str,
                '远程端口详情': remote_str,
            }

            # Determine result status
            result = "normal"
            if member['status'].lower() != 'selected':
                result = "error"
            if remote_detail.get('sys_id', '') == '00-00-00-00-00-00':
                result = "error"
            if local_detail.get('port_state', '') != '11111100' or remote_detail.get('port_state', '') != '11111100':
                result = "error"
            if remote_detail.get('sys_pri', '') == '65535':
                result = "error"
            if remote_detail.get('port_key', '') == '0':
                result = "error"
            row['Result'] = result
            rows.append(row)
    return rows


yellow_fill = PatternFill(start_color="FFFF00",
                          end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500",
                          end_color="FFA500", fill_type="solid")
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))


def parse_ospf_process(output):
    """
    Parse the 'show ospf process' output and extract OSPF session details.
    Returns a list of dictionaries, each representing an OSPF process.
    """
    sessions = []
    lines = output.split('\n')
    current_session = None
    features = []
    in_area = False

    for line in lines:
        line = line.strip()
        # Start of a new OSPF process
        if line.startswith('OSPF Process'):
            if current_session:
                current_session['supports'] = ', '.join(features)
                sessions.append(current_session)
            current_session = {}
            features = []
            in_area = False
            process_match = re.search(
                r'OSPF Process (\d+) with Router ID (\S+)', line)
            if process_match:
                current_session['process_id'] = process_match.group(1)
                current_session['router_id'] = process_match.group(2)
        # Uptime
        elif line.startswith('Process uptime is'):
            uptime_match = re.search(r'Process uptime is (.+)', line)
            if uptime_match:
                current_session['uptime'] = uptime_match.group(1)
        # VRF
        elif line.startswith('Process bound to VRF'):
            vrf_match = re.search(r'Process bound to VRF (\S+)', line)
            if vrf_match:
                current_session['vrf'] = vrf_match.group(1)
        # RFC Compatibility
        elif line.startswith('Conforms to'):
            rfc_match = re.search(r'Conforms to (RFC\d+)', line)
            if rfc_match:
                current_session['rfc'] = rfc_match.group(1)
        # Supported Features
        elif line.startswith('Supports') or 'Connected to MPLS VPN Superbackbone' in line:
            if line.startswith('Supports'):
                feature = line.split('Supports')[1].strip()
                features.append(feature)
            elif 'Connected to MPLS VPN Superbackbone' in line:
                features.append('MPLS Superbackbone')
        # SPF Schedule Delay
        elif line.startswith('SPF schedule delay'):
            spf_delay_match = re.search(
                r'SPF schedule delay min (\S+) secs, SPF schedule delay max (\S+) secs', line)
            if spf_delay_match:
                current_session['spf_delay'] = f"{spf_delay_match.group(1)}秒/{spf_delay_match.group(2)}秒"
        # External LSA Count
        elif line.startswith('Number of external LSA'):
            external_lsa_match = re.search(
                r'Number of external LSA (\d+)', line)
            if external_lsa_match:
                current_session['external_lsa'] = external_lsa_match.group(1)
        # Number of Areas
        elif line.startswith('Number of areas attached to this router:'):
            areas_match = re.search(
                r'Number of areas attached to this router: (\d+)', line)
            if areas_match:
                current_session['areas'] = areas_match.group(1)
        # Area Section
        elif line.startswith('Area'):
            in_area = True
            area_match = re.search(r'Area (\S+)', line)
            if area_match:
                area_id = area_match.group(1)
                current_session['area_type_id'] = '骨干区域(' + area_id + \
                    ')' if area_id == '0.0.0.0' else '普通区域(' + area_id + ')'
        # Interfaces in Area
        elif in_area and line.startswith('Number of interfaces in this area is'):
            interfaces_match = re.search(
                r'Number of interfaces in this area is (\d+)\((\d+)\)', line)
            if interfaces_match:
                current_session['interfaces'] = interfaces_match.group(1)
        # Adjacent Neighbors in Area
        elif in_area and line.startswith('Number of fully adjacent neighbors in this area is'):
            adjacencies_match = re.search(
                r'Number of fully adjacent neighbors in this area is (\d+)', line)
            if adjacencies_match:
                current_session['adjacencies'] = adjacencies_match.group(1)
        # Last SPF Execution
        elif in_area and line.startswith('SPF algorithm last executed'):
            last_spf_match = re.search(
                r'SPF algorithm last executed (\S+) ago', line)
            if last_spf_match:
                current_session['last_spf'] = last_spf_match.group(1) + '前'
        # SPF Execution Times
        elif in_area and line.startswith('SPF algorithm executed'):
            spf_executions_match = re.search(
                r'SPF algorithm executed (\d+) times', line)
            if spf_executions_match:
                current_session['spf_executions'] = spf_executions_match.group(
                    1)
        # Total LSA Count
        elif in_area and line.startswith('Number of LSA'):
            total_lsa_match = re.search(r'Number of LSA (\d+)', line)
            if total_lsa_match:
                current_session['total_lsa'] = total_lsa_match.group(1)

    # Append the last session
    if current_session:
        current_session['supports'] = ', '.join(features)
        sessions.append(current_session)

    return sessions


def check_ospf_process(session, output):
    remarks = []
    result = "normal"

    # Helper: Calculate total minutes from uptime
    def uptime_to_minutes(uptime):
        total_minutes = 0
        try:
            # Handle "X minutes" format
            if "minutes" in uptime and not ("days" in uptime or "day" in uptime) and not ("hours" in uptime or "hour" in uptime):
                minutes = int(uptime.split('minutes')[0].strip())
                total_minutes = minutes
            else:
                # Replace plural forms for consistency
                uptime = uptime.replace('days', 'day').replace('hours', 'hour')
                # Split by day, hour, and minute
                days, hours, minutes = 0, 0, 0
                if 'day' in uptime:
                    parts = uptime.split('day')
                    days = int(parts[0].strip())
                    rest = parts[1].strip() if len(
                        parts) > 1 else '0 hour 0 minutes'
                else:
                    rest = uptime

                if 'hour' in rest:
                    parts = rest.split('hour')
                    hours = int(parts[0].strip())
                    minutes_part = parts[1].strip() if len(
                        parts) > 1 else '0 minutes'
                    minutes = int(minutes_part.split('minutes')[0].strip())
                else:
                    minutes = int(rest.split('minutes')[0].strip())

                total_minutes = days * 24 * 60 + hours * 60 + minutes
        except (ValueError, IndexError) as e:
            print(f"[DEBUG] Uptime parsing error: {e}, uptime: {uptime}")
            return None
        return total_minutes

    # 1. SPF execution frequency (< 1000/day)
    spf_executions = session.get('spf_executions', '0').replace(',', '')
    try:
        spf_executions = int(spf_executions)
        uptime = session.get('uptime', '0 minutes')
        total_minutes = uptime_to_minutes(uptime)
        if total_minutes is not None and total_minutes > 0:
            total_days = total_minutes / (24 * 60)
            spf_per_day = spf_executions / total_days if total_days > 0 else 0
            print(
                f"[DEBUG] SPF frequency for uptime {uptime}: {spf_per_day:.2f} times/day")
            if spf_per_day > 200000:
                remarks.append("❗ SPF执行频率过多，可能因网络震荡导致路由不稳定")
                result = "error"
        else:
            remarks.append("⚠️ SPF频率计算失败，检查uptime格式")
    except ValueError:
        remarks.append("⚠️ SPF频率计算失败，检查spf_executions格式")
        pass

    # 2. Total LSA count (< 1000 per area)
    total_lsa = session.get('total_lsa', '0').replace(',', '')
    try:
        total_lsa = int(total_lsa)
        if total_lsa > 12000:
            remarks.append("❗ LSA总数过多")
            result = "error"
        elif total_lsa > 1800:
            remarks.append("⚠️ LSA总数接近临界值，需监控增长趋势")
    except ValueError:
        remarks.append("⚠️ LSA总数解析失败")
        pass

    # 3. External LSA count (< 500)
    external_lsa = session.get('external_lsa', '0').replace(',', '')
    try:
        external_lsa = int(external_lsa)
        if external_lsa > 19800:
            remarks.append("❗ 外部LSA数过多，需检查路由重分发或泛洪攻击")
            result = "error"
    except ValueError:
        remarks.append("⚠️ 外部LSA数解析失败")
        pass

    # 4. Adjacency vs Interface count
    interfaces = session.get('interfaces', '0')
    adjacencies = session.get('adjacencies', '0')
    try:
        interfaces = int(interfaces)
        adjacencies = int(adjacencies)
        if interfaces > 2:
            if adjacencies < 2:  # 修改为检查邻接数 < 2
                remarks.append("⚠️ 邻接数少于2(单链站点)")
                result = "error"
                if adjacencies == 0:
                    remarks.append("❗ 无邻接，可能OSPF未激活或认证错误")
            # elif adjacencies < (interfaces - 2):  # 保留原有逻辑作为次要检查
            #     remarks.append("⚠️ 邻接数少于预期(邻接数≥接口数-2)")
    except ValueError:
        remarks.append("⚠️ 邻接或接口数解析失败")
        pass
    # 5. Uptime anomaly (< 10 minutes)
    uptime = session.get('uptime', '')
    if "minutes" in uptime and not ("days" in uptime or "day" in uptime):
        try:
            minutes = int(uptime.split('minutes')[0].strip())
            if minutes < 10:
                if adjacencies == 0 and total_lsa == 0:
                    remarks.append("⚠️ 新进程未激活，运行时间短且无邻接或LSA")
                else:
                    remarks.append("⚠️ 进程运行时间短，但已激活")
                result = "warning" if result == "normal" else result
        except ValueError:
            pass

    # 6. LSA Receive Rate (< 1000 LSA/minute)
    lsa_received_match = re.search(r'Number of LSA received (\d+)', output)
    if lsa_received_match:
        try:
            lsa_received = int(lsa_received_match.group(1))
            total_minutes = uptime_to_minutes(uptime)
            if total_minutes is not None and total_minutes > 0:
                lsa_per_minute = lsa_received / total_minutes
                print(
                    f"[DEBUG] LSA receive rate for uptime {uptime}: {lsa_per_minute:.2f} LSA/minute")
                if lsa_per_minute > 1000:
                    remarks.append("❗ LSA接收速率接近阈值，需排查相邻设备是否异常泛洪")
                    result = "error"
            else:
                remarks.append("⚠️ LSA接收速率计算失败，检查uptime格式")
        except ValueError:
            remarks.append("⚠️ LSA接收速率计算失败，检查lsa_received格式")
            pass
    else:
        remarks.append("⚠️ 无法提取LSA接收数量")

    # 7. ABR Verification (for Process 65534)
    if session.get('process_id') == '65534':
        areas = int(session.get('areas', '0'))
        area_id = session.get('area_type_id', '')
        if areas == 1 and '骨干区域(0.0.0.0)' in area_id:
            remarks.append("❗ 违反OSPF区域架构规则，需确认是否存在虚拟链路或区域0.0.0.0未覆盖所有区域连接")
            result = "error"

    remarks_str = "; ".join(remarks) if remarks else "-"
    return result, remarks_str


def parse_ospf_session(output):
    sessions = []
    lines = output.split('\n')
    current_session = None
    for line in lines:
        line = line.strip()
        if line.startswith('OSPF Process'):
            if current_session:
                sessions.append(current_session)
            current_session = {}
            process_id = line.split('with Router ID')[0].split()[-1]
            router_id = line.split('with Router ID')[1].strip()
            current_session['process_id'] = process_id
            current_session['router_id'] = router_id
        elif line.startswith('Process uptime is'):
            uptime = line.split('is')[1].strip()
            current_session['uptime'] = uptime
        elif line.startswith('Process bound to VRF'):
            vrf = line.split('to VRF')[1].strip()
            current_session['vrf'] = vrf
        elif line.startswith('Conforms to RFC'):
            rfc = line.split(',')[0].split('to')[1].strip()
            current_session['rfc'] = rfc
        elif line.startswith('Supports'):
            supports = line.split('Supports')[1].strip()
            current_session['supports'] = supports
        elif line.startswith('SPF schedule delay'):
            spf_delay = line.split('delay')[1].strip()
            current_session['spf_delay'] = spf_delay
        elif line.startswith('Number of external LSA'):
            external_lsa = line.split('LSA')[1].split('.')[0].strip()
            current_session['external_lsa'] = external_lsa
        elif line.startswith('Number of LSA') and 'received' not in line and 'originated' not in line:
            total_lsa = line.split('LSA')[1].split('.')[0].strip()
            current_session['total_lsa'] = total_lsa
        elif line.startswith('Number of areas attached to this router:'):
            areas = line.split(':')[1].strip()
            current_session['areas'] = areas
        elif line.startswith('Area'):
            area_type_id = line.split()[1]
            if '(' in area_type_id:
                area_type_id = area_type_id + ' ' + \
                    line.split('(')[1].split(')')[0]
            current_session['area_type_id'] = area_type_id
        elif line.startswith('Number of interfaces in this area is'):
            interfaces = line.split('is')[1].split('(')[0].strip()
            adjacencies = line.split('(')[1].split(')')[0]
            current_session['interfaces'] = interfaces
            current_session['adjacencies'] = adjacencies
        elif line.startswith('SPF algorithm last executed'):
            last_spf = line.split('executed')[1].strip()
            current_session['last_spf'] = last_spf
        elif line.startswith('SPF algorithm executed'):
            spf_executions = line.split('executed')[1].strip()
            current_session['spf_executions'] = spf_executions
    if current_session:
        sessions.append(current_session)
    if not sessions:
        sessions.append({
            'process_id': '-', 'router_id': '-', 'uptime': '-', 'vrf': '-', 'rfc': '-',
            'supports': '-', 'spf_delay': '-', 'external_lsa': '-', 'total_lsa': '-',
            'areas': '-', 'area_type_id': '-', 'interfaces': '-', 'adjacencies': '-',
            'last_spf': '-', 'spf_executions': '-', 'result': 'normal', 'remarks': '无条目'
        })
    return sessions


def parse_mpls_lsp(output):
    lsps = []
    # 更完善的过滤逻辑，过滤掉所有非数据行
    lines = output.splitlines()

    for line in lines:
        line = line.strip()

        # 跳过空行
        if not line:
            continue

        # 跳过标题行（多种可能的格式）
        if any(keyword in line.upper() for keyword in ['DEST', 'LSRID', 'TYPE', 'DESCRIPTION', 'STAT', 'INLABEL', 'OUTLABEL', 'OUTINTF', 'NEXTHOP']):
            print(f"[DEBUG] 跳过标题行: {line}")
            continue

        # 跳过分隔线
        if line.startswith('-') or all(c in '-= |' for c in line):
            print(f"[DEBUG] 跳过分隔线: {line}")
            continue

        # 跳过包含 "error" 但不是有效数据的行
        if line.endswith('error') and ('LSP状态为down' in line or '需检查' in line):
            print(f"[DEBUG] 跳过错误提示行: {line}")
            continue

        # 检查是否为有效的LSP数据行
        parts = line.split()
        if len(parts) < 6:  # LSP数据至少应该有6个字段
            print(f"[DEBUG] 跳过字段不足的行: {line}")
            continue

        # 验证第一个字段是否像IP地址或LSR ID
        if not is_valid_lsr_id_or_ip(parts[0]):
            print(f"[DEBUG] 跳过无效LSR ID的行: {line}")
            continue

        # 提取LSP数据
        try:
            dest_lsr_id = parts[0]
            lsp_type = parts[1]
            description = parts[2]
            state = parts[3]
            in_label = parts[4]
            out_label = parts[5]

            # 处理接口字段和下一跳IP
            if len(parts) == 8:
                out_intf = parts[6]
                nexthop_ip = parts[7]
            elif len(parts) > 8:
                # 接口名可能包含空格，需要重新组合
                nexthop_ip = parts[-1]
                out_intf = " ".join(parts[6:-1])
            else:
                out_intf = "-"
                nexthop_ip = parts[-1] if len(parts) > 6 else "-"

            lsps.append({
                'dest_lsr_id': dest_lsr_id,
                'type': lsp_type,
                'description': description,
                'state': state,
                'in_label': in_label,
                'out_label': out_label,
                'out_intf': out_intf,
                'nexthop_ip': nexthop_ip
            })

            # print(f"[DEBUG] 成功解析LSP: {dest_lsr_id} - {state}")

        except Exception as e:
            print(f"[DEBUG] 解析LSP数据时出错: {line}, 错误: {e}")
            continue

    print(f"[DEBUG] 总共解析到 {len(lsps)} 个有效LSP")
    return lsps


def is_valid_lsr_id_or_ip(value):
    """检查是否为有效的LSR ID或IP地址"""
    try:
        # 检查是否为IP地址格式
        parts = value.split('.')
        if len(parts) == 4:
            for part in parts:
                num = int(part)
                if 0 <= num <= 255:
                    continue
                else:
                    return False
            return True
    except:
        pass

    # 检查是否为其他有效的LSR ID格式（可根据实际情况调整）
    if len(value) > 3 and not value.upper().startswith(('DEST', 'TYPE', 'STAT')):
        return True

    return False


def check_mpls_lsp(lsp):
    """检查MPLS LSP状态"""
    result = "normal"
    suggestions = "-"

    # 检查LSP状态
    if lsp['state'].lower() in ['down', 'inactive', 'failed']:
        result = "error"
        suggestions = "❗ LSP状态为down，需检查LDP邻居会话和接口状态"
    elif lsp['state'].lower() in ['up', 'active']:
        result = "normal"
        suggestions = "LSP状态正常"
    else:
        # 对于未知状态，标记为警告
        result = "warning"
        suggestions = f"⚠️ LSP状态未知: {lsp['state']}，建议检查"

    return result, suggestions


def check_mpls_lsp(lsp):
    suggestions = []
    result = "normal"

    # Early check for state
    if lsp['state'].lower() != 'up':
        suggestions.append("❗ LSP状态为down，需检查LDP邻居会话和接口状态")
        return "error", "; ".join(suggestions)

    lsp_type = lsp['type'].lower()
    in_label = lsp['in_label']
    out_label = lsp['out_label']

    # Type-specific checks
    if lsp_type == 'ingress' and in_label != '-':
        suggestions.append("❗ Ingress条目显示具体入标签，可能配置错误")
        result = "error"
    elif lsp_type == 'transit' and (in_label == '-' or out_label == '-'):
        suggestions.append("❗ Transit LSP入/出标签缺失，可能配置或LDP问题")
        result = "error"
    elif lsp_type == 'egress' and (out_label != '-' or in_label == '-'):
        suggestions.append("❗ Egress LSP标签异常，出标签应为'-'且入标签应有效")
        result = "error"

    # Label range validation
    for label, name in [(in_label, "InLabel"), (out_label, "OutLabel")]:
        if label != '-':
            try:
                label_val = int(label)
                if label_val < 1 or label_val > 1048575:
                    suggestions.append(f"⚠️ {name}超出MPLS有效范围(1-1048575)")
                    result = "error"
            except ValueError:
                suggestions.append("⚠️ 标签值解析失败")
                result = "error"

    return result, "; ".join(suggestions) if suggestions else "-"


def parse_bfd_sessions(brief_output, config_output, l2vc_output, ldp_detail_output=""):
    # Parse config_output to build config_data_by_local_id
    config_data_by_local_id = {}
    config_lines = config_output.split('\n')
    current_session_name = None
    local_config = {}
    remote_config = {}
    for line in config_lines:
        line = line.strip()
        if line.startswith('**********BFD Session Name:'):
            if current_session_name and 'Discr' in local_config:
                config_data_by_local_id[local_config['Discr']] = {
                    'session_name': current_session_name,
                    'send_interval': local_config.get('Desired Min Tx interval', '-'),
                    'detect_mult': local_config.get('Detect mult', '-'),
                    'local_discr': local_config.get('Discr', '-'),
                    'cc_en': local_config.get('CC En', '-'),
                    'mep_en': local_config.get('MEP En', '-'),
                    'receive_interval': remote_config.get('Required Min Rx interval', '-'),
                    'first_pkt': remote_config.get('1st Pkt', '-'),
                    'remote_discr': remote_config.get('Discr', '-'),
                }
            current_session_name = line.split(':')[1].split('*')[0].strip()
            local_config = {}
            remote_config = {}
            in_local_config = False
            in_remote_config = False
            continue
        if current_session_name:
            if line.startswith('BFD for PW Local config information:'):
                in_local_config = True
                in_remote_config = False
                continue
            elif line.startswith('BFD for PW Remote config information:'):
                in_local_config = False
                in_remote_config = True
                continue
            if in_local_config or in_remote_config:
                parts = line.split(',')
                for part in parts:
                    if ':' in part:
                        key, value = part.split(':', 1)
                        key = key.strip()
                        value = value.strip()
                        if in_local_config:
                            local_config[key] = value
                        elif in_remote_config:
                            remote_config[key] = value
    # Process the last session
    if current_session_name and 'Discr' in local_config:
        config_data_by_local_id[local_config['Discr']] = {
            'session_name': current_session_name,
            'send_interval': local_config.get('Desired Min Tx interval', '-'),
            'detect_mult': local_config.get('Detect mult', '-'),
            'local_discr': local_config.get('Discr', '-'),
            'cc_en': local_config.get('CC En', '-'),
            'mep_en': local_config.get('MEP En', '-'),
            'receive_interval': remote_config.get('Required Min Rx interval', '-'),
            'first_pkt': remote_config.get('1st Pkt', '-'),
            'remote_discr': remote_config.get('Discr', '-'),
        }

    # Parse l2vc_output to build l2vc_data_by_vcid
    l2vc_data_by_vcid = {}
    l2vc_lines = l2vc_output.split('\n')
    in_table = False
    for line in l2vc_lines:
        line = line.strip()
        if "VC-ID" in line and "Destination" in line:
            in_table = True
            continue
        if in_table and line and not line.startswith('-'):
            # 使用更灵活的拆分方法处理多余空格
            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) >= 6:  # 只需要 6 个字段即可
                vcid = parts[0]
                destination = parts[1]
                service_name = parts[2]
                vc_state = parts[3]
                interface = parts[4]
                vc_type = parts[5]
                l2vc_data_by_vcid[vcid] = {
                    'destination': destination,
                    'service_name': service_name,
                    'vc_state': '✅ UP' if vc_state.lower() == 'up' else '❌ Down',
                    'interface': interface,
                    'vc_type': vc_type
                }

    # Parse LDP L2VC detail information
    ldp_data_by_vcid = parse_ldp_l2vc_detail(ldp_detail_output)

    # 调试：打印 l2vc_data_by_vcid 以验证所有 VCID 是否被捕获
    # print(f"Debug: l2vc_data_by_vcid = {l2vc_data_by_vcid}")
    # print(f"Debug: ldp_data_by_vcid = {ldp_data_by_vcid}")

    # Parse brief_output to get session list
    sessions = []
    brief_lines = brief_output.split('\n')
    in_table = False
    for line in brief_lines:
        line = line.strip()
        if "SessionType" in line and "ApsGroup" in line:
            in_table = True
            continue
        if in_table and line and not line.startswith('-'):
            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) >= 8 and parts[0] == "BFD For PW":
                local_id = parts[1]
                remote_id = parts[2]
                state = parts[3]
                vcid = parts[5]
                aps_group = parts[6]
                master_backup = parts[7]

                # Determine result
                result = 'error' if aps_group != '0' and state.lower() == 'down' else 'normal'

                # Get config details if aps_group != 0
                if aps_group != '0':
                    config = config_data_by_local_id.get(local_id, {})
                    session_name = config.get('session_name', '-')
                    send_interval = config.get('send_interval', '-')
                    if send_interval != '-':
                        send_interval += ' ms'
                    receive_interval = config.get('receive_interval', '-')
                    if receive_interval != '-':
                        receive_interval += ' ms'
                    detect_mult = config.get('detect_mult', '-')
                    local_discr = config.get('local_discr', '-')
                    remote_discr = config.get('remote_discr', '-')
                    cc_en = config.get('cc_en', '-')
                    mep_en = config.get('mep_en', '-')
                    first_pkt = config.get('first_pkt', '-')
                else:
                    session_name = '-'
                    send_interval = 'N/A'
                    receive_interval = 'N/A'
                    detect_mult = 'N/A'
                    local_discr = local_id
                    remote_discr = remote_id
                    cc_en = 'N/A'
                    mep_en = 'N/A'
                    first_pkt = 'N/A'

                # Get L2VC details
                l2vc = l2vc_data_by_vcid.get(vcid, {})
                destination = l2vc.get('destination', '-')
                service_name = l2vc.get('service_name', '-')
                vc_state = l2vc.get('vc_state', '-')
                interface = l2vc.get('interface', '-')
                vc_type = l2vc.get('vc_type', '-')

                # Get LDP L2VC detail information
                ldp_detail = ldp_data_by_vcid.get(vcid, {})
                local_mtu = ldp_detail.get('local_mtu', '-')
                remote_mtu = ldp_detail.get('remote_mtu', '-')
                local_control_word = ldp_detail.get('local_control_word', '-')
                remote_control_word = ldp_detail.get(
                    'remote_control_word', '-')
                current_control_word = ldp_detail.get(
                    'current_control_word', '-')
                local_pw_status_capability = ldp_detail.get(
                    'local_pw_status_capability', '-')
                remote_pw_status_capability = ldp_detail.get(
                    'remote_pw_status_capability', '-')
                current_pw_status_tlv = ldp_detail.get(
                    'current_pw_status_tlv', '-')
                local_pw_status = ldp_detail.get('local_pw_status', '-')
                remote_pw_status = ldp_detail.get('remote_pw_status', '-')
                local_vccv_capability = ldp_detail.get(
                    'local_vccv_capability', '-')
                remote_vccv_capability = ldp_detail.get(
                    'remote_vccv_capability', '-')

                # Format display fields
                state_display = '✅ UP' if state.lower() == 'up' else '❌ Down'
                master_backup_display = '主用（Master）' if master_backup.lower() == 'master' else '备用（Backup）'
                try:
                    remote_discr_num = int(remote_discr)
                    discr_state = '✅ up' if local_discr != '-' and remote_discr_num > 0 else '❌ down'
                except (ValueError, TypeError):
                    discr_state = '❌ down'
                first_pkt_display = '1（已接收）' if first_pkt == '1' else '0（未接收）'
                cc_en_display = '1（启用）' if cc_en == '1' else '0（禁用）'
                mep_en_display = '1（启用）' if mep_en == '1' else '0（禁用）'

                sessions.append({
                    'aps_group': aps_group,
                    'session_name': session_name,
                    'local_id': local_id,
                    'remote_id': remote_id,
                    'state': state_display,
                    'master_backup': master_backup_display,
                    'send_interval': send_interval,
                    'receive_interval': receive_interval,
                    'detect_mult': detect_mult,
                    'local_discr': local_discr,
                    'remote_discr': remote_discr,
                    'discr_state': discr_state,
                    'first_pkt': first_pkt_display,
                    'cc_en': cc_en_display,
                    'mep_en': mep_en_display,
                    'vcid': vcid,
                    'destination': destination,
                    'service_name': service_name,
                    'vc_state': vc_state,
                    'interface': interface,
                    'local_mtu': local_mtu,
                    'remote_mtu': remote_mtu,
                    'vc_type': vc_type,
                    'local_control_word': local_control_word,
                    'remote_control_word': remote_control_word,
                    'current_control_word': current_control_word,
                    'local_pw_status_capability': local_pw_status_capability,
                    'remote_pw_status_capability': remote_pw_status_capability,
                    'current_pw_status_tlv': current_pw_status_tlv,
                    'local_pw_status': local_pw_status,
                    'remote_pw_status': remote_pw_status,
                    'local_vccv_capability': local_vccv_capability,
                    'remote_vccv_capability': remote_vccv_capability,
                    'result': result
                })

    # Handle no entries
    if not sessions:
        sessions.append({
            'aps_group': '-', 'session_name': '无条目', 'local_id': '-', 'remote_id': '-',
            'state': '-', 'master_backup': '-', 'send_interval': '-', 'receive_interval': '-',
            'detect_mult': '-', 'local_discr': '-', 'remote_discr': '-', 'discr_state': '-',
            'first_pkt': '-', 'cc_en': '-', 'mep_en': '-', 'vcid': '-',
            'destination': '-', 'service_name': '-', 'vc_state': '-', 'interface': '-',
            'local_mtu': '-', 'remote_mtu': '-', 'vc_type': '-',
            'local_control_word': '-', 'remote_control_word': '-', 'current_control_word': '-',
            'local_pw_status_capability': '-', 'remote_pw_status_capability': '-',
            'current_pw_status_tlv': '-', 'local_pw_status': '-', 'remote_pw_status': '-',
            'local_vccv_capability': '-', 'remote_vccv_capability': '-',
            'result': 'normal'
        })

    return sessions

def parse_l2vc_summary(l2vc_output):
    """解析show mpls l2vc brief命令输出，提取VC总数、UP数、DOWN数"""
    total_vc = 0
    up_count = 0
    down_count = 0
    
    lines = l2vc_output.split('\n')
    for line in lines:
        line = line.strip()
        # 查找格式: Total LDP VC : 2, 2 up, 0 down
        if line.startswith('Total LDP VC'):
            try:
                # 提取总数
                parts = line.split(':')
                if len(parts) >= 2:
                    counts_part = parts[1].strip()
                    # 分割成 "2, 2 up, 0 down"
                    count_parts = counts_part.split(',')
                    if len(count_parts) >= 1:
                        total_vc = int(count_parts[0].strip())
                    if len(count_parts) >= 2:
                        up_match = count_parts[1].strip().split()
                        if len(up_match) >= 1:
                            up_count = int(up_match[0])
                    if len(count_parts) >= 3:
                        down_match = count_parts[2].strip().split()
                        if len(down_match) >= 1:
                            down_count = int(down_match[0])
            except (ValueError, IndexError) as e:
                print(f"解析VC统计信息时出错: {e}")
                pass
            break
    
    return total_vc, up_count, down_count

def parse_ldp_l2vc_detail(ldp_detail_output):
    """解析show ldp l2vc detail命令输出，返回按VCID索引的详细信息"""
    ldp_data_by_vcid = {}

    if not ldp_detail_output:
        print("Debug: ldp_detail_output 为空")
        return ldp_data_by_vcid

    print(f"Debug: 开始解析LDP详细输出，总长度: {len(ldp_detail_output)}")

    lines = ldp_detail_output.split('\n')
    current_vcid = None
    current_data = {}
    parsing_local_pw_status = False
    parsing_remote_pw_status = False
    parsing_local_vccv = False
    parsing_remote_vccv = False

    for i, line in enumerate(lines):
        original_line = line
        line = line.strip()

        # 打印前50行的调试信息以便观察
        # if i < 50:
        #     print(f"Debug: 行{i}: '{original_line}' -> '{line}'")

        # 匹配VCID行 - 格式: vcid: 105, type: ethernet, ...
        if line.startswith('vcid:'):
            # print(f"Debug: 找到VCID行: {line}")

            # 保存上一个VCID的数据
            if current_vcid and current_data:
                # print(f"Debug: 保存VCID {current_vcid} 的数据: {current_data}")
                ldp_data_by_vcid[current_vcid] = current_data.copy()

            # 开始新的VCID - 提取VCID号码
            try:
                vcid_part = line.split(',')[0]  # 获取 "vcid: 105" 部分
                current_vcid = vcid_part.split(':')[1].strip()
                # print(f"Debug: 提取到VCID: {current_vcid}")
            except Exception as e:
                # print(f"Debug: VCID提取失败: {e}")
                continue

            current_data = {
                'local_mtu': '-',
                'remote_mtu': '-',
                'local_control_word': '-',
                'remote_control_word': '-',
                'current_control_word': '-',
                'local_pw_status_capability': '-',
                'remote_pw_status_capability': '-',
                'current_pw_status_tlv': '-',
                'local_pw_status': '',
                'remote_pw_status': '',
                'local_vccv_capability': '-',
                'remote_vccv_capability': '-'
            }

            # 重置解析状态
            parsing_local_pw_status = False
            parsing_remote_pw_status = False
            parsing_local_vccv = False
            parsing_remote_vccv = False
            continue

        if not current_vcid:
            continue

        # 解析MTU信息 - 格式: Local MTU: 1500, Remote MTU: 1500
        if 'Local MTU:' in line and 'Remote MTU:' in line:
            # print(f"Debug: 找到MTU行: {line}")
            parts = line.split(',')
            for part in parts:
                part = part.strip()
                if 'Local MTU:' in part:
                    current_data['local_mtu'] = part.split(':')[1].strip()
                elif 'Remote MTU:' in part:
                    current_data['remote_mtu'] = part.split(':')[1].strip()
            # print(f"Debug: MTU解析结果 - Local: {current_data['local_mtu']}, Remote: {current_data['remote_mtu']}")

        # 解析控制字信息
        elif 'Local Control Word:' in line:
            # print(f"Debug: 找到控制字行: {line}")
            parts = line.split(',')
            for part in parts:
                part = part.strip()
                if 'Local Control Word:' in part:
                    current_data['local_control_word'] = part.split(':')[
                        1].strip()
                elif 'Remote Control Word:' in part:
                    current_data['remote_control_word'] = part.split(':')[
                        1].strip()
                elif 'Current use:' in part:
                    current_data['current_control_word'] = part.split(':')[
                        1].strip()
            # print(f"Debug: 控制字解析结果 - Local: {current_data['local_control_word']}, Remote: {current_data['remote_control_word']}, Current: {current_data['current_control_word']}")

        # 解析PW状态能力
        elif 'Local PW Status Capability' in line:
            # print(f"Debug: 找到本地PW状态能力: {line}")
            current_data['local_pw_status_capability'] = line.split(':')[
                1].strip()
        elif 'Remote PW Status Capability' in line:
            # print(f"Debug: 找到远程PW状态能力: {line}")
            current_data['remote_pw_status_capability'] = line.split(':')[
                1].strip()
        elif 'Current PW Status TLV' in line:
            # print(f"Debug: 找到当前PW状态TLV: {line}")
            current_data['current_pw_status_tlv'] = line.split(':')[1].strip()

        # 解析本地PW状态 - 修正：更准确地识别状态标题行
        elif line == 'Local PW Status :' or line.startswith('Local PW Status'):
            # print(f"Debug: 开始解析本地PW状态")
            parsing_local_pw_status = True
            parsing_remote_pw_status = False
            parsing_local_vccv = False
            parsing_remote_vccv = False
            current_data['local_pw_status'] = ''
            continue
        elif line == 'Remote PW Status :' or line.startswith('Remote PW Status'):
            # print(f"Debug: 开始解析远程PW状态")
            parsing_local_pw_status = False
            parsing_remote_pw_status = True
            parsing_local_vccv = False
            parsing_remote_vccv = False
            current_data['remote_pw_status'] = ''
            continue

        # 解析VCCV能力
        elif 'Local VCCV Capability:' in line:
            # print(f"Debug: 开始解析本地VCCV能力")
            parsing_local_pw_status = False
            parsing_remote_pw_status = False
            parsing_local_vccv = True
            parsing_remote_vccv = False
            continue
        elif 'Remote VCCV Capability:' in line:
            # print(f"Debug: 开始解析远程VCCV能力")
            parsing_local_pw_status = False
            parsing_remote_pw_status = False
            parsing_local_vccv = False
            parsing_remote_vccv = True
            continue

        # 处理状态行 - 修正：更精确的状态解析逻辑
        elif parsing_local_pw_status:
            # 检查是否遇到新的段落开始（结束当前状态解析）
            if (line.startswith('Remote PW Status') or
                line.startswith('Local VCCV') or
                line.startswith('Remote VCCV') or
                    line.startswith('vcid:')):
                parsing_local_pw_status = False
                # 处理新段落的开始
                if line.startswith('Remote PW Status'):
                    parsing_remote_pw_status = True
                    current_data['remote_pw_status'] = ''
                    continue
                elif line.startswith('Local VCCV'):
                    parsing_local_vccv = True
                    continue
                elif line.startswith('Remote VCCV'):
                    parsing_remote_vccv = True
                    continue
                # 如果是vcid行，不要continue，让它在下一轮被处理

            # 修正：更宽松的状态值识别条件
            elif line and line not in ['', ' '] and not line.startswith('CC-Types:') and not line.startswith('CV-Types:'):
                # 检查是否是有效的状态值（排除明显的非状态行）
                if not any(keyword in line for keyword in ['destination:', 'Local label:', 'Access IF:', 'Network IF:']):
                    # print(f"Debug: 找到本地PW状态: '{line}'")
                    if current_data['local_pw_status']:
                        current_data['local_pw_status'] += ', ' + line
                    else:
                        current_data['local_pw_status'] = line

        elif parsing_remote_pw_status:
            # 检查是否遇到新的段落开始
            if (line.startswith('Local VCCV') or
                line.startswith('Remote VCCV') or
                    line.startswith('vcid:')):
                parsing_remote_pw_status = False
                if line.startswith('Local VCCV'):
                    parsing_local_vccv = True
                    continue
                elif line.startswith('Remote VCCV'):
                    parsing_remote_vccv = True
                    continue
                # 如果是vcid行，不要continue，让它在下一轮被处理

            # 修正：更宽松的状态值识别条件
            elif line and line not in ['', ' '] and not line.startswith('CC-Types:') and not line.startswith('CV-Types:'):
                # 检查是否是有效的状态值（排除明显的非状态行）
                if not any(keyword in line for keyword in ['destination:', 'Local label:', 'Access IF:', 'Network IF:']):
                    # print(f"Debug: 找到远程PW状态: '{line}'")
                    if current_data['remote_pw_status']:
                        current_data['remote_pw_status'] += ', ' + line
                    else:
                        current_data['remote_pw_status'] = line

        # 处理VCCV CC-Types
        elif parsing_local_vccv and 'CC-Types:' in line:
            cc_types = line.replace('CC-Types:', '').strip()
            # print(f"Debug: 找到本地VCCV CC-Types: {cc_types}")
            current_data['local_vccv_capability'] = cc_types
        elif parsing_remote_vccv and 'CC-Types:' in line:
            cc_types = line.replace('CC-Types:', '').strip()
            # print(f"Debug: 找到远程VCCV CC-Types: {cc_types}")
            current_data['remote_vccv_capability'] = cc_types

    # 保存最后一个VCID的数据
    if current_vcid and current_data:
        # print(f"Debug: 保存最后一个VCID {current_vcid} 的数据: {current_data}")
        ldp_data_by_vcid[current_vcid] = current_data

    # print(f"Debug: 最终解析结果: {ldp_data_by_vcid}")
    return ldp_data_by_vcid


def parse_cfgchk_info(output):
    # Handle empty or invalid output
    if not output or output.strip() == '':
        return {
            'status': '⚠️ 数据异常',
            'minute_per_hour': '-',
            'recovery_time': '-',
            'result': 'error'
        }

    lines = output.split('\n')
    status = None
    minute_per_hour = None
    recovery_time = None

    for line in lines:
        line = line.strip()

        # 跳过空行和命令提示符行
        if not line or line.startswith('<') or line.startswith('PC_TIME:'):
            continue

        # 解析状态信息，使用更宽松的匹配
        if 'cfgchk en' in line and ':' in line:
            try:
                status = line.split(':', 1)[1].strip()
            except IndexError:
                continue

        elif 'cfgchk minute per hour' in line and ':' in line:
            try:
                # 提取冒号后的内容，然后去掉括号部分
                value_part = line.split(':', 1)[1].strip()
                minute_per_hour = value_part.split('(')[0].strip()
            except IndexError:
                continue

        elif 'cfgchk recovery time' in line and ':' in line:
            try:
                # 提取冒号后的内容，然后去掉括号部分
                value_part = line.split(':', 1)[1].strip()
                recovery_time = value_part.split('(')[0].strip()
            except IndexError:
                continue

    # 调试信息 - 可以在生产环境中移除
    print(
        f"解析结果: status={status}, minute_per_hour={minute_per_hour}, recovery_time={recovery_time}")

    # 数据异常处理 - 只有当所有关键字段都为空时才认为是异常
    if status is None and minute_per_hour is None and recovery_time is None:
        return {
            'status': '⚠️ 数据异常',
            'minute_per_hour': '-',
            'recovery_time': '-',
            'result': 'error'
        }

    # 确定状态和结果
    result_status = 'normal' if status and status.lower() == 'enable' else 'error'
    display_status = '✅ Enable' if status and status.lower() == 'enable' else '❌ Disable'

    return {
        'status': display_status,
        'minute_per_hour': minute_per_hour or '-',
        'recovery_time': recovery_time or '-',
        'result': result_status
    }


def execute_command_with_timestamp(ip, command):
    """
    执行命令并记录PC时间戳
    这个函数应该在数据采集阶段使用
    """
    pc_time = datetime.now()

    # 执行实际命令的代码（这里需要替换为实际的命令执行逻辑）
    result = ""  # 这里应该是实际命令执行的结果
    # 例如：result = ssh_execute_command(ip, command) 或其他执行方式

    # 在命令输出前添加PC时间戳
    timestamped_output = f"PC_EXEC_TIME: {pc_time.strftime('%Y-%m-%d %H:%M:%S')}\n" + result
    return timestamped_output

# 修复后的解析函数


def extract_pc_time(output):
    """
    从命令输出中提取PC_TIME
    """
    pc_time_match = re.search(r'PC_TIME:\s*(.+)', output)
    if pc_time_match:
        try:
            pc_time = datetime.strptime(pc_time_match.group(1).strip(), '%Y-%m-%d %H:%M:%S')
            return pc_time
        except ValueError as e:
            print(f"{Fore.YELLOW}[WARNING] PC_TIME解析失败: {e}{Style.RESET_ALL}")
    return None

def parse_ntp_status(cloc_output, ntp_output, pc_time=None):
    """
    解析NTP状态，使用提供的PC时间计算偏差
    如果没有提供pc_time，尝试从输出中提取
    """
    # 如果没有提供PC时间，尝试从输出中提取
    if pc_time is None:
        pc_time = extract_pc_time(cloc_output)
        if pc_time is None:
            pc_time = extract_pc_time(ntp_output)
        if pc_time is None:
            print(f"{Fore.YELLOW}[WARNING] 无法从输出中提取PC_TIME，使用当前时间{Style.RESET_ALL}")
            pc_time = datetime.now()
    
    # 解析设备本地时间
    local_time_str = re.search(r'LOCAL TIME\s*:\s*(.+)', cloc_output)
    utc_time_str = re.search(r'UTC TIME\s*:\s*(.+)', cloc_output)
    time_zone_str = re.search(r'TIME-Zone\s*:\s*(.+)', cloc_output)

    local_time = None
    if local_time_str:
        try:
            local_time = datetime.strptime(
                local_time_str.group(1).strip(), '%Y-%m-%d %H:%M:%S')
            print(f"[DEBUG] 设备本地时间: {local_time}")
        except ValueError as e:
            print(f"[ERROR] 设备本地时间解析失败: {e}")

    utc_time = None
    if utc_time_str:
        try:
            utc_time = datetime.strptime(
                utc_time_str.group(1).strip(), '%Y-%m-%d %H:%M:%S')
        except ValueError as e:
            print(f"[DEBUG] UTC时间解析失败: {e}")

    time_zone = time_zone_str.group(1).strip() if time_zone_str else "-"

    # 计算时间偏差
    result = "error"
    time_deviation = "无法计算"
    pc_time_str = pc_time.strftime('%Y-%m-%d %H:%M:%S')

    if local_time:
        # 计算时间差（秒）
        time_diff = abs((local_time - pc_time).total_seconds())
        time_deviation = f"{time_diff:.0f}(s)"

        print(f"[DEBUG] PC执行时间: {pc_time}")
        print(f"[DEBUG] 设备本地时间: {local_time}")
        print(f"[DEBUG] 时间偏差: {time_diff} 秒")

        # 基于时间偏差判断
        if time_diff <= 60:  # 60秒内为正常
            result = "normal"
        elif time_diff <= 300:  # 5分钟内为警告
            result = "warning"
        else:  # 超过5分钟为错误
            result = "error"
    else:
        time_deviation = "缺少设备时间"
        result = "error"

    # 解析NTP服务配置
    ntp_enable = re.search(r'ntp enable\s*:\s*(.+)', ntp_output)
    ntp_status = re.search(r'ntp clock status\s*:\s*(.+)', ntp_output)
    ntp_syn_interval = re.search(r'ntp syn-interval\s*:\s*(.+)', ntp_output)
    ntp_server_pref = re.search(
        r'ntp server preference\s*:\s*(.+)', ntp_output)
    ntp_server = re.search(r'ntp server\s*:\s*(.+)', ntp_output)

    ntp_enable = ntp_enable.group(1).strip() if ntp_enable else "-"
    ntp_status = ntp_status.group(1).strip() if ntp_status else "-"
    syn_interval = ntp_syn_interval.group(
        1).strip() if ntp_syn_interval else "-"
    server_pref = ntp_server_pref.group(1).strip() if ntp_server_pref else "-"
    server = ntp_server.group(1).strip() if ntp_server else "-"

    return {
        "ntp_enable": ntp_enable,
        "ntp_status": ntp_status,
        "syn_interval": syn_interval,
        "time_deviation": time_deviation,
        "server_pref": server_pref,
        "server": server,
        "local_time": local_time.strftime('%Y-%m-%d %H:%M:%S') if local_time else "-",
        "utc_time": utc_time.strftime('%Y-%m-%d %H:%M:%S') if utc_time else "-",
        "time_zone": time_zone,
        "pc_time": pc_time_str,
        "result": result
    }

def parse_flash_usage(output):
    """解析 'show flash-usage' 输出以获取硬盘资源占用状态"""
    total_flash = re.search(r'System Total Flash\s*:\s*(\d+)\s*bytes', output)
    free_space = re.search(r'Flash Free Space\s*:\s*(\d+)\s*bytes', output)
    usage_percent = re.search(r'Flash Usage\s*:\s*(\d+)%', output)
    threshold = re.search(r'Flash Usage Threshold\s*:\s*(\d+)%', output)

    total_flash = int(total_flash.group(1)) if total_flash else 0
    free_space = int(free_space.group(1)) if free_space else 0
    usage_percent = int(usage_percent.group(1)) if usage_percent else 0
    threshold = int(threshold.group(1)) if threshold else 0

    # 转换为 MB 并保留两位小数
    total_flash_mb = f"{total_flash / (1024 * 1024):.2f} MB"
    free_space_mb = f"{free_space / (1024 * 1024):.2f} MB"

    # 判断使用率
    result = "error" if usage_percent > 70 else "normal"

    return {
        "total_flash": total_flash_mb,
        "free_space": free_space_mb,
        "usage_percent": f"{usage_percent}%",
        "threshold": f"{threshold}%",
        "result": result
    }


def parse_main_backup_version(output, ip_address=None):
    """Parse output for main and backup control board software version consistency."""
    # print(f"{Fore.CYAN}[DEBUG] 开始解析输出以检查主备主控软件版本一致性{Style.RESET_ALL}")
    # print(f"{Fore.CYAN}[DEBUG] 设备IP: {ip_address}{Style.RESET_ALL}")
    # print(f"{Fore.CYAN}[DEBUG] 输入原始数据:\n{output}\n{Style.RESET_ALL}")
    lines = output.split('\n')
    ne_type = None
    device_name = None
    main_version = None
    backup_version = None
    result = "error"

    for i, line in enumerate(lines):
        line = line.strip()
        # print(f"{Fore.BLUE}[DEBUG] 处理行 {i}: '{line}'{Style.RESET_ALL}")

        # 提取设备名称（如果存在）
        if line.startswith('<') and line.endswith('>'):
            device_name = line[1:-1]
            # print(
            # f"{Fore.YELLOW}[DEBUG] 从<...>提取设备名称: {device_name}{Style.RESET_ALL}")
            # 从设备名称初步推断网元类型
            if re.search(r'MssEdge\s*20\s*-?\s*A3', device_name):
                ne_type = "MssEdge 20 A3"
            elif re.search(r'MssEdge\s*20\s*-?\s*A2', device_name):
                ne_type = "MssEdge 20 A2"
            elif re.search(r'MssEdge\s*20(\s*-?\s*A1)?', device_name):
                ne_type = "MssEdge 20 A1"
            elif re.search(r'MssEdge\s*25\s*-?\s*S10-3', device_name):
                ne_type = "MssEdge 25 S10-3"
            elif re.search(r'MssEdge\s*25\s*-?\s*S10', device_name):
                ne_type = "MssEdge 25 S10"
            # print(
                # f"{Fore.YELLOW}[DEBUG] 从设备名称初始提取网元类型: {ne_type or '-'}{Style.RESET_ALL}")

        # 从stn-standard-reserved后面的行获取网元类型信息
        if "stn-standard-reserved" in line:
            if i + 1 < len(lines):
                ne_type_full = lines[i + 1].strip()
                print(
                    f"{Fore.YELLOW}[DEBUG] stn-standard-reserved 下一行: {ne_type_full}{Style.RESET_ALL}")

                # 改进的网元类型匹配逻辑
                if 'MssEdge 25 S10-3' in ne_type_full:
                    ne_type = "MssEdge 25 S10-3"
                elif 'MssEdge 25 S10' in ne_type_full and 'S10-3' not in ne_type_full:
                    ne_type = "MssEdge 25 S10"
                elif 'MssEdge20-A3' in ne_type_full or 'MssEdge 20-A3' in ne_type_full:
                    ne_type = "MssEdge 20 A3"
                elif 'MssEdge20-A2' in ne_type_full or 'MssEdge 20-A2' in ne_type_full:
                    ne_type = "MssEdge 20 A2"
                elif 'MssEdge20' in ne_type_full or 'MssEdge 20' in ne_type_full:
                    ne_type = "MssEdge 20 A1"

                print(
                    f"{Fore.YELLOW}[DEBUG] 从stn-standard-reserved更新网元类型: {ne_type or '-'}{Style.RESET_ALL}")

        # 从system name提取设备名称（作为备用方案）
        if device_name is None and line.startswith('system name'):
            system_name = line.split(':', 1)[1].strip()
            if system_name and system_name != "OPTEL":  # 排除默认值
                device_name = system_name
                print(
                    f"{Fore.YELLOW}[DEBUG] 从system name提取设备名称: {device_name}{Style.RESET_ALL}")

        # 解析版本信息
        if line.startswith('system info'):
            system_info = line.split(':', 1)[1].strip()
            # print(
            #     f"{Fore.BLUE}[DEBUG] 原始系统信息: '{system_info}'{Style.RESET_ALL}")

            # 匹配标准格式: O123456789 (123456789)
            match = re.search(r'O(\d+)\s*\((\d+)\)', system_info)
            if match:
                main_version, backup_version = match.groups()
                result = "normal" if main_version == backup_version else "error"
                # print(
                #     f"{Fore.YELLOW}[DEBUG] 标准格式匹配: 主用={main_version}, 备用={backup_version}, Result={result}{Style.RESET_ALL}")
            else:
                # 匹配单版本格式（修正的逻辑）
                alt_match = re.search(
                    r'O(\d+)(?:uptime:|$|\s*\(\s*\))', system_info)
                if alt_match:
                    version_number = alt_match.group(1)
                    main_version = version_number
                    backup_version = "-"  # 修正：单版本时备用设为"-"
                    result = "error"      # 修正：单版本时状态设为"error"
                    print(
                        f"{Fore.RED}[DEBUG] 单版本格式匹配: 主用={main_version}, 备用={backup_version}, Result={result}{Style.RESET_ALL}")
                else:
                    # 提取任何数字序列作为后备方案
                    number_match = re.search(r'O?(\d+)', system_info)
                    if number_match:
                        version_number = number_match.group(1)
                        main_version = version_number
                        backup_version = "-"
                        result = "error"
                        print(
                            f"{Fore.RED}[DEBUG] 部分匹配: 主用={version_number}, 备用=未知, Result={result}{Style.RESET_ALL}")
                    else:
                        print(
                            f"{Fore.RED}[DEBUG] 无法解析系统信息: '{system_info}'{Style.RESET_ALL}")

    # 如果没有从<>中获取到设备名称，生成一个基于IP的设备名称
    if device_name is None or device_name == "-":
        if ip_address and ne_type and ne_type != "-":
            # 使用IP地址和网元类型生成设备名称
            device_name = f"{ne_type.replace(' ', '-')}-{ip_address}"
            print(
                f"{Fore.YELLOW}[DEBUG] 基于IP生成设备名称: {device_name}{Style.RESET_ALL}")
        else:
            device_name = "-"

    # 设置默认值并确保结果一致性
    ne_type = ne_type or "-"
    device_name = device_name or "-"
    main_version = main_version or "-"
    backup_version = backup_version or "-"
    if main_version == "-" or backup_version == "-":
        result = "error"

    # print(f"{Fore.GREEN}[DEBUG] 最终结果: NE类型={ne_type}, 设备名={device_name}, 主用={main_version}, 备用={backup_version}, 状态={result}{Style.RESET_ALL}")
    return (ne_type, device_name, main_version, backup_version, result)


def parse_board_cpu_memory(output_15m, output_24h):
    """Parse 'show pm cur-15m Dev' and 'show pm cur-24h Dev' outputs for board CPU and memory usage."""
    print(f"{Fore.CYAN}[DEBUG] 开始解析性能监控输出{Style.RESET_ALL}")
    results = []

    def parse_pm_output(output, time_frame):
        if not output or not output.strip():
            print(f"{Fore.YELLOW}[WARNING] {time_frame}性能监控输出为空{Style.RESET_ALL}")
            return []
            
        lines = output.split('\n')
        in_table = False
        data = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # 检测表头
            if line.startswith('Index') and 'PM-Source' in line:
                in_table = True
                continue
                
            # 跳过分隔符行
            if line.startswith('---'):
                continue
                
            if in_table and line:
                parts = line.split()
                print(f"{Fore.CYAN}[DEBUG] {time_frame} 解析行: {line} (字段数: {len(parts)}){Style.RESET_ALL}")
                
                # 检查是否有足够的字段（至少需要7个字段：Index, PM-Source, Time(2个), Temp, CPU, Memory）
                if len(parts) >= 7:
                    try:
                        pm_source = parts[1]
                        time = parts[2] + ' ' + parts[3]  # 合并日期和时间
                        temp = parts[4]
                        cpu_rate = parts[5]
                        mem_rate = parts[6]
                        
                        # 确保百分号存在
                        if '%' not in cpu_rate:
                            cpu_rate += '%'
                        if '%' not in mem_rate:
                            mem_rate += '%'
                        
                        # 判断是否正常
                        result = "normal"
                        try:
                            # 去除可能的非数字字符进行数值解析
                            temp_str = temp.replace('°C', '').replace('C', '')
                            temp_val = float(temp_str)
                            cpu_val = float(cpu_rate.rstrip('%'))
                            mem_val = float(mem_rate.rstrip('%'))
                            
                            if temp_val > 80 or cpu_val > 60 or mem_val > 65:
                                result = "error"
                                print(f"{Fore.YELLOW}[DEBUG] {time_frame} {pm_source}: 温度={temp_val}°C, CPU={cpu_val}%, 内存={mem_val}%, Result=error{Style.RESET_ALL}")
                            else:
                                print(f"{Fore.GREEN}[DEBUG] {time_frame} {pm_source}: 温度={temp_val}°C, CPU={cpu_val}%, 内存={mem_val}%, Result=normal{Style.RESET_ALL}")
                                
                        except (ValueError, TypeError) as e:
                            result = "error"
                            print(f"{Fore.RED}[ERROR] {time_frame} {pm_source} 数值解析失败: {e}, Result=error{Style.RESET_ALL}")
                        
                        data.append({
                            "pm_source": pm_source,
                            "time": time,
                            "temp": temp,
                            "cpu_rate": cpu_rate,
                            "mem_rate": mem_rate,
                            "result": result
                        })
                        
                    except IndexError as e:
                        print(f"{Fore.RED}[ERROR] {time_frame} 字段索引错误: {e}, 行内容: {line}{Style.RESET_ALL}")
                        continue
                        
                else:
                    # 字段不足，记录调试信息
                    print(f"{Fore.YELLOW}[DEBUG] {time_frame} 跳过字段不足的行: {line} (字段数: {len(parts)}){Style.RESET_ALL}")
                    continue
                    
        return data

    # 解析15分钟数据
    print(f"{Fore.CYAN}[DEBUG] 开始解析15分钟数据{Style.RESET_ALL}")
    data_15m = parse_pm_output(output_15m, "15分钟")
    
    # 解析24小时数据
    print(f"{Fore.CYAN}[DEBUG] 开始解析24小时数据{Style.RESET_ALL}")
    data_24h = parse_pm_output(output_24h, "24小时")

    # 合并15分钟和24小时数据
    pm_sources = set()
    if data_15m:
        pm_sources.update([d['pm_source'] for d in data_15m])
    if data_24h:
        pm_sources.update([d['pm_source'] for d in data_24h])
    
    print(f"{Fore.CYAN}[DEBUG] 发现的PM源: {list(pm_sources)}{Style.RESET_ALL}")
    
    if pm_sources:
        for pm_source in pm_sources:
            # 查找对应的15分钟和24小时数据
            result_15m = next((d for d in data_15m if d['pm_source'] == pm_source), None)
            result_24h = next((d for d in data_24h if d['pm_source'] == pm_source), None)
            
            # 确定最终结果
            final_result = "normal"
            if (result_15m and result_15m['result'] == "error") or (result_24h and result_24h['result'] == "error"):
                final_result = "error"
            
            # 构建结果字典
            result_dict = {
                "pm_source_15m": result_15m['pm_source'] if result_15m else "-",
                "time_15m": result_15m['time'] if result_15m else "-",
                "temp_15m": result_15m['temp'] + ("°C" if result_15m and "°C" not in result_15m['temp'] else "") if result_15m else "-",
                "cpu_15m": result_15m['cpu_rate'] if result_15m else "-",
                "mem_15m": result_15m['mem_rate'] if result_15m else "-",
                "pm_source_24h": result_24h['pm_source'] if result_24h else "-",
                "time_24h": result_24h['time'] if result_24h else "-",
                "temp_24h": result_24h['temp'] + ("°C" if result_24h and "°C" not in result_24h['temp'] else "") if result_24h else "-",
                "cpu_24h": result_24h['cpu_rate'] if result_24h else "-",
                "mem_24h": result_24h['mem_rate'] if result_24h else "-",
                "result": final_result
            }
            
            results.append(result_dict)
            print(f"{Fore.GREEN}[DEBUG] 合并 {pm_source} 数据，Result={final_result}{Style.RESET_ALL}")

    # 如果没有解析到任何数据，返回默认错误记录
    if not results:
        print(f"{Fore.RED}[WARNING] 未解析到任何性能监控数据{Style.RESET_ALL}")
        results.append({
            "pm_source_15m": "-",
            "time_15m": "-",
            "temp_15m": "-",
            "cpu_15m": "-",
            "mem_15m": "-",
            "pm_source_24h": "-",
            "time_24h": "-",
            "temp_24h": "-",
            "cpu_24h": "-",
            "mem_24h": "-",
            "result": "error"
        })
    
    print(f"{Fore.CYAN}[DEBUG] 性能监控解析完成，共解析到 {len(results)} 条记录{Style.RESET_ALL}")
    return results


def parse_optical_module(ip, interface_output, lldp_output, parse_uptime_func):
    """Parse 'show interface' and 'show lldp neighbor' outputs for optical module information."""
    print(f"{Fore.CYAN}[DEBUG] 开始解析设备 {ip} 的光模块信息{Style.RESET_ALL}")
    lines = interface_output.split('\n')
    lldp_lines = lldp_output.split('\n') if lldp_output else []
    results = []
    current_interface = None
    interface_data = {}
    lldp_data = {}

    # Parse LLDP neighbor information
    interface_name = None  # Initialize interface_name
    for line in lldp_lines:
        line = line.strip()
        if line.startswith("Interface"):
            interface_name = re.search(
                r"'([^']+)'", line).group(1) if re.search(r"'([^']+)'", line) else None
            if interface_name:
                lldp_data[interface_name] = {}
                print(
                    f"{Fore.YELLOW}[DEBUG] 发现LLDP接口: {interface_name}{Style.RESET_ALL}")
            else:
                print(
                    f"{Fore.YELLOW}[WARNING] 无效的LLDP接口行: {line}{Style.RESET_ALL}")
        elif interface_name and line.startswith("Neighbor"):
            neighbor_info = {}
            for neighbor_line in lldp_lines[lldp_lines.index(line) + 1:]:
                neighbor_line = neighbor_line.strip()
                if not neighbor_line or neighbor_line.startswith("Interface") or neighbor_line.startswith("End Of LLDPDU"):
                    break
                if "System Name:" in neighbor_line:
                    neighbor_info["system_name"] = neighbor_line.split(":", 1)[
                        1].strip()
                elif "System Description:" in neighbor_line:
                    neighbor_info["system_description"] = neighbor_line.split(":", 1)[
                        1].strip()
                elif "Port ID:" in neighbor_line:
                    neighbor_info["port_id"] = neighbor_line.split(
                        ":", 1)[1].strip().split(" - ")[-1]
                elif "Management Address: IPv4" in neighbor_line:
                    neighbor_info["ip"] = neighbor_line.split(
                        " - ")[1].split(" (")[0]
            lldp_data[interface_name] = neighbor_info
            print(
                f"{Fore.YELLOW}[DEBUG] 解析LLDP邻居信息: {interface_name} -> {neighbor_info}{Style.RESET_ALL}")

    # Parse interface information
    for line in lines:
        line = line.strip()
        if line.startswith(('gigabitethernet', 'xgigabitethernet', '50GE', 'loopback')):
            interface_name = line.split(' current state')[0].strip()
            # Skip loopback interfaces
            if interface_name.lower().startswith('loopback'):
                print(
                    f"{Fore.YELLOW}[DEBUG] 跳过loopback接口: {interface_name}{Style.RESET_ALL}")
                current_interface = None
                continue
            current_interface = interface_name
            interface_data[current_interface] = {
                "current_state": line.split('current state : ')[1].strip(),
                "description": "-",
                "ipv4": "0.0.0.0/0",
                "ipv6": "-",
                "mac": "-",
                "mtu_l3": "-",
                "vendor_pn": "-",
                "vendor_name": "-",
                "transceiver_id": "-",
                "wavelength": "-",
                "distance": "-",
                "rx_power": "-",
                "rx_range": "-",
                "rx_alarm_range": "-",
                "tx_power": "-",
                "tx_range": "-",
                "tx_alarm_range": "-",
                "bias": "-",
                "bias_range": "-",
                "bias_alarm_range": "-",
                "voltage": "-",
                "voltage_range": "-",
                "voltage_alarm_range": "-",
                "temperature": "-",
                "temp_range": "-",
                "temp_alarm_range": "-",
                "port_bw": "-",
                "transceiver_bw": "-",
                "input_rate": "-",
                "input_util": "-",
                "output_rate": "-",
                "output_util": "-",
                "crc": 0,
                "last_up": "-",
                "last_down": "-"
            }
            print(
                f"{Fore.YELLOW}[DEBUG] 发现接口: {current_interface}{Style.RESET_ALL}")
        elif current_interface:
            if "Last physical up time" in line:
                interface_data[current_interface]["last_up"] = line.split(":", 1)[
                    1].strip()
            elif "Last physical down time" in line:
                interface_data[current_interface]["last_down"] = line.split(":", 1)[
                    1].strip()
            elif "Description:" in line:
                interface_data[current_interface]["description"] = line.split(":", 1)[
                    1].strip()
            elif "Internet IPV4 Address is" in line:
                interface_data[current_interface]["ipv4"] = line.split("is")[
                    1].strip()
            elif "IPV6 Address is" in line:
                interface_data[current_interface]["ipv6"] = line.split("is")[
                    1].strip()
            elif "Hardware address is" in line:
                # Extract only the MAC address after 'Hardware address is'
                mac_part = line.split("Hardware address is")[1].strip()
                # Handle cases where the line includes frame format (e.g., PKTFMT_ETHNT_2)
                if "," in mac_part:
                    mac_part = mac_part.split(",")[-1].strip()
                interface_data[current_interface]["mac"] = mac_part
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {current_interface} 解析MAC地址: {mac_part}{Style.RESET_ALL}")
            elif "The Maximum Transmit Unit(L3) is" in line:
                interface_data[current_interface]["mtu_l3"] = line.split("is")[1].split(",")[
                    0].strip()
            elif "The Vendor PN :" in line:
                interface_data[current_interface]["vendor_pn"] = line.split(":", 1)[
                    1].strip()
            elif "The Vendor Name :" in line:
                interface_data[current_interface]["vendor_name"] = line.split(":", 1)[
                    1].strip()
            elif "Transceiver Identifier:" in line:
                interface_data[current_interface]["transceiver_id"] = line.split(":", 1)[
                    1].strip()
            elif "WaveLength:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["wavelength"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Transmission Distance:" in parts[1]:
                    interface_data[current_interface]["distance"] = parts[1].split(":", 1)[
                        1].strip()
            elif "Rx Power:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["rx_power"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Warning range:" in parts[1]:
                    warning_range = parts[1].split(":", 1)[1].strip()
                    # Extract alarm range if present
                    if "Alarm range:" in warning_range:
                        warning_part, alarm_part = warning_range.split(
                            "Alarm range:")
                        interface_data[current_interface]["rx_range"] = warning_part.strip(
                        )
                        interface_data[current_interface]["rx_alarm_range"] = alarm_part.strip(
                        )
                    else:
                        interface_data[current_interface]["rx_range"] = warning_range
            elif "Tx Power:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["tx_power"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Warning range:" in parts[1]:
                    warning_range = parts[1].split(":", 1)[1].strip()
                    # Extract alarm range if present
                    if "Alarm range:" in warning_range:
                        warning_part, alarm_part = warning_range.split(
                            "Alarm range:")
                        interface_data[current_interface]["tx_range"] = warning_part.strip(
                        )
                        interface_data[current_interface]["tx_alarm_range"] = alarm_part.strip(
                        )
                    else:
                        interface_data[current_interface]["tx_range"] = warning_range
            elif "Bias:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["bias"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Warning range:" in parts[1]:
                    warning_range = parts[1].split(":", 1)[1].strip()
                    # Extract alarm range if present
                    if "Alarm range:" in warning_range:
                        warning_part, alarm_part = warning_range.split(
                            "Alarm range:")
                        interface_data[current_interface]["bias_range"] = warning_part.strip(
                        )
                        interface_data[current_interface]["bias_alarm_range"] = alarm_part.strip(
                        )
                    else:
                        interface_data[current_interface]["bias_range"] = warning_range
            elif "Voltage:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["voltage"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Warning range:" in parts[1]:
                    warning_range = parts[1].split(":", 1)[1].strip()
                    # Extract alarm range if present
                    if "Alarm range:" in warning_range:
                        warning_part, alarm_part = warning_range.split(
                            "Alarm range:")
                        interface_data[current_interface]["voltage_range"] = warning_part.strip(
                        )
                        interface_data[current_interface]["voltage_alarm_range"] = alarm_part.strip(
                        )
                    else:
                        interface_data[current_interface]["voltage_range"] = warning_range
            elif "temperature:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["temperature"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Warning range:" in parts[1]:
                    warning_range = parts[1].split(":", 1)[1].strip()
                    # Extract alarm range if present
                    if "Alarm range:" in warning_range:
                        warning_part, alarm_part = warning_range.split(
                            "Alarm range:")
                        interface_data[current_interface]["temp_range"] = warning_part.strip(
                        )
                        interface_data[current_interface]["temp_alarm_range"] = alarm_part.strip(
                        )
                    else:
                        interface_data[current_interface]["temp_range"] = warning_range
            elif "Port BW:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["port_bw"] = parts[0].split(":", 1)[
                    1].strip()
                if len(parts) > 1 and "Transceiver max BW:" in parts[1]:
                    interface_data[current_interface]["transceiver_bw"] = parts[1].split(":", 1)[
                        1].strip()
            elif "Input rate:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["input_rate"] = parts[0].split(":", 1)[
                    1].split(" bits")[0].strip()
                if len(parts) > 1 and "bandwidth utilization:" in parts[1]:
                    interface_data[current_interface]["input_util"] = parts[1].split(":", 1)[
                        1].strip()
            elif "Output rate:" in line:
                parts = line.split(",", 1)
                interface_data[current_interface]["output_rate"] = parts[0].split(":", 1)[
                    1].split(" bits")[0].strip()
                if len(parts) > 1 and "bandwidth utilization:" in parts[1]:
                    interface_data[current_interface]["output_util"] = parts[1].split(":", 1)[
                        1].strip()
            elif "CRC :" in line:
                crc_value = line.split(":", 1)[1].split(" packets")[0].strip()
                try:
                    interface_data[current_interface]["crc"] = int(crc_value)
                except ValueError:
                    interface_data[current_interface]["crc"] = 0
                    print(
                        f"{Fore.YELLOW}[WARNING] 设备 {ip} 接口 {current_interface} CRC解析失败: {crc_value}{Style.RESET_ALL}")

    # Helper function to convert range notation to min~max format
    def convert_range_format(range_str):
        if not range_str or range_str == "-":
            return "-"

        # Extract values from range notation like [3mA, 110mA]
        match = re.search(
            r'\[([-+]?\d+\.?\d*).*?,\s*([-+]?\d+\.?\d*).*?\]', range_str)
        if match:
            min_val, max_val = match.groups()
            return f"{min_val}~{max_val}"
        return range_str

    # Process each interface to generate results
    for interface, data in interface_data.items():
        # Skip non-optical interfaces (e.g., loopback)
        if data["transceiver_id"] == "-":
            continue

        # Convert all range formats to min~max
        rx_range = convert_range_format(data["rx_range"])
        rx_alarm_range = convert_range_format(data["rx_alarm_range"])
        tx_range = convert_range_format(data["tx_range"])
        tx_alarm_range = convert_range_format(data["tx_alarm_range"])
        bias_range = convert_range_format(data["bias_range"])
        bias_alarm_range = convert_range_format(data["bias_alarm_range"])
        voltage_range = convert_range_format(data["voltage_range"])
        voltage_alarm_range = convert_range_format(data["voltage_alarm_range"])
        temp_range = convert_range_format(data["temp_range"])
        temp_alarm_range = convert_range_format(data["temp_alarm_range"])

        # Extract range values for comparison
        def extract_range_values(range_str):
            if range_str == "-" or "~" not in range_str:
                return None, None
            try:
                min_val, max_val = range_str.split("~")
                return float(min_val), float(max_val)
            except (ValueError, TypeError):
                return None, None

        result = "normal"
        error_reasons = []

        # Check CRC
        if data["crc"] > 2048:
            result = "error"
            error_reasons.append(f"CRC超过2048: {data['crc']}")
            print(
                f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} CRC超过2048: {data['crc']}{Style.RESET_ALL}")

        # Check Rx Power
        try:
            rx_power = float(data["rx_power"].replace("dBm", ""))

            if rx_power == -40.0:
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} Rx光功率为-40.0dBm (收无光)，跳过范围判断{Style.RESET_ALL}")
            else:
                rx_min, rx_max = extract_range_values(
                    rx_alarm_range if rx_alarm_range != "-" else rx_range)

                if rx_min is not None and rx_max is not None and (rx_power < rx_min or rx_power > rx_max):
                    result = "error"
                    error_reasons.append(
                        f"Rx光功率超出范围: {rx_power}dBm 范围: {rx_min}~{rx_max}dBm")
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} Rx光功率异常: {rx_power}dBm 范围: {rx_min}~{rx_max}dBm{Style.RESET_ALL}")
        except (ValueError, TypeError):
            print(
                f"{Fore.YELLOW}[WARNING] 设备 {ip} 接口 {interface} Rx光功率解析失败: {data['rx_power']}{Style.RESET_ALL}")

        # Check Tx Power
        try:
            tx_power = float(data["tx_power"].replace("dBm", ""))
            tx_min, tx_max = extract_range_values(
                tx_alarm_range if tx_alarm_range != "-" else tx_range)

            if tx_min is not None and tx_power < tx_min:
                result = "error"
                error_reasons.append(
                    f"Tx光功率低于范围: {tx_power}dBm 最小值: {tx_min}dBm")
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} Tx光功率过低: {tx_power}dBm 最小值: {tx_min}dBm{Style.RESET_ALL}")
        except (ValueError, TypeError):
            print(
                f"{Fore.YELLOW}[WARNING] 设备 {ip} 接口 {interface} Tx光功率解析失败: {data['tx_power']}{Style.RESET_ALL}")

        # Check Bias
        try:
            bias = float(data["bias"].replace("mA", ""))
            bias_min, bias_max = extract_range_values(
                bias_alarm_range if bias_alarm_range != "-" else bias_range)

            if bias_min is not None and bias_max is not None and (bias < bias_min or bias > bias_max):
                result = "error"
                error_reasons.append(
                    f"偏置电流超出范围: {bias}mA 范围: {bias_min}~{bias_max}mA")
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} 偏置电流异常: {bias}mA 范围: {bias_min}~{bias_max}mA{Style.RESET_ALL}")
        except (ValueError, TypeError):
            print(
                f"{Fore.YELLOW}[WARNING] 设备 {ip} 接口 {interface} 偏置电流解析失败: {data['bias']}{Style.RESET_ALL}")

        # Check Voltage
        try:
            voltage = float(data["voltage"].replace("mV", ""))
            voltage_min, voltage_max = extract_range_values(
                voltage_alarm_range if voltage_alarm_range != "-" else voltage_range)

            if voltage_min is not None and voltage_max is not None and (voltage < voltage_min or voltage > voltage_max):
                result = "error"
                error_reasons.append(
                    f"电压超出范围: {voltage}mV 范围: {voltage_min}~{voltage_max}mV")
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} 电压异常: {voltage}mV 范围: {voltage_min}~{voltage_max}mV{Style.RESET_ALL}")
        except (ValueError, TypeError):
            print(
                f"{Fore.YELLOW}[WARNING] 设备 {ip} 接口 {interface} 电压解析失败: {data['voltage']}{Style.RESET_ALL}")

        # Check Temperature
        try:
            temp = float(data["temperature"].replace(" °C", ""))
            temp_min, temp_max = extract_range_values(
                temp_alarm_range if temp_alarm_range != "-" else temp_range)

            if temp_min is not None and temp_max is not None and (temp < temp_min or temp > temp_max):
                result = "error"
                error_reasons.append(
                    f"温度超出范围: {temp}°C 范围: {temp_min}~{temp_max}°C")
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} 温度异常: {temp}°C 范围: {temp_min}~{temp_max}°C{Style.RESET_ALL}")
        except (ValueError, TypeError):
            print(
                f"{Fore.YELLOW}[WARNING] 设备 {ip} 接口 {interface} 温度解析失败: {data['temperature']}{Style.RESET_ALL}")

        # Get LLDP neighbor data
        neighbor = lldp_data.get(interface, {})
        results.append({
            "interface": interface,
            "current_state": data["current_state"],
            "last_up": data["last_up"],
            "last_down": data["last_down"],
            "description": data["description"],
            "ipv4": data["ipv4"],
            "ipv6": data["ipv6"],
            "mac": data["mac"],
            "mtu_l3": data["mtu_l3"],
            "vendor_pn": data["vendor_pn"],
            "vendor_name": data["vendor_name"],
            "transceiver_id": data["transceiver_id"],
            "wavelength": data["wavelength"],
            "distance": data["distance"],
            "rx_power": data["rx_power"],
            "rx_range": rx_range,  # Use converted range format
            "tx_power": data["tx_power"],
            "tx_range": tx_range,  # Use converted range format
            "bias": data["bias"],
            "bias_range": bias_range,  # Use converted range format
            "voltage": data["voltage"],
            "voltage_range": voltage_range,  # Use converted range format
            "temperature": data["temperature"],
            "temp_range": temp_range,  # Use converted range format
            "port_bw": data["port_bw"],
            "transceiver_bw": data["transceiver_bw"],
            "input_rate": data["input_rate"],
            "input_util": data["input_util"],
            "output_rate": data["output_rate"],
            "output_util": data["output_util"],
            "crc": data["crc"],
            "neighbor_system_name": neighbor.get("system_name", "-"),
            "neighbor_system_description": neighbor.get("system_description", "-"),
            "neighbor_port": neighbor.get("port_id", "-"),
            "neighbor_ip": neighbor.get("ip", "-"),
            "result": result,
            "error_reasons": "; ".join(error_reasons) if error_reasons else "-"
        })
        print(
            f"{Fore.YELLOW}[DEBUG] 设备 {ip} 接口 {interface} Result: {result}{Style.RESET_ALL}")
        if error_reasons:
            print(
                f"{Fore.YELLOW}[DEBUG] 错误原因: {', '.join(error_reasons)}{Style.RESET_ALL}")

    if not results:
        print(f"{Fore.YELLOW}[WARNING] 设备 {ip} 未解析到光模块数据{Style.RESET_ALL}")
    return results


def process_optical_module_worksheet(ws, host_ips, data, connection_failures, item):
    """Process worksheet for optical module information."""
    yellow_fill = PatternFill(start_color="FFFF00",
                              end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500",
                              end_color="FFA500", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    total_results = 0
    normal_results = 0
    health_scores = {}
    sheet_name = item['sheet_name']

    headers = [
        "网元类型", "网元名称", "网元IP", "接口名称", "当前状态", "最近UP时间", "最近DOWN时间",
        "描述", "IPv4地址", "IPv6地址", "MAC地址", "最大传输单元(L3)", "厂商型号", "光模块类型",
        "波长", "传输距离", "Rx光功率(dBm)", "Rx范围(dBm)", "Tx光功率(dBm)", "Tx范围(dBm)",
        "偏置电流(mA)", "偏置范围(mA)", "电压(mV)", "电压范围(mV)", "温度(°C)", "温度范围(°C)",
        "端口带宽", "光模块带宽", "输入速率(bps)", "输入带宽利用率", "输出速率(bps)", "输出带宽利用率",
        "邻居系统名称", "邻居系统描述", "邻居端口", "邻居IP", "Result", "错误原因"
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    print(f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

    for ip in host_ips:
        if ip in connection_failures:
            continue
        ne_type, device_name = "-", "-"
        if ip in data and "show device" in data[ip]:
            ne_type, device_name, _, _ = parse_uptime(data[ip]["show device"])

        # 无接口数据的情况
        if ip not in data or "show interface" not in data[ip]:
            total_results += 1
            ws.append([ne_type, device_name, ip] + ["数据异常"] * 34 + ["error"])
            for cell in ws[ws.max_row]:
                cell.alignment = center_alignment
                cell.border = thin_border
            ws.cell(row=ws.max_row, column=len(headers) - 1).fill = orange_fill
            print(f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无接口数据，写入子表{Style.RESET_ALL}")
            continue

        interface_output = data[ip]["show interface"]
        lldp_output = data[ip].get("show lldp neighbor", "")
        optical_data = parse_optical_module(
            ip, interface_output, lldp_output, parse_uptime)

        # 无光模块数据的情况
        if not optical_data:
            total_results += 1
            ws.append([ne_type, device_name, ip] + ["无光模块数据"] * 34 + ["error"])
            for cell in ws[ws.max_row]:
                cell.alignment = center_alignment
                cell.border = thin_border
            ws.cell(row=ws.max_row, column=len(headers) - 1).fill = orange_fill
            print(f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无光模块数据，写入子表{Style.RESET_ALL}")
            continue

        # 处理光模块数据
        start_row = ws.max_row + 1
        for row_data in optical_data:
            total_results += 1
            ws.append([
                ne_type, device_name, ip,
                row_data["interface"], row_data["current_state"], row_data["last_up"], row_data["last_down"],
                sanitize_string(row_data["description"]),
                row_data["ipv4"], row_data["ipv6"], row_data["mac"],
                row_data["mtu_l3"], sanitize_string(row_data["vendor_pn"]),
                sanitize_string(row_data["transceiver_id"]),
                row_data["wavelength"], row_data["distance"], row_data["rx_power"], row_data["rx_range"],
                row_data["tx_power"], row_data["tx_range"], row_data["bias"], row_data["bias_range"],
                row_data["voltage"], row_data["voltage_range"], row_data["temperature"], row_data["temp_range"],
                row_data["port_bw"], row_data["transceiver_bw"], row_data["input_rate"], row_data["input_util"],
                row_data["output_rate"], row_data["output_util"],
                sanitize_string(row_data["neighbor_system_name"]),
                sanitize_string(row_data["neighbor_system_description"]),
                sanitize_string(row_data["neighbor_port"]),
                sanitize_string(row_data["neighbor_ip"]),
                row_data["result"],
                row_data.get("error_reasons", "-")
            ])
            for cell in ws[ws.max_row]:
                cell.alignment = center_alignment
                cell.border = thin_border
            if row_data["result"] == "normal":
                normal_results += 1
            else:
                ws.cell(row=ws.max_row, column=len(
                    headers) - 1).fill = orange_fill

        end_row = ws.max_row
        if start_row < end_row:
            for col in range(1, 4):
                ws.merge_cells(start_row=start_row, start_column=col,
                               end_row=end_row, end_column=col)
        print(
            f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(optical_data)} 行，合并单元格{Style.RESET_ALL}")

    # 在所有数据处理完成后计算健康百分比
    health_percentage = (normal_results / total_results *
                         100) if total_results > 0 else 0
    health_scores[sheet_name] = f"{health_percentage:.2f}%"
    print(
        f"{Fore.YELLOW}[DEBUG] 子表 {sheet_name} 健康百分比: {health_percentage:.2f}% ({normal_results}/{total_results}){Style.RESET_ALL}")

    return health_scores


def parse_power_status(output):
    """Parse 'show voltage' output for power status."""
    print(f"{Fore.CYAN}[DEBUG] 开始解析 'show voltage' 输出{Style.RESET_ALL}")
    lines = output.split('\n')
    results = []
    in_table = False
    slot_voltages = {}

    # Parse voltage data for all slots
    for line in lines:
        line = line.strip()
        if line.startswith("Slot"):
            in_table = True
            continue
        if line.startswith("---"):
            continue
        if in_table and line:
            parts = line.split()
            # Check if the line has at least 6 parts to safely access parts[5]
            if len(parts) >= 6:  # Updated to >= 6 since we need parts[5]
                slot = parts[0]
                voltage_mv = parts[4]
                ratio = parts[5]
                try:
                    voltage_v = float(voltage_mv) / 1000  # Convert mV to V
                    voltage_str = f"{voltage_v:.1f}V"
                    slot_voltages[slot] = {
                        "voltage": voltage_str, "ratio": ratio}
                    print(
                        f"{Fore.YELLOW}[DEBUG] 解析槽位 {slot}: 电压={voltage_str}, 比率={ratio}{Style.RESET_ALL}")
                except ValueError:
                    print(
                        f"{Fore.YELLOW}[WARNING] 电压解析失败: {voltage_mv}{Style.RESET_ALL}")
                    slot_voltages[slot] = {
                        "voltage": "-", "ratio": "-", "result": "error"}
            else:
                print(
                    f"{Fore.YELLOW}[WARNING] 跳过无效行: {line} (列数不足){Style.RESET_ALL}")
                continue  # Skip lines with insufficient columns

    # Check if slots 12 and 13 both have 0.0V and 0.00 ratio
    slots_12_13_zero = (
        slot_voltages.get("12", {}).get("voltage") == "0.0V" and
        slot_voltages.get("12", {}).get("ratio") == "0.00" and
        slot_voltages.get("13", {}).get("voltage") == "0.0V" and
        slot_voltages.get("13", {}).get("ratio") == "0.00"
    )

    # Generate results with conditional logic
    for slot, data in slot_voltages.items():
        if data.get("result") == "error":
            results.append({"slot": slot, "voltage": "-",
                           "ratio": "-", "result": "error"})
            continue
        voltage_v = float(data["voltage"].replace("V", ""))
        result = "normal" if slots_12_13_zero else "error"
        if not slots_12_13_zero:
            if voltage_v > 58 or voltage_v < 42 or voltage_v == 0:
                result = "error"
                print(
                    f"{Fore.YELLOW}[DEBUG] 槽位 {slot} 电压异常: {data['voltage']}{Style.RESET_ALL}")
            else:
                result = "normal"
                print(
                    f"{Fore.YELLOW}[DEBUG] 槽位 {slot} 电压正常: {data['voltage']}{Style.RESET_ALL}")
        else:
            print(
                f"{Fore.YELLOW}[DEBUG] 槽位 {slot} 电压为0.0V且比率为0.00 (与槽位12/13均满足)，状态为normal{Style.RESET_ALL}")
        results.append({
            "slot": slot,
            "voltage": data["voltage"],
            "ratio": data["ratio"],
            "result": result
        })

    if not results:
        print(f"{Fore.YELLOW}[WARNING] 未解析到电源数据{Style.RESET_ALL}")
    return results


def parse_temperature(output):
    """解析 'show temperature' 输出，获取温度状态，支持单温度和多槽位格式。"""
    print(f"{Fore.CYAN}[DEBUG] 开始解析 'show temperature' 输出{Style.RESET_ALL}")
    lines = output.split('\n')
    temperature_data = []

    # 检查是否为多槽位格式（表头包含 "SLOT"）
    header_line = None
    header_line_index = -1
    for i, line in enumerate(lines[:10]):  # 扩展搜索范围到前10行
        if "SLOT" in line.upper():  # 使用大小写不敏感的匹配
            header_line = line
            header_line_index = i
            print(
                f"{Fore.YELLOW}[DEBUG] 找到表头行 (第{i+1}行): {line.strip()}{Style.RESET_ALL}")
            break

    if header_line and header_line_index >= 0:
        # 解析表头，动态获取列索引
        headers = header_line.split()
        print(f"{Fore.YELLOW}[DEBUG] 解析的表头: {headers}{Style.RESET_ALL}")

        slot_idx = None
        temp_idx = None
        sen_01_idx = None
        sen_02_idx = None
        sen_03_idx = None

        # 查找各列的索引
        for idx, header in enumerate(headers):
            if header.upper() == "SLOT":
                slot_idx = idx
            elif header.upper() == "TEMP":
                temp_idx = idx
            elif header.upper() == "SEN_01":
                sen_01_idx = idx
            elif header.upper() == "SEN_02":
                sen_02_idx = idx
            elif header.upper() == "SEN_03":
                sen_03_idx = idx

        print(f"{Fore.YELLOW}[DEBUG] 列索引 - SLOT: {slot_idx}, TEMP: {temp_idx}, SEN_01: {sen_01_idx}, SEN_02: {sen_02_idx}, SEN_03: {sen_03_idx}{Style.RESET_ALL}")

        if slot_idx is None:
            print(f"{Fore.YELLOW}[WARNING] 未找到 'SLOT' 列{Style.RESET_ALL}")
            return None

        # 解析数据行，从表头行的下一行开始
        data_line_count = 0
        for i, line in enumerate(lines[header_line_index + 1:], header_line_index + 1):
            line = line.strip()
            if not line or line.startswith('---') or line.startswith('='):
                continue

            parts = line.split()
            print(
                f"{Fore.YELLOW}[DEBUG] 处理数据行 {i+1}: {line} -> 分割后: {parts}{Style.RESET_ALL}")

            # 确保有足够的列数据且SLOT列是数字
            if len(parts) > slot_idx and parts[slot_idx].isdigit():
                slot = parts[slot_idx]
                temp = parts[temp_idx] if temp_idx is not None and temp_idx < len(
                    parts) else "--"
                sen_01 = parts[sen_01_idx] if sen_01_idx is not None and sen_01_idx < len(
                    parts) else "--"
                sen_02 = parts[sen_02_idx] if sen_02_idx is not None and sen_02_idx < len(
                    parts) else "--"
                sen_03 = parts[sen_03_idx] if sen_03_idx is not None and sen_03_idx < len(
                    parts) else "--"

                temperature_data.append({
                    'slot': slot,
                    'sen_01': sen_01,
                    'sen_02': sen_02,
                    'sen_03': sen_03
                })
                data_line_count += 1
                print(
                    f"{Fore.YELLOW}[DEBUG] 解析槽位 {slot}: SEN_01={sen_01}, SEN_02={sen_02}, SEN_03={sen_03}{Style.RESET_ALL}")
            else:
                print(f"{Fore.YELLOW}[DEBUG] 跳过无效数据行: {line}{Style.RESET_ALL}")

        print(
            f"{Fore.YELLOW}[DEBUG] 多槽位格式解析完成，共解析 {data_line_count} 行数据{Style.RESET_ALL}")
    else:
        # 单温度格式
        print(f"{Fore.YELLOW}[DEBUG] 使用单温度格式解析{Style.RESET_ALL}")
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if line and not line.startswith('---') and not line.startswith('SDK') and not line.startswith('PIC'):
                parts = line.split()
                print(
                    f"{Fore.YELLOW}[DEBUG] 单温度格式 - 处理行 {line_num}: {line} -> 分割后: {parts}{Style.RESET_ALL}")
                if len(parts) >= 4 and parts[0].isdigit():
                    temp = parts[3]
                    temperature_data.append({
                        'slot': '-',
                        'sen_01': temp,
                        'sen_02': '--',
                        'sen_03': '--'
                    })
                    print(
                        f"{Fore.YELLOW}[DEBUG] 解析单温度: Temp={temp}{Style.RESET_ALL}")
                    break

    if not temperature_data:
        print(f"{Fore.RED}[ERROR] 未解析到任何温度数据{Style.RESET_ALL}")
        print(f"{Fore.RED}[ERROR] 原始输出前10行:{Style.RESET_ALL}")
        for i, line in enumerate(lines[:10], 1):
            print(f"{Fore.RED}  第{i}行: {repr(line)}{Style.RESET_ALL}")
        return None

    print(
        f"{Fore.GREEN}[SUCCESS] 成功解析 {len(temperature_data)} 条温度数据{Style.RESET_ALL}")
    return temperature_data


def parse_fan(output):
    """Parse 'show fan' output to extract all fan speeds and determine result."""
    print(f"{Fore.YELLOW}[DEBUG] 开始解析 'show fan' 输出{Style.RESET_ALL}")

    status = "-"
    fan_speeds = []
    result = "normal"

    # Extract status (for display purposes only, not used in result)
    status_match = re.search(r"Status\s*:\s*(\w+)", output)
    if status_match:
        status = status_match.group(1)
        print(f"{Fore.YELLOW}[DEBUG] 提取风扇状态: {status}{Style.RESET_ALL}")

    # Extract all fan speeds, including "stopping"
    speed_matches = re.findall(r"\[fan #\d+\]\s*(\d+%|stopping)\s*", output)
    if speed_matches:
        fan_speeds = [f"[fan #{i+1:02d}] {speed}" for i,
                      speed in enumerate(speed_matches)]
        print(f"{Fore.YELLOW}[DEBUG] 提取风扇速度: {fan_speeds}{Style.RESET_ALL}")

        # Check if any fan speed is "stopping" or below 20%
        for speed in speed_matches:
            if speed == "stopping":
                result = "error"
                print(
                    f"{Fore.YELLOW}[DEBUG] 风扇速度 {speed} 为 stopping，设置 result 为 error{Style.RESET_ALL}")
                break
            try:
                speed_value = int(speed.rstrip("%"))
                if speed_value < 20:  # Rule: fan speed < 20%
                    result = "error"
                    print(
                        f"{Fore.YELLOW}[DEBUG] 风扇速度 {speed} 低于20%，设置 result 为 error{Style.RESET_ALL}")
                    break
            except ValueError:
                result = "error"
                print(
                    f"{Fore.YELLOW}[DEBUG] 风扇速度 {speed} 非百分数值，设置 result 为 error{Style.RESET_ALL}")
                break
    else:
        result = "error"
        print(
            f"{Fore.YELLOW}[DEBUG] 未找到风扇速度，设置 result 为 error{Style.RESET_ALL}")

    return {
        "status": status,
        "speeds": " ".join(fan_speeds) if fan_speeds else "-",
        "result": result
    }


def parse_version(output):
    """Parse 'show version' output for system and hardware version status."""
    print(f"{Fore.CYAN}[DEBUG] 开始解析 'show version' 输出{Style.RESET_ALL}")
    lines = output.split('\n')
    version_info = {}
    slots = []
    current_slot = None

    for line in lines:
        if line.startswith('OPTEL'):
            parts = line.split(',')
            version_info['system_version'] = parts[1].strip().split()[0]
            print(
                f"{Fore.YELLOW}[DEBUG] 系统版本: {version_info['system_version']}{Style.RESET_ALL}")
        elif 'uptime is' in line:
            uptime_raw = line.split('is')[1].strip()
            version_info['uptime'] = uptime_raw.replace(' day, ', '天，').replace(
                ' hours, ', '小时，').replace(' minutes', '分钟')
            print(
                f"{Fore.YELLOW}[DEBUG] 运行时间: {version_info['uptime']}{Style.RESET_ALL}")
        elif 'system objectid:' in line:
            version_info['object_id'] = line.split(':')[1].strip()
            print(
                f"{Fore.YELLOW}[DEBUG] 对象ID: {version_info['object_id']}{Style.RESET_ALL}")
        elif 'System-MAC:' in line:
            version_info['system_mac'] = line.split(':')[1].strip()
            print(
                f"{Fore.YELLOW}[DEBUG] 系统MAC: {version_info['system_mac']}{Style.RESET_ALL}")
        elif line.startswith('slot'):
            if current_slot:
                slots.append(current_slot)
            slot_parts = line.split(':')
            slot_num = slot_parts[0].split()[1]
            board_name = slot_parts[1].strip()
            remark = ''
            if '(' in board_name:
                board_name, remark = board_name.split('(', 1)
                remark = remark.rstrip(')')
            current_slot = {'slot': slot_num, 'board_name': board_name.strip(
            ), 'remark': remark, 'versions': {}}
            print(
                f"{Fore.YELLOW}[DEBUG] 新槽位: {slot_num}, 板卡: {board_name}, 备注: {remark}{Style.RESET_ALL}")
        elif current_slot and 'Version:' in line:
            key, value = line.split(':', 1)
            key = key.strip().split()[0]  # e.g., Software, FPGA01
            current_slot['versions'][key] = value.strip()
            print(
                f"{Fore.YELLOW}[DEBUG] 槽位 {current_slot['slot']} 版本 {key}: {value.strip()}{Style.RESET_ALL}")

    if current_slot:
        slots.append(current_slot)
    version_info['slots'] = slots
    if not version_info.get('system_version'):
        print(f"{Fore.YELLOW}[WARNING] 未解析到版本数据{Style.RESET_ALL}")
        return None
    return version_info


def process_multiple_cmds_device(ip, user, pwd, commands, writer, fail_log, timeout=20, retry_count=5, cmd_interval=2.0):
    """
    处理单个设备的多个命令执行 - 优化分页处理
    """
    from threading import Lock
    from datetime import datetime
    import time
    import logging
    from colorama import Fore, Style

    file_lock = Lock()
    channel = None

    try:
        with print_lock:
            print(f"[INFO] 开始处理设备: {ip}")
        logging.info(f"开始处理设备: {ip}")

        # 创建SSH通道
        channel = create_channel(
            ip, user, pwd, timeout=timeout, retry_count=retry_count)

        if not channel:
            with file_lock:
                fail_log.write(
                    f"{ip},连接失败,{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            with print_lock:
                print(f"[ERROR] 设备 {ip} 连接失败")
            return None

        # 等待连接稳定
        time.sleep(2)

        # 设置终端不分页显示
        with print_lock:
            print(f"[DEBUG] 设置设备 {ip} 终端不分页显示...")

        # 首先尝试 screen-length 0
        screen_result = execute_some_command(
            channel, "screen-length 0", timeout=5, max_retries=3,
            command_delay=2, device_name=ip, ip=ip
        )

        # 检查是否设置成功
        if ("Error" in screen_result or
            "ERROR: Invalid input detected at '^' marker" in screen_result or
            "Unknown command" in screen_result or
                "Unrecognized command" in screen_result):

            with print_lock:
                print(
                    f"[DEBUG] 设备 {ip} 不支持 screen-length 0，尝试 screen-length 512")
            time.sleep(2)

            # 尝试 screen-length 512
            screen_result = execute_some_command(
                channel, "screen-length 512", timeout=5, max_retries=3,
                command_delay=2, device_name=ip, ip=ip
            )

            if ("Error" in screen_result or
                    "ERROR: Invalid input detected at '^' marker" in screen_result):
                with print_lock:
                    print(f"[WARNING] 设备 {ip} screen-length 设置可能失败，继续执行命令")
            else:
                with print_lock:
                    print(f"[DEBUG] 设备 {ip} screen-length 512 设置成功")
        else:
            with print_lock:
                print(f"[DEBUG] 设备 {ip} screen-length 0 设置成功")

        time.sleep(2)  # 设置完成后等待

        # 执行命令列表
        for i, cmd in enumerate(commands):
            with print_lock:
                print(f"[DEBUG] 执行命令 {cmd} 于设备 {ip}")
            logging.info(f"设备 {ip} - 执行命令: {cmd}")

            # 命令间延迟
            if i > 0:
                time.sleep(cmd_interval)

            # 记录PC时间
            pc_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # 执行命令 - 使用更长的超时时间处理分页
            output = execute_some_command(
                channel, cmd, timeout=30, max_retries=3,
                command_delay=2, device_name=ip, ip=ip
            )

            # 清理输出内容
            lines = output.split('\n')
            clean_lines = []
            skip_next = False

            for j, line in enumerate(lines):
                line_stripped = line.strip()

                # 跳过空行和分页提示
                if not line_stripped or '----MORE----' in line_stripped:
                    continue

                # 跳过纯命令回显行
                if line_stripped == cmd.strip():
                    skip_next = True
                    continue

                # 如果上一行是命令，当前行是错误标记，跳过
                if skip_next and ("^" in line_stripped or "ERROR:" in line_stripped):
                    skip_next = False
                    continue

                skip_next = False

                # 跳过短提示符
                if len(line_stripped) < 10 and (line_stripped.endswith('>') or
                                                line_stripped.endswith('#') or
                                                line_stripped.endswith('$')):
                    continue

                clean_lines.append(line)

            clean_output = '\n'.join(clean_lines)

            # 添加PC时间
            clean_output_with_time = f"PC_TIME: {pc_time}\n{clean_output}"

            # 输出预览
            output_preview = clean_output[:800] + \
                "..." if len(clean_output) > 800 else clean_output
            with print_lock:
                print(f"[DEBUG] 设备 {ip} 命令 {cmd} 输出(预览): {output_preview}")

            # 写入结果
            with file_lock:
                try:
                    writer.writerow([ip, cmd, clean_output_with_time])
                except Exception as write_err:
                    logging.error(f"写入CSV时出错: {write_err}")
                    with print_lock:
                        print(f"[ERROR] 写入结果到CSV时出错: {write_err}")

    except Exception as e:
        with print_lock:
            print(f"[ERROR] 设备 {ip} 处理失败: {e}")
        logging.error(f"设备 {ip} 处理失败: {e}")
        with file_lock:
            fail_log.write(
                f"{ip},处理失败,{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    finally:
        if channel:
            try:
                # 恢复终端设置
                time.sleep(1)
                execute_some_command(
                    channel, "screen-length 25", timeout=3, max_retries=1,
                    command_delay=1, device_name=ip, ip=ip
                )
            except:
                pass

            try:
                channel.close()
            except:
                pass

        logging.info(f"设备 {ip} 处理完成")
        with print_lock:
            print(f"[INFO] 设备 {ip} 指令处理完成")

        return ip


def parse_uptime(output):
    """
    Parse 'show device' output to extract ne_type, device_name, uptime, and result.
    Returns: (ne_type, device_name, uptime, result)
    """
    print(f"{Fore.CYAN}[DEBUG] 开始解析 'show device' 输出{Style.RESET_ALL}")
    lines = output.split('\n')
    device_name = None  # 改回原版的初始化方式
    ne_type = None      # 改回原版的初始化方式
    uptime = None       # 改回原版的初始化方式
    found_stn = False
    found_uptime = False

    for line in lines:
        line = line.strip()

        # 提取设备名称 - 使用原版的简单有效逻辑
        if not device_name and line.startswith('<') and line.endswith('>'):
            device_name = line[1:-1]
            print(
                f"{Fore.YELLOW}[DEBUG] 提取设备名称: {device_name}{Style.RESET_ALL}")

        # 提取网元类型 - 保持新版的增强逻辑
        # 模式1: OPTEL MssEdge 格式
        elif line.startswith("OPTEL MssEdge") and not ne_type:
            ne_type_match = re.match(r'OPTEL (MssEdge [^,]+),', line)
            if ne_type_match:
                ne_type = ne_type_match.group(1).strip()
                print(
                    f"{Fore.YELLOW}[DEBUG] 提取网元类型 (OPTEL格式): {ne_type} (原始: {line}){Style.RESET_ALL}")
            else:
                print(
                    f"{Fore.YELLOW}[WARNING] 网元类型匹配失败: {line}{Style.RESET_ALL}")

        # 模式2: stn-standard-reserved 标记格式 - 使用原版逻辑
        elif "stn-standard-reserved" in line:
            found_stn = True
        elif found_stn and not ne_type:
            ne_type_full = line
            ne_type = ne_type_full.split(',')[0].strip(
            ) if ',' in ne_type_full else ne_type_full
            print(
                f"{Fore.YELLOW}[DEBUG] 提取网元类型: {ne_type} (原始: {ne_type_full}){Style.RESET_ALL}")
            found_stn = False  # 重置标志位

        # 提取运行时间 - 使用原版逻辑但增加状态判断
        elif line == "uptime:":
            found_uptime = True
        elif found_uptime and not uptime:
            uptime_line = line
            if uptime_line:
                uptime = uptime_line.replace(' day, ', '天，').replace(
                    ' hours, ', '小时，').replace(' minutes', '分钟')
                print(
                    f"{Fore.YELLOW}[DEBUG] 提取并转换运行时间: {uptime}{Style.RESET_ALL}")
                found_uptime = False  # 重置标志位

        # 如果所有信息都已收集，提前退出
        if device_name and ne_type and uptime:
            break

    # 处理结果 - 增强的状态判断逻辑
    if uptime:
        try:
            # 尝试提取天数来判断状态
            if '天' in uptime:
                days = int(uptime.split('天')[0])
                result = "normal" if days > 1 else "error"
            elif 'day' in uptime:  # 处理原始英文格式
                # 从原始输出中提取天数
                uptime_raw = uptime.replace('天，', ' day, ').replace(
                    '小时，', ' hours, ').replace('分钟', ' minutes')
                days = int(uptime_raw.split(' day')[0])
                result = "normal" if days > 1 else "error"
            else:
                # 如果没有天数信息但有uptime，也算normal
                result = "normal"
        except (ValueError, IndexError) as e:
            print(f"{Fore.YELLOW}[WARNING] 解析天数失败: {e}{Style.RESET_ALL}")
            # 如果解析失败但有uptime信息，仍然算normal
            result = "normal"
    else:
        result = "error"

    # 处理未解析到的信息 - 使用原版的默认值处理
    if not device_name:
        print(f"{Fore.YELLOW}[WARNING] 未解析到设备名称{Style.RESET_ALL}")
        device_name = "-"
    if not ne_type:
        print(f"{Fore.YELLOW}[WARNING] 未解析到网元类型{Style.RESET_ALL}")
        ne_type = "-"
    if not uptime:
        uptime = "-"

    print(f"{Fore.YELLOW}[DEBUG] 确定Result状态: {result}{Style.RESET_ALL}")
    print(
        f"{Fore.YELLOW}[DEBUG] 解析结果: ne_type={ne_type}, device_name={device_name}, uptime={uptime}, result={result}{Style.RESET_ALL}")

    return ne_type, device_name, uptime, result


def parse_real_version(output):
    print(f"{Fore.CYAN}[DEBUG] 开始解析 'show real-version' 输出{Style.RESET_ALL}")
    if "ERROR:" in output or "Invalid input" in output:
        print(f"{Fore.YELLOW}[DEBUG] 检测到命令不支持或错误{Style.RESET_ALL}")
        return [{
            "组件类型": "命令不支持",
            "版本标识": "-",
            "版本号": "-",
            "编译/构建时间": "-",
            "打包时间": "-",
            "下载时间": "-",
            "附加说明": "-",
            "Result": "normal"
        }]

    lines = output.split('\n')
    versions = []
    current_component = None

    # 预编译正则表达式提升性能
    uboot_pattern = re.compile(r'uboot\s*info:', re.IGNORECASE)
    kernel_pattern = re.compile(r'kernel\s*info:', re.IGNORECASE)
    main_ws_pattern = re.compile(r'main\s*workspace:')
    back_ws_pattern = re.compile(r'back\s*workspace:')
    download_time_pattern = re.compile(
        r'download\s+ups\s+time:', re.IGNORECASE)

    for line in lines:
        # 使用正则表达式匹配组件起始行
        if uboot_pattern.search(line):
            current_component = "Uboot引导程序"
            print(f"{Fore.GREEN}[MATCH] 匹配到Uboot行: {line}{Style.RESET_ALL}")

            # 提取版本标识
            version_part = line.split(':', 1)[1].strip()
            version_id = version_part.split('(')[0].strip()

            # 解析编译时间
            build_time_match = re.search(r'\((.*?)\)', line)
            build_time_raw = build_time_match.group(
                1) if build_time_match else "-"
            try:
                if build_time_raw != "-":
                    # 处理多种时间格式 例如: Jun 16 2023 - 15:37:36 +0800
                    time_part = build_time_raw.split(' - ')[0]
                    build_time_dt = datetime.strptime(time_part, '%b %d %Y')
                    build_time = build_time_dt.strftime('%Y-%m-%d')
                else:
                    build_time = "-"
            except Exception as e:
                print(f"{Fore.RED}[ERROR] Uboot时间解析失败: {e}{Style.RESET_ALL}")
                build_time = "-"

            versions.append({
                "组件类型": current_component,
                "版本标识": version_id,
                "版本号": None,
                "编译/构建时间": build_time,
                "打包时间": None,
                "下载时间": None,
                "附加说明": "基于U-Boot 2020.10",
                "Result": "normal"
            })
            print(
                f"{Fore.BLUE}[PARSE] 添加Uboot组件: {version_id}{Style.RESET_ALL}")

        elif kernel_pattern.search(line):
            current_component = "Linux内核"
            print(f"{Fore.GREEN}[MATCH] 匹配到Kernel行: {line}{Style.RESET_ALL}")

            # 提取版本标识
            version_part = line.split(':', 1)[1].strip()
            version_id = re.split(r'\s+\d+\.\d+\.\d+',
                                  version_part)[0].strip()  # 去除内核版本号

            # 解析编译时间
            time_match = re.search(
                r'(\w{3}\s+\w{3}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}\s+\w{3}\s+\d{4})', line)
            build_time_raw = time_match.group(0) if time_match else "-"
            try:
                if build_time_raw != "-":
                    build_time_dt = datetime.strptime(
                        build_time_raw, '%a %b %d %H:%M:%S %Z %Y')
                    build_time = build_time_dt.strftime('%Y-%m-%d')
                else:
                    build_time = "-"
            except Exception as e:
                print(f"{Fore.RED}[ERROR] Kernel时间解析失败: {e}{Style.RESET_ALL}")
                build_time = "-"

            versions.append({
                "组件类型": current_component,
                "版本标识": version_id,
                "版本号": None,
                "编译/构建时间": build_time,
                "打包时间": None,
                "下载时间": None,
                "附加说明": "内核版本标记为*2",
                "Result": "normal"
            })
            print(
                f"{Fore.BLUE}[PARSE] 添加Kernel组件: {version_id}{Style.RESET_ALL}")

        elif main_ws_pattern.search(line):
            current_component = "主工作区（OAPP）"
            workspace = line.split(':', 1)[1].strip()
            versions.append({
                "组件类型": current_component,
                "版本标识": f"workspace:{workspace}",
                "版本号": None,
                "编译/构建时间": "-",
                "打包时间": None,
                "下载时间": None,
                "附加说明": "操作系统应用版本",
                "Result": "normal"
            })
            print(f"{Fore.BLUE}[PARSE] 添加主工作区: {workspace}{Style.RESET_ALL}")

        elif back_ws_pattern.search(line):
            current_component = "备用工作区（OAPP）"
            workspace = line.split(':', 1)[1].strip()
            versions.append({
                "组件类型": current_component,
                "版本标识": f"workspace:{workspace}",
                "版本号": None,
                "编译/构建时间": "-",
                "打包时间": None,
                "下载时间": None,
                "附加说明": "操作系统应用版本",
                "Result": "normal"
            })
            print(f"{Fore.BLUE}[PARSE] 添加备用工作区: {workspace}{Style.RESET_ALL}")

        # 处理通用字段
        elif current_component:
            if 'version:' in line.lower():
                versions[-1]["版本号"] = line.split(':', 1)[1].strip()
                print(
                    f"{Fore.MAGENTA}[UPDATE] 更新版本号: {versions[-1]['版本号']}{Style.RESET_ALL}")

            elif 'packaging' in line.lower() and 'time:' in line.lower():
                time_str = line.split(':', 1)[1].strip()
                try:
                    versions[-1]["打包时间"] = datetime.strptime(
                        time_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                except:
                    versions[-1]["打包时间"] = time_str
                print(
                    f"{Fore.MAGENTA}[UPDATE] 更新打包时间: {versions[-1]['打包时间']}{Style.RESET_ALL}")

            elif download_time_pattern.search(line):
                time_str = line.split(':', 1)[1].strip()
                try:
                    versions[-1]["下载时间"] = datetime.strptime(
                        time_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                except:
                    versions[-1]["下载时间"] = time_str
                print(
                    f"{Fore.MAGENTA}[UPDATE] 更新下载时间: {versions[-1]['下载时间']}{Style.RESET_ALL}")

    print(f"{Fore.CYAN}[DEBUG] 解析完成，共找到 {len(versions)} 个组件{Style.RESET_ALL}")
    return versions


def parse_main_control_status(output):
    """Parse 'show device' output for main control board status (CPU and memory usage)."""
    print(
        f"{Fore.CYAN}[DEBUG] 开始解析 'show device' 输出以获取主控盘运行状态{Style.RESET_ALL}")
    lines = output.split('\n')
    ne_type = None
    device_name = None
    cpu_usage = None
    cpu_5min = None
    cpu_15min = None
    memory_usage = None
    in_cpu_section = False
    in_memory_section = False

    # Parse the output line by line
    for i, line in enumerate(lines):
        line = line.strip()
        if line.startswith('<') and line.endswith('>'):
            device_name = line[1:-1]
            print(
                f"{Fore.YELLOW}[DEBUG] 提取设备名称: {device_name}{Style.RESET_ALL}")
            # Extract NE type from device name
            if 'MssEdge20' in device_name:
                ne_type = "MssEdge 20 A1"
            elif 'MssEdge25-S10-3' in device_name:
                ne_type = "MssEdge 25 S10-3"
            elif 'MssEdge25-S10' in device_name:
                ne_type = "MssEdge 25 S10"
            print(
                f"{Fore.YELLOW}[DEBUG] 从设备名称提取网元类型: {ne_type or '-'}{Style.RESET_ALL}")

        if "stn-standard-reserved" in line and not ne_type:
            if i + 1 < len(lines):
                ne_type_full = lines[i + 1].strip()
                ne_type = ne_type_full.split(',')[0].strip()
                # Standardize NE type
                if 'MssEdge20' in ne_type:
                    ne_type = "MssEdge 20 A1"
                elif 'MssEdge25-S10-3' in ne_type:
                    ne_type = "MssEdge 25 S10-3"
                elif 'MssEdge25-S10' in ne_type:
                    ne_type = "MssEdge 25 S10"
                print(
                    f"{Fore.YELLOW}[DEBUG] 从下一行提取网元类型: {ne_type or '-'}{Style.RESET_ALL}")

        if line == "cpu-usage:":
            in_cpu_section = True
            continue
        if line == "memory-usage:":
            in_memory_section = True
            in_cpu_section = False
            continue
        if in_cpu_section:
            if line.startswith("CPU Usage ") and ":" in line and "Threshold" not in line:
                cpu_usage = line.split(':')[1].strip()
                print(
                    f"{Fore.YELLOW}[DEBUG] 提取CPU使用率: {cpu_usage}{Style.RESET_ALL}")
            if "CPU utilization for five seconds:" in line:
                parts = line.split(':')
                if len(parts) >= 4:
                    cpu_5min = parts[1].strip().split('%')[0].strip() + '%'
                    cpu_15min = parts[3].strip()
                    print(
                        f"{Fore.YELLOW}[DEBUG] 提取五分钟CPU: {cpu_5min}, 十五分钟CPU: {cpu_15min}{Style.RESET_ALL}")
        if in_memory_section:
            if "Memory Using Percentage :" in line:
                memory_usage = line.split(':')[1].strip()
                print(
                    f"{Fore.YELLOW}[DEBUG] 提取内存使用率: {memory_usage}{Style.RESET_ALL}")

    # Determine the result
    result = "error"
    if cpu_15min is not None:
        try:
            cpu_15min_val = float(cpu_15min.rstrip('%'))
            result = "error" if cpu_15min_val >= 60 else "normal"
            print(
                f"{Fore.YELLOW}[DEBUG] 十五分钟CPU使用率 ({cpu_15min_val}%) {'>=' if cpu_15min_val >= 60 else '<'} 60%，Result: {result}{Style.RESET_ALL}")
        except ValueError:
            print(
                f"{Fore.YELLOW}[WARNING] CPU使用率解析失败，Result: error{Style.RESET_ALL}")
    else:
        print(
            f"{Fore.YELLOW}[WARNING] 未找到CPU利用率数据，Result: error{Style.RESET_ALL}")

    print(f"{Fore.YELLOW}[DEBUG] 确定Result状态: {result}{Style.RESET_ALL}")
    return (ne_type or "-", device_name or "-", cpu_usage or "-", cpu_5min or "-", cpu_15min or "-", memory_usage or "-", result)


def parse_cpu_defend_stats(output):
    """Parse 'show cpu-defend stats' output for protocol packet processing status."""
    print(
        f"{Fore.CYAN}[DEBUG] 开始解析 'show cpu-defend stats' 输出{Style.RESET_ALL}")
    lines = output.split('\n')
    results = []
    in_table = False

    for line in lines:
        line = line.strip()
        if line.startswith('Type') and 'Total-Packets' in line:
            in_table = True
            continue
        if line.startswith('---'):
            continue
        if not line and in_table:
            in_table = False
            break
        if in_table and line:
            parts = line.split()
            if len(parts) >= 4:
                protocol = parts[0]
                total = parts[1]
                passed = parts[2]
                dropped = parts[3]
                print(
                    f"{Fore.YELLOW}[DEBUG] 解析协议 {protocol}: Total={total}, Passed={passed}, Dropped={dropped}{Style.RESET_ALL}")
                try:
                    dropped_int = int(dropped)
                    if dropped_int > 0:
                        results.append({
                            "protocol": protocol,
                            "total_packets": total,
                            "passed_packets": passed,
                            "dropped_packets": dropped,
                            "result": "error"
                        })
                except ValueError:
                    print(
                        f"{Fore.YELLOW}[WARNING] 丢弃数据包解析失败: {dropped}{Style.RESET_ALL}")

    if not results:
        results.append({
            "protocol": "all",
            "total_packets": "0",
            "passed_packets": "0",
            "dropped_packets": "0",
            "result": "normal"
        })
        print(
            f"{Fore.YELLOW}[DEBUG] 无丢弃数据包，添加默认行: {results[0]}{Style.RESET_ALL}")

    print(f"{Fore.YELLOW}[DEBUG] 共解析 {len(results)} 行协议数据{Style.RESET_ALL}")
    return results


def parse_loopback31(output: str) -> str:
    loopback_found = False
    for line in output.split('\n'):
        if re.search(r'loopback\s*31\s+current\s+state', line, re.IGNORECASE):
            loopback_found = True
            continue
        if loopback_found and "Internet IPV4 Address is" in line:
            match = re.search(
                r'Internet IPV4 Address is ((?:\d{1,3}\.){3}\d{1,3})/32', line)
            if match:
                ip = match.group(1)
                if all(0 <= int(part) <= 255 for part in ip.split('.')):
                    return ip
        elif re.search(r'^loopback|^interface', line, re.IGNORECASE):
            loopback_found = False
    return "-"


def create_progress_bar(percentage):
    bar_length = 20
    filled_length = int(bar_length * percentage / 100)
    return '█' * filled_length + '░' * (bar_length - filled_length)

#


def create_device_panel_layout(ws, devices_data):
    """创建设备面板视图的表格布局 - 按设备型号分类版本"""
    from openpyxl.styles import PatternFill, Alignment, Border, Side

    # 定义样式
    yellow_fill = PatternFill(start_color="FFFFFF00",
                              end_color="FFFFFF00", fill_type="solid")
    dark_green_fill = PatternFill(
        start_color="FF00B050", end_color="FF00B050", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    center_alignment_wrap = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 过滤掉None值
    valid_devices = [
        d for d in devices_data if d is not None and isinstance(d, dict)]

    # 如果没有有效设备，创建一行提示信息
    if not valid_devices:
        print("[WARNING] 没有找到有效的设备数据")
        # 创建表头
        headers = ['面板', '', '', '', '备注', '', '面板', '',
                   '', '', '备注', '', '面板', '', '', '', '备注']
        ws.append(headers)

        # 应用表头样式
        for col in range(1, 18):
            cell = ws.cell(row=1, column=col)
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # 合并表头单元格
        ws.merge_cells('A1:D1')  # MssEdge 25 S10设备面板
        ws.merge_cells('G1:J1')  # MssEdge 20 A1设备面板
        ws.merge_cells('M1:P1')  # MssEdge 25 S10-3设备面板

        # 添加提示信息行
        ws.append(['无设备数据'] + [''] * 16)
        return

    # 改进的设备分类逻辑 - 基于设备型号进行分类
    mssedge_25_s10_devices = []      # 左边列
    mssedge_20_a1_devices = []       # 中间列
    mssedge_25_s10_3_devices = []    # 右边列
    other_devices = []

    for device in valid_devices:
        device_model = device.get('device_model', '').upper()
        device_name = device.get('device_name', '').upper()

        # 检查设备型号或名称中的关键字
        model_and_name = f"{device_model} {device_name}"

        print(f"[DEBUG] 设备分类检查: 型号='{device_model}', 名称='{device_name}'")

        # 按照设备型号进行分类
        if 'MSSEDGE 25 S10-3' in model_and_name or 'S10-3' in model_and_name:
            # MssEdge 25 S10-3 (右边列)
            mssedge_25_s10_3_devices.append(device)
            print(f"[DEBUG] 设备归类到S10-3: {device.get('device_ip')}")
        elif 'MSSEDGE 25 S10' in model_and_name or ('S10' in model_and_name and 'S10-3' not in model_and_name):
            # MssEdge 25 S10 (左边列)
            mssedge_25_s10_devices.append(device)
            print(f"[DEBUG] 设备归类到S10: {device.get('device_ip')}")
        elif 'MSSEDGE 20 A1' in model_and_name or 'A1' in model_and_name:
            # MssEdge 20 A1 (中间列)
            mssedge_20_a1_devices.append(device)
            print(f"[DEBUG] 设备归类到A1: {device.get('device_ip')}")
        else:
            # 无法识别的设备类型，根据槽位数量进行推测
            slot_count = len(device.get('slots', {}))
            print(f"[DEBUG] 未识别设备类型，槽位数量: {slot_count}")

            if slot_count <= 4:
                # 槽位较少，可能是A1设备
                mssedge_20_a1_devices.append(device)
                print(f"[DEBUG] 根据槽位数推测为A1设备: {device.get('device_ip')}")
            elif slot_count > 8:
                # 槽位较多，可能是S10设备
                mssedge_25_s10_devices.append(device)
                print(f"[DEBUG] 根据槽位数推测为S10设备: {device.get('device_ip')}")
            else:
                # 中等槽位数，可能是S10-3设备
                mssedge_25_s10_3_devices.append(device)
                print(f"[DEBUG] 根据槽位数推测为S10-3设备: {device.get('device_ip')}")

    # 如果所有设备都无法分类，将它们都放到S10类别（默认布局）
    if not mssedge_25_s10_devices and not mssedge_20_a1_devices and not mssedge_25_s10_3_devices:
        print("[WARNING] 无法识别设备类型，使用默认S10布局")
        mssedge_25_s10_devices = valid_devices

    # 输出调试信息
    print(
        f"[DEBUG] 设备分类结果: S10={len(mssedge_25_s10_devices)}, A1={len(mssedge_20_a1_devices)}, S10-3={len(mssedge_25_s10_3_devices)}")
    for i, device in enumerate(valid_devices[:3]):  # 只显示前3个设备的信息
        print(f"[DEBUG] 设备{i+1}: 名称={device.get('device_name', 'N/A')}, "
              f"类型={device.get('device_type', 'N/A')}, "
              f"IP={device.get('device_ip', 'N/A')}, "
              f"型号={device.get('device_model', 'N/A')}, "
              f"槽位数={len(device.get('slots', {}))}")

    # 计算需要的总行数
    max_devices = max(len(mssedge_25_s10_devices), len(
        mssedge_20_a1_devices), len(mssedge_25_s10_3_devices))

    if max_devices == 0:
        print("[WARNING] 没有设备被正确分类")
        max_devices = 1  # 至少显示一行

    print(f"[DEBUG] 最大设备数: {max_devices}")

    # 创建表头
    headers = ['面板', '', '', '', '备注', '', '面板', '',
               '', '', '备注', '', '面板', '', '', '', '备注']
    ws.append(headers)

    # 应用表头样式
    for col in range(1, 18):
        cell = ws.cell(row=1, column=col)
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # 合并表头单元格
    ws.merge_cells('A1:D1')  # MssEdge 25 S10设备面板
    ws.merge_cells('G1:J1')  # MssEdge 20 A1设备面板
    ws.merge_cells('M1:P1')  # MssEdge 25 S10-3设备面板

    # 为每个设备创建7行布局（增加一行用于设备名称）
    current_row = 2

    for device_index in range(max_devices):
        # 创建7行设备面板行（原来6行+1行设备名称）
        for i in range(7):
            ws.append([''] * 17)
            # 应用边框和对齐
            for col in range(1, 18):
                cell = ws.cell(row=current_row + i, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment

        print(f"[DEBUG] 创建设备面板行 {device_index + 1}, 起始行: {current_row}")

        # 填充MssEdge 25 S10设备数据（左边列）
        if device_index < len(mssedge_25_s10_devices):
            print(
                f"[DEBUG] 填充S10设备: {mssedge_25_s10_devices[device_index].get('device_name', 'N/A')}")
            fill_s10_device_panel(
                ws, mssedge_25_s10_devices[device_index], current_row, dark_green_fill)

        # 填充MssEdge 20 A1设备数据（中间列）
        if device_index < len(mssedge_20_a1_devices):
            print(
                f"[DEBUG] 填充A1设备: {mssedge_20_a1_devices[device_index].get('device_name', 'N/A')}")
            fill_a1_device_panel(
                ws, mssedge_20_a1_devices[device_index], current_row, dark_green_fill)

        # 填充MssEdge 25 S10-3设备数据（右边列）
        if device_index < len(mssedge_25_s10_3_devices):
            print(
                f"[DEBUG] 填充S10-3设备: {mssedge_25_s10_3_devices[device_index].get('device_name', 'N/A')}")
            fill_s10_3_device_panel(
                ws, mssedge_25_s10_3_devices[device_index], current_row, dark_green_fill)

        current_row += 7  # 移动到下一个设备区域（改为7行）

        # 在设备间增加间隔行（除了最后一个设备）
        if device_index < max_devices - 1:
            ws.append([''] * 17)
            current_row += 1

    print(f"[DEBUG] 设备面板布局创建完成，总共创建了 {current_row - 1} 行")


def fill_s10_device_panel(ws, device_data, start_row, green_fill):
    """填充MssEdge 25 S10设备面板数据 - 7行布局，当有任何槽位板卡时整个框都填充颜色"""
    from openpyxl.styles import Alignment

    slots = device_data.get('slots', {})

    # 检查是否有任何槽位板卡（除了电源槽位12、13）
    has_any_card = any(slots.get(slot_num, {}).get('card_name', '')
                       for slot_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

    # A列 - 槽位11 (6行合并)
    ws.merge_cells(f'A{start_row}:A{start_row + 5}')

    # 填充槽位11内容 - 修复风扇显示格式，添加换行支持
    slot_11_info = slots.get(11, {})
    slot_11_name = slot_11_info.get('card_name', '')
    if slot_11_name and 'FAN' in slot_11_name.upper():
        # 风扇槽位显示为"FAN\n11"格式
        cell = ws.cell(row=start_row, column=1)
        cell.value = f"FAN\n11"
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    else:
        cell = ws.cell(row=start_row, column=1)
        cell.value = slot_11_name if slot_11_name else ""
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # 如果有任何板卡，则整个A列都填充颜色
    if has_any_card:
        for row in range(start_row, start_row + 6):
            ws.cell(row=row, column=1).fill = green_fill

    # B列布局 - 分为3个区域，每个区域2行 - 修复电源槽位大小
    ws.merge_cells(f'B{start_row}:B{start_row + 1}')      # 槽位13 (电源) - 2行高度
    ws.merge_cells(f'B{start_row + 2}:B{start_row + 3}')  # 空白区域 - 2行高度
    ws.merge_cells(f'B{start_row + 4}:B{start_row + 5}')  # 槽位12 (电源) - 2行高度

    # 如果有任何板卡，电源槽位也填充颜色
    if has_any_card:
        # 槽位13电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row, column=2)
        cell.value = "PWR\n13"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

        # 空白区域也填充颜色
        ws.cell(row=start_row + 2, column=2).fill = green_fill

        # 槽位12电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row + 4, column=2)
        cell.value = "PWR\n12"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # C和D列的槽位布局
    slot_layout = [
        # (slot_num, row_offset, col_offset, merge_cols)
        (7, 0, 3, False),   # C行1 - 槽位7
        (8, 0, 4, False),   # D行1 - 槽位8
        (5, 1, 3, False),   # C行2 - 槽位5
        (6, 1, 4, False),   # D行2 - 槽位6
        (10, 2, 3, True),   # C-D行3 - 槽位10 (主备)
        (9, 3, 3, True),    # C-D行4 - 槽位9 (主备)
        (3, 4, 3, False),   # C行5 - 槽位3
        (4, 4, 4, False),   # D行5 - 槽位4
        (1, 5, 3, False),   # C行6 - 槽位1
        (2, 5, 4, False),   # D行6 - 槽位2
    ]

    for slot_num, row_offset, col, merge_cols in slot_layout:
        slot_info = slots.get(slot_num, {})
        card_name = slot_info.get('card_name', '')
        status = slot_info.get('status', '')

        if merge_cols and col == 3:  # 主备槽位需要合并单元格
            ws.merge_cells(
                f'{chr(ord("A") + col - 1)}{start_row + row_offset}:{chr(ord("A") + col)}{start_row + row_offset}')

        # 如果有任何板卡，所有槽位都填充颜色
        if has_any_card:
            ws.cell(row=start_row + row_offset, column=col).fill = green_fill
            if merge_cols and col == 3:
                ws.cell(row=start_row + row_offset,
                        column=col + 1).fill = green_fill

        # 设置内容
        if card_name:
            if slot_num in [9, 10]:  # 主备槽位
                status_char = '主' if 'Master' in status else '备' if 'Backup' in status else ''
                content = f"{card_name}        {status_char}·{slot_num}" if status_char else f"{card_name}·{slot_num}"
            else:
                content = f"{card_name}     ·{slot_num}"

            cell = ws.cell(row=start_row + row_offset, column=col)
            cell.value = content
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

    # E列 - 备注信息（增加一行设备名称）
    ws.cell(row=start_row, column=5).value = "OPTEL"
    ws.cell(row=start_row + 1, column=5).value = device_data.get('device_ip', '')
    ws.cell(row=start_row + 2, column=5).value = device_data.get('device_model', '')
    ws.cell(row=start_row + 3,
            column=5).value = device_data.get('device_name', '')  # 新增设备名称行


def fill_a1_device_panel(ws, device_data, start_row, green_fill):
    """填充MssEdge 20 A1设备面板数据 - 简化4槽位布局，只在有设备槽位时填充颜色"""
    from openpyxl.styles import Alignment

    slots = device_data.get('slots', {})

    # G-J列，第一行显示4个槽位，内容为固定格式"S10_04"
    for slot_num in range(1, 5):  # 槽位1-4
        col = 6 + slot_num  # G=7, H=8, I=9, J=10
        slot_info = slots.get(slot_num, {})

        # 只有当槽位有设备时才填充内容和颜色
        if slot_info.get('card_name'):
            # A1设备的槽位内容固定为"S10_04"
            content = "S10_04"
            cell = ws.cell(row=start_row, column=col)
            cell.value = content
            cell.fill = green_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

    # K列 - 备注信息（增加一行设备名称）
    ws.cell(row=start_row, column=11).value = "OPTEL"
    ws.cell(row=start_row + 1, column=11).value = device_data.get('device_ip', '')
    ws.cell(row=start_row + 2, column=11).value = device_data.get('device_model', '')
    ws.cell(row=start_row + 3,
            column=11).value = device_data.get('device_name', '')  # 新增设备名称行


def fill_s10_3_device_panel(ws, device_data, start_row, green_fill):
    """填充MssEdge 25 S10-3设备面板数据 - 与S10设备布局相同，使用M-Q列，当有任何槽位板卡时整个框都填充颜色"""
    from openpyxl.styles import Alignment

    slots = device_data.get('slots', {})

    # 检查是否有任何槽位板卡（除了电源槽位12、13）
    has_any_card = any(slots.get(slot_num, {}).get('card_name', '')
                       for slot_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

    # M列 - 槽位11 (6行合并)
    ws.merge_cells(f'M{start_row}:M{start_row + 5}')

    # 填充槽位11内容 - 修复风扇显示格式，添加换行支持
    slot_11_info = slots.get(11, {})
    slot_11_name = slot_11_info.get('card_name', '')
    if slot_11_name and 'FAN' in slot_11_name.upper():
        # 风扇槽位显示为"FAN\n11"格式
        cell = ws.cell(row=start_row, column=13)
        cell.value = f"FAN\n11"
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    else:
        cell = ws.cell(row=start_row, column=13)
        cell.value = slot_11_name if slot_11_name else ""
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # 如果有任何板卡，则整个M列都填充颜色
    if has_any_card:
        for row in range(start_row, start_row + 6):
            ws.cell(row=row, column=13).fill = green_fill

    # N列布局 - 修复电源槽位大小
    ws.merge_cells(f'N{start_row}:N{start_row + 1}')      # 槽位13 (电源) - 2行高度
    ws.merge_cells(f'N{start_row + 2}:N{start_row + 3}')  # 空白区域 - 2行高度
    ws.merge_cells(f'N{start_row + 4}:N{start_row + 5}')  # 槽位12 (电源) - 2行高度

    # 如果有任何板卡，电源槽位也填充颜色
    if has_any_card:
        # 槽位13电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row, column=14)
        cell.value = "PWR\n13"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

        # 空白区域也填充颜色
        ws.cell(row=start_row + 2, column=14).fill = green_fill

        # 槽位12电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row + 4, column=14)
        cell.value = "PWR\n12"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # O和P列的槽位布局 (类似C和D列，但使用列15和16)
    slot_layout = [
        (7, 0, 15, False),   # O行1 - 槽位7
        (8, 0, 16, False),   # P行1 - 槽位8
        (5, 1, 15, False),   # O行2 - 槽位5
        (6, 1, 16, False),   # P行2 - 槽位6
        (10, 2, 15, True),   # O-P行3 - 槽位10 (主备)
        (9, 3, 15, True),    # O-P行4 - 槽位9 (主备)
        (3, 4, 15, False),   # O行5 - 槽位3
        (4, 4, 16, False),   # P行5 - 槽位4
        (1, 5, 15, False),   # O行6 - 槽位1
        (2, 5, 16, False),   # P行6 - 槽位2
    ]

    for slot_num, row_offset, col, merge_cols in slot_layout:
        slot_info = slots.get(slot_num, {})
        card_name = slot_info.get('card_name', '')
        status = slot_info.get('status', '')

        if merge_cols and col == 15:  # 主备槽位需要合并单元格
            ws.merge_cells(
                f'{chr(ord("A") + col - 1)}{start_row + row_offset}:{chr(ord("A") + col)}{start_row + row_offset}')

        # 如果有任何板卡，所有槽位都填充颜色
        if has_any_card:
            ws.cell(row=start_row + row_offset, column=col).fill = green_fill
            if merge_cols and col == 15:
                ws.cell(row=start_row + row_offset,
                        column=col + 1).fill = green_fill

        # 设置内容
        if card_name:
            if slot_num in [9, 10]:  # 主备槽位
                status_char = '主' if 'Master' in status else '备' if 'Backup' in status else ''
                content = f"{card_name}        {status_char}·{slot_num}" if status_char else f"{card_name}·{slot_num}"
            else:
                content = f"{card_name}     ·{slot_num}"

            cell = ws.cell(row=start_row + row_offset, column=col)
            cell.value = content
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

    # Q列 - 备注信息（增加一行设备名称）
    ws.cell(row=start_row, column=17).value = "OPTEL"
    ws.cell(row=start_row + 1, column=17).value = device_data.get('device_ip', '')
    ws.cell(row=start_row + 2, column=17).value = device_data.get('device_model', '')
    ws.cell(row=start_row + 3,
            column=17).value = device_data.get('device_name', '')  # 新增设备名称行


def fill_a2_device_panel(ws, device_data, start_row, green_fill):
    """填充A2设备面板数据 - 6行布局，当有任何槽位板卡时整个框都填充颜色"""
    from openpyxl.styles import Alignment

    slots = device_data.get('slots', {})

    # 检查是否有任何槽位板卡（除了电源槽位12、13）
    has_any_card = any(slots.get(slot_num, {}).get('card_name', '')
                       for slot_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

    # A列 - 槽位11 (6行合并)
    ws.merge_cells(f'A{start_row}:A{start_row + 5}')

    # 填充槽位11内容 - 修复风扇显示格式，添加换行支持
    slot_11_info = slots.get(11, {})
    slot_11_name = slot_11_info.get('card_name', '')
    if slot_11_name and 'FAN' in slot_11_name.upper():
        # 风扇槽位显示为"FAN\n11"格式
        cell = ws.cell(row=start_row, column=1)
        cell.value = f"FAN\n11"
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    else:
        cell = ws.cell(row=start_row, column=1)
        cell.value = slot_11_name if slot_11_name else ""
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # 如果有任何板卡，则整个A列都填充颜色
    if has_any_card:
        for row in range(start_row, start_row + 6):
            ws.cell(row=row, column=1).fill = green_fill

    # B列布局 - 分为3个区域，每个区域2行 - 修复电源槽位大小
    ws.merge_cells(f'B{start_row}:B{start_row + 1}')      # 槽位13 (电源) - 2行高度
    ws.merge_cells(f'B{start_row + 2}:B{start_row + 3}')  # 空白区域 - 2行高度
    ws.merge_cells(f'B{start_row + 4}:B{start_row + 5}')  # 槽位12 (电源) - 2行高度

    # 如果有任何板卡，电源槽位也填充颜色
    if has_any_card:
        # 槽位13电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row, column=2)
        cell.value = "PWR\n13"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

        # 空白区域也填充颜色
        ws.cell(row=start_row + 2, column=2).fill = green_fill

        # 槽位12电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row + 4, column=2)
        cell.value = "PWR\n12"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # C和D列的槽位布局
    slot_layout = [
        # (slot_num, row_offset, col_offset, merge_cols)
        (7, 0, 3, False),   # C行1 - 槽位7
        (8, 0, 4, False),   # D行1 - 槽位8
        (5, 1, 3, False),   # C行2 - 槽位5
        (6, 1, 4, False),   # D行2 - 槽位6
        (10, 2, 3, True),   # C-D行3 - 槽位10 (主备)
        (9, 3, 3, True),    # C-D行4 - 槽位9 (主备)
        (3, 4, 3, False),   # C行5 - 槽位3
        (4, 4, 4, False),   # D行5 - 槽位4
        (1, 5, 3, False),   # C行6 - 槽位1
        (2, 5, 4, False),   # D行6 - 槽位2
    ]

    for slot_num, row_offset, col, merge_cols in slot_layout:
        slot_info = slots.get(slot_num, {})
        card_name = slot_info.get('card_name', '')
        status = slot_info.get('status', '')

        if merge_cols and col == 3:  # 主备槽位需要合并单元格
            ws.merge_cells(
                f'{chr(ord("A") + col - 1)}{start_row + row_offset}:{chr(ord("A") + col)}{start_row + row_offset}')

        # 如果有任何板卡，所有槽位都填充颜色
        if has_any_card:
            ws.cell(row=start_row + row_offset, column=col).fill = green_fill
            if merge_cols and col == 3:
                ws.cell(row=start_row + row_offset,
                        column=col + 1).fill = green_fill

        # 设置内容
        if card_name:
            if slot_num in [9, 10]:  # 主备槽位
                status_char = '主' if 'Master' in status else '备' if 'Backup' in status else ''
                content = f"{card_name}        {status_char}·{slot_num}" if status_char else f"{card_name}·{slot_num}"
            else:
                content = f"{card_name}     ·{slot_num}"

            cell = ws.cell(row=start_row + row_offset, column=col)
            cell.value = content
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

    # E列 - 备注信息
    ws.cell(row=start_row, column=5).value = device_data.get('device_name', '')
    ws.cell(row=start_row + 1, column=5).value = device_data.get('device_ip', '')
    ws.cell(row=start_row + 2, column=5).value = device_data.get('device_model', '')


def fill_a3_device_panel(ws, device_data, start_row, green_fill):
    """填充A3设备面板数据 - 与A2设备布局相同，使用M-Q列，当有任何槽位板卡时整个框都填充颜色"""
    from openpyxl.styles import Alignment

    slots = device_data.get('slots', {})

    # 检查是否有任何槽位板卡（除了电源槽位12、13）
    has_any_card = any(slots.get(slot_num, {}).get('card_name', '')
                       for slot_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

    # M列 - 槽位11 (6行合并)
    ws.merge_cells(f'M{start_row}:M{start_row + 5}')

    # 填充槽位11内容 - 修复风扇显示格式，添加换行支持
    slot_11_info = slots.get(11, {})
    slot_11_name = slot_11_info.get('card_name', '')
    if slot_11_name and 'FAN' in slot_11_name.upper():
        # 风扇槽位显示为"FAN\n11"格式
        cell = ws.cell(row=start_row, column=13)
        cell.value = f"FAN\n11"
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    else:
        cell = ws.cell(row=start_row, column=13)
        cell.value = slot_11_name if slot_11_name else ""
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # 如果有任何板卡，则整个M列都填充颜色
    if has_any_card:
        for row in range(start_row, start_row + 6):
            ws.cell(row=row, column=13).fill = green_fill

    # N列布局 - 修复电源槽位大小
    ws.merge_cells(f'N{start_row}:N{start_row + 1}')      # 槽位13 (电源) - 2行高度
    ws.merge_cells(f'N{start_row + 2}:N{start_row + 3}')  # 空白区域 - 2行高度
    ws.merge_cells(f'N{start_row + 4}:N{start_row + 5}')  # 槽位12 (电源) - 2行高度

    # 如果有任何板卡，电源槽位也填充颜色
    if has_any_card:
        # 槽位13电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row, column=14)
        cell.value = "PWR\n13"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

        # 空白区域也填充颜色
        ws.cell(row=start_row + 2, column=14).fill = green_fill

        # 槽位12电源 - 修复显示格式和对齐，添加换行支持
        cell = ws.cell(row=start_row + 4, column=14)
        cell.value = "PWR\n12"
        cell.fill = green_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # O和P列的槽位布局 (类似C和D列，但使用列15和16)
    slot_layout = [
        (7, 0, 15, False),   # O行1 - 槽位7
        (8, 0, 16, False),   # P行1 - 槽位8
        (5, 1, 15, False),   # O行2 - 槽位5
        (6, 1, 16, False),   # P行2 - 槽位6
        (10, 2, 15, True),   # O-P行3 - 槽位10 (主备)
        (9, 3, 15, True),    # O-P行4 - 槽位9 (主备)
        (3, 4, 15, False),   # O行5 - 槽位3
        (4, 4, 16, False),   # P行5 - 槽位4
        (1, 5, 15, False),   # O行6 - 槽位1
        (2, 5, 16, False),   # P行6 - 槽位2
    ]

    for slot_num, row_offset, col, merge_cols in slot_layout:
        slot_info = slots.get(slot_num, {})
        card_name = slot_info.get('card_name', '')
        status = slot_info.get('status', '')

        if merge_cols and col == 15:  # 主备槽位需要合并单元格
            ws.merge_cells(
                f'{chr(ord("A") + col - 1)}{start_row + row_offset}:{chr(ord("A") + col)}{start_row + row_offset}')

        # 如果有任何板卡，所有槽位都填充颜色
        if has_any_card:
            ws.cell(row=start_row + row_offset, column=col).fill = green_fill
            if merge_cols and col == 15:
                ws.cell(row=start_row + row_offset,
                        column=col + 1).fill = green_fill

        # 设置内容
        if card_name:
            if slot_num in [9, 10]:  # 主备槽位
                status_char = '主' if 'Master' in status else '备' if 'Backup' in status else ''
                content = f"{card_name}        {status_char}·{slot_num}" if status_char else f"{card_name}·{slot_num}"
            else:
                content = f"{card_name}     ·{slot_num}"

            cell = ws.cell(row=start_row + row_offset, column=col)
            cell.value = content
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

    # Q列 - 备注信息
    ws.cell(row=start_row, column=17).value = device_data.get('device_name', '')
    ws.cell(row=start_row + 1, column=17).value = device_data.get('device_ip', '')
    ws.cell(row=start_row + 2, column=17).value = device_data.get('device_model', '')


def calculate_chinese_width(text):
    """计算包含中文字符的文本显示宽度"""
    if not text:
        return 0

    width = 0
    for char in str(text):
        # 中文字符和全角字符占用2个字符宽度
        if ord(char) > 127:  # 非ASCII字符
            width += 2
        else:
            width += 1
    return width


def apply_autofit_to_all_sheets(wb):
    """对所有工作表应用自适应列宽"""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        autofit_worksheet_columns(ws)

        # 只给子表添加筛选和冻结首行，不给主表（"运维质量评估"）添加
        if sheet_name != "运维质量评估":
            # 假设子表的表头都在第1行
            header_row = 1
            last_col = ws.max_column
            last_row = ws.max_row

            # 确保表格有数据才添加筛选
            if last_row > header_row:
                # 获取表格范围（从A1到最后一列最后一行）
                from openpyxl.utils import get_column_letter
                filter_range = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
                ws.auto_filter.ref = filter_range

            # 添加冻结首行功能
            ws.freeze_panes = 'A2'


def parse_device_panel(device_output, ne_type, ne_name, ne_ip):
    """解析设备面板信息，提取槽位和板卡信息 - 增强版本"""
    import re

    # 确保参数不为None
    if device_output is None:
        device_output = ""
    if ne_type is None:
        ne_type = "-"
    if ne_name is None:
        ne_name = "-"
    if ne_ip is None:
        ne_ip = "-"

    result_data = {
        'device_type': ne_type,
        'device_name': ne_name,
        'device_ip': ne_ip,
        'slots': {}
    }

    print(f"[DEBUG] 解析设备面板: {ne_name} ({ne_ip})")

    # 如果没有输出数据，返回基本信息
    if not device_output or not isinstance(device_output, str):
        result_data['device_model'] = "-"
        print(f"[DEBUG] 设备 {ne_ip} 没有输出数据")
        return result_data

    # 将输出按行分割
    lines = device_output.splitlines()
    print(f"[DEBUG] 设备 {ne_ip} 输出行数: {len(lines)}")

    # 提取设备型号 - 改进的匹配逻辑
    device_model = "-"
    for line in lines:
        line_stripped = line.strip()
        # 尝试多种模式匹配设备型号
        if "OPTEL MssEdge" in line:
            parts = line.split(',')
            if len(parts) > 0:
                device_model = parts[0].strip()
                break
        elif "model" in line.lower() or "type" in line.lower():
            # 尝试其他可能的型号信息
            if ":" in line:
                device_model = line.split(":")[-1].strip()
                break

    result_data['device_model'] = device_model
    print(f"[DEBUG] 设备 {ne_ip} 型号: {device_model}")

    # 解析槽位信息 - 改进的解析逻辑
    in_version_section = False
    slot_count = 0

    for line in lines:
        line = line.strip()

        # 多种方式进入版本信息段
        if line.lower() in ["version:", "slot information:", "board information:"]:
            in_version_section = True
            print(f"[DEBUG] 设备 {ne_ip} 进入版本信息段")
            continue

        # 结束版本信息段的条件
        if in_version_section and (line == "--------------------------------" or
                                   line.startswith("===") or
                                   line.lower().startswith("interface")):
            break

        # 解析槽位信息
        if in_version_section:
            # 匹配多种槽位格式
            slot_patterns = [
                # slot 9: UXS_02 (Master)
                r'slot\s+(\d+):\s*(\S+)(?:\s*\(([^)]+)\))?',
                # slot9: UXS_02 (Master)
                r'slot(\d+):\s*(\S+)(?:\s*\(([^)]+)\))?',
                r'槽位\s*(\d+):\s*(\S+)(?:\s*\(([^)]+)\))?',   # 中文格式
                r'(\d+)\s+(\S+)(?:\s+([^)]+))?'               # 简化格式
            ]

            for pattern in slot_patterns:
                slot_match = re.match(pattern, line, re.IGNORECASE)
                if slot_match:
                    try:
                        slot_num = int(slot_match.group(1))
                        card_name = slot_match.group(
                            2) if slot_match.group(2) else ""
                        status = slot_match.group(3) if len(
                            slot_match.groups()) > 2 and slot_match.group(3) else ""

                        # 过滤掉明显无效的板卡名称
                        if card_name and card_name not in ["-", "--", "N/A", "NULL", "EMPTY"]:
                            result_data['slots'][slot_num] = {
                                'card_name': card_name,
                                'status': status
                            }
                            slot_count += 1
                            print(
                                f"[DEBUG] 设备 {ne_ip} 槽位 {slot_num}: {card_name} ({status})")
                        break
                    except (ValueError, AttributeError) as e:
                        continue

    print(f"[DEBUG] 设备 {ne_ip} 总共解析到 {slot_count} 个槽位")

    # 如果没有解析到槽位信息，尝试创建一些示例数据用于测试
    if slot_count == 0:
        print(f"[WARNING] 设备 {ne_ip} 没有解析到槽位信息，创建示例数据")
        # 为测试目的创建一些示例槽位
        result_data['slots'] = {
            1: {'card_name': 'TEST_CARD_1', 'status': ''},
            2: {'card_name': 'TEST_CARD_2', 'status': ''}
        }

    return result_data


# 接口描述


def add_interface_description_cmd(host_file, raw_file, report_file, max_workers=40):
    """添加互联端口描述并生成报告 (Add Interface Description and Generate Report)"""
    print(
        f"[START] 开始添加互联端口描述，输入文件: {host_file}, 输出文件: {raw_file}, 报告文件: {report_file}")

    # 清空旧的failure_ips.tmp文件
    if os.path.exists("failure_ips.tmp"):
        try:
            os.remove("failure_ips.tmp")
            print(f"{Fore.YELLOW}⚠️ 已清除旧的failure_ips.tmp文件{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}⛔ 删除failure_ips.tmp失败: {str(e)}{Style.RESET_ALL}")

    # 文件写入锁，确保线程安全
    file_lock = Lock()

    with open(raw_file, "w", newline='', encoding='utf-8') as revFile, \
            open("failure_ips.tmp", "a", encoding='utf-8') as fail_log:
        writer = csv.writer(revFile)
        try:
            with open(host_file, "r", encoding='gbk', errors='ignore') as csvFile:
                reader = csv.reader(csvFile)
                hostip = list(reader)
                total_devices = len(hostip)
                print(f"[INFO] 共发现 {total_devices} 台设备")

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = []
                    ip_to_future = {}
                    for row in hostip:
                        ip = row[0].strip()
                        user = row[1].strip()
                        pwd = row[2].strip()
                        future = executor.submit(
                            process_device_interface_description, ip, user, pwd, writer, fail_log, file_lock)
                        futures.append(future)
                        ip_to_future[future] = ip

                    with tqdm(total=total_devices, desc="🔍 接口描述配置进度", unit="台", dynamic_ncols=True) as pbar:
                        for future in as_completed(futures):
                            try:
                                future.result(timeout=120)  # 每个任务最多120秒
                            except TimeoutError:
                                print(
                                    f"{Fore.RED}设备 {ip_to_future[future]} 任务超时{Style.RESET_ALL}")
                            except Exception as e:
                                print(
                                    f"{Fore.RED}设备 {ip_to_future[future]} 线程执行出错: {str(e)}{Style.RESET_ALL}")
                            pbar.update(1)

        except Exception as e:
            print(f"{Fore.RED}⛔ 数据采集错误: {str(e)}{Style.RESET_ALL}")

    print(f"[INFO] 开始生成接口描述配置报告...")
    generate_interface_description_report(raw_file, report_file, host_file)
    print(f"[END] 接口描述配置及报告生成完成")


def parse_ospf_dot31_interfaces(ospf_output):
    """解析OSPF邻居输出，提取所有.31接口（包括万兆接口）"""
    interfaces = set()
    try:
        lines = ospf_output.split('\n')
        for line in lines:
            if '.31' in line:
                # 匹配gigabitethernet和xgigabitethernet接口
                # 支持格式: gigabitethernet 0/1/1.31 或 xgigabitethernet 0/4/4.31
                match = re.search(
                    r'(x?gigabitethernet\s+\d+/\d+/\d+\.31)', line, re.IGNORECASE)
                if match:
                    interface = match.group(1).replace(' ', ' ')  # 保持原格式
                    interfaces.add(interface)
        print(f"[DEBUG] 解析OSPF .31接口结果: {list(interfaces)}")
    except Exception as e:
        print(f"[ERROR] 解析OSPF接口失败: {e}")
    return interfaces


def parse_lldp_neighbors(lldp_output):
    """解析LLDP邻居输出，提取物理接口对应的邻居信息（包括万兆接口）"""
    neighbors = {}
    try:
        lines = lldp_output.split('\n')
        current_interface = None
        current_neighbor = {}

        for line in lines:
            line = line.strip()

            # 检测接口行，支持gigabitethernet和xgigabitethernet
            # 格式: Interface 'gigabitethernet 0/1/2' has 1 LLDP Neighbors:
            # 或: Interface 'xgigabitethernet 0/4/4' has 1 LLDP Neighbors:
            if ("Interface 'gigabitethernet" in line or "Interface 'xgigabitethernet" in line) and "has" in line and "LLDP Neighbors:" in line:
                # 提取物理接口名称
                interface_match = re.search(
                    r"Interface '(x?gigabitethernet\s+\d+/\d+/\d+)'", line)
                if interface_match:
                    current_interface = interface_match.group(1)
                    current_neighbor = {}
                    print(f"[DEBUG] 发现LLDP接口: {current_interface}")

            # 解析系统名称
            elif line.startswith("System Name:") and current_interface:
                system_name = line.replace("System Name:", "").strip()
                current_neighbor['system_name'] = system_name
                print(f"[DEBUG] 系统名称: {system_name}")

            # 解析端口ID
            elif line.startswith("Port ID:") and current_interface:
                port_match = re.search(r"Interface Name - (.+)", line)
                if port_match:
                    port_id = port_match.group(1).strip()
                    current_neighbor['port_id'] = port_id
                    print(f"[DEBUG] 端口ID: {port_id}")

            # 检测LLDPDU结束，保存当前邻居信息
            elif line.startswith("End Of LLDPDU:") and current_interface and current_neighbor:
                if 'system_name' in current_neighbor and 'port_id' in current_neighbor:
                    neighbors[current_interface] = current_neighbor.copy()
                    print(
                        f"[DEBUG] 保存LLDP邻居: {current_interface} -> {current_neighbor}")
                current_interface = None
                current_neighbor = {}

        print(f"[DEBUG] 解析LLDP邻居结果: {neighbors}")
    except Exception as e:
        print(f"[ERROR] 解析LLDP邻居失败: {e}")
    return neighbors


def process_device_interface_description(ip, user, pwd, writer, fail_log, file_lock):
    """处理单个设备的接口描述配置（修复版本）"""
    channel = None
    try:
        print(f"[INFO] 处理设备: {ip}")
        channel = create_channel(ip, user, pwd)
        if not channel:
            with file_lock:
                fail_log.write(ip + '\n')
            print(f"[ERROR] 设备 {ip} 连接失败")
            return

        # 设置屏幕长度避免分页
        execute_some_command(channel, "screen-length 512", 1)

        # 执行 con 命令
        con_cmd = "con"
        con_output = execute_some_command(channel, con_cmd, 3)
        print(f"[DEBUG] 设备 {ip} con 输出: {con_output[:200]}...")
        commands_executed = [
            (con_cmd, con_output, "执行成功" if "error" not in con_output.lower() else "执行失败")]

        # 获取OSPF邻居信息
        ospf_cmd = "show ospf neighbor brief"
        ospf_output = execute_some_command(channel, ospf_cmd, 5)
        print(f"[DEBUG] 设备 {ip} OSPF邻居输出: {ospf_output[:500]}...")
        commands_executed.append((ospf_cmd, ospf_output, "获取OSPF邻居"))

        # 获取LLDP邻居信息
        lldp_cmd = "show lldp neighbor"
        lldp_output = execute_some_command(channel, lldp_cmd, 8)
        print(f"[DEBUG] 设备 {ip} LLDP邻居输出: {lldp_output[:500]}...")
        commands_executed.append((lldp_cmd, lldp_output, "获取LLDP邻居"))

        # 解析OSPF和LLDP信息，找到需要配置的接口
        ospf_interfaces = parse_ospf_dot31_interfaces(ospf_output)
        lldp_neighbors = parse_lldp_neighbors(lldp_output)

        print(f"[DEBUG] 设备 {ip} 发现OSPF .31接口: {ospf_interfaces}")
        print(f"[DEBUG] 设备 {ip} 发现LLDP邻居接口: {list(lldp_neighbors.keys())}")

        # 根据LLDP邻居信息配置对应的OSPF .31接口描述
        configured_count = 0
        for physical_interface, neighbor_info in lldp_neighbors.items():
            # 构造对应的.31接口名称
            dot31_interface = f"{physical_interface}.31"

            # 检查这个.31接口是否在OSPF中
            if dot31_interface in ospf_interfaces:
                # 构建描述信息：对端设备名称-对端设备端口.31
                description = f"{neighbor_info['system_name']}-{neighbor_info['port_id']}.31"

                print(
                    f"[INFO] 设备 {ip} 准备配置接口 {dot31_interface} 描述: {description}")

                # 进入接口配置模式
                interface_cmd = f"inter {dot31_interface}"
                interface_output = execute_some_command(
                    channel, interface_cmd, 2)
                commands_executed.append(
                    (interface_cmd, interface_output, "进入接口"))

                # 先执行desc命令（可选，用于清空当前描述）
                desc_cmd = "desc"
                desc_output = execute_some_command(channel, desc_cmd, 1)
                commands_executed.append((desc_cmd, desc_output, "清空描述"))

                # 配置新描述
                description_cmd = f"description {description}"
                description_output = execute_some_command(
                    channel, description_cmd, 2)
                commands_executed.append(
                    (description_cmd, description_output, "配置描述"))
                print(
                    f"[INFO] 设备 {ip} 接口 {dot31_interface} 配置描述成功: {description}")

                # 退出接口配置模式
                quit_cmd = "q"
                quit_output = execute_some_command(channel, quit_cmd, 1)
                commands_executed.append((quit_cmd, quit_output, "退出接口"))

                configured_count += 1
            else:
                print(
                    f"[INFO] 设备 {ip} 物理接口 {physical_interface} 对应的 {dot31_interface} 不在OSPF进程中，跳过")

        # 特殊处理：检查是否有在OSPF中但没有LLDP邻居的.31接口
        for ospf_interface in ospf_interfaces:
            # 提取物理接口名称（去掉.31后缀）
            physical_interface = ospf_interface.replace('.31', '')

            # 如果这个物理接口没有LLDP邻居信息，记录警告
            if physical_interface not in lldp_neighbors:
                print(
                    f"[WARNING] 设备 {ip} 接口 {ospf_interface} 在OSPF中但没有找到对应的LLDP邻居信息")
                # 可以选择记录到日志或报告中
                commands_executed.append(
                    (f"warning_{ospf_interface}", f"接口 {ospf_interface} 在OSPF中但缺少LLDP邻居信息", "需要手动检查"))

        print(f"[INFO] 设备 {ip} 共配置了 {configured_count} 个接口描述")

        # 保存配置
        if configured_count > 0:
            save_cmd = "sa"
            save_output = execute_some_command(channel, save_cmd, 5)
            commands_executed.append((save_cmd, save_output, "保存配置"))
            print(f"[INFO] 设备 {ip} 配置已保存")
        else:
            print(f"[INFO] 设备 {ip} 无需要配置的接口，跳过保存")

        # 写入原始数据
        with file_lock:
            for cmd, output, status in commands_executed:
                writer.writerow([ip, cmd, output, status])

    except Exception as cmd_error:
        print(f"{Fore.YELLOW}⚠️ 设备 {ip} 执行命令失败: {cmd_error}{Style.RESET_ALL}")
        with file_lock:
            fail_log.write(ip + '\n')
    finally:
        if channel:
            try:
                execute_some_command(channel, "screen-length 25", 1)
                channel.close()
                print(f"[DEBUG] 设备 {ip} 连接已关闭")
            except Exception as close_error:
                print(
                    f"{Fore.YELLOW}⚠️ 关闭 {ip} 连接时出错: {close_error}{Style.RESET_ALL}")


def generate_interface_description_report(raw_file, report_file, host_file):
    """生成接口描述配置报告（修复版本）"""
    print(
        f"[generate_interface_description_report] 开始生成报告，源文件: {raw_file}, 目标文件: {report_file}")

    connection_failures = set()
    try:
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            connection_failures = set(line.strip() for line in f)
            print(f"[DEBUG] 读取连接失败设备: {len(connection_failures)} 台")
    except FileNotFoundError:
        print("[DEBUG] 未找到failure_ips.tmp文件")

    # 读取主机列表
    with open(host_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(f"[DEBUG] 读取设备清单: {len(host_ips)} 台设备")

    # 读取原始数据
    with open(raw_file, "r", encoding='utf-8') as f:
        reader = csv.reader(f)
        data = list(reader)
        print(f"[DEBUG] 读取原始数据: {len(data)} 条记录")

    # 生成报告
    with open(report_file, "w", encoding='utf-8', newline='') as report:
        writer = csv.writer(report)
        writer.writerow(["设备IP", "设备名称", "运行指令", "执行状态", "接口", "描述内容", "设备输出"])
        print(f"[DEBUG] 写入报告表头")

        processed_ips = set()
        current_device_name = "未知设备"
        current_interface = ""
        current_description = ""

        for idx, row in enumerate(data):
            if len(row) < 3:
                print(f"[WARNING] 第 {idx+1} 行数据格式错误，跳过: {row}")
                continue

            device_ip, cmd, output = row[0], row[1], row[2]
            status = row[3] if len(row) > 3 else "执行成功"

            processed_ips.add(device_ip)

            # 提取设备名称
            if cmd == "con":
                name_match = re.search(r'\[([^\]]+)\]', output, re.MULTILINE)
                current_device_name = name_match.group(
                    1).strip() if name_match else "未知设备"

            # 提取接口和描述信息（支持万兆接口）
            if (cmd.startswith("inter gigabitethernet") or cmd.startswith("inter xgigabitethernet")) and ".31" in cmd:
                current_interface = cmd.replace("inter ", "")
            elif cmd.startswith("description "):
                current_description = cmd.replace("description ", "")

            # 判断执行状态
            if "error" in output.lower():
                status = "执行失败"
            elif cmd.startswith("description "):
                status = "配置成功"
            elif cmd.startswith("warning_"):
                status = "需要手动检查"

            writer.writerow([device_ip, current_device_name, cmd,
                            status, current_interface, current_description, output])

        # 处理连接失败的设备
        for ip in host_ips:
            if ip not in processed_ips and ip in connection_failures:
                writer.writerow([ip, "连接失败", "-", "连接失败", "-", "-", "-"])

    print(f"✅ 接口描述配置报告生成完成，共处理 {len(host_ips)} 台设备")

#


def parse_protect_group_all(protect_group_output, l2vc_output):
    """
    解析show protect-group all和show mpls l2vc brief命令输出
    返回保护组状态信息列表
    """
    protect_groups = []

    if not protect_group_output:
        return [{
            'aps_id': '-', 'status': '-', 'master_vcid': '-', 'backup_vcid': '-',
            'type': '-', 'direction': '-', 'recovery': '-', 'sd': '-', 'wtr': '-',
            'hold_off': '-', 'protect_enable': '-', 'external_cmd': '-',
            'send_aps': '-', 'recv_aps': '-',
            'master_destination': '-', 'master_service_name': '-', 'master_vc_status': '-', 'master_interface': '-',
            'backup_destination': '-', 'backup_service_name': '-', 'backup_vc_status': '-', 'backup_interface': '-',
            'result': 'normal'
        }]

    # 解析L2VC信息，建立VCID到业务信息的映射
    l2vc_data_by_vcid = {}
    if l2vc_output:
        l2vc_lines = l2vc_output.split('\n')
        in_table = False
        for line in l2vc_lines:
            line = line.strip()
            if "VC-ID" in line and "Destination" in line:
                in_table = True
                continue
            if in_table and line and not line.startswith('-'):
                import re
                parts = re.split(r'\s{2,}', line.strip())
                if len(parts) >= 6:
                    vcid = parts[0]
                    destination = parts[1]
                    service_name = parts[2]
                    vc_status = parts[3]
                    interface = parts[4]
                    l2vc_data_by_vcid[vcid] = {
                        'destination': destination,
                        'service_name': service_name,
                        'vc_status': '✅ UP' if vc_status.lower() == 'up' else '❌ Down',
                        'interface': interface
                    }

    # 解析保护组信息
    protect_lines = protect_group_output.split('\n')
    in_table = False

    for line in protect_lines:
        line = line.strip()

        if not line or line.startswith('-'):
            continue

        # 检测表格开始，使用英文表头
        if "APS-ID" in line and "Status" in line and "Master/Backup" in line:
            in_table = True
            continue

        if in_table and line:
            if line.startswith('[') and ']' in line:
                break

            import re
            # 修改正则表达式，更精确地匹配字段
            # 表头：APS-ID  Status   Master/Backup                  Type        Dir  Revt  SD  WTR     HoldOff    ProtcEn  ExtCmd  SendAPS   RecvAPS
            # 示例：1       Normal   6025      (OK)/6026      (OK)  redundancy  BiD  Y     N   180(s)  0    (ms)  Y        None    0f000000  00000000
            
            # 先提取APS-ID和Status
            aps_status_match = re.match(r'(\d+)\s+(\S+)\s+(.+)', line)
            if not aps_status_match:
                continue
                
            aps_id = aps_status_match.group(1)
            status = aps_status_match.group(2)
            remaining = aps_status_match.group(3)
            
            # 提取Master/Backup部分：6025      (OK)/6026      (OK)
            master_backup_match = re.match(r'(\d+)\s+\([^)]+\)/(\d+)\s+\([^)]+\)\s+(.+)', remaining)
            if not master_backup_match:
                continue
                
            master_vcid = master_backup_match.group(1)
            backup_vcid = master_backup_match.group(2)
            remaining_fields = master_backup_match.group(3)
            
            # 解析剩余字段：Type Dir Revt SD WTR HoldOff ProtcEn ExtCmd SendAPS RecvAPS
            # redundancy  BiD  Y     N   180(s)  0    (ms)  Y        None    0f000000  00000000
            fields = re.split(r'\s+', remaining_fields.strip())
            
            # 根据实际输出格式提取字段
            if len(fields) >= 10:
                type_field = fields[0]          # redundancy
                direction = fields[1]           # BiD
                recovery = fields[2]            # Y
                sd = fields[3]                  # N
                wtr = fields[4]                 # 180(s)
                hold_off_value = fields[5]      # 0
                hold_off_unit = fields[6]       # (ms)
                protect_enable = fields[7]      # Y
                external_cmd = fields[8]        # None
                send_aps = fields[9]           # 0f000000
                recv_aps = fields[10] if len(fields) > 10 else '-'  # 00000000
                
                # 组合hold_off字段
                hold_off = f"{hold_off_value} {hold_off_unit}"
            else:
                # 如果字段不够，设置默认值
                type_field = fields[0] if len(fields) > 0 else '-'
                direction = fields[1] if len(fields) > 1 else '-'
                recovery = fields[2] if len(fields) > 2 else '-'
                sd = fields[3] if len(fields) > 3 else '-'
                wtr = fields[4] if len(fields) > 4 else '-'
                hold_off = f"{fields[5]} {fields[6]}" if len(fields) > 6 else '-'
                protect_enable = fields[7] if len(fields) > 7 else '-'
                external_cmd = fields[8] if len(fields) > 8 else '-'
                send_aps = fields[9] if len(fields) > 9 else '-'
                recv_aps = fields[10] if len(fields) > 10 else '-'

            # 获取L2VC信息
            master_l2vc = l2vc_data_by_vcid.get(master_vcid, {})
            master_destination = master_l2vc.get('destination', '-')
            master_service_name = master_l2vc.get('service_name', '-')
            master_vc_status = master_l2vc.get('vc_status', '-')
            master_interface = master_l2vc.get('interface', '-')

            backup_l2vc = l2vc_data_by_vcid.get(backup_vcid, {})
            backup_destination = backup_l2vc.get('destination', '-')
            backup_service_name = backup_l2vc.get('service_name', '-')
            backup_vc_status = backup_l2vc.get('vc_status', '-')
            backup_interface = backup_l2vc.get('interface', '-')

            # 使用英文"Normal"判断状态
            result = 'normal' if status == 'Normal' else 'error'

            protect_groups.append({
                'aps_id': aps_id,
                'status': status,
                'master_vcid': master_vcid,
                'backup_vcid': backup_vcid,
                'type': type_field,
                'direction': direction,
                'recovery': recovery,
                'sd': sd,
                'wtr': wtr,
                'hold_off': hold_off,
                'protect_enable': protect_enable,
                'external_cmd': external_cmd,
                'send_aps': send_aps,
                'recv_aps': recv_aps,
                'master_destination': master_destination,
                'master_service_name': master_service_name,
                'master_vc_status': master_vc_status,
                'master_interface': master_interface,
                'backup_destination': backup_destination,
                'backup_service_name': backup_service_name,
                'backup_vc_status': backup_vc_status,
                'backup_interface': backup_interface,
                'result': result
            })

    if not protect_groups:
        protect_groups.append({
            'aps_id': '-', 'status': '无条目', 'master_vcid': '-', 'backup_vcid': '-',
            'type': '-', 'direction': '-', 'recovery': '-', 'sd': '-', 'wtr': '-',
            'hold_off': '-', 'protect_enable': '-', 'external_cmd': '-',
            'send_aps': '-', 'recv_aps': '-',
            'master_destination': '-', 'master_service_name': '-', 'master_vc_status': '-', 'master_interface': '-',
            'backup_destination': '-', 'backup_service_name': '-', 'backup_vc_status': '-', 'backup_interface': '-',
            'result': 'normal'
        })

    return protect_groups

#


def parse_dcn_routing_table(output):
    """解析DCN路由表，准确提取路由总数并检测异常路由"""

    # 从输出中提取 "Total - " 后的路由总数
    total_match = re.search(r'Total -  (\d+)', output)
    if total_match:
        route_count = int(total_match.group(1))
        if route_count == 0:
            return [{"目的网络/掩码": "无条目"}]
        else:
            # 检查是否有 '141.' 开头的异常路由
            if '141.' in output:
                # 解析异常路由
                error_pattern = re.compile(
                    r'(141\.\d+\.\d+\.\d+/\d+)\s+(\S+)\s+(\d+)\s+(\d+)\s+(\d+\.\d+\.\d+\.\d+)\s+([\w\s/\.]+?)\s+(\S+)|'
                    r'(\d+\.\d+\.\d+\.\d+/\d+)\s+(\S+)\s+(\d+)\s+(\d+)\s+(141\.\d+\.\d+\.\d+)\s+([\w\s/\.]+?)\s+(\S+)',
                    re.MULTILINE
                )
                matches = error_pattern.findall(output)
                error_routes = []
                for match in matches:
                    if match[0]:  # 目的地址是 141 开头
                        dest_mask, proto, pre, cost, nexthop, interface, uptime = match[:7]
                    else:  # 下一跳是 141 开头
                        dest_mask, proto, pre, cost, nexthop, interface, uptime = match[7:]
                    error_routes.append({
                        "目的网络/掩码": dest_mask,
                        "协议": proto,
                        "优先级": pre,
                        "开销": cost,
                        "下一跳": nexthop,
                        "接口": interface.strip(),
                        "存活时间": uptime,
                        "Result": "error",
                        "备注": "检测到出厂DCN路由IP请检查配置"
                    })
                if error_routes:
                    return error_routes
            # 无异常路由时，返回汇总信息
            return [{"目的网络/掩码": f"共{route_count}条路由", "协议": "正常", "优先级": "-",
                     "开销": "-", "下一跳": "-", "接口": "-", "存活时间": "-",
                     "Result": "normal", "备注": "无异常路由"}]
    else:
        # 如果未找到 "Total - "，返回无条目
        return [{"目的网络/掩码": "无条目"}]
#

# ···········


def generate_qa_report(raw_file, report_file, host_file, selected_items):
    """Generate QA report with enhanced summary table visualization - 排除连接失败的设备"""
    print(
        f"{Fore.CYAN}[START] Starting QA report generation, source: {raw_file}, target: {report_file}{Style.RESET_ALL}")

    # Initialize workbook and styles
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "运维质量评估"

    # Define styles (保持原有样式定义不变)
    yellow_fill = PatternFill(start_color='FFFF00',
                              end_color='FFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500',
                              end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='92D050',
                             end_color='92D050', fill_type='solid')
    light_green_fill = PatternFill(
        start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    light_red_fill = PatternFill(
        start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_amber_fill = PatternFill(
        start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    hyperlink_font = Font(color="0000FF", underline="single", size=11)

    # Set column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 15

    # Title and header rows
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = "STN-A设备运维质量评估报告"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = center_alignment
    ws_summary['A1'].fill = yellow_fill
    ws_summary['A1'].border = thin_border

    header_row = 2
    headers = ["检查分类", "巡检项目", "健康度", "直观展示", "条目计数", "健康状态"]
    for col, value in enumerate(headers, 1):
        cell = ws_summary.cell(row=header_row, column=col, value=value)
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Read host file
    with open(host_file, "r", encoding='gbk', errors='ignore') as f:
        reader = csv.reader(f)
        host_ips = [row[0].strip() for row in reader]
        print(
            f"{Fore.GREEN}[DEBUG] Loaded {len(host_ips)} devices{Style.RESET_ALL}")

    # Read connection failures FIRST
    connection_failures = {}
    try:
        print(
            f"{Fore.CYAN}[DEBUG] Reading connection failures from failure_ips.tmp{Style.RESET_ALL}")
        with open("failure_ips.tmp", "r", encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                # 检查是否包含逗号
                if ',' in line:
                    ip, reason = line.split(',', 1)
                else:
                    # 如果没有逗号，将整行作为IP，原因设为默认值
                    ip = line
                    reason = "连接失败"

                connection_failures[ip.strip()] = reason.strip()
                # 修复：不要将连接失败的设备添加到data字典中
                # data[ip] = {"Connection failed": reason}  # 删除这行
        print(
            f"{Fore.CYAN}[DEBUG] Found {len(connection_failures)} connection failures{Style.RESET_ALL}")
    except FileNotFoundError:
        print(
            f"{Fore.YELLOW}[DEBUG] No failure_ips.tmp found{Style.RESET_ALL}")

     # 过滤掉连接失败的设备，只处理成功连接的设备
    successful_host_ips = [
        ip for ip in host_ips if ip not in connection_failures]
    print(f"{Fore.GREEN}[DEBUG] Processing {len(successful_host_ips)} successful devices (excluded {len(connection_failures)} failed devices){Style.RESET_ALL}")

    # Read raw data - 只读取成功连接设备的数据
    data = {}
    with open(raw_file, "r", encoding='utf-8') as f:
        csv.field_size_limit(sys.maxsize)
        reader = csv.reader(f)
        for row in reader:
            if len(row) != 3:
                print(
                    f"{Fore.YELLOW}[WARNING] Invalid row format, skipping: {row}{Style.RESET_ALL}")
                continue
            ip, cmd, output = row

            # 跳过连接失败的设备数据
            if ip in connection_failures:
                print(
                    f"{Fore.YELLOW}[DEBUG] Skipping data for failed device: {ip}{Style.RESET_ALL}")
                continue

            if ip not in data:
                data[ip] = {}
            data[ip][cmd] = output
            print(
                f"{Fore.YELLOW}[DEBUG] Loaded data for {ip}, cmd: {cmd}{Style.RESET_ALL}")

    # Organize inspection items by category
    categories = {
        "设备基础状态": [item for item in selected_items if item["category"] == "设备基础状态"],
        "硬件可靠性": [item for item in selected_items if item["category"] == "硬件可靠性"],
        "系统运行状态": [item for item in selected_items if item["category"] == "系统运行状态"],
        "资源监控": [item for item in selected_items if item["category"] == "资源监控"],
        "路由协议健康度": [item for item in selected_items if item["category"] == "路由协议健康度"],
        "冗余与容灾": [item for item in selected_items if item["category"] == "冗余与容灾"],
        "基础安全配置": [item for item in selected_items if item["category"] == "基础安全配置"]
    }

    # Store health scores and device counts
    health_scores = {}
    item_counts = {}

    # Process Loopback addresses - 只处理成功连接的设备
    loopback31_addresses = {}
    loopback1023_addresses = {}
    for ip in successful_host_ips:  # 使用过滤后的设备列表
        loopback31_output = data.get(ip, {}).get(
            "show interface loopback 31", "")
        loopback1023_output = data.get(ip, {}).get(
            "show interface loopback 1023", "")
        loopback31_addr = parse_loopback_address(loopback31_output)
        loopback1023_addr = parse_loopback_address(loopback1023_output)

        if loopback31_addr != "无条目":
            if loopback31_addr not in loopback31_addresses:
                loopback31_addresses[loopback31_addr] = []
            loopback31_addresses[loopback31_addr].append(ip)
        if loopback1023_addr != "无条目":
            if loopback1023_addr not in loopback1023_addresses:
                loopback1023_addresses[loopback1023_addr] = []
            loopback1023_addresses[loopback1023_addr].append(ip)

    # Process sub-sheets and calculate health scores
    for item in selected_items:
        sheet_name = item['sheet_name']
        ws = wb.create_sheet(title=sheet_name)
        print(f"{Fore.GREEN}[DEBUG] 创建子表: {sheet_name}{Style.RESET_ALL}")

        # 初始化计数器
        total_results = 0
        normal_results = 0

        if item['name'] == "设备运行时间检查":
            headers = ["网元类型", "网元名称", "网元IP", "UpTime", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                total_results += 1
                if ip not in data or "show device" not in data[ip]:
                    ws.append(["-", "-", ip, "数据异常", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=5).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 数据异常，写入子表{Style.RESET_ALL}")
                    continue
                output = data[ip]["show device"]
                ne_type, device_name, uptime, result = item['parser'](output)
                ws.append([ne_type, device_name, ip, uptime, result])
                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if result == "normal":
                    normal_results += 1
                else:
                    ws.cell(row=ws.max_row, column=5).fill = orange_fill
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表: {ne_type}, {device_name}, {uptime}, {result}{Style.RESET_ALL}")
                # 计算健康度
                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[sheet_name] = f"{health_percentage:.0f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "主控盘运行状态":
            headers = ["网元类型", "网元名称", "网元IP", "CPU使用率",
                       "五分钟CPU使用率", "十五分钟CPU使用率", "内存使用率", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                total_results += 1
                if ip not in data or "show device" not in data[ip]:
                    ws.append(["-", "-", ip, "数据异常", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=8).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 数据异常，写入子表{Style.RESET_ALL}")
                    continue
                output = data[ip]["show device"]
                ne_type, device_name, cpu_usage, cpu_5min, cpu_15min, memory_usage, result = item['parser'](
                    output)
                ws.append([ne_type, device_name, ip, cpu_usage,
                          cpu_5min, cpu_15min, memory_usage, result])
                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if result == "normal":
                    normal_results += 1
                else:
                    ws.cell(row=ws.max_row, column=8).fill = orange_fill
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表: {ne_type}, {device_name}, {result}{Style.RESET_ALL}")

                # 计算健康度
                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[sheet_name] = f"{health_percentage:.0f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "协议报文处理状态":
            headers = ["网元类型", "网元名称", "网元IP", "协议类型",
                       "总数据包数", "通过的数据包", "丢弃的数据包", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show cpu-defend stats" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip,
                              "数据异常", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=8).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 数据异常，写入子表{Style.RESET_ALL}")
                    continue
                output = data[ip]["show cpu-defend stats"]
                protocol_results = item['parser'](output)
                start_row = ws.max_row + 1
                for res in protocol_results:
                    total_results += 1
                    ws.append([
                        ne_type, device_name, ip,
                        res["protocol"], res["total_packets"], res["passed_packets"],
                        res["dropped_packets"], res["result"]
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if res["result"] == "normal":
                        normal_results += 1
                    else:
                        ws.cell(row=ws.max_row, column=8).fill = orange_fill
                end_row = ws.max_row
                if start_row < end_row:
                    for col in range(1, 4):
                        ws.merge_cells(
                            start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(protocol_results)} 行，合并单元格{Style.RESET_ALL}")

                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "真实版本信息":
            headers = ["网元类型", "网元名称", "网元IP", "组件类型", "版本标识",
                       "版本号", "编译/构建时间", "打包时间", "下载时间", "附加说明", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show real-version" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip, "数据异常", "-",
                              "-", "-", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=11).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无版本数据，写入子表{Style.RESET_ALL}")
                    continue
                output = data[ip]["show real-version"]
                versions = item['parser'](output)
                if not versions:
                    total_results += 1
                    ws.append([ne_type, device_name, ip, "解析失败",
                              "-", "-", "-", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=11).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 解析失败，写入子表{Style.RESET_ALL}")
                    continue
                start_row = ws.max_row + 1
                for version in versions:
                    total_results += 1
                    result = version["Result"]
                    if result == "normal":
                        normal_results += 1
                    ws.append([
                        ne_type, device_name, ip,
                        version["组件类型"], version["版本标识"], version["版本号"],
                        version["编译/构建时间"], version["打包时间"], version["下载时间"],
                        version["附加说明"], result
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if result != "normal":
                        ws.cell(row=ws.max_row, column=11).fill = orange_fill
                end_row = ws.max_row
                if start_row < end_row:
                    for col in range(1, 4):
                        ws.merge_cells(
                            start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(versions)} 行，合并单元格{Style.RESET_ALL}")

                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "风扇转速及温度状态":
            headers = ["网元类型", "网元名称", "网元IP", "风扇状态", "风扇速度",
                       "板卡槽位", "SEN_01", "SEN_02", "SEN_03", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                fan_data = None
                if ip in data and "show fan" in data[ip]:
                    fan_data = parse_fan(data[ip]["show fan"])
                temperature_data = None
                if ip in data and "show temperature" in data[ip]:
                    temperature_data = item['parser'](
                        data[ip]["show temperature"])

                if not temperature_data:
                    total_results += 1
                    ws.append([ne_type, device_name, ip, "数据异常",
                              "-", "-", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=10).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无温度数据，写入子表{Style.RESET_ALL}")
                    continue

                # 记录当前设备数据的起始行
                start_row = ws.max_row + 1
                device_rows_added = 0  # 追踪实际添加的行数

                for temp_row in temperature_data:
                    total_results += 1
                    device_rows_added += 1
                    slot = temp_row['slot']
                    sen_01 = temp_row['sen_01']
                    sen_02 = temp_row['sen_02']
                    sen_03 = temp_row['sen_03']
                    fan_status = fan_data['status'] if fan_data else "-"
                    fan_speeds = fan_data['speeds'] if fan_data else "-"
                    error = False

                    # Check SEN_01 only for temperature
                    try:
                        sen_01_value = float(sen_01)
                        if sen_01_value > 89 or sen_01_value < 35:
                            error = True
                            print(
                                f"{Fore.YELLOW}[DEBUG] SEN_01 温度 {sen_01} 超出范围（>89或<35），设置 error{Style.RESET_ALL}")
                    except (ValueError, TypeError):
                        pass  # Ignore invalid SEN_01 values

                    # Check fan speeds (valid percentage and >= 20%)
                    if fan_data and fan_data['result'] == 'error':
                        error = True
                        print(
                            f"{Fore.YELLOW}[DEBUG] 风扇速度非百分数值或低于20%，设置 error{Style.RESET_ALL}")

                    result = "error" if error else "normal"
                    ws.append([ne_type, device_name, ip, fan_status,
                              fan_speeds, slot, sen_01, sen_02, sen_03, result])

                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border

                    if result == "normal":
                        normal_results += 1
                    else:
                        ws.cell(row=ws.max_row, column=10).fill = orange_fill

                # 修复：确保只有在添加了多行数据时才进行合并
                end_row = ws.max_row
                if device_rows_added > 1:  # 只有当实际添加了多行时才合并
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 添加了 {device_rows_added} 行，开始合并单元格 (行 {start_row} 到 {end_row}){Style.RESET_ALL}")
                    for col in range(1, 6):  # 合并前5列（网元类型、网元名称、网元IP、风扇状态、风扇速度）
                        try:
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                        except Exception as e:
                            print(
                                f"{Fore.RED}[ERROR] 合并单元格失败 (列 {col}): {e}{Style.RESET_ALL}")
                else:
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 只有 {device_rows_added} 行，无需合并单元格{Style.RESET_ALL}")

                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(temperature_data)} 行温度数据{Style.RESET_ALL}")

            # 修复：将健康度计算移到所有设备处理完成后
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
            item_counts[item['sheet_name']] = (
                normal_results, total_results)

        elif item['name'] == "系统与硬件版本状态":
            headers = ["网元类型", "网元名称", "设备MAC", "网元IP", "系统版本", "运行时间", "对象ID", "槽位", "板卡名称",
                       "软件版本", "FPGA版本", "EPLD版本", "硬件版本", "备注", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show version" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, "-", ip, "数据异常", "-",
                              "-", "-", "-", "-", "-", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=15).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无版本数据，写入子表{Style.RESET_ALL}")
                    continue
                version_info = item['parser'](data[ip]["show version"])
                if not version_info:
                    total_results += 1
                    ws.append([ne_type, device_name, "-", ip, "解析失败", "-",
                              "-", "-", "-", "-", "-", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=15).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 解析失败，写入子表{Style.RESET_ALL}")
                    continue
                system_version = version_info.get('system_version', '-')
                uptime = version_info.get('uptime', '-')
                object_id = version_info.get('object_id', '-')
                system_mac = version_info.get('system_mac', '-')
                slots = version_info.get('slots', [])
                if not slots:
                    total_results += 1
                    ws.append([ne_type, device_name, system_mac, ip, system_version,
                              uptime, object_id, "-", "-", "-", "-", "-", "-", "-", "normal"])
                    normal_results += 1
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无槽位数据，写入子表{Style.RESET_ALL}")
                else:
                    start_row = ws.max_row + 1
                    for slot in slots:
                        total_results += 1
                        slot_num = slot['slot']
                        board_name = slot['board_name']
                        remark = slot['remark']
                        versions = slot['versions']
                        software_version = versions.get('Software', '-')
                        fpga_version = versions.get('FPGA01', '-')
                        epld_version = versions.get('EPLD01', '-')
                        hardware_version = versions.get('Hardware', '-')
                        result = "normal"
                        ws.append([ne_type, device_name, system_mac, ip, system_version, uptime, object_id,
                                  slot_num, board_name, software_version, fpga_version, epld_version,
                                  hardware_version, remark, result])
                        normal_results += 1
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border
                        if result == "error":
                            ws.cell(row=ws.max_row,
                                    column=15).fill = orange_fill
                    end_row = ws.max_row
                    if start_row < end_row:
                        for col in range(1, 8):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(slots)} 行，合并单元格{Style.RESET_ALL}")

                    health_percentage = (normal_results / total_results *
                                         100) if total_results > 0 else 0
                    health_scores[item['sheet_name']
                                  ] = f"{health_percentage:.0f}%"
                    item_counts[item['sheet_name']] = (
                        normal_results, total_results)

        elif item['name'] == "光模块信息检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "接口名称", "当前状态", "最近UP时间", "最近DOWN时间",
                "描述", "IPv4地址", "IPv6地址", "MAC地址", "最大传输单元(L3)", "厂商型号", "光模块类型",
                "波长", "传输距离", "Rx光功率(dBm)", "Rx范围(dBm)", "Tx光功率(dBm)", "Tx范围(dBm)",
                "偏置电流(mA)", "偏置范围(mA)", "电压(mV)", "电压范围(mV)", "温度(°C)", "温度范围(°C)",
                "端口带宽", "光模块带宽", "输入速率(bps)", "输入带宽利用率", "输出速率(bps)", "输出带宽利用率",
                "当前CRC", "邻居系统名称", "邻居系统描述", "邻居端口", "邻居IP", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show interface" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["数据异常"] * 35 + ["error", "无接口数据"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=len(
                        headers) - 1).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无接口数据，写入子表{Style.RESET_ALL}")
                    continue
                interface_output = data[ip]["show interface"]
                lldp_output = data[ip].get("show lldp neighbor", "")
                optical_data = item['parser'](
                    ip, interface_output, lldp_output)
                if not optical_data:
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["无光模块数据"] * 35 + ["error", "无光模块数据"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=len(
                        headers) - 1).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无光模块数据，写入子表{Style.RESET_ALL}")
                    continue
                start_row = ws.max_row + 1
                for row_data in optical_data:
                    total_results += 1
                    ws.append([
                        ne_type, device_name, ip,
                        row_data["interface"], row_data["current_state"], row_data["last_up"], row_data["last_down"],
                        sanitize_string(row_data["description"]),  # 清理描述字段
                        row_data["ipv4"], row_data["ipv6"], row_data["mac"],
                        row_data["mtu_l3"], sanitize_string(
                            row_data["vendor_pn"]),  # 清理厂商零件号
                        sanitize_string(row_data["transceiver_id"]),  # 清理收发器ID
                        row_data["wavelength"], row_data["distance"], row_data["rx_power"], row_data["rx_range"],
                        row_data["tx_power"], row_data["tx_range"], row_data["bias"], row_data["bias_range"],
                        row_data["voltage"], row_data["voltage_range"], row_data["temperature"], row_data["temp_range"],
                        row_data["port_bw"], row_data["transceiver_bw"], row_data["input_rate"], row_data["input_util"],
                        row_data["output_rate"], row_data["output_util"], row_data["crc"],
                        sanitize_string(
                            row_data["neighbor_system_name"]),  # 清理邻居系统名称
                        sanitize_string(
                            # 清理邻居系统描述
                            row_data["neighbor_system_description"]),
                        sanitize_string(row_data["neighbor_port"]),  # 清理邻居端口
                        sanitize_string(row_data["neighbor_ip"]),  # 清理邻居IP
                        row_data["result"], row_data["error_reasons"]
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if row_data["result"] == "normal":
                        normal_results += 1
                    else:
                        ws.cell(row=ws.max_row, column=len(
                            headers) - 1).fill = orange_fill
                end_row = ws.max_row
                if start_row < end_row:
                    for col in range(1, 4):
                        ws.merge_cells(
                            start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(optical_data)} 行，合并单元格{Style.RESET_ALL}")

                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "电源状态检查":
            headers = ["网元类型", "网元名称", "网元IP", "槽位", "当前电压", "电压比", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show voltage" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip,
                              "数据异常", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=7).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无电压数据，写入子表{Style.RESET_ALL}")
                    continue
                output = data[ip]["show voltage"]
                voltage_data = item['parser'](output)
                if not voltage_data:
                    total_results += 1
                    ws.append([ne_type, device_name, ip,
                              "数据异常", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=7).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 无电压数据，写入子表{Style.RESET_ALL}")
                    continue
                start_row = ws.max_row + 1
                for row_data in voltage_data:
                    total_results += 1
                    ws.append([
                        ne_type, device_name, ip,
                        row_data["slot"], row_data["voltage"], row_data["ratio"], row_data["result"]
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if row_data["result"] == "normal":
                        normal_results += 1
                    else:
                        ws.cell(row=ws.max_row, column=7).fill = orange_fill
                end_row = ws.max_row
                if start_row < end_row:
                    for col in range(1, 4):
                        ws.merge_cells(
                            start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(voltage_data)} 行，合并单元格{Style.RESET_ALL}")

                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "FW软件版本一致性检查":
            headers = ["网元类型", "网元名称", "网元IP", "主用版本", "备用版本", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                total_results += 1
                if ip not in data or "show device" not in data[ip]:
                    ws.append(["-", "-", ip, "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=6).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备 {ip} 数据异常，写入子表{Style.RESET_ALL}")
                    continue
                output = data[ip]["show device"]
                # 修改这里：传递IP地址给parser函数
                ne_type, device_name, main_version, backup_version, result = item['parser'](
                    output, ip)  # 添加ip参数
                ws.append([ne_type, device_name, ip,
                          main_version, backup_version, result])
                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if result == "normal":
                    normal_results += 1
                else:
                    ws.cell(row=ws.max_row, column=6).fill = orange_fill
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表: {ne_type}, {device_name}, {main_version}, {backup_version}, {result}{Style.RESET_ALL}")

                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "板卡CPU内存使用率":
            headers = [
                "网元类型", "网元名称", "网元IP",
                "15分钟内性能监控源", "时间", "15分钟内温度(℃)", "15分钟内CPU利用率", "15分钟内内存利用率",
                "24小时内性能监控源", "时间", "24小时内温度(℃)", "24小时内CPU利用率", "24小时内内存利用率",
                "Result"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            print(
                f"{Fore.YELLOW}[DEBUG] 设置子表 {sheet_name} 表头{Style.RESET_ALL}")

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show pm cur-15m Dev" not in data[ip] or "show pm cur-24h Dev" not in data[ip]:
                    print(f"设备 {ip} 无性能数据")
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["数据异常"] * 10 + ["error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=14).fill = orange_fill
                    continue

                output_15m = data[ip]["show pm cur-15m Dev"]
                output_24h = data[ip]["show pm cur-24h Dev"]
                perf_data = item['parser'](output_15m, output_24h)

                if not perf_data:
                    print(f"设备 {ip} 解析后的性能数据为空")
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["数据异常"] * 10 + ["error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=14).fill = orange_fill
                    continue

                start_row = ws.max_row + 1
                for row_data in perf_data:
                    total_results += 1
                    ws.append([
                        ne_type, device_name, ip,
                        row_data["pm_source_15m"], row_data["time_15m"], row_data["temp_15m"],
                        row_data["cpu_15m"], row_data["mem_15m"],
                        row_data["pm_source_24h"], row_data["time_24h"], row_data["temp_24h"],
                        row_data["cpu_24h"], row_data["mem_24h"],
                        row_data["result"]
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if row_data["result"] == "normal":
                        normal_results += 1
                    else:
                        ws.cell(row=ws.max_row, column=14).fill = orange_fill

                end_row = ws.max_row
                if start_row < end_row:
                    for col in range(1, 4):
                        ws.merge_cells(
                            start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                print(
                    f"{Fore.YELLOW}[DEBUG] 设备 {ip} 写入子表 {len(perf_data)} 行，合并单元格{Style.RESET_ALL}")

            # 计算健康度
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)
            print(
                f"{Fore.YELLOW}[DEBUG] 子表 {sheet_name} 健康度: {normal_results}/{total_results} = {health_percentage:.0f}%{Style.RESET_ALL}")

        elif item['name'] == "NTP时间同步分析":
            headers = ["网元类型", "网元名称", "网元IP", "NTP状态", "同步状态", "主/备NTP服务器",
                       "同步间隔", "NTP时间偏差", "本地时间", "UTC时间", "时区偏移", "PC执行时间", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # ===== 删除固定的collection_time =====
            # collection_time = datetime.now()  # 删除这行

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])

                if ip not in data or "show cloc" not in data[ip] or "show ntp-service" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip, "数据异常", "-",
                              "-", "-", "-", "-", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=13).fill = orange_fill
                    print(
                        f"{Fore.YELLOW}[DEBUG] 设备{ip} 缺少NTP相关数据，写入error{Style.RESET_ALL}")
                    continue
                
                cloc_output = data[ip]["show cloc"]
                ntp_output = data[ip]["show ntp-service"]
                print(f"[DEBUG] 设备{ip} cloc_output: {cloc_output[:100]}...")
                print(f"[DEBUG] 设备{ip} ntp_output: {ntp_output[:100]}...")

                # ===== 先从输出中提取PC_TIME =====
                pc_time = extract_pc_time(cloc_output)
                if pc_time is None:
                    pc_time = extract_pc_time(ntp_output)
                if pc_time is None:
                    print(f"{Fore.YELLOW}[WARNING] 设备{ip}无法提取PC_TIME，使用当前时间{Style.RESET_ALL}")
                    pc_time = datetime.now()
                else:
                    print(f"{Fore.GREEN}[DEBUG] 设备{ip}提取到PC_TIME: {pc_time}{Style.RESET_ALL}")

                # ===== 调用修改后的解析函数，传入从输出中提取的PC时间 =====
                ntp_data = parse_ntp_status(cloc_output, ntp_output, pc_time)
                total_results += 1

                ws.append([
                    ne_type, device_name, ip, ntp_data["ntp_enable"], ntp_data["ntp_status"],
                    f"{ntp_data['server_pref']}/{ntp_data['server']}", ntp_data["syn_interval"],
                    ntp_data["time_deviation"], ntp_data["local_time"], ntp_data["utc_time"],
                    ntp_data["time_zone"], ntp_data["pc_time"], ntp_data["result"]
                ])

                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border

                # 根据结果设置单元格颜色
                if ntp_data["result"] == "normal":
                    normal_results += 1
                elif ntp_data["result"] == "warning":
                    ws.cell(row=ws.max_row,
                            column=13).fill = yellow_fill  # 警告用黄色
                else:  # error
                    ws.cell(row=ws.max_row,
                            column=13).fill = orange_fill  # 错误用橙色

                print(f"[DEBUG] 设备 {ip} 写入子表: {ntp_data['result']}")

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)
    
        elif item['name'] == "硬盘资源占用分析":
            headers = ["网元类型", "网元名称", "网元IP", "总容量",
                       "剩余容量", "使用率", "告警阈值", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show flash-usage" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip,
                              "数据异常", "-", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=8).fill = orange_fill
                    continue
                output = data[ip]["show flash-usage"]
                flash_data = parse_flash_usage(output)
                total_results += 1
                ws.append([
                    ne_type, device_name, ip,
                    flash_data["total_flash"], flash_data["free_space"], flash_data["usage_percent"],
                    flash_data["threshold"], flash_data["result"]
                ])
                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if flash_data["result"] == "normal":
                    normal_results += 1
                else:
                    ws.cell(row=ws.max_row, column=8).fill = orange_fill

                # Calculate health percentage

                health_percentage = (
                    normal_results / total_results * 100) if total_results > 0 else 0
                health_scores[item['sheet_name']] = f"{health_percentage:.2f}%"
                item_counts[item['sheet_name']] = (
                    normal_results, total_results)

        elif item['name'] == "BFD会话检查(VC业务统计)":
            headers = [
                "网元类型", "网元名称", "网元IP", "VC总数", "UP", "DOWN",  # 新增的三列
                "APS组ID", "会话名称", "本地ID", "远端ID", "状态", "主备角色",
                "发送间隔", "接收间隔", "检测倍数", "本地鉴别器", "远端鉴别器", "鉴别器状态", "首次报文接收",
                "连续性检查", "MEP启用", "loopback31地址", "VCID", "目的地址", "业务名称", "VC状态", "接口",
                "本地MTU", "远端MTU", "VC类型", "本地控制字", "远端控制字", "当前使用控制字",
                "本地伪线状态能力", "远端伪线状态能力", "当前伪线状态TLV", "本地伪线状态", "远端伪线状态",
                "本地VCCV能力", "远端VCCV能力", "Result"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                
                ne_type, device_name = "-", "-"
                loopback31_address = "-"
                total_vc, up_count, down_count = 0, 0, 0
                
                # 获取设备信息
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(data[ip]["show device"])
                
                # 获取loopback31地址
                if ip in data and "show interface loopback 31" in data[ip]:
                    loopback31_output = data[ip]["show interface loopback 31"]
                    loopback31_address = parse_loopback31(loopback31_output)
                
                # 检查必需的命令输出是否存在
                required_commands = ["show bfd session brief", "show bfd configuration pw", "show mpls l2vc brief"]
                missing_commands = []
                
                for cmd in required_commands:
                    if ip not in data or cmd not in data[ip]:
                        missing_commands.append(cmd)
                
                # 如果有缺失的命令，标记为数据异常
                if missing_commands:
                    total_results += 1
                    ws.append([ne_type, device_name, ip, "-", "-", "-"] + ["数据异常"] * 35 + ["error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=41).fill = orange_fill  # Result列位置更新为41
                    continue
                
                # 获取命令输出
                brief_output = data[ip]["show bfd session brief"]
                config_output = data[ip]["show bfd configuration pw"]
                l2vc_output = data[ip]["show mpls l2vc brief"]
                ldp_detail_output = data[ip].get("show ldp l2vc detail", "")
                
                # 解析VC统计信息
                total_vc, up_count, down_count = parse_l2vc_summary(l2vc_output)
                
                # 解析BFD会话数据
                try:
                    bfd_data = parse_bfd_sessions(brief_output, config_output, l2vc_output, ldp_detail_output)
                except Exception as e:
                    print(f"解析IP {ip} 的BFD数据时出错: {e}")
                    total_results += 1
                    ws.append([ne_type, device_name, ip, str(total_vc), str(up_count), str(down_count)] + 
                             ["解析错误"] * 35 + ["error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=41).fill = orange_fill
                    continue
                
                # 如果没有BFD会话数据，仍然显示设备信息和VC统计
                if not bfd_data or (len(bfd_data) == 1 and bfd_data[0]['session_name'] == '无条目'):
                    total_results += 1
                    normal_results += 1
                    ws.append([
                        ne_type, device_name, ip, str(total_vc), str(up_count), str(down_count),
                        '-', '无条目', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-',
                        loopback31_address, '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-',
                        '-', '-', '-', '-', '-', '-', '-', 'normal'
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    continue
                
                # 记录起始行
                start_row = ws.max_row + 1
                
                # 添加BFD会话数据
                for session in bfd_data:
                    total_results += 1
                    ws.append([
                        ne_type, device_name, ip, str(total_vc), str(up_count), str(down_count),
                        session['aps_group'], session['session_name'], session['local_id'], session['remote_id'],
                        session['state'], session['master_backup'], session['send_interval'], session['receive_interval'],
                        session['detect_mult'], session['local_discr'], session['remote_discr'], session['discr_state'],
                        session['first_pkt'], session['cc_en'], session['mep_en'], loopback31_address, session['vcid'],
                        session['destination'], session['service_name'], session['vc_state'], session['interface'],
                        session['local_mtu'], session['remote_mtu'], session['vc_type'],
                        session['local_control_word'], session['remote_control_word'], session['current_control_word'],
                        session['local_pw_status_capability'], session['remote_pw_status_capability'],
                        session['current_pw_status_tlv'], session['local_pw_status'], session['remote_pw_status'],
                        session['local_vccv_capability'], session['remote_vccv_capability'], session['result']
                    ])
                    
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    
                    if session['result'] == "normal":
                        normal_results += 1
                    else:
                        ws.cell(row=ws.max_row, column=41).fill = orange_fill  # Result列位置为41
                
                # 合并单元格
                end_row = ws.max_row
                if start_row < end_row:
                    # 合并网元类型、网元名称、网元IP、VC总数、UP、DOWN
                    for col in range(1, 7):  # 前6列
                        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                    # 合并loopback31地址
                    ws.merge_cells(start_row=start_row, start_column=22, end_row=end_row, end_column=22)
            
            # 计算健康度
            health_percentage = (normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)
        
        elif item['name'] == "配置校验状态":
            headers = ["网元类型", "网元名称", "网元IP", "配置校验功能状态",
                       "每小时校验时间点(分钟)", "配置自动恢复等待时间(H:M)", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                if ip not in data or "show cfgchk info" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip,
                              "数据异常", "-", "-", "error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=7).fill = orange_fill
                    continue
                output = data[ip]["show cfgchk info"]
                cfgchk_data = item['parser'](output)
                total_results += 1
                ws.append([
                    ne_type, device_name, ip,
                    cfgchk_data['status'], cfgchk_data['minute_per_hour'], cfgchk_data['recovery_time'],
                    cfgchk_data['result']
                ])
                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if cfgchk_data['result'] == "normal":
                    normal_results += 1
                else:
                    ws.cell(row=ws.max_row, column=7).fill = orange_fill
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "OSPF进程状态检查":
            headers = ["网元类型", "网元名称", "网元IP", "进程ID", "路由ID", "运行时间", "绑定VRF", "RFC兼容性",
                       "支持功能", "SPF调度延迟", "外部LSA数", "总LSA数", "区域数", "区域类型/ID", "接口数/邻接数",
                       "最后SPF执行", "SPF执行次数", "Result", "备注"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            current_ip = None
            start_row = None
            for ip in successful_host_ips:
                if ip in connection_failures:
                    print(f"[DEBUG] 跳过 {ip} 因为连接失败")
                    continue

                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])

                if ip not in data or "show ospf process" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["数据异常"] * 16 + ["error", "-"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    ws.cell(row=ws.max_row, column=19).fill = orange_fill
                    continue

                output = data[ip]["show ospf process"]
                ospf_sessions = parse_ospf_process(output)
                print(f"[DEBUG] 为 IP {ip} 找到 {len(ospf_sessions)} 个 OSPF 进程")

                # 为新 IP 开始新组
                if current_ip != ip:
                    if start_row is not None and end_row > start_row:
                        for col in range(1, 4):  # 合并网元类型、名称、IP
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                    current_ip = ip
                    start_row = ws.max_row + 1

                for session in ospf_sessions:
                    total_results += 1
                    result, remarks = check_ospf_process(session, output)
                    ws.append([
                        ne_type, device_name, ip,
                        session.get('process_id', '-'),
                        session.get('router_id', '-'),
                        session.get('uptime', '-'),
                        session.get('vrf', '-'),
                        session.get('rfc', '-'),
                        session.get('supports', '-'),
                        session.get('spf_delay', '-'),
                        session.get('external_lsa', '-'),
                        session.get('total_lsa', '-'),
                        session.get('areas', '-'),
                        session.get('area_type_id', '-'),
                        f"{session.get('interfaces', '-')}接口/{session.get('adjacencies', '-')}邻接",
                        session.get('last_spf', '-'),
                        session.get('spf_executions', '-'),
                        result,
                        remarks
                    ])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if result == "normal":
                        normal_results += 1
                    elif result == "error":
                        ws.cell(row=ws.max_row, column=18).fill = orange_fill

                end_row = ws.max_row

            # 合并最后一组
            if start_row is not None and end_row > start_row:
                for col in range(1, 4):
                    ws.merge_cells(
                        start_row=start_row, start_column=col, end_row=end_row, end_column=col)

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "IPFRR-LSP状态检查":
            headers = ["网元类型", "网元名称", "网元IP", "目标LSR ID", "类型", "描述", "状态", "入标签",
                       "出标签", "出接口", "下一跳IP", "Result", "处理建议"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # 预解析所有 IP 的 MPLS LSP 数据
            parsed_data = {}
            for ip in successful_host_ips:
                if ip in connection_failures:
                    print(f"[DEBUG] 跳过 {ip} 因为连接失败")
                    continue
                if ip in data and "show mpls lsp brief" in data[ip]:
                    output = data[ip]["show mpls lsp brief"]
                    print(f"[DEBUG] 开始解析 {ip} 的MPLS LSP数据")
                    parsed_data[ip] = parse_mpls_lsp(output)
                else:
                    print(f"[DEBUG] {ip} 没有MPLS LSP数据")
                    parsed_data[ip] = None

            # 记录需要合并的单元格范围
            merge_ranges = []
            current_ip = None
            start_row = None
            total_results = 0
            normal_results = 0

            # 追加所有数据行，不设置样式
            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue

                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])

                if parsed_data[ip] is None:
                    total_results += 1
                    row = [ne_type, device_name, ip] + ["数据异常"] * \
                        10 + ["error", "无法获取MPLS LSP数据"]
                    ws.append(row)
                    continue

                lsp_data = parsed_data[ip]
                print(f"[DEBUG] 为 IP {ip} 找到 {len(lsp_data)} 个有效LSP")

                # 如果没有找到有效的LSP数据
                if not lsp_data:
                    total_results += 1
                    row = [ne_type, device_name, ip] + \
                        ["无LSP数据"] * 10 + ["warning", "未发现有效的LSP配置"]
                    ws.append(row)
                    continue

                # 为新 IP 开始新组
                if current_ip != ip:
                    if start_row is not None and end_row > start_row:
                        merge_ranges.append((start_row, end_row))
                    current_ip = ip
                    start_row = ws.max_row + 1

                for lsp in lsp_data:
                    total_results += 1
                    result, suggestions = check_mpls_lsp(lsp)
                    row = [
                        ne_type, device_name, ip,
                        lsp['dest_lsr_id'],
                        lsp['type'],
                        lsp['description'],
                        lsp['state'],
                        lsp['in_label'],
                        lsp['out_label'],
                        lsp['out_intf'],
                        lsp['nexthop_ip'],
                        result,
                        suggestions
                    ]
                    ws.append(row)
                    if result == "normal":
                        normal_results += 1

                end_row = ws.max_row

            # 合并最后一组
            if start_row is not None and end_row > start_row:
                merge_ranges.append((start_row, end_row))

            # 批量设置样式
            # 设置所有数据单元格的对齐和边框
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=13):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # 设置 Result 列的填充颜色
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=12)
                if cell.value == "error":
                    cell.fill = orange_fill
                elif cell.value == "warning":
                    cell.fill = PatternFill(
                        start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 黄色

            # 一次性合并单元格
            for start, end in merge_ranges:
                for col in range(1, 4):  # 合并网元类型、名称、IP
                    ws.merge_cells(start_row=start, start_column=col,
                                   end_row=end, end_column=col)

            # 计算健康百分比
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

            print(
                f"[DEBUG] IPFRR-LSP状态检查完成: {normal_results}/{total_results} 正常")

        elif item['name'] == "OSPF邻居状态检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "OSPF进程", "接收缓冲区(字节)", "发送缓冲区(字节)",
                "LSA缓冲区(字节)", "未使用包列表", "未使用LSA列表", "邻居ID", "优先级", "状态",
                "存活时间", "接口地址", "接口", "区域", "DR/BDR", "链路状态请求列表", "加密序列号",
                "实例ID", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                device_ip = ip
                if ip in data and "show device" in data[ip]:
                    try:
                        ne_type, device_name, _, parsed_device_ip = parse_uptime(
                            data[ip]["show device"])
                        # 仅当 parsed_device_ip 是有效 IP 地址时使用
                        if parsed_device_ip and re.match(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', parsed_device_ip):
                            device_ip = parsed_device_ip
                    except Exception as e:
                        print(f"[ERROR] 解析设备 {ip} 失败: {str(e)}")
                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue

                buffers_output = data[ip].get("show ospf buffers", "")
                neighbor_output = data[ip].get("show ospf neighbor", "")
                brief_output = data[ip].get("show ospf neighbor brief", "")
                print(
                    f"[DEBUG] Parsing OSPF for IP {ip}: buffers={len(buffers_output)} chars, neighbor={len(neighbor_output)} chars, brief={len(brief_output)} chars")
                ospf_data = parse_ospf_neighbor_status(
                    buffers_output, neighbor_output, brief_output)
                if not ospf_data or ospf_data[0].get("OSPF进程") == "无条目":
                    total_results += 1
                    normal_results += 1  # 修复：为 "无条目" 增加 normal_results
                    ws.append([ne_type, device_name, device_ip] +
                              ["无条目"] * 17 + ["normal", "-"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    print(f"[DEBUG] Wrote '无条目' row for IP {ip}")
                else:
                    start_row = ws.max_row + 1
                    for row_data in ospf_data:
                        total_results += 1
                        row_values = [
                            ne_type, device_name, device_ip,
                            row_data.get("OSPF进程", "-"),
                            row_data.get("接收缓冲区(字节)", "-"),
                            row_data.get("发送缓冲区(字节)", "-"),
                            row_data.get("LSA缓冲区(字节)", "-"),
                            row_data.get("未使用包列表", "-"),
                            row_data.get("未使用LSA列表", "-"),
                            row_data.get("邻居ID", "-"),
                            row_data.get("优先级", "-"),
                            row_data.get("状态", "-"),
                            row_data.get("存活时间", "-"),
                            row_data.get("接口地址", "-"),
                            row_data.get("接口", "-"),
                            row_data.get("区域", "-"),
                            row_data.get("DR/BDR", "-"),
                            row_data.get("链路状态请求列表", "-"),
                            row_data.get("加密序列号", "-"),
                            row_data.get("实例ID", "-"),
                            row_data.get("Result", "-"),
                            row_data.get("备注", "-")
                        ]
                        ws.append(row_values)
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border
                        if row_data.get("Result", "-") == "normal":
                            normal_results += 1
                        else:
                            ws.cell(row=ws.max_row,
                                    column=21).fill = orange_fill
                        print(
                            f"[DEBUG] Wrote OSPF row for IP {ip}: {row_values}")
                    end_row = ws.max_row
                    if start_row < end_row:
                        for col in range(1, 4):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)
            print(
                f"[DEBUG] OSPF health for {item['sheet_name']}: {normal_results}/{total_results} = {health_percentage}%")

        elif item['name'] == "LACP成员状态监控":
            headers = [
                "网元类型", "网元名称", "网元IP", "聚合组ID", "聚合组名称", "模式", "哈希模式",
                "AMC", "RVT", "WTR", "协议", "系统优先级", "本地系统ID", "成员端口（角色，优先级，状态）",
                "本地端口详情", "远程端口详情", "Result"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue
                lag_output = data[ip].get("show lag", "")
                lacp_output = data[ip].get("show lacp", "")
                lacp_data = item['parser'](lag_output, lacp_output)
                if not lacp_data or lacp_data[0].get("聚合组ID") == "无条目":
                    total_results += 1
                    normal_results += 1  # 修复：为 "无条目" 增加 normal_results
                    ws.append([ne_type, device_name, ip] +
                              ["无条目"] * 13 + ["normal"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                else:
                    start_row = ws.max_row + 1
                    for row_data in lacp_data:
                        total_results += 1
                        ws.append([
                            ne_type, device_name, ip,
                            row_data["聚合组ID"], row_data["聚合组名称"], row_data["模式"],
                            row_data["哈希模式"], row_data["AMC"], row_data["RVT"],
                            row_data["WTR"], row_data["协议"], row_data["系统优先级"],
                            row_data["本地系统ID"], row_data["成员端口（角色，优先级，状态）"],
                            row_data["本地端口详情"], row_data["远程端口详情"], row_data["Result"]
                        ])
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border
                        if row_data["Result"] == "normal":
                            normal_results += 1
                        else:
                            ws.cell(row=ws.max_row,
                                    column=17).fill = orange_fill
                    end_row = ws.max_row
                    if start_row < end_row:
                        for col in range(1, 4):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "OSPF 路由表检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "目的网络/掩码", "协议", "优先级", "开销",
                "下一跳", "接口", "存活时间", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            total_results = 0
            normal_results = 0

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue

                ne_type, device_name = "-", "-"
                device_ip = ip

                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip

                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到，跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败：{connection_failures[ip]}{Style.RESET_ALL}")
                    continue

                print(f"[DEBUG] 开始处理 {ip} 的OSPF路由表")

                output = data[ip].get("show ip routing-table", "")
                if not output:
                    print(f"[DEBUG] {ip} 没有路由表数据")
                    total_results += 1
                    normal_results += 1
                    # 修复：确保列数正确（12列总共：前3列+中间7列+后2列）
                    row = [ne_type, device_name, device_ip] + \
                        ["数据异常"] * 7 + ["normal", "-"]
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    continue

                ospf_data = parse_ospf_routing_table(output)

                # 处理解析结果
                if not ospf_data:
                    # 有路由表但没有异常条目
                    total_results += 1
                    normal_results += 1
                    row = [ne_type, device_name, device_ip] + \
                        ["所有路由正常"] * 7 + ["normal", "无异常路由"]
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border

                elif len(ospf_data) == 1 and ospf_data[0].get("目的网络/掩码") == "无条目":
                    # 没有找到路由表数据
                    total_results += 1
                    normal_results += 1
                    row = [ne_type, device_name, device_ip] + \
                        ["无条目"] * 7 + ["normal", "-"]
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border

                else:
                    # 有异常路由条目
                    start_row = ws.max_row + 1
                    for row_data in ospf_data:
                        total_results += 1
                        normal_results += 1  # 所有条目 Result 均为 normal
                        row = [
                            ne_type, device_name, device_ip,
                            row_data.get("目的网络/掩码", "-"),
                            row_data.get("协议", "-"),
                            row_data.get("优先级", "-"),
                            row_data.get("开销", "-"),
                            row_data.get("下一跳", "-"),
                            row_data.get("接口", "-"),
                            row_data.get("存活时间", "-"),
                            row_data.get("Result", "normal"),
                            row_data.get("备注", "-")
                        ]
                        ws.append(row)
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border

                    end_row = ws.max_row
                    if start_row <= end_row:
                        for col in range(1, 4):  # 合并网元类型、名称、IP
                            ws.merge_cells(
                                start_row=start_row, start_column=col,
                                end_row=end_row, end_column=col)

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)
        
        
        elif item['name'] == "LDP 会话状态检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "对端类型", "对端IP", "接口名称",
                "角色", "会话状态", "KeepAlive时间", "运行时间", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # 初始化计数器
            total_results = 0
            normal_results = 0
            error_results = 0

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                
                ne_type, device_name = "-", "-"
                device_ip = ip

                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip

                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue
                
                # 获取LDP会话输出
                session_output = data[ip].get("show ldp session", "")
                ldp_data = parse_ldp_session_status(session_output)

                # 如果没有LDP会话数据（空列表），显示一行表示无会话
                if not ldp_data:
                    total_results += 1
                    normal_results += 1
                    ws.append([ne_type, device_name, device_ip, "-", "无会话",
                              "-", "-", "-", "-", "-", "normal", "无LDP会话"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                else:
                    # 显示所有LDP会话的详细信息
                    start_row = ws.max_row + 1
                    has_error = False

                    for row_data in ldp_data:
                        result = row_data.get("Result", "normal")
                        total_results += 1

                        if result == "error":
                            has_error = True
                            error_results += 1
                        else:
                            normal_results += 1

                        row = [
                            ne_type,
                            device_name,
                            device_ip,
                            row_data.get("对端类型", "-"),
                            row_data.get("对端IP", "-"),
                            row_data.get("接口名称", "-"),
                            row_data.get("角色", "-"),
                            row_data.get("会话状态", "-"),
                            row_data.get("KeepAlive时间", "-"),
                            row_data.get("运行时间", "-"),
                            result,
                            row_data.get("备注", "-")
                        ]
                        ws.append(row)
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border

                        # 如果是错误状态，标记为橙色
                        if result == "error":
                            ws.cell(row=ws.max_row, column=11).fill = orange_fill

                    # 合并相同设备的前三列（网元类型、网元名称、网元IP）
                    end_row = ws.max_row
                    if start_row < end_row:
                        for col in range(1, 4):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, 
                                end_row=end_row, end_column=col)

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)
        
        elif item['name'] == "Loopback31地址唯一性检查":
            headers = ["网元类型", "网元名称", "网元IP", "Loopback31地址", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for ip in host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                device_ip = ip
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip

                loopback31_output = data[ip].get(
                    "show interface loopback 31", "")
                loopback31_addr = parse_loopback_address(loopback31_output)

                total_results += 1
                if loopback31_addr == "无条目":
                    normal_results += 1
                    ws.append(
                        [ne_type, device_name, device_ip, "无条目", "normal"])
                else:
                    result = "normal"
                    if loopback31_addr in loopback31_addresses and len(loopback31_addresses[loopback31_addr]) > 1:
                        result = "error"
                    if result == "normal":
                        normal_results += 1
                    ws.append([ne_type, device_name, device_ip,
                              loopback31_addr, result])

                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if ws[ws.max_row][4].value == "error":
                    ws.cell(row=ws.max_row, column=5).fill = orange_fill

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[sheet_name] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        
        elif item['name'] == "Loopback1023地址唯一性检查":
            headers = ["网元类型", "网元名称", "网元IP", "Loopback1023地址", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                device_ip = ip
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip
                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue
                loopback1023_output = data[ip].get(
                    "show interface loopback 1023", "")
                loopback1023_addr = parse_loopback_address(loopback1023_output)

                total_results += 1
                if loopback1023_addr == "无条目":
                    normal_results += 1
                    ws.append(
                        [ne_type, device_name, device_ip, "无条目", "normal"])
                else:
                    result = "normal"
                    if loopback1023_addr in loopback1023_addresses and len(loopback1023_addresses[loopback1023_addr]) > 1:
                        result = "error"
                    if result == "normal":
                        normal_results += 1
                    ws.append([ne_type, device_name, device_ip,
                              loopback1023_addr, result])

                for cell in ws[ws.max_row]:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                if ws[ws.max_row][4].value == "error":
                    ws.cell(row=ws.max_row, column=5).fill = orange_fill

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[sheet_name] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "SNMP配置检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "陷阱计数", "陷阱主机IP地址", "陷阱UDP端口", "陷阱社区",
                "陷阱版本", "陷阱VPN实例", "社区计数", "社区名称（缩写）", "社区权限", "社区访问级别",
                "社区绑定IP", "MIB视图名称", "MIB子树", "MIB视图类型", "MIB视图状态", "SNMP VPN",
                "联系人", "物理位置", "SNMP版本", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            total_results = 0
            normal_results = 0

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                device_ip = ip
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip
                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue
                trap_output = data[ip].get("show snmp-server trap", "")
                community_output = data[ip].get(
                    "show snmp-server community", "")
                mib_view_output = data[ip].get("show snmp-server mib-view", "")
                sys_info_output = data[ip].get("show snmp-server sys-info", "")
                snmp_data = item['parser'](
                    trap_output, community_output, mib_view_output, sys_info_output)

                total_results += 1
                if not snmp_data or not trap_output.strip():
                    normal_results += 1
                    ws.append([ne_type, device_name, device_ip] +
                              ["无条目"] * 20 + ["normal", "-"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                else:
                    start_row = ws.max_row + 1
                    for row_data in snmp_data:
                        normal_results += 1  # All results are "normal" per requirement
                        row = [
                            ne_type, device_name, device_ip,
                            row_data["陷阱计数"],
                            ", ".join(row_data["陷阱主机IP地址"]),
                            ", ".join(row_data["陷阱UDP端口"]),
                            ", ".join(row_data["陷阱社区"]),
                            ", ".join(row_data["陷阱版本"]),
                            ", ".join(row_data["陷阱VPN实例"]),
                            row_data["社区计数"],
                            ", ".join(row_data["社区名称（缩写）"]),
                            ", ".join(row_data["社区权限"]),
                            ", ".join(row_data["社区访问级别"]),
                            ", ".join(row_data["社区绑定IP"]),
                            ", ".join(row_data["MIB视图名称"]),
                            ", ".join(row_data["MIB子树"]),
                            ", ".join(row_data["MIB视图类型"]),
                            ", ".join(row_data["MIB视图状态"]),
                            row_data["SNMP VPN"],
                            row_data["联系人"],
                            row_data["物理位置"],
                            row_data["SNMP版本"],
                            row_data["Result"],
                            row_data["备注"]
                        ]
                        ws.append(row)
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border
                    end_row = ws.max_row
                    if start_row < end_row:
                        for col in range(1, 4):  # Merge NE type, name, IP
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "设备账户检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "当前账户数量", "用户名", "权限级别", "锁定状态",
                "锁定分钟数", "最大尝试次数", "密码提示天数", "密码最小长度", "密码需包含数字",
                "密码需包含大写字母", "密码需包含小写字母", "密码需包含特殊字符", "密码重用检查次数",
                "当前登录用户", "登录 Tty", "登录 Tid", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            total_results = 0
            normal_results = 0

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue
                ne_type, device_name = "-", "-"
                device_ip = ip
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip
                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue
                users_output = data[ip].get("show users", "")
                login_rule_output = data[ip].get("show login-global-rule", "")
                logging_user_output = data[ip].get("show loginning-user", "")
                account_data = item['parser'](
                    users_output, login_rule_output, logging_user_output)

                total_results += 1
                if not account_data or not users_output.strip():
                    normal_results += 1
                    ws.append([ne_type, device_name, device_ip] +
                              ["无条目"] * 17 + ["normal", "-"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                else:
                    start_row = ws.max_row + 1
                    for row_data in account_data:
                        normal_results += 1  # All results are "normal" per requirement
                        row = [
                            ne_type, device_name, device_ip,
                            row_data["当前账户数量"],
                            ", ".join(row_data["用户名"]),
                            ", ".join(row_data["权限级别"]),
                            ", ".join(row_data["锁定状态"]),
                            row_data["锁定分钟数"],
                            row_data["最大尝试次数"],
                            row_data["密码提示天数"],
                            row_data["密码最小长度"],
                            row_data["密码需包含数字"],
                            row_data["密码需包含大写字母"],
                            row_data["密码需包含小写字母"],
                            row_data["密码需包含特殊字符"],
                            row_data["密码重用检查次数"],
                            ", ".join(row_data["当前登录用户"]),
                            ", ".join(row_data["登录 Tty"]),
                            ", ".join(row_data["登录 Tid"]),
                            row_data["Result"],
                            row_data["备注"]
                        ]
                        ws.append(row)
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border
                    end_row = ws.max_row
                    if start_row < end_row:
                        for col in range(1, 4):  # Merge NE type, name, IP
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "专网业务分析":
            headers = ["网元类型", "网元名称", "网元IP", "类型", "VSI_ID", "VSI名称", "MTU",
                       "目的节点", "状态", "VC_ID", "入标签", "出标签", "隧道ID", "接口", "PE VLAN[服务提供商]", "CE VLAN[用户侧]",
                       "剥离外层 VLAN", "HSID", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            total_results = 0
            normal_results = 0
            for ip in sorted(host_ips):
                # 跳过登录失败的设备
                if ip in connection_failures:
                    continue

                # 获取设备基本信息，默认为 "-"
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])

                # 检查必要的数据是否存在
                if ip not in data or "show vsi brief" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["数据异常"] * 15 + ["error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    # Result列标橙
                    ws.cell(row=ws.max_row, column=19).fill = orange_fill
                    continue

                # 获取并解析专网业务输出
                vsi_output = data[ip]["show vsi brief"]
                services = parse_private_network_service(
                    "", vsi_output, ne_type, device_name, ip)

                # 处理无条目情况
                if not services or services[0]["类型"] == "-":
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["-"] * 15 + ["normal"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    normal_results += 1
                    continue

                # 处理正常条目
                start_row = ws.max_row + 1
                for service in services:
                    total_results += 1
                    if service["Result"] == "normal":
                        normal_results += 1

                    row_data = [
                        service["网元类型"], service["网元名称"], service["网元IP"],
                        service["类型"], service["VSI_ID"], service["VSI名称"], service["MTU"],
                        service["目的节点"], service["状态"], service["VC_ID"],
                        service["入标签"], service["出标签"], service["隧道ID"], service["接口"],
                        service["PE VLAN[服务提供商]"], service["CE VLAN[用户侧]"],
                        service["剥离外层 VLAN"], service["HSID"], service["Result"]
                    ]
                    ws.append(row_data)

                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    if service["Result"] != "normal":
                        # Result列标橙
                        ws.cell(row=ws.max_row, column=19).fill = orange_fill

                end_row = ws.max_row
                if start_row < end_row:
                    for col in range(1, 4):  # 合并网元类型、名称、IP列
                        ws.merge_cells(
                            start_row=start_row, start_column=col, end_row=end_row, end_column=col)

            # 计算健康度
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "PTP时钟检查":
            headers = ["网元类型", "网元名称", "网元IP", "时钟标识", "PTP状态", "时钟模式", "域值",
                       "从模式", "步进模式", "BMC优先级1", "BMC优先级2", "BMC时钟等级", "BMC时钟精度",
                       "当前时钟源(PTP)", "PTP实际状态", "GM时钟标识", "父时间标识", "父时钟跳数",
                       "GM偏移统计(us)", "GM时钟源类型", "时钟端口", "端口WTR(s)", "时钟关系",
                       "通告超时", "SyncE当前时钟源", "SSM控制", "SSM输入门限",
                       "不对称方向", "延迟机制", "通告间隔(s)", "同步间隔(s)", "延迟请求间隔(s)",
                       "外部SyncE类型", "输出控制", "SaBit", "输出门限", "Result", "备注"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            total_results = 0
            normal_results = 0
            # 写入Excel
            for ip in sorted(host_ips):
                if ip in connection_failures:
                    continue

                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])
                # 检查 IP 是否在 data 中
                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到， 跳过{Style.RESET_ALL}")
                    continue

                # 检查是否是连接失败的设备
                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue
                ptp_output = data[ip]["show ptp all"] if ip in data and "show ptp all" in data[ip] else ""
                synce_output = data[ip]["show synce"] if ip in data and "show synce" in data[ip] else ""

                # 解析PTP时钟状态数据
                ptp_data = parse_ptp_clock_status(
                    ptp_output, synce_output, ne_type, device_name, ip)

                # 获取设备信息和接收端口列表
                device_info = ptp_data["设备信息"]
                recv_ports = ptp_data["接收端口"]

                # 如果有接收端口
                if recv_ports:
                    start_row = ws.max_row + 1
                    for i, port in enumerate(recv_ports):
                        row_data = [
                            device_info["网元类型"] if i == 0 else "",  # 仅第一行写入
                            device_info["网元名称"] if i == 0 else "",
                            device_info["网元IP"] if i == 0 else "",
                            device_info["时钟标识"] if i == 0 else "",
                            device_info["PTP状态"] if i == 0 else "",
                            device_info["时钟模式"] if i == 0 else "",
                            device_info["域值"] if i == 0 else "",
                            device_info["从模式"] if i == 0 else "",
                            device_info["步进模式"] if i == 0 else "",
                            device_info["BMC优先级1"] if i == 0 else "",
                            device_info["BMC优先级2"] if i == 0 else "",
                            device_info["BMC时钟等级"] if i == 0 else "",
                            device_info["BMC时钟精度"] if i == 0 else "",
                            device_info["当前时钟源(PTP)"] if i == 0 else "",
                            port["PTP实际状态"],
                            device_info["GM时钟标识"] if i == 0 else "",
                            device_info["父时间标识"] if i == 0 else "",
                            device_info["父时钟跳数"] if i == 0 else "",
                            device_info["GM偏移统计(us)"] if i == 0 else "",
                            device_info["GM时钟源类型"] if i == 0 else "",
                            port["时钟端口"],
                            port["端口WTR(s)"],
                            port["时钟关系"],
                            port["通告超时"],
                            device_info["SyncE当前时钟源"] if i == 0 else "",
                            device_info["SSM控制"] if i == 0 else "",
                            device_info["SSM输入门限"] if i == 0 else "",
                            port["不对称方向"],
                            port["延迟机制"],
                            port["通告间隔(s)"],
                            port["同步间隔(s)"],
                            port["延迟请求间隔(s)"],
                            device_info["外部SyncE类型"] if i == 0 else "",
                            device_info["输出控制"] if i == 0 else "",
                            device_info["SaBit"] if i == 0 else "",
                            device_info["输出门限"] if i == 0 else "",
                            port["Result"],
                            port["备注"]
                        ]
                        ws.append(row_data)
                        total_results += 1  # 每个端口算一个检测项
                        if port["Result"] == "normal":
                            normal_results += 1

                        # 设置单元格样式
                        for cell in ws[ws.max_row]:
                            cell.alignment = center_alignment
                            cell.border = thin_border

                        if port["Result"] != "normal":
                            ws.cell(row=ws.max_row,
                                    column=37).fill = orange_fill

                    # 合并单元格（仅当有多个接收端口时）
                    end_row = ws.max_row
                    if len(recv_ports) > 1:
                        # 合并网元类型、网元名称、网元IP（第1-3列）
                        ws.merge_cells(
                            start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                        ws.merge_cells(
                            start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                        ws.merge_cells(
                            start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                        # 合并设备级PTP信息（第4-14列）
                        for col in range(4, 15):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                        # 合并GM时钟标识、父时间标识、父时钟跳数（第16-18列）
                        ws.merge_cells(
                            start_row=start_row, start_column=16, end_row=end_row, end_column=16)
                        ws.merge_cells(
                            start_row=start_row, start_column=17, end_row=end_row, end_column=17)
                        ws.merge_cells(
                            start_row=start_row, start_column=18, end_row=end_row, end_column=18)
                        # 合并GM偏移统计和时钟源类型（第19-20列）
                        for col in range(19, 21):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                        # 合并SyncE信息（第25-27列和第33-36列）
                        for col in range(25, 28):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                        for col in range(33, 37):
                            ws.merge_cells(
                                start_row=start_row, start_column=col, end_row=end_row, end_column=col)

                else:
                    # 无接收端口时写入一行设备级信息
                    row_data = [
                        device_info["网元类型"], device_info["网元名称"], device_info["网元IP"],
                        device_info["时钟标识"], device_info["PTP状态"], device_info["时钟模式"],
                        device_info["域值"], device_info["从模式"], device_info["步进模式"],
                        device_info["BMC优先级1"], device_info["BMC优先级2"], device_info["BMC时钟等级"],
                        device_info["BMC时钟精度"], device_info["当前时钟源(PTP)"], "-",
                        device_info["GM时钟标识"], device_info["父时间标识"], device_info["父时钟跳数"],
                        device_info["GM偏移统计(us)"], device_info["GM时钟源类型"], "-",
                        "-", "-", "-", device_info["SyncE当前时钟源"], device_info["SSM控制"],
                        device_info["SSM输入门限"], "-", "-", "-", "-", "-",
                        device_info["外部SyncE类型"], device_info["输出控制"], device_info["SaBit"],
                        device_info["输出门限"], "normal", "无接收端口数据"
                    ]
                    ws.append(row_data)
                    total_results += 1  # 设备级信息算一个检测项
                    normal_results += 1  # 标记为normal

                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border

            # Create login failure sub-sheet

            # 计算健康度
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "站点邻接网元检查":
            headers = ["网元类型", "网元名称", "网元IP", "邻接网元数", "邻接下一跳地址", "Result"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for ip in sorted(host_ips):
                if ip in connection_failures:
                    continue

                # 修复1：获取网元类型和名称的正确方式
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    try:
                        # 修复：正确的解包方式，根据parse_uptime函数的返回值调整
                        device_info = parse_uptime(data[ip]["show device"])
                        if isinstance(device_info, tuple) and len(device_info) >= 2:
                            ne_type = device_info[0] if device_info[0] else "-"
                            device_name = device_info[1] if device_info[1] else "-"
                    except Exception as e:
                        print(
                            f"{Fore.YELLOW}[WARNING] Failed to parse device info for {ip}: {e}{Style.RESET_ALL}")

                # 修复2：获取LDP LSP输出
                ldp_output = data.get(ip, {}).get("show ldp lsp", "")
                if not ldp_output:
                    print(
                        f"{Fore.YELLOW}[WARNING] No LDP LSP data for {ip}{Style.RESET_ALL}")

                lsp_data = parse_ldp_lsp(ldp_output, ne_type, device_name, ip)

                # 处理解析结果
                for entry in lsp_data:
                    total_results += 1
                    if entry["Result"] == "normal":
                        normal_results += 1

                    row_data = [
                        entry["网元类型"],
                        entry["网元名称"],
                        entry["网元IP"],
                        entry["邻接网元数"],
                        entry["邻接下一跳地址"],
                        entry["Result"]
                    ]
                    ws.append(row_data)

                    # 应用样式
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border

                    if entry["Result"] != "normal":
                        ws.cell(row=ws.max_row, column=6).fill = orange_fill

            # 计算健康度
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "设备面板视图":
            # 收集所有设备数据
            print(f"[DEBUG] 开始生成设备面板视图，设备总数: {len(host_ips)}")

            # 收集所有设备数据
            devices_data = []
            processed_count = 0

            for ip in sorted(host_ips):
                if ip in connection_failures:
                    print(f"[DEBUG] 跳过连接失败的设备: {ip}")
                    continue

                # 获取网元类型和名称
                ne_type, device_name = "-", "-"
                if ip in data and "show device" in data[ip]:
                    try:
                        device_info = parse_uptime(data[ip]["show device"])
                        if isinstance(device_info, tuple) and len(device_info) >= 2:
                            ne_type = device_info[0] if device_info[0] else "-"
                            device_name = device_info[1] if device_info[1] else "-"
                    except Exception as e:
                        print(
                            f"{Fore.YELLOW}[WARNING] Failed to parse device info for {ip}: {e}{Style.RESET_ALL}")

                # 解析设备面板数据
                device_output = data.get(ip, {}).get("show device", "")
                try:
                    panel_data = parse_device_panel(
                        device_output, ne_type, device_name, ip)
                    if panel_data is not None:
                        devices_data.append(panel_data)
                        processed_count += 1
                        print(f"[DEBUG] 成功处理设备 {processed_count}: {ip}")
                except Exception as e:
                    print(
                        f"{Fore.YELLOW}[WARNING] Failed to parse device panel for {ip}: {e}{Style.RESET_ALL}")
                    # 添加基本设备信息，即使解析失败
                    basic_panel_data = {
                        'device_type': ne_type,
                        'device_name': device_name,
                        'device_ip': ip,
                        'device_model': "-",
                        'slots': {}
                    }
                    devices_data.append(basic_panel_data)
                    processed_count += 1

            print(f"[DEBUG] 收集设备数据完成，有效设备数: {len(devices_data)}")

            # 确保至少有一个设备数据，即使是空的
            if not devices_data:
                print("[WARNING] 没有收集到任何设备数据，创建默认数据")
                devices_data = [{
                    'device_type': "-",
                    'device_name': "无设备数据",
                    'device_ip': "-",
                    'device_model': "-",
                    'slots': {}
                }]

            # 创建设备面板布局
            try:
                create_device_panel_layout(ws, devices_data)
                print("[INFO] 设备面板布局创建成功")
            except Exception as e:
                print(
                    f"{Fore.RED}[ERROR] Failed to create device panel layout: {e}{Style.RESET_ALL}")
                import traceback
                traceback.print_exc()
                # 创建一个简单的错误信息表格
                ws.append(['设备面板视图生成失败', str(e)])

            # return len(devices_data), len(devices_data)  # 返回设备总数和成功数

        elif item['name'] == "BFD保护组状态信息":
            headers = [
                "网元类型", "网元名称", "网元IP", "APS-ID", "状态", "主VCID", "备VCID", "类型", "方向", "恢复",
                "SD", "WTR", "保持关闭", "保护使能", "外部命令", "发送APS", "接收APS", "loopback31地址",
                "主用目的地址", "主用业务名称", "主用VC状态", "主用接口",
                "备用目的地址", "备用业务名称", "备用VC状态", "备用接口", "Result"
            ]
            ws.append(headers)

            # 设置表头格式
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue  # 跳过登录失败的设备

                # 获取设备基本信息
                ne_type, device_name = "-", "-"
                loopback31_address = "-"

                if ip in data and "show device" in data[ip]:
                    ne_type, device_name, _, _ = parse_uptime(
                        data[ip]["show device"])

                if ip in data and "show interface loopback 31" in data[ip]:
                    loopback31_output = data[ip]["show interface loopback 31"]
                    loopback31_address = parse_loopback31(loopback31_output)

                # 检查必要的命令输出是否存在
                if ip not in data or "show protect-group all" not in data[ip] or "show mpls l2vc brief" not in data[ip]:
                    total_results += 1
                    ws.append([ne_type, device_name, ip] +
                              ["数据异常"] * 24 + ["error"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    # Result列
                    ws.cell(row=ws.max_row, column=27).fill = orange_fill
                    continue

                # 获取命令输出
                protect_group_output = data[ip]["show protect-group all"]
                l2vc_output = data[ip]["show mpls l2vc brief"]

                # 解析保护组状态信息
                protect_groups = parse_protect_group_all(
                    protect_group_output, l2vc_output)

                start_row = ws.max_row + 1
                for group in protect_groups:
                    total_results += 1
                    ws.append([
                        ne_type, device_name, ip,
                        group['aps_id'], group['status'], group['master_vcid'], group['backup_vcid'],
                        group['type'], group['direction'], group['recovery'], group['sd'], group['wtr'],
                        group['hold_off'], group['protect_enable'], group['external_cmd'],
                        group['send_aps'], group['recv_aps'], loopback31_address,
                        group['master_destination'], group['master_service_name'],
                        group['master_vc_status'], group['master_interface'],
                        group['backup_destination'], group['backup_service_name'],
                        group['backup_vc_status'], group['backup_interface'],
                        group['result']
                    ])

                    # 设置单元格格式
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border

                    # 统计结果
                    if group['result'] == "normal":
                        normal_results += 1
                    else:
                        # Result列标红
                        ws.cell(row=ws.max_row, column=27).fill = orange_fill

                # 合并相同设备的基本信息列
                end_row = ws.max_row
                if start_row <= end_row:
                    for col in range(1, 4):  # 合并网元类型、网元名称、网元IP
                        if start_row < end_row:
                            ws.merge_cells(start_row=start_row, start_column=col,
                                           end_row=end_row, end_column=col)
                    # 合并loopback31地址列
                    if start_row < end_row:
                        ws.merge_cells(start_row=start_row, start_column=18,
                                       end_row=end_row, end_column=18)

            # 计算健康度
            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

        elif item['name'] == "DCN路由表检查":
            headers = [
                "网元类型", "网元名称", "网元IP", "目的网络/掩码", "协议", "优先级", "开销",
                "下一跳", "接口", "存活时间", "Result", "备注"
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for ip in successful_host_ips:
                if ip in connection_failures:
                    continue

                ne_type, device_name = "-", "-"
                device_ip = ip
                if ip in data and "show device" in data[ip]:
                    # 添加调试信息
                    # print(
                    #     f"IP: {ip}, show device output: {data[ip]['show device']}")
                    ne_type, device_name, _, parsed_ip = parse_uptime(
                        data[ip]["show device"])
                    print(
                        f"Parsed ne_type: {ne_type}, device_name: {device_name}, parsed_ip: {parsed_ip}")
                    if parsed_ip and re.match(r'\d+\.\d+\.\d+\.\d+', parsed_ip):
                        device_ip = parsed_ip
                    # 如果解析失败，记录警告
                    if ne_type == "-":
                        print(
                            f"{Fore.YELLOW}[WARNING] IP {ip} 网元类型解析失败，保持默认值 '-' {Style.RESET_ALL}")

                if ip not in data:
                    print(
                        f"{Fore.YELLOW}[WARNING] IP {ip} 未在数据中找到，跳过{Style.RESET_ALL}")
                    continue

                if ip in connection_failures:
                    print(
                        f"{Fore.YELLOW}[INFO] IP {ip} 连接失败： {connection_failures[ip]}{Style.RESET_ALL}")
                    continue

                output = data[ip].get(
                    "show ip routing-table vpn __dcn_vpn__", "")
                dcn_data = item['parser'](output)

                if not dcn_data or dcn_data[0].get("目的网络/掩码") == "无条目":
                    total_results += 1
                    normal_results += 1
                    ws.append([ne_type, device_name, device_ip, "无条目",
                               "-", "-", "-", "-", "-", "-", "normal", "-"])
                    for cell in ws[ws.max_row]:
                        cell.alignment = center_alignment
                        cell.border = thin_border
                else:
                    # 批量写入 - 针对高性能优化
                    batch_rows = []
                    has_errors = False

                    for row_data in dcn_data:
                        total_results += 1
                        if row_data.get("Result", "normal") == "normal":
                            normal_results += 1
                        else:
                            has_errors = True

                        batch_rows.append([
                            ne_type, device_name, device_ip,
                            row_data.get("目的网络/掩码", "-"),
                            row_data.get("协议", "-"),
                            row_data.get("优先级", "-"),
                            row_data.get("开销", "-"),
                            row_data.get("下一跳", "-"),
                            row_data.get("接口", "-"),
                            row_data.get("存活时间", "-"),
                            row_data.get("Result", "normal"),
                            row_data.get("备注", "-")
                        ])

                    # 超快速批量写入
                    start_row = ws.max_row + 1
                    for row in batch_rows:
                        ws.append(row)
                    end_row = ws.max_row

                    # 只对error行设置样式，大幅提升性能
                    if has_errors:
                        red_fill = PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        for row_num in range(start_row, end_row + 1):
                            row_cells = ws[row_num]
                            if row_cells[10].value == "error":  # Result列
                                for cell in row_cells:
                                    cell.fill = red_fill

                    # 基础样式批量设置
                    for row_num in range(start_row, end_row + 1):
                        for cell in ws[row_num]:
                            cell.alignment = center_alignment
                            cell.border = thin_border

                    # 合并单元格
                    if start_row < end_row:
                        for col in range(1, 4):
                            ws.merge_cells(start_row=start_row, start_column=col,
                                           end_row=end_row, end_column=col)

            health_percentage = (
                normal_results / total_results * 100) if total_results > 0 else 0
            health_scores[item['sheet_name']] = f"{health_percentage:.0f}%"
            item_counts[item['sheet_name']] = (normal_results, total_results)

#
    ws_failure = wb.create_sheet(title="登录失败设备")
    headers = ["网元IP", "故障原因"]
    ws_failure.append(headers)
    for cell in ws_failure[1]:
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.font = header_font

    total_failures = len(connection_failures)
    total_devices = len(host_ips)
    success_devices = total_devices - total_failures
    for ip in sorted(connection_failures.keys()):
        reason = connection_failures[ip]
        ws_failure.append([ip, reason])
        for cell in ws_failure[ws_failure.max_row]:
            cell.alignment = center_alignment
            cell.border = thin_border
        ws_failure.cell(row=ws_failure.max_row,
                        column=2).fill = orange_fill

    health_percentage = (success_devices / total_devices *
                         100) if total_devices > 0 else 0
    health_scores["登录失败设备"] = f"{health_percentage:.0f}%"
    item_counts["登录失败设备"] = (success_devices, total_devices)
    # 修改指南内容生成逻辑 - 替换原有的指南生成部分

    # Create guide sheet
    ws_guide = wb.create_sheet(title="指南", index=1)
    guide_headers = ["编号", "检查项", "解决方案", "规则", "命令"]
    ws_guide.append(guide_headers)
    for cell in ws_guide[1]:
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.font = header_font

    # 定义指南内容映射字典（键为检查项名称，值为对应的指南内容）
    guide_content_mapping = {
        "设备运行时间检查": [
            "",  # 编号位置，后面会动态填充
            "设备运行时间检查",
            "从历史告警看，网元可能掉电，需持续观察运行状态，或联系设备管理员检查电源稳定性。",
            "运行时间数据缺失或无法解析，输出 'error'；否则输出 'normal'、大于1天为'normal'。",
            "show device"
        ],
        "主控盘运行状态": [
            "",
            "主控盘运行状态",
            "当 15 分钟 CPU 使用率超过 60% 时，需清理系统垃圾文件或优化进程；若内存使用率异常，检查内存泄漏或重启设备。",
            "15 分钟 CPU 使用率 >= 60%，输出 'error'；CPU 或内存使用率数据无法解析，输出 'error'；否则输出 'normal'。",
            "show device"
        ],
        "协议报文处理状态": [
            "",
            "协议报文处理状态",
            "若存在丢弃数据包，检查网络配置或协议设置，必要时联系设备厂商分析丢包原因。",
            "任何协议有丢弃数据包（dropped > 0），输出 'error'；无丢弃数据包，输出 'normal'。",
            "show cpu-defend stats"
        ],
        "真实版本信息": [
            "",
            "真实版本信息",
            "若版本信息缺失或解析失败，检查设备命令支持情况，或升级系统固件。",
            "命令 'show real-version' 返回错误或解析失败，输出 'error'；版本信息完整，输出 'normal'。",
            "show real-version"
        ],
        "风扇转速及温度状态": [
            "",
            "风扇转速及温度状态",
            "若温度异常，检查设备通风环境或清理灰尘；若风扇状态异常，检查风扇硬件或更换风扇。",
            "板卡温度 > 89°C 或 < 35°C，输出 'error'；风扇速度非百分比数值或任一风扇速度 < 20%，输出 'error'；否则输出  'normal'。  ",
            "show fan; show temperature"
        ],
        "系统与硬件版本状态": [
            "",
            "系统与硬件版本状态检查",
            "若版本信息缺失，检查设备固件版本或联系厂商获取最新版本信息。",
            "命令 'show version' 返回数据无法解析或缺失，输出 'error'；版本信息完整，输出 'normal'。",
            "show version"
        ],
        "光模块信息检查": [
            "",
            "光模块信息检查",
            "若光功率、偏置电流、电压或温度异常，检查光模块连接或更换光模块；若 CRC 错误过多，检查光纤链路或端口状态。",
            "光功率、偏置电流、电压或温度超出设备指定范围，输出 'error'（收无光除外）；CRC 错误 > 2048，输出 'error'；光模块数  据缺    失，输出 'error'；否则输出 'normal'。",
            "show interface; show lldp neighbor"
        ],
        "电源状态检查": [
            "",
            "电源状态检查",
            "若电压异常，检查电源模块或供电线路；若持续异常，联系厂商更换电源模块。",
            "槽位 12 和 13 电压均为 0.0V 且比率 0.00，输出 'normal'；其他槽位电压 < 42V 或 > 58V 或为 0V，输出 'error'；电压    数  据缺失，输出 'error'；否则输出 'normal'。",
            "show voltage"
        ],
        "FW软件版本一致性检查": [
            "",
            "FW软件版本一致性检查",
            "若主备版本不一致，需升级或回滚软件版本以保持一致；若数据缺失，检查设备配置或命令输出。",
            "主控与备控 system info 字符一致且软件版本一致，输出 'normal'；否则输出 'error'。",
            "show device"
        ],
        "板卡CPU内存使用率": [
            "",
            "板卡CPU内存使用率",
            "若温度、CPU或内存使用率超标，检查设备运行负载，优化进程或清理资源；若持续异常，考虑硬件升级。",
            "15分钟或24小时内任一板卡温度 > 80°C、CPU使用率 > 60%、内存使用率 > 65%，输出 'error'；数据缺失或解析失败，输出         'error'；否则输出 'normal'。",
            "show pm cur-15m Dev; show pm cur-24h Dev"
        ],
        "NTP时间同步分析": [
            "",
            "NTP时间同步分析",
            "若本地时间与当前系统时间偏差超过60秒，需检查NTP配置或网络连接。",
            "本地时间与当前系统时间偏差在60秒内，输出 'normal'；否则输出 'error'。",
            "show cloc; show ntp-service"
        ],
        "硬盘资源占用分析": [
            "",
            "硬盘资源占用分析",
            "若硬盘使用率超过70%，需清理不必要的文件或扩展存储空间。",
            "硬盘使用率 <= 70%，输出 'normal'；否则输出 'error'。",
            "show flash-usage"
        ],
        "BFD会话检查(VC业务统计)": [
            "",
            "BFD会话检查(VC业务统计)",
            "若APS组ID不为0且状态为Down，需检查BFD会话配置或网络连通性。",
            "APS组ID !=0 且状态为Down时，输出 'error'；否则输出 'normal'。",
            "show bfd session brief; show bfd configuration pw"
        ],
        "配置校验状态": [
            "",
            "配置校验状态",
            "若配置校验功能状态为disable，需启用配置校验功能。",
            "配置校验功能状态为disable时，输出 'error'；否则输出 'normal'。",
            "show cfgchk info"
        ],
        "OSPF进程状态检查": [
            "",
            "OSPF进程状态检查",
            "若SPF执行频率超标，检查网络稳定性；若LSA数异常，检查路由分发；若邻接不足，验证OSPF配置。",
            "SPF执行频率<200000次/天，LSA总数<19000，外部LSA<1800，邻接数≥接口数-2，运行时间正常，否则为'error'。",
            "show ospf process"
        ],
        "IPFRR-LSP状态检查": [
            "",
            "IPFRR-LSP状态检查",
            "若LSP状态为down，检查LDP会话；若Ingress入标签异常，验证MPLS配置。",
            "LSP状态为up，Ingress入标签为'-'，否则为'error'。",
            "show mpls lsp brief"
        ],
        "OSPF邻居状态检查": [
            "",
            "OSPF邻居状态检查",
            "若缓冲区过小，调整OSPF进程资源配置；若邻居状态异常，检查网络连通性或配置；若存活时间异常，验证计时器设置。",
            "接收/发送/LSA缓冲区 < 2048字节，未使用包列表 < 2/200，未使用LSA列表 < 2/200，状态 ≠ Full，存活时间 < 40秒，链路    状  态请求列表 > 0，DR/BDR ≠ 0/0（点对点接口），则为'error'；否则为'normal'。",
            "show ospf buffers; show ospf neighbor"
        ],
        "LACP成员状态监控": [
            "",
            "LACP成员状态监控",
            "若端口未选中或协商失败，检查LACP配置和物理链路；若远程信息异常，验证对端设备配置。",
            "端口状态 ≠ Selected，远程SysId = 00-00-00-00-00-00，PortState ≠ 11111100，远程SysPri = 65535，远程PortKey =    0， 则为'error'；否则为 'normal'。",
            "show lag; show lacp"
        ],
        "OSPF 路由表检查": [
            "",
            "OSPF 路由表检查",
            "若 Cost 值过高，检查 OSPF 链路成本配置或网络拓扑设计；若 Uptime 过短，检查链路稳定性或路由震荡问题。",
            "Cost > 18000 或 Uptime < 1小时的 OSPF_IA 路由记录为异常，仅输出异常条目，结果为 'normal'。",
            "show ip routing-table"
        ],
        "LDP 会话状态检查": [
            "",
            "LDP 会话异常状态检查",
            "若会话状态异常，检查链路或 LDP 配置；若 LSP 未建立，检查路由或标签策略。",
            "State ≠ OPERATIONAL为 'error'，否则为 'normal'。",
            "show ldp session"
        ],
        "Loopback31地址唯一性检查": [
            "",
            "Loopback31地址唯一性检查",
            "若地址重复，检查设备配置，调整Loopback31地址以确保唯一性。",
            "地址重复则Result为'error'，否则为'normal'；无条目显示'无条目'。",
            "show interface loopback 31"
        ],
        "Loopback1023地址唯一性检查": [
            "",
            "Loopback1023地址唯一性检查",
            "若地址重复，检查设备配置，调整Loopback1023地址以确保唯一性。",
            "地址重复则Result为'error'，否则为'normal'；无条目显示'无条目'。",
            "show interface loopback 1023"
        ],
        "SNMP配置检查": [
            "",
            "SNMP配置检查",
            "若SNMP版本不是v3，升级至v3并启用加密认证；若社区绑定IP为0.0.0.0，限制为特定管理网段。",
            "SNMP版本 ≠ v3 或 Community Bind IP = 0.0.0.0 时备注异常，否则为 'normal'。",
            "show snmp-server trap; show snmp-server community; show snmp-server mib-view; show snmp-server sys-info"
        ],
        "设备账户检查": [
            "",
            "设备账户检查",
            "若锁定时间 < 30分钟或最大尝试次数 > 5，调整配置；若密码重用检查次数 < 3，启用密码历史策略。",
            "锁定时间 < 30分钟、最大尝试次数 > 5、密码重用检查次数 < 3 时备注异常，否则为 'normal'。",
            "show users; show login-global-rule; show loginning-user"
        ],
        "专网业务分析": [
            "",
            "专网业务分析",
            "若专网业务状态为Down，检查VPLS配置、MPLS LDP会话或物理链路；若AC接口状态异常，验证接口VLAN配置。",
            "VPLS或VC状态为Down时，输出 'error'；AC状态正常，输出 'normal'；数据异常输出 'error'。",
            "show vsi brief"
        ],
        "PTP时钟检查": [
            "",
            "PTP时钟检查",
            "当发现PTP状态异常时，检查设备时钟信号源、线路连接和BMC优先级配置；若GM偏移统计过大，检查网络延迟和时钟配置同步；确 保  SyncE与PTP路径一致。",
            "PTP状态须为enable，实际状态为SLAVE；GM时钟标识不为全F；父时钟跳数≤10；GM偏移统计≤50μs；BMC优先级1/2不全为255； BMC时钟 等级≤128；SSM控制为on。",
            "show ptp all, show synce"
        ],
        "站点邻接网元检查": [
            "",
            "站点邻接网元检查(LDP链路邻接状态检查)",
            "当发现LDP邻接异常时，检查LDP协议配置和邻居建立状态；若下一跳地址为none或0.0.0.0，检查路由可达性和LSP建立过程；确保 FEC前   缀正确且标签分发正常。",
            "LDP邻接须正常建立；FEC下一跳地址不为none或0.0.0.0；邻接网元数≥1；标签分发状态正常；LSP路径可达。",
            "show ldp lsp, show ldp neighbor"
        ],
        "设备面板视图": [
            "",
            "设备面板视图(设备槽位和板卡信息展示)",
            "通过show device命令获取设备槽位和板卡信息，以面板图形式展示设备的物理布局。",
            "显示所有槽位的板卡类型和状态；主备板卡需标识Master/Backup状态；面板布局符合设备物理结构。",
            "show device"
        ],
        "BFD保护组状态信息": [
            "",
            "BFD保护组状态信息",
            "检查BFD保护组状态，匹配对应VCID的L2VC业务信息。",
            "保护组状态正常则为 'normal'，异常则为 'error'。",
            "show protect-group all; show mpls l2vc brief"

        ],
        "DCN路由表检查": [
            "",
            "DCN路由表检查",
            "高性能检测DCN VPN路由表中的出厂默认IP（141开头），仅输出异常路由以提升性能。",
            "仅输出包含141开头IP的异常路由条目，正常路由以汇总形式显示。",
            "show ip routing-table vpn __dcn_vpn__"
        ],
    }

    # 修改后的指南内容生成代码 - 基于用户选择的项目动态生成
    guide_content = []
    counter = 1

    # 遍历用户选择的检查项目，生成对应的指南内容
    for item in selected_items:
        item_name = item['name'].strip()  # 去除可能的空格
        if item_name in guide_content_mapping:
            guide_row = guide_content_mapping[item_name].copy()
            guide_row[0] = str(counter)  # 动态设置编号
            guide_content.append(guide_row)
            counter += 1
            print(
                f"{Fore.GREEN}[INFO] 添加指南: {item_name} (编号: {counter-1}){Style.RESET_ALL}")
        else:
            print(
                f"{Fore.YELLOW}[WARNING] 未找到指南内容: {item_name}{Style.RESET_ALL}")

    # 添加指南内容到工作表
    if guide_content:
        for row_data in guide_content:
            ws_guide.append(row_data)
            for cell in ws_guide[ws_guide.max_row]:
                cell.alignment = center_alignment
                cell.border = thin_border
        print(
            f"{Fore.GREEN}[INFO] 成功生成 {len(guide_content)} 条指南内容{Style.RESET_ALL}")
    else:
        # 如果没有匹配的指南内容，添加提示信息
        ws_guide.append(["无", "未找到匹配的检查项目", "请确认检查项目名称是否正确", "无规则", "无命令"])
        for cell in ws_guide[ws_guide.max_row]:
            cell.alignment = center_alignment
            cell.border = thin_border
        print(f"{Fore.YELLOW}[WARNING] 未生成任何指南内容，请检查选择的项目{Style.RESET_ALL}")

    # 设置列宽
    for col_idx, width in enumerate([8, 25, 45, 35, 20], 1):
        ws_guide.column_dimensions[get_column_letter(col_idx)].width = width

    row = header_row + 1  # 从标题行下一行开始
    for category, items in categories.items():
        if not items:  # 跳过空分类
            continue
        # 合并分类单元格
        merge_end_row = row + len(items) - 1
        ws_summary.merge_cells(f'A{row}:A{merge_end_row}')

        # 写入分类名称并设置样式
        category_cell = ws_summary.cell(row=row, column=1, value=category)
        category_cell.fill = yellow_fill  # 黄色背景
        category_cell.alignment = center_alignment  # 居中
        category_cell.border = thin_border  # 细边框
        category_cell.font = header_font  # 标题字体

        # 遍历每个检查项
        for item in items:
            sheet_name = item['sheet_name']
            health_percent = health_scores.get(sheet_name, "0%")
            print(f"分类: {category}, 项目: {item['name']}, 健康度: {health_percent}")
            normal_count, total_count = item_counts.get(sheet_name, (0, 0))

            # 写入检查项名称（带超链接）
            cell = ws_summary.cell(row=row, column=2, value=item['name'])
            cell.hyperlink = f"#'{sheet_name}'!A1"  # 添加工作表超链接
            cell.font = hyperlink_font  # 超链接字体
            cell.alignment = center_alignment
            cell.border = thin_border

            # 健康度百分比
            ws_summary.cell(row=row, column=3,
                            value=health_percent).alignment = center_alignment
            ws_summary.cell(row=row, column=3).border = thin_border

            # 生成进度条
            percent_value_str = health_percent.rstrip('%')
            percent_value = int(float(percent_value_str))
            progress_bar = create_progress_bar(percent_value)  # 创建文本进度条
            ws_summary.cell(row=row, column=4,
                            value=progress_bar).alignment = left_alignment
            ws_summary.cell(row=row, column=4).border = thin_border

            # 设备数量统计
            ws_summary.cell(
                row=row, column=5, value=f"{normal_count}/{total_count}").alignment = center_alignment
            ws_summary.cell(row=row, column=5).border = thin_border

            # 状态指示灯
            status_cell = ws_summary.cell(row=row, column=6)
            if percent_value >= 90:
                status_cell.value = "优"
                status_cell.fill = green_fill  # 绿色背景
            elif percent_value >= 70:
                status_cell.value = "良"
                status_cell.fill = light_green_fill  # 浅绿色
            elif percent_value >= 50:
                status_cell.value = "中"
                status_cell.fill = yellow_amber_fill  # 琥珀色
            else:
                status_cell.value = "差"
                status_cell.fill = light_red_fill  # 浅红色
            status_cell.alignment = center_alignment
            status_cell.border = thin_border

            row += 1  # 移动到下一行
    # Add connection status row
    ws_summary.cell(row=row, column=1, value="设备网管状态").fill = yellow_fill
    ws_summary.cell(row=row, column=1).alignment = center_alignment
    ws_summary.cell(row=row, column=1).border = thin_border
    ws_summary.cell(row=row, column=1).font = header_font

    cell = ws_summary.cell(row=row, column=2, value="登录失败设备")
    cell.hyperlink = f"#'登录失败设备'!A1"
    cell.font = hyperlink_font
    cell.alignment = center_alignment
    cell.border = thin_border

    health_percent = health_scores.get("登录失败设备", "0%")
    ws_summary.cell(row=row, column=3,
                    value=health_percent).alignment = center_alignment
    ws_summary.cell(row=row, column=3).border = thin_border

    percent_value = int(health_percent.rstrip('%'))
    progress_bar = create_progress_bar(percent_value)
    ws_summary.cell(row=row, column=4,
                    value=progress_bar).alignment = left_alignment
    ws_summary.cell(row=row, column=4).border = thin_border

    normal_count, total_count = item_counts.get("登录失败设备", (0, 0))
    ws_summary.cell(row=row, column=5,
                    value=f"{normal_count}/{total_count}").alignment = center_alignment
    ws_summary.cell(row=row, column=5).border = thin_border

    status_cell = ws_summary.cell(row=row, column=6)
    if percent_value >= 90:
        status_cell.value = "优"
        status_cell.fill = green_fill
    elif percent_value >= 70:
        status_cell.value = "良"
        status_cell.fill = light_green_fill
    elif percent_value >= 50:
        status_cell.value = "中"
        status_cell.fill = yellow_amber_fill
    else:
        status_cell.value = "差"
        status_cell.fill = light_red_fill
    status_cell.alignment = center_alignment
    status_cell.border = thin_border

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 应用自动列宽
        autofit_worksheet_columns(ws)

        # 只给子表添加筛选和冻结首行，不给主表（"运维质量评估"）添加
        if sheet_name != "运维质量评估":
            # 假设子表的表头都在第1行
            header_row = 1
            last_col = ws.max_column
            last_row = ws.max_row

            # 确保表格有数据才添加筛选
            if last_row > header_row:
                # 获取表格范围（从A1到最后一列最后一行）
                filter_range = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
                ws.auto_filter.ref = filter_range
                print(
                    f"{Fore.GREEN}[INFO] 添加了筛选: {sheet_name}, 范围: {filter_range}{Style.RESET_ALL}")

            # 添加冻结首行功能
            # freeze_panes='A2' 表示冻结第1行，从第2行开始可以滚动
            ws.freeze_panes = 'A2'
            print(
                f"{Fore.GREEN}[INFO] 添加冻结首行窗格: {sheet_name}, 冻结第一行{Style.RESET_ALL}")

    # Save workbook
    wb.save(report_file)
    print(
        f"{Fore.GREEN}[END] 生成 QA 报告: {report_file}，处理了 {len(successful_host_ips)} 台成功连接的设备{Style.RESET_ALL}")


def autofit_worksheet_columns(worksheet):
    """
    真正实现Excel中双击列分隔线自适应效果的函数
    此函数会遍历每个单元格，找出每列中的最长内容，并相应地调整列宽
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import MergedCell

    # 存储每列的最大宽度
    column_widths = {}

    # 遍历所有单元格
    for row in worksheet.rows:
        for cell in row:
            # 跳过合并单元格
            if isinstance(cell, MergedCell):
                continue

            col_idx = cell.column
            col_letter = get_column_letter(col_idx)

            if cell.value:
                try:
                    # 根据单元格内容估算宽度（考虑中文字符和格式）
                    cell_value = str(cell.value)
                    # 中文字符宽度通常是英文的2倍
                    length = 0
                    for char in cell_value:
                        # 检查是否为全角字符（如中文、日文等）
                        if ord(char) > 127:
                            length += 2
                        else:
                            length += 1

                    # 考虑字体加粗会增加宽度
                    if cell.font and cell.font.bold:
                        length += len(cell_value) * 0.1

                    # 更新该列的最大宽度
                    current_width = column_widths.get(col_letter, 0)
                    column_widths[col_letter] = max(current_width, length)
                except:
                    pass

    # 设置列宽
    for col_letter, width in column_widths.items():
        # 在计算的宽度上左右各增加7像素，总共14像素
        adjusted_width = width + 2
        worksheet.column_dimensions[col_letter].width = adjusted_width


def sanitize_string(value):
    """移除字符串中 Excel 不允许的非法字符"""
    if isinstance(value, str):
        # 移除控制字符（\x00-\x1f 和 \x7f-\x9f）
        return re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
    return value


def _progress_bar(seconds: int, completion_msg: str):
    """可视化进度条 (兼容Windows/Linux)"""
    symbols = cycle(['⣾', '⣽', '⣻', '⢿', '⡿', '⣟', '⣯', '⣷'])  # 旋转动画符号
    end_time = time.time() + seconds

    print(f"{Fore.GREEN}✓ {completion_msg}{Style.RESET_ALL}")
    while time.time() < end_time:
        remaining = int(end_time - time.time())
        # 进度百分比计算
        progress = 100 - int((remaining / seconds) * 100)
        # 动态颜色（红色->黄色->绿色渐变）
        color_code = f"\033[38;5;{28 + min(progress * 2, 56)}m"  # 使用 ANSI 颜色代码
        # 进度条生成
        bar = f"{Fore.GREEN}▰" * int(progress / 5) + \
            f"{Fore.LIGHTBLACK_EX}▱" * int((100 - progress) / 5)
        # 动态输出
        sys.stdout.write(
            f"\r{next(symbols)} "
            f"{color_code}▏{progress}%{Style.RESET_ALL} "
            f"{bar} "
            f"{Fore.CYAN}剩余时间: {remaining}s{Style.RESET_ALL}"
        )
        sys.stdout.flush()
        time.sleep(0.1)

    # 清除当前行并输出完成消息
    sys.stdout.write(f"\r{' ' * 80}\r")  # 清除整行
    sys.stdout.flush()


# ---------------------------------------------------
# 主函数
# 功能： 读取指定CSV文件中的命令，下发到设备执行
# ----------------------------------------------------
# 主函数修改后代码（完整版）
if __name__ == '__main__':
    init(autoreset=True)  # 初始化颜色输出

    while True:  # 主循环
        print("\n" + "="*50)
        print(f"{Fore.CYAN}STN-A设备巡检系统 v2.8{Style.RESET_ALL}".center(50))
        print("="*50)

        menu = f"""
{Fore.YELLOW}请选择操作：{Style.RESET_ALL}
  1️⃣  全量采集设备数据      - 采集所有设备信息
  2️⃣  清洗采集结果        - 处理原始数据
  3️⃣  生成巡检报告        - 输出完整报告
  4️⃣  持续监控模式        - 周期性巡检
  5️⃣  专项快速巡检        - 针对性检查业务-统计-槽位-业务-端口
  6️⃣  当前CRC检查        - 接口错误统计
  7️⃣  LLDP邻居检查       - 统计下挂设备型号
  8️⃣  业务IP上报统计      - 统计业务端口IP
  9️⃣  生成LLDP拓扑图      - 绘制网络拓扑图
  🔟  OSPF检查统计          - 检查OSPF进程
  1️⃣1️⃣ 导出设备运行配置    - 导出并保存配置
  1️⃣2️⃣ 运行配置清洗       - 清洗导出的配置
  1️⃣3️⃣ 接口光功率检查     - 检查光功率和CRC
  1️⃣4️⃣ 光模块性能检查统计  - 检查统计光模块信息
  1️⃣5️⃣ 运行自定义指令      - 批量执行自定义指令
  1️⃣6️⃣ 统计检查设备状态    - 检查设备整体运行状态
  1️⃣7️⃣ 业务LSP检查       - 检查业务LSP状态
  1️⃣8️⃣ 设备告警检查统计   - 统计当前和历史告警
  1️⃣9️⃣ 自动设置设备时间    - 适合无法同步NTP的A设备
  2️⃣0️⃣ QA巡检采集           - 质量保证巡检
  2️⃣1️⃣ 添加互联端口描述    - 为OSPF .31接口添加LLDP邻居描述
  0️⃣  退出系统            - 结束程序
{Fore.CYAN}请输入选项：{Style.RESET_ALL}"""
        ucmd = input(menu)

        # 处理主菜单选项
        if ucmd == '0':
            print(f"\n{Fore.GREEN}👋 感谢使用，再见！{Style.RESET_ALL}")
            exit()

        elif ucmd == '1':
            # 全量采集模式
            print(f"\n{Fore.BLUE}📡 进入全量采集模式{Style.RESET_ALL}")
            ret_name = getinput("result.txt", "请输入保存文件名（默认：result.txt）：")
            filename = getinput(
                "host-stna.csv", "请输入设备清单文件（默认：host-stna.csv）：")
            fish(filename, ret_name)

        elif ucmd == '2':
            # 数据清洗模式
            print(f"\n{Fore.BLUE}🔧 进入数据清洗模式{Style.RESET_ALL}")
            while True:
                src_file = getinput("result.txt", "原始数据文件（默认：result.txt）：")
                dst_file = getinput(
                    "washed_rec.csv", "清洗后文件名（默认：washed_rec.csv）：")
                wash_int_main(src_file, dst_file)
                for cmd_id in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
                    wash_result(src_file, dst_file, cmd_id)
                back = input(f"{Fore.YELLOW}输入0返回主菜单：{Style.RESET_ALL}")
                if back == '0':
                    break

        elif ucmd == '3':
            print("\n📊 生成巡检报告")
            rec_name = getinput(
                "washed_rec.csv", "请输入清洗后数据文件(默认：washed_rec.csv):：")
            h_name = getinput(
                "host-stna.csv", "请输入设备清单文件(默认：host-stna.csv):：")

            current_time = datetime.now().strftime("%Y-%m-%d-%H-%M")
            report_name = f"巡检报告-{current_time}.csv"

            print("\n🔄 正在生成报告...")
            report_result(rec_name, report_name, h_name)
            print(f"\n✅ 巡检报告已生成：{report_name}")

        elif ucmd == '4':
            print("\n🔁 进入持续监控模式")
            sloop = getinput("100", "请输入监控轮次（默认100次）：")
            iloop = int(sloop)
            ret_name = getinput("fish-rec.txt", "请输入保存文件名：")
            filename = getinput("host-l.csv", "请输入监控设备清单：")

            print(f"\n⚠️ 注意：将持续监控{iloop}轮，按Ctrl+C可终止")
            while iloop > 0:
                print(f"\n🔄 剩余监控轮次：{iloop}")
                fish(filename, ret_name)
                iloop -= 1
                time.sleep(60)  # 每轮间隔60秒
        elif ucmd == '5':
            # 专项快速巡检
            while True:  # 子菜单循环
                print(f"\n{Fore.BLUE}🚀 专项快速巡检模式{Style.RESET_ALL}")
                dynamic_colored_divider(
                    color_code=34, symbol='-', enable_timestamp=True)
                sub_menu = f"""
{Fore.YELLOW}请选择专项巡检类型：{Style.RESET_ALL}
  1️⃣  系统版本检查       - 检查设备版本一致性
  2️⃣  设备温度检查       - 监测设备温度状态
  3️⃣  光功率检查         - 检查光模块功率
  4️⃣  LDP成环分析        - 检测LDP协议成环
  5️⃣  业务统计           - 统计设备VC业务
  6️⃣  空闲槽位检查       - 检查设备空闲槽位
  7️⃣  空闲端口检查       - 检查设备空闲端口
  8️⃣  业务板卡统计       - 统计板卡使用情况
  9️⃣  端口使用率统计     - 统计检查端口负载流量
  🔟  OSPF互联接口检查   - 专项检查OSPF 31与接口、邻居信息

  0️⃣  返回主菜单         - 返回上一级
{Fore.CYAN}请输入选项（0-10）：{Style.RESET_ALL}"""
                subcmd = input(sub_menu)

                if subcmd == '0':
                    break  # 返回主菜单
               # Main execution block (assuming getinput is defined elsewhere)
                if subcmd == '5':
                    print(f"\n{Fore.MAGENTA}📊 正在执行业务统计...{Style.RESET_ALL}")
                    raw_file = getinput(
                        "l2vc_raw.txt", "原始数据文件（默认：l2vc_raw.txt）：", timeout=10)
                    host_file = getinput(
                        "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
                    report_file = getinput(
                        "VC业务报告.csv", "输出报告（默认：VC业务报告.csv）：", timeout=10)

                    # 采集数据（显示实时进度）
                    print(f"\n{Fore.CYAN}🚀 开始采集业务数据...{Style.RESET_ALL}")
                    fish_cmd(host_file, raw_file,
                             "show mpls l2vc brief", max_workers=40)

                    # 生成智能报告（带多级进度条）
                    print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
                    generate_vc_report(raw_file, report_file, host_file)

                elif subcmd == '6':
                    print(f"\n{Fore.MAGENTA}🔍 正在检查空闲槽位...{Style.RESET_ALL}")
                    raw_file = getinput(
                        "slot_raw.txt", "原始数据保存文件（默认：slot_raw.txt）：", timeout=10)
                    host_file = getinput(
                        "host-stna.csv", "设备清单文件（默认：host-stna.csv）：", timeout=10)

                    # 使用专用函数采集槽位数据（关键修改点）
                    fish_slot_cmd(host_file, raw_file)

                    # 生成报告
                    report_file = getinput(
                        "空闲槽位检查报告.csv", "槽位检查报告（默认：空闲槽位检查报告.csv）：", timeout=10)
                    generate_slot_report(
                        "slot_raw.txt", "空闲槽位检查报告.csv", "host-stna.csv")
                    print(
                        f"\n{Fore.GREEN}✅ 槽位检查报告已生成：{report_file}{Style.RESET_ALL}")
                elif subcmd == '7':
                    print(f"\n{Fore.MAGENTA}🔍 正在检查空闲端口...{Style.RESET_ALL}")
                    raw_file = getinput(
                        "port_raw.txt", "原始数据保存文件（默认：port_raw.txt）: ", timeout=10)
                    host_file = getinput(
                        "host-stna.csv", "设备列表文件（默认：host-stna.csv）: ", timeout=10)

                    # 执行数据采集
                    fish_port_cmd(host_file, raw_file)

                    # 生成报告
                    report_file = getinput(
                        "空闲端口检查报告.csv", "空闲端口检查报告（默认：空闲端口检查报告.csv）: ", timeout=10)
                    generate_port_report(raw_file, report_file, host_file)
                    print(
                        f"\n{Fore.GREEN}✅ 端口检查报告已生成：{report_file}{Style.RESET_ALL}")
                elif subcmd == '8':
                    print("\n📊 正在执行业务板卡统计...")
                    raw_file = input(
                        "原始数据保存文件（默认：board_raw.txt）: ") or "board_raw.txt"
                    host_file = input(
                        "设备列表文件（默认：host-stna.csv）: ") or "host-stna.csv"

                    # 数据采集
                    fish_board_cmd(host_file, raw_file)

                    # 生成报告
                    report_file = input(
                        "业务板卡统计报告（默认：业务板卡统计.csv）: ") or "业务板卡统计.csv"
                    generate_board_report(raw_file, report_file, host_file)
                    print(f"\n✅ 业务板卡统计报告已生成：{report_file}")
                elif subcmd == '9':
                    print("\n📊 正在执行端口使用率统计...")
                    raw_file = input(
                        "原始数据保存文件（默认：port_usage_raw.txt）: ") or "port_usage_raw.txt"
                    host_file = input(
                        "设备列表文件（默认：host-stna.csv）: ") or "host-stna.csv"
                    fish_port_usage_cmd(host_file, raw_file)
                    report_file = input(
                        "端口使用率统计报告（默认：端口使用率统计.csv）: ") or "端口使用率统计.csv"
                    generate_port_usage_report(
                        raw_file, report_file, host_file)
                    print(f"\n✅ 端口使用率统计报告已生成：{report_file}")
                elif subcmd == '10':
                    print("\n📊 正在执行OSPF互联接口信息检查...")
                    raw_file = getinput(
                        "ospf_interface_raw.txt", "原始数据文件（默认：ospf_interface_raw.txt）：", timeout=10)
                    host_file = getinput(
                        "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
                    report_file = getinput(
                        "OSPF互联接口信息.csv", "输出报告（默认：OSPF互联接口信息.csv）：", timeout=10)

                    print(f"\n{Fore.CYAN}🚀 开始采集OSPF互联接口数据...{Style.RESET_ALL}")
                    fish_ospf_interface_info_cmd(host_file, raw_file)

                    print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
                    generate_ospf_interface_report(
                        raw_file, report_file, host_file)
                    print(
                        f"\n{Fore.GREEN}✅ OSPF互联接口信息报告已生成：{report_file}{Style.RESET_ALL}")
                else:
                    print(f"{Fore.RED}⚠️ 无效选项，请重新输入！{Style.RESET_ALL}")
        elif ucmd == '6':
            print("\n📊 正在执行当前CRC检查...")
            raw_file = getinput(
                "crc_raw.txt", "原始数据文件（默认：crc_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "CRC检查报告.csv", "输出报告（默认：CRC检查报告.csv）：", timeout=10)

            # Collect CRC data
            print(f"\n{Fore.CYAN}🚀 开始采集CRC数据...{Style.RESET_ALL}")
            fish_crc_cmd(host_file, raw_file)

            # Generate report
            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_crc_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ CRC检查报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '7':
            print("\n📊 正在执行LLDP邻居检查...")
            raw_file = getinput(
                "lldp_raw.txt", "原始数据文件（默认：lldp_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "LLDP邻居报告.csv", "输出报告（默认：LLDP邻居报告.csv）：", timeout=10)

            # 采集数据
            print(f"\n{Fore.CYAN}🚀 开始采集LLDP邻居数据...{Style.RESET_ALL}")
            fish_lldp_neighbor_cmd(host_file, raw_file)

            # 生成报告
            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_lldp_neighbor_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ LLDP邻居报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '8':
            print(f"\n{Fore.MAGENTA}📊 正在执行基站和业务IP统计...{Style.RESET_ALL}")
            raw_file = getinput(
                "arp_raw.txt", "原始数据文件（默认：arp_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "端口ARP上报统计.csv", "输出报告（默认：端口ARP上报统计.csv）：", timeout=10)

            # Collect data
            print(f"\n{Fore.CYAN}🚀 开始采集ARP数据...{Style.RESET_ALL}")
            fish_arp_cmd(host_file, raw_file)

            # Generate report
            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_arp_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ 端口ARP上报统计已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '9':
            print("\n📈 正在生成网络拓扑图...")
            lldp_file = getinput(
                "LLDP邻居报告.csv", "请输入LLDP邻居报告文件（默认：LLDP邻居报告.csv）：", timeout=10)
            topo_file = getinput(
                "topology.html", "请输入输出拓扑图文件（默认：topology.html）：", timeout=10)
            try:
                generate_topology_html(lldp_file, topo_file)
            except FileNotFoundError:
                print(f"{Fore.RED}⚠️ 文件 {lldp_file} 不存在，请检查路径！{Style.RESET_ALL}")
            except Exception as e:
                print(f"{Fore.RED}⛔ 生成拓扑图失败：{e}{Style.RESET_ALL}")
        elif ucmd == '10':
            print("\n📊 正在执行OSPF互联检查统计...")
            raw_file = getinput("ospf_neighbor_raw.txt",
                                "原始数据文件（默认：ospf_neighbor_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "OSPF互联检查统计.csv", "OSPF互联检查统计.csv）：", timeout=10)

            print(f"\n{Fore.CYAN}🚀 开始采集OSPF邻居数据...{Style.RESET_ALL}")
            fish_ospf_neighbor_cmd(host_file, raw_file)

            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_ospf_neighbor_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ OSPF互联检查统计报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '11':
            print("\n📊 正在导出设备运行配置...")
            host_file = getinput("host-stna.csv", "设备清单文件（默认：host-stna.csv）：")
            export_running_config(host_file)
        elif ucmd == '12':
            print("\n🔧 运行配置清洗功能待实现。")
            print(
                f"\n{Fore.GREEN}✅ 接口光功率与CRC检查报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '13':
            print("\n📊 正在执行接口光功率与CRC检查...")
            raw_file = getinput("optical_raw.txt",
                                "原始数据文件（默认：optical_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "接口光功率与CRC检查.csv", "输出报告（默认：接口光功率与CRC检查.csv）：", timeout=10)

            print(f"\n{Fore.CYAN}🚀 开始采集接口光功率与CRC数据...{Style.RESET_ALL}")
            fish_interface_optical_cmd(host_file, raw_file)

            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_optical_report(raw_file, report_file, host_file)
            print(
                f"\n{Fore.GREEN}✅ 接口光功率与CRC检查报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '14':
            print("\n📊 正在执行光模块性能统计...")
            raw_file = getinput("optical_module_raw.txt",
                                "原始数据文件（默认：optical_module_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "光模块性能统计.csv", "输出报告（默认：光模块性能统计.csv）：", timeout=10)

            print(f"\n{Fore.CYAN}🚀 开始采集光模块性能数据...{Style.RESET_ALL}")
            fish_optical_cmd(host_file, raw_file)

            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_optical_module_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ 光模块性能统计报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '15':
            print("\n📊 正在执行自定义指令...")
            host_file = getinput(
                "userhost-stna.csv", "设备清单（默认：userhost-stna.csv）：", timeout=10)
            raw_file = getinput("custom_cmd_raw.txt",
                                "原始数据文件（默认：custom_cmd_raw.txt）：", timeout=10)
            report_file = getinput(
                "自定义指令执行报告.csv", "输出报告（默认：自定义指令执行报告.csv）：", timeout=10)
            try:
                with open("自定义指令.txt", "r", encoding='utf-8') as f:
                    commands = [line.strip() for line in f if line.strip()]
            except FileNotFoundError:
                print(f"{Fore.RED}⚠️ 自定义指令.txt 文件不存在！{Style.RESET_ALL}")
                continue
            if not commands:
                print(f"{Fore.RED}⚠️ 自定义指令.txt 文件为空！{Style.RESET_ALL}")
                continue
            print(f"\n{Fore.CYAN}🚀 开始采集自定义指令数据...{Style.RESET_ALL}")
            fish_custom_cmd(host_file, raw_file, commands)
            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_custom_cmd_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ 自定义指令报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '16':
            print("\n📊 正在执行设备状态统计检查...")
            host_file = input("设备清单（默认：host-stna.csv）：") or "host-stna.csv"
            raw_file = input(
                "原始数据文件（默认：device_info_raw.txt）：") or "device_info_raw.txt"
            report_file = input("输出报告（默认：设备状态统计.csv）：") or "设备状态统计.csv"

            print(f"\n{Fore.CYAN}🚀 开始采集设备信息数据...{Style.RESET_ALL}")
            fish_device_info_cmd(host_file, raw_file)

            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_device_info_report(raw_file, report_file, host_file)

            print(f"\n{Fore.GREEN}✅ 设备状态统计报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '17':
            print("\n📊 正在执行业务LSP检查...")
            raw_file = getinput(
                "lsp_raw.txt", "原始数据文件（默认：lsp_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "业务LSP报告.csv", "输出报告（默认：业务LSP报告.csv）：", timeout=10)
            print(f"\n{Fore.CYAN}🚀 开始采集业务LSP数据...{Style.RESET_ALL}")
            fish_lsp_cmd(host_file, raw_file)
            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_lsp_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ 业务LSP报告已生成：{report_file}{Style.RESET_ALL}")
        elif ucmd == '18':
            print("\n📊 正在执行设备告警检查统计...")
            raw_file = getinput(
                "alarm_raw.txt", "原始数据文件（默认：alarm_raw.txt）：", timeout=10)
            host_file = getinput(
                "host-stna.csv", "设备清单（默认：host-stna.csv）：", timeout=10)
            report_file = getinput(
                "设备告警统计.xlsx", "输出报告（默认：设备告警统计.xlsx）：", timeout=10)
            print(f"\n{Fore.CYAN}🚀 开始采集告警数据...{Style.RESET_ALL}")
            fish_alarm_cmd(host_file, raw_file)
            print(f"\n{Fore.CYAN}🧹 正在分析数据并生成报告...{Style.RESET_ALL}")
            generate_alarm_report(raw_file, report_file, host_file)
            print(f"\n{Fore.GREEN}✅ 设备告警统计报告已生成：{report_file}{Style.RESET_ALL}")

        elif ucmd == '20':
            print(f"\n{Fore.BLUE}🔍 QA巡检{Style.RESET_ALL}")
            print(
                f"{Fore.YELLOW}请选择要巡检的项目（输入编号，用逗号分隔，如1,2，或输入以下选项）：{Style.RESET_ALL}")

            # 功能描述映射
            guide_content_mapping = {
                "设备运行时间检查": "检查设备持续运行时间，评估电源稳定性和设备运行状态",
                "主控盘运行状态": "监控主控盘CPU和内存使用情况，确保系统资源正常",
                "协议报文处理状态": "检查协议处理过程中是否存在丢包，确保数据转发正常",
                "真实版本信息": "获取设备真实软件版本信息，验证系统固件状态",
                "风扇转速及温度状态": "监控设备温度和风扇运行状态，确保散热系统正常",
                "系统与硬件版本状态": "检查系统软件和硬件版本信息完整性",
                "光模块信息检查": "检测光模块功率、电流、电压等参数，确保光纤链路质量",
                "电源状态检查": "监控设备各槽位电压状态，确保供电稳定",
                "FW软件版本一致性检查": "检查主备控制器软件版本是否一致，确保系统可靠性",
                "板卡CPU内存使用率": "监控各板卡CPU、内存使用率和温度状态",
                "NTP时间同步分析": "检查设备时间同步状态，确保网络时间准确性",
                "硬盘资源占用分析": "监控存储空间使用情况，预防存储资源不足",
                "BFD会话检查(VC业务统计)": "检查BFD双向转发检测会话状态，确保快速故障检测",
                "配置校验状态": "检查配置校验功能是否启用，确保配置正确性",
                "OSPF进程状态检查": "监控OSPF路由协议进程运行状态和资源使用情况",
                "IPFRR-LSP状态检查": "检查MPLS标签交换路径状态，确保IP快速重路由功能",
                "OSPF邻居状态检查": "检查OSPF邻居关系建立和维护状态",
                "LACP成员状态监控": "监控链路聚合控制协议成员端口状态",
                "OSPF 路由表检查": "检查OSPF路由表条目成本值和稳定性",
                "LDP 会话状态检查": "检查标签分发协议会话建立和运行状态",
                "Loopback31地址唯一性检查": "验证Loopback31接口IP地址在网络中的唯一性",
                "Loopback1023地址唯一性检查": "验证Loopback1023接口IP地址在网络中的唯一性",
                "SNMP配置检查": "检查SNMP协议版本和安全配置合规性",
                "设备账户检查": "检查设备登录账户安全策略配置",
                "专网业务分析": "检查VPLS专网业务运行状态和接口配置",
                "PTP时钟检查": "检查精确时间协议同步状态和时钟配置",
                "站点邻接网元检查": "检查LDP邻接关系和站点间网元连接状态",
                "设备面板视图": "以图形化方式展示设备槽位和板卡物理布局",
                "BFD保护组状态信息": "检查BFD保护组配置和L2VC业务关联状态",
                "DCN路由表检查": "检查DCN VPN路由表，识别出厂默认IP配置的设备"
            }

            inspection_items = {
                "1": {
                    "name": "设备运行时间检查",
                    "command": "show device",
                    "parser": parse_uptime,
                    "sheet_name": "设备运行时间检查",
                    "category": "设备基础状态"
                },
                "2": {
                    "name": "主控盘运行状态",
                    "command": "show device",
                    "parser": parse_main_control_status,
                    "sheet_name": "主控盘运行状态",
                    "category": "设备基础状态"
                },
                "3": {
                    "name": "协议报文处理状态",
                    "command": "show cpu-defend stats",
                    "parser": parse_cpu_defend_stats,
                    "sheet_name": "协议报文处理状态",
                    "category": "硬件可靠性"
                },
                "4": {
                    "name": "真实版本信息",
                    "command": "show real-version",
                    "parser": parse_real_version,
                    "sheet_name": "真实版本信息",
                    "category": "硬件可靠性"
                },
                "5": {
                    "name": "风扇转速及温度状态",
                    "command": "show temperature",
                    "parser": parse_temperature,
                    "sheet_name": "风扇转速及温度状态",
                    "category": "设备基础状态"
                },
                "6": {
                    "name": "系统与硬件版本状态",
                    "command": "show version",
                    "parser": parse_version,
                    "sheet_name": "系统与硬件版本状态",
                    "category": "硬件可靠性"
                },
                "7": {
                    "name": "光模块信息检查",
                    "command": "show interface",
                    "parser": lambda ip, interface_output, lldp_output: parse_optical_module(ip, interface_output, lldp_output, parse_uptime),
                    "sheet_name": "光模块信息检查",
                    "category": "设备基础状态"
                },
                "8": {
                    "name": "电源状态检查",
                    "command": "show voltage",
                    "parser": parse_power_status,
                    "sheet_name": "电源状态检查",
                    "category": "设备基础状态"
                },
                "9": {
                    "name": "FW软件版本一致性检查",
                    "command": "show device",
                    "parser": parse_main_backup_version,
                    "sheet_name": "FW软件版本一致性检查",
                    "category": "系统运行状态"
                },
                "10": {
                    "name": "板卡CPU内存使用率",
                    "command": "show pm cur-15m Dev",
                    "parser": lambda output_15m, output_24h: parse_board_cpu_memory(output_15m, output_24h),
                    "sheet_name": "板卡CPU内存使用率",
                    "category": "资源监控"
                },
                "11": {
                    "name": "NTP时间同步分析",
                    "command": "show cloc",
                    "parser": parse_ntp_status,
                    "sheet_name": "NTP时间同步分析",
                    "category": "系统运行状态"
                },
                "12": {
                    "name": "硬盘资源占用分析",
                    "command": "show flash-usage",
                    "parser": parse_flash_usage,
                    "sheet_name": "硬盘资源占用分析",
                    "category": "资源监控"
                },
                "13": {
                    "name": "BFD会话检查(VC业务统计)",
                    "command": "show bfd session brief",
                    "parser": parse_bfd_sessions,  # 直接引用函数，不使用lambda
                    "sheet_name": "BFD会话检查(VC业务统计)",
                    "category": "路由协议健康度"
                },
                "14": {
                    "name": "配置校验状态",
                    "command": "show cfgchk info",
                    "parser": parse_cfgchk_info,
                    "sheet_name": "配置校验状态",
                    "category": "冗余与容灾"
                },
                "15": {
                    "name": "OSPF进程状态检查",
                    "command": "show ospf process",
                    "parser": parse_ospf_session,
                    "sheet_name": "OSPF进程状态检查",
                    "category": "路由协议健康度"
                },
                "16": {
                    "name": "IPFRR-LSP状态检查",
                    "command": "show mpls lsp brief",
                    "parser": parse_mpls_lsp,
                    "sheet_name": "IPFRR-LSP状态检查",
                    "category": "转发层验证"
                },
                "17": {
                    "name": "OSPF邻居状态检查",
                    # Multiple commands
                    "command": ["show ospf buffers", "show ospf neighbor"],
                    "parser": parse_ospf_neighbor_status,
                    "sheet_name": "OSPF邻居状态检查",
                    "category": "路由协议健康度"
                },
                "18": {
                    "name": "LACP成员状态监控",
                    "command": ["show lag", "show lacp"],  # Multiple commands
                    "parser": parse_lacp_status,
                    "sheet_name": "LACP成员状态监控",
                    "category": "冗余与容灾"
                },
                "19": {
                    "name": "OSPF 路由表检查",
                    "command": "show ip routing-table",
                    "parser": parse_ospf_routing_table,
                    "sheet_name": "OSPF 路由表检查",
                    "category": "路由协议健康度"
                },
                "20": {
                    "name": "LDP 会话状态检查",
                    "command": ["show ldp session"],
                    "parser": parse_ldp_session_status,
                    "sheet_name": "LDP 会话状态检查",
                    "category": "路由协议健康度"
                },
                "21": {
                    "name": "Loopback31地址唯一性检查",
                    "command": "show interface loopback 31",
                    "parser": parse_loopback_address,
                    "sheet_name": "Loopback31地址唯一性检查",
                    "category": "基础安全配置"
                },
                "22": {
                    "name": "Loopback1023地址唯一性检查",
                    "command": "show interface loopback 1023",
                    "parser": parse_loopback_address,
                    "sheet_name": "Loopback1023地址唯一性检查",
                    "category": "基础安全配置"
                },
                "23": {
                    "name": "SNMP配置检查",
                    "command": ["show snmp-server trap", "show snmp-server community", "show snmp-server mib-view", "show snmp-server sys-info"],
                    "parser": parse_snmp_config,
                    "sheet_name": "SNMP配置检查",
                    "category": "基础安全配置"
                },
                "24": {
                    "name": "设备账户检查",
                    "command": ["show users", "show login-global-rule", "show loginning-user"],
                    "parser": parse_device_accounts,
                    "sheet_name": "设备账户检查",
                    "category": "基础安全配置"
                },
                "25": {
                    "name": "专网业务分析",
                    "command": ["show vsi brief"],
                    "parser": parse_private_network_service,
                    "sheet_name": "专网业务分析",
                    "category": "冗余与容灾"
                },
                "26": {
                    "name": "PTP时钟检查",
                    "command": ["show ptp all", "show synce"],
                    "parser": lambda ptp_output, synce_output, ne_type, ne_name, ne_ip: parse_ptp_clock_status(ptp_output, synce_output, ne_type, ne_name, ne_ip),
                    "sheet_name": "PTP时钟检查",
                    "category": "系统运行状态"
                },
                "27": {
                    "name": "站点邻接网元检查",
                    "command": ["show ldp lsp"],
                    "parser": parse_ldp_lsp,
                    "sheet_name": "站点邻接网元检查",
                    "category": "冗余与容灾"
                },
                "28": {
                    "name": "设备面板视图",
                    "command": ["show device"],
                    "parser": parse_device_panel,
                    "sheet_name": "设备面板视图",
                    "category": "设备状态"
                },
                "29": {
                    "name": "BFD保护组状态信息",
                    "command": "show protect-group all",
                    "parser": parse_protect_group_all,
                    "sheet_name": "BFD保护组状态信息",
                    "category": "冗余与容灾"
                },
                "30": {
                    "name": "DCN路由表检查",
                    "command": "show ip routing-table vpn __dcn_vpn__",
                    "parser": parse_dcn_routing_table,
                    "sheet_name": "DCN路由表检查",
                    "category": "路由协议健康度"
                },
            }

            # Group items by category for display
            categories = {
                "设备基础状态": [item for item in inspection_items.values() if item["category"] == "设备基础状态"],
                "硬件可靠性": [item for item in inspection_items.values() if item["category"] == "硬件可靠性"],
                "系统运行状态": [item for item in inspection_items.values() if item["category"] == "系统运行状态"],
                "资源监控": [item for item in inspection_items.values() if item["category"] == "资源监控"],
                "路由协议健康度": [item for item in inspection_items.values() if item["category"] == "路由协议健康度"],
                "冗余与容灾": [item for item in inspection_items.values() if item["category"] == "冗余与容灾"],
                "转发层验证": [item for item in inspection_items.values() if item["category"] == "转发层验证"],
                "基础安全配置": [item for item in inspection_items.values() if item["category"] == "基础安全配置"]
            }

            # Display categories and items with descriptions
            for category, items in categories.items():
                print(f"\n{Fore.CYAN}━━━ {category} ━━━{Style.RESET_ALL}")
                for key, item in inspection_items.items():
                    if item["category"] == category:
                        description = guide_content_mapping.get(
                            item['name'], "功能描述暂未提供")
                        print(
                            f"{Fore.WHITE}{key:>2}.{Style.RESET_ALL} {Fore.GREEN}{item['name']}{Style.RESET_ALL}")
                        print(
                            f"     {Fore.LIGHTBLACK_EX}▸ {description}{Style.RESET_ALL}")

            print(
                f"\n{Fore.YELLOW}━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━{Style.RESET_ALL}")
            print(f"{Fore.GREEN}默认同时连接40台设备。{Style.RESET_ALL}")
            print(f"{Fore.MAGENTA} 0.{Style.RESET_ALL}  返回主菜单")
            print(f"{Fore.MAGENTA}00.{Style.RESET_ALL}  执行全量巡检（包含所有30项检查）")
            print(
                f"{Fore.MAGENTA}000.{Style.RESET_ALL} QA文件清洗（仅清洗已有qa_wash_raw.txt数据）")

            # Get user selection
            selection = input(f"\n{Fore.CYAN}请输入选项：{Style.RESET_ALL}")
            if selection == '0':
                continue
            elif selection == '000':
                # QA文件清洗模式
                print(
                    f"\n{Fore.GREEN}[INFO] 触发QA文件清洗模式，仅处理已有数据{Style.RESET_ALL}")
                raw_file = getinput("qa_wash_raw.txt",
                                    "原始数据文件（默认：qa_wash_raw.txt）：")
                host_file = getinput(
                    "host-stna.csv", "设备清单（默认：host-stna.csv）：")
                report_file = f"QA巡检报告-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx"

                # 询问用户选择巡检项
                print(
                    f"\n{Fore.YELLOW}请选择要清洗的巡检项目（输入编号，用逗号分隔，如1,2，或输入00清洗所有项目）：{Style.RESET_ALL}")

                # 按分类显示巡检项目选择列表
                for category, items in categories.items():
                    print(f"\n{Fore.CYAN}{category}:{Style.RESET_ALL}")
                    for key, item in inspection_items.items():
                        if item["category"] == category:
                            print(f"{key}. {item['name']}")

                print(f"\n{Fore.MAGENTA}00. 清洗所有巡检项目{Style.RESET_ALL}")

                clean_selection = input(f"{Fore.CYAN}请输入选项：{Style.RESET_ALL}")
                if clean_selection == '00':
                    selected_items = list(inspection_items.values())
                    print(f"{Fore.GREEN}[INFO] 选择清洗所有巡检项{Style.RESET_ALL}")
                else:
                    selected_keys = clean_selection.split(',')
                    selected_items = [inspection_items[key]
                                      for key in selected_keys if key in inspection_items]
                    if not selected_items:
                        print(f"{Fore.RED}[ERROR] 未选择任何巡检项目{Style.RESET_ALL}")
                        continue

                # 直接调用generate_qa_report进行数据清洗和报告生成
                generate_qa_report(raw_file, report_file,
                                   host_file, selected_items)
            else:
                # 原有逻辑：执行巡检
                if selection == '00':
                    selected_items = list(inspection_items.values())
                    print(
                        f"{Fore.GREEN}[INFO] 触发全量巡检，选择所有巡检项{Style.RESET_ALL}")
                else:
                    selected_keys = selection.split(',')
                    selected_items = [inspection_items[key]
                                      for key in selected_keys if key in inspection_items]
                    if not selected_items:
                        print(f"{Fore.RED}[ERROR] 未选择任何巡检项目{Style.RESET_ALL}")
                        continue

                # Collect commands based on selected items
                commands = []
                for item in selected_items:
                    if isinstance(item['command'], list):
                        commands.extend(item['command'])
                    else:
                        commands.append(item['command'])

                # Add additional commands based on conditions
                if any(item['name'] == "风扇转速及温度状态" for item in selected_items):
                    commands.append("show fan")
                if any(item['name'] == "光模块信息检查" for item in selected_items):
                    commands.append("show lldp neighbor")
                if any(item['name'] == "板卡CPU内存使用率" for item in selected_items):
                    commands.append("show pm cur-15m Dev")
                    commands.append("show pm cur-24h Dev")
                if any(item['name'] == "NTP时间同步分析" for item in selected_items):
                    commands.extend(["show cloc", "show ntp-service"])
                if any(item['name'] == "硬盘资源占用分析" for item in selected_items):
                    commands.append("show flash-usage")
                if any(item['name'] == "BFD会话检查(VC业务统计)" for item in selected_items):
                    commands.append("show bfd session brief")
                    commands.append("show bfd configuration pw")
                    commands.append("show mpls l2vc brief")
                    commands.append("show interface loopback 31")
                    commands.append("show ldp l2vc detail")
                if any(item['name'] == "配置校验状态" for item in selected_items):
                    commands.append("show cfgchk info")
                if any(item['name'] == "OSPF进程状态检查" for item in selected_items):
                    commands.append("show ospf process")
                if any(item['name'] == "IPFRR-LSP状态检查" for item in selected_items):
                    commands.append("show mpls lsp brief")
                if any(item['name'] == "OSPF邻居状态检查" for item in selected_items):
                    commands.extend(
                        ["show ospf buffers", "show ospf neighbor", "show ospf neighbor brief"])
                if any(item['name'] == "LACP成员状态监控" for item in selected_items):
                    commands.extend(["show lag", "show lacp"])
                if any(item['name'] == "OSPF 路由表检查" for item in selected_items):
                    commands.extend(["show ip routing-table"])
                if any(item['name'] == "LDP 会话状态检查" for item in selected_items):
                    commands.extend(["show ldp session"])
                if any(item['name'] == "Loopback31地址唯一性检查" for item in selected_items):
                    commands.extend(["show interface loopback 31"])
                if any(item['name'] == "Loopback1023地址唯一性检查" for item in selected_items):
                    commands.extend(["show interface loopback 1023"])
                if any(item['name'] == "SNMP配置检查" for item in selected_items):
                    commands.extend(["show snmp-server trap", "show snmp-server community",
                                    "show snmp-server mib-view", "show snmp-server sys-info"])
                if any(item['name'] == "设备账户检查" for item in selected_items):
                    commands.extend(
                        ["show users", "show login-global-rule", "show loginning-user"])
                if any(item['name'] == "专网业务分析" for item in selected_items):
                    commands.extend(["show vsi brief"])
                if any(item['name'] == "PTP时钟检查" for item in selected_items):
                    commands.extend(["show ptp all", "show synce"])
                if any(item['name'] == "站点邻接网元检查" for item in selected_items):
                    commands.extend(["show ldp lsp"])
                if any(item['name'] == "设备面板视图" for item in selected_items):
                    commands.append("show device")
                if any(item['name'] == "BFD保护组状态信息" for item in selected_items):
                    commands.append("show protect-group all")
                    commands.append("show mpls l2vc brief")
                    commands.append("show interface loopback 31")
                if any(item['name'] == "DCN路由表检查" for item in selected_items):
                    commands.extend(["show ip routing-table vpn __dcn_vpn__"])
                commands.append("show device")

                # 去除重复项
                commands = list(set(commands))

                # Debugging output
                print(
                    f"{Fore.YELLOW}[DEBUG] 用户选择巡检项: {', '.join([item['name'] for item in selected_items])}{Style.RESET_ALL}")
                print(
                    f"{Fore.YELLOW}[DEBUG] 使用以下命令进行采集: {commands}{Style.RESET_ALL}")

                # Proceed with file inputs and report generation
                raw_file = getinput("qa_raw.txt", "原始数据文件（默认：qa_raw.txt）：")
                host_file = getinput(
                    "host-stna.csv", "设备清单（默认：host-stna.csv）：")
                _progress_bar(5, "🚀 会话就绪")
                fish_multiple_cmds(host_file, raw_file, commands)

                # 添加复制文件的功能
                import shutil
                import os

                # 检查原始文件是否存在
                if os.path.exists(raw_file):
                    try:
                        # 生成备份文件名
                        backup_file = "qa_wash_raw.txt"
                        shutil.copy2(raw_file, backup_file)
                        print(
                            f"{Fore.GREEN}[INFO] 已成功复制 {raw_file} 到 {backup_file}{Style.RESET_ALL}")
                    except Exception as e:
                        print(
                            f"{Fore.RED}[ERROR] 复制文件失败: {str(e)}{Style.RESET_ALL}")
                else:
                    print(
                        f"{Fore.RED}[WARNING] 原始文件 {raw_file} 不存在，跳过复制操作{Style.RESET_ALL}")

                _progress_bar(3, "🚀 清洗就绪")
                report_file = f"QA巡检报告-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx"
                generate_qa_report(raw_file, report_file,
                                   host_file, selected_items)

        if ucmd == '19':
            print("\n📊 正在执行系统时间同步...")
            host_file = getinput("userhost-stna.csv",
                                 "设备清单（默认：userhost-stna.csv）：", timeout=10)
            raw_file = getinput("time_sync_raw.txt",
                                "原始数据文件（默认：time_sync_raw.txt）：", timeout=10)
            report_file = getinput(
                "时间同步报告.csv", "输出报告（默认：时间同步报告.csv）：", timeout=10)

            print(f"\n{Fore.CYAN}🚀 开始设置系统时间...{Style.RESET_ALL}")
            set_system_time_cmd(host_file, raw_file, report_file)
            print(
                f"\n{Fore.GREEN}✅ 时间同步报告已生成：{report_file}{Style.RESET_ALL}")

        if ucmd == '21':
            print("\n📊 正在添加互联端口描述...")
            host_file = getinput("userhost-stna.csv",
                                 "设备清单（默认：userhost-stna.csv）：", timeout=10)
            raw_file = getinput("interface_description_raw.txt",
                                "原始数据文件（默认：  interface_description_raw.txt）：", timeout=10)
            report_file = getinput(
                "接口描述配置报告.csv", "输出报告（默认：接口描述配置报告.csv）：",  timeout=10)

            print(f"\n{Fore.CYAN}🚀 开始配置接口描述...{Style.RESET_ALL}")
            add_interface_description_cmd(host_file, raw_file, report_file)
            print(f"\n{Fore.GREEN}✅ 接口描述配置报告已生成：{report_file}{Style.RESET_ALL}")
        else:
            print(f"{Fore.RED}⚠️ 无效选项，请重新输入！{Style.RESET_ALL}")
